import tkinter as tk
import winsound
from tkinter import ttk, messagebox, simpledialog
from tkinter.filedialog import asksaveasfilename
import os
from datetime import datetime, date, timedelta
import sys
import pyodbc
import bcrypt
from calendar import monthrange
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side, DEFAULT_FONT, NamedStyle
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.page import PageMargins
import win32com.client as win32
from ttkwidgets.autocomplete import AutocompleteEntry
from matplotlib.figure import Figure
import matplotlib.dates as mdates
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg



# DEFINE GLOBAL VARs FIRST TIME
error_log = "error.log"
config_ini = "config.ini"
settings_ini = "settings.ini"
test_done = "check.ini"
if "SQLSERVERNAME" not in globals(): SQLSERVERNAME = None
if "SQLSERVERNAME" not in globals(): SQLUSERNAME = None
if "SQLPASSWORD" not in globals(): SQLPASSWORD = None
if "SQLDATABASE" not in globals(): SQLDATABASE = None
if "SQLDRIVER" not in globals(): SQLDRIVER = None
if "conf" not in globals(): conf = None
if "sets" not in globals(): sets = None
if 'security' not in globals(): security = None
if 'login_var' not in globals(): login_var = None
if 'password_var' not in globals(): password_var = None
if 'Parking_Var' not in globals(): Parking_Var = 0
if 'H_Parking_Var' not in globals(): H_Parking_Var = 0
if 'Company_Var' not in globals(): Company_Var = None
if "Current_Company_obj" not in globals(): Current_Company_obj = None
if 'Truck_Var' not in globals(): Truck_Var = None
if "Current_Truck_obj" not in globals(): Current_Truck_obj = None
if 'Trailer_Var' not in globals(): Trailer_Var = None
if "Current_Trailer_obj" not in globals(): Current_Trailer_obj = None
if "Tenant_Event" not in globals(): Tenant_Event = None
if "Tenant_Comment" not in globals(): Tenant_Comment = None
if "GN_Truck_Var" not in globals(): GN_Truck_Var = None
if "Current_GN_Truck_obj" not in globals(): Current_GN_Truck_obj = None
if "GN_Trailer_Var" not in globals(): GN_Trailer_Fb_Var = None
if "Current_GN_Trailer_Fb_obj" not in globals(): Current_GN_Trailer_Fb_obj = None
if "GN_Trigger" not in globals(): GN_Trigger = 1
if "GN_Other_Carrier_Var" not in globals(): GN_Other_Carrier_Var = "GN"
if "GN_Trigger_LU" not in globals(): GN_Trigger_LU = None
if "GN_Comment_Var" not in globals(): GN_Comment_Var = None
if "GN_Event" not in globals(): GN_Event = None
if "GN_Trk_or_Fb_Var" not in globals(): GN_Trk_or_Fb_Var = None
if "Visitor_Company_Var" not in globals(): Visitor_Company_Var = None
if "Current_Visitor_Company_obj" not in globals(): Current_Visitor_Company_obj = None
if "VIS_Company_Var" not in globals(): VIS_Company_Var = None
if "VIS_Plates_Var" not in globals(): VIS_Plates_Var = None
if "VIS_Car_Var" not in globals(): VIS_Car_Var = None
if "VIS_Name_Var" not in globals(): VIS_Name_Var = None
if "VIS_Comment" not in globals(): VIS_Comment = None
if "Current_Visitor_Unit" not in globals(): Current_Visitor_Unit = None
if "Check_Menu_Var" not in globals(): Check_Menu_Var = None
if "chk_Company_Var" not in globals(): chk_Company_Var = None
if "Current_chk_Company_obj" not in globals(): Current_chk_Company_obj = None
if "Current_chk_T_T_obj" not in globals(): Current_chk_T_T_obj = None
if "chk_T_T_Var" not in globals(): chk_T_T_Var = None
if "Current_chk_T_Tr_obj" not in globals(): Current_chk_T_Tr_obj = None
if "chk_T_Tr_Var" not in globals(): chk_T_Tr_Var = None
if "Current_chk_GN_T_obj" not in globals(): Current_chk_GN_T_obj = None
if "chk_GN_T_Var" not in globals(): chk_GN_T_Var = None
if "Current_chk_GN_Tr_Fb_obj" not in globals(): Current_chk_GN_Tr_Fb_obj = None
if "chk_GN_Tr_Fb_Var" not in globals(): chk_GN_Tr_Fb_Var = None
if "GN_Menu_Var" not in globals(): GN_Menu_Var = None
if "Admin_Menu_Var" not in globals(): Admin_Menu_Var = None
if "Admin_Company_Entries" not in globals(): Admin_Company_Entries = None
if"Admin_Company_Quantity_Var" not in globals(): Admin_Company_Quantity_Var = 0
if"Admin_Company_Quantity_D_Var" not in globals(): Admin_Company_Quantity_D_Var = 0
if"Admin_Company_Quantity_R_Var" not in globals(): Admin_Company_Quantity_R_Var = 0
if"Admin_Company_Quantity_T_Var" not in globals(): Admin_Company_Quantity_T_Var = 0
if"Admin_Company_Quantity_Tr_Var" not in globals(): Admin_Company_Quantity_Tr_Var = 0
if"Admin_Company_Quantity_V_Var" not in globals(): Admin_Company_Quantity_V_Var = 0
if "Adm_Company_Var" not in globals(): Adm_Company_Var = None
if "Adm_Company_obj" not in globals(): Adm_Company_obj = None
if "Adm_Truck_Var" not in globals(): Adm_Truck_Var = None
if "Adm_Trailer_Var" not in globals(): Adm_Trailer_Var = None
if "Adm_Unit_obj" not in globals(): Adm_Unit_obj = None
if "GN_combo_filter_var" not in globals(): GN_combo_filter_var = None
if "Adm_GN_Truck_Var" not in globals(): Adm_GN_Truck_Var = None
if "Adm_GN_Truck_obj" not in globals(): Adm_GN_Truck_obj = None
if "Adm_GN_Trailer_Var" not in globals(): Adm_GN_Trailer_Var = None
if "Adm_GN_Trailer_obj" not in globals(): Adm_GN_Trailer_obj = None
if "Adm_GN_Fb_Var" not in globals(): Adm_GN_Fb_Var = None
if "Adm_GN_Fb_obj" not in globals(): Adm_GN_Fb_obj = None
if "Adm_Vis_Company_Var" not in globals(): Adm_Vis_Company_Var = None
if "Adm_Vis_Company_obj" not in globals(): Adm_Vis_Company_obj = None
if "Current_Adm_Visitor_Unit" not in globals(): Current_Adm_Visitor_Unit = None
if "Admin_Vendor_Var" not in globals(): Admin_Vendor_Var = None
if "Admin_Vendor_obj" not in globals(): Admin_Vendor_obj = None
if "Admin_Vendor_MaxID" not in globals(): Admin_Vendor_MaxID = None
if "chart_scale" not in globals(): chart_scale = None
if "db_data" not in globals(): db_data = None
if "City_Data" not in globals(): City_Data = None


#default config.ini str
config_string = """font_name=Arial
font_size=22
font_color=#FFFFFF
window_bg=#2e2e2e
window_topbar_font=Georgia Italic
window_topbar_size=26
window_topbar_fg=#ae95c1
window_topbar_sel_bg=#ba1111
notebook_bg=#2e2e2e
notebook_tab_unsel_bg=#999999
notebook_tab_sel_bg=#b6b8b9
notebook_tab_unsel_fg=#0b2232
notebook_tab_sel_fg=#0b2232
notebook_tab_font=Arial
notebook_tab_size=12
submenu_bg=#a4a4ad
submenu_sel_bg=#bec0c1
submenu_fg=#34344d
submenu_sel_fg=#34344d
submenu_font=Arial
submenu_size=26
header_bg=#314d4c
header_fg=#6edddd
header_font=Arial
header_size=20
history_font=Arial
history_size=20
widget_bg=#5a5a5a
widget_sel_bg=#9b9d9e
widget_fg=#e0e0e0
widget_sel_fg=#813b3b
widget_font=Arial
widget_size=22
entry_bg=#5a5a5a
entry_fg=yellow
entry_sel_frame=#a71360
entry_unsel_frame=#5a5a5a
entry_font=Arial
entry_size=28
in_button_bg=#9b9d9e
in_button_sel_bg=green
in_button_fg=green
in_button_sel_fg=white
out_button_bg=#9b9d9e
out_button_sel_bg=red
out_button_fg=red
out_button_sel_fg=white
func_button_bg=#9b9d9e
func_button_sel_bg=#5a5a5a
func_button_fg=blue
func_button_sel_fg=#813b3b
func_button_font=Arial
func_button_size=28
status_fg=#EB5559
status_font=Arial
status_size=28
p_button_w=10
p_t_company_w=400
p_t_truck_h=200
p_t_manual_h=70
p_t_last_event_h=40
p_g_truck_h=200
p_g_trailer_h=240
p_g_fb_h=200
p_g_feature_w=500
p_g_bottom_h=190
p_g_feature_h=80
expired_date=red
on_parking=yellow
UNREG_bg=#2d5ba6
storage_fg=#95C267
designated_period=30
chk_filter_frame=200
admin_company_button_height=40
chart_tenant=blue
chart_gn=red
chart_ax=#9b9d9e
chart_title=white"""
#default settings.ini str
settings_string = """SQL_Server_Name=DEV-BOX\SQLEXPRESS
SQL_DB_Name=PARKING
SQL_Driver={ODBC Driver 18 for SQL Server}
SQL_Login=Parking_Hawk
SQL_Password=SQl12345
Overparking_Timeout=2
Year_List=2023|2024|2025|2026|2027|2028|2029|2030|2031|2032|2033|2034|2035
Month_List=January|February|March|April|May|June|July|August|September|October|November|December
SQL_path=Z:\\Max\\Overparking\\
chk_path=Z:\\Max\\CheckYard\\
archive_path=Z:\\Max\\Archive\\
QR_path=Z:\\Max\\QR\\
chk_datetime=None"""
#Default database structure
DataBaseStructure = {
"statistics": [("date", "date"), ("tenant_amount", "smallint"), ("gn_amount", "smallint")],
"visitors_history": [("company_ID", "int"), ("plates", "varchar(50)"), ("car_model", "varchar(50)"), ("driver_name", "varchar(50)"), ("datetime_event", "smalldatetime"), ("comment", "varchar(50)"), ("status", "bit"), ("full_name", "varchar(50)")],
"Tenant_Trucks_UNREG": [("company_ID", "smallint"), ("truck_number", "varchar(50)"), ("status", "bit"), ("last_date", "smalldatetime")],
"Tenant_Trailers_UNREG": [("company_ID", "smallint"), ("trailer_number", "varchar(50)"), ("storage", "bit"), ("status", "bit"), ("last_date", "smalldatetime")],
"visitors_UNREG": [("company_ID", "smallint"), ("plates", "varchar(50)"), ("driver_name", "varchar(50)"), ("car_model", "varchar(50)"), ("car_color", "varchar(50)"), ("expiration", "date"), ("private", "bit"), ("status", "bit"), ("last_date", "smalldatetime")],
"OVERPARKING": [("date", "date"), ("company_ID", "smallint"), ("over_count", "smallint"), ("trucks_onyard", "varchar(MAX)"), ("trailers_onyard", "varchar(MAX)"), ("vehicles_onyard", "varchar(50)"), ("last_time", "time(7)"), ("over_time", "varchar(MAX)"), ("last_over_count", "smallint"), ("last_trucks_onyard", "varchar(MAX)"), ("last_trailers_onyard", "varchar(MAX)")],
"GN_Trucks": [("truck_number", "varchar(50)"), ("status", "bit"), ("last_date", "smalldatetime"), ("city", "bit")],
"Car_Vendors": [("ID", "smallint"), ("Vendor", "varchar(50)")],
"GN_Flatbed": [("fb_number", "varchar(50)"), ("storage", "bit"), ("status", "bit"), ("last_date", "smalldatetime"), ("LU", "bit")],
"GN_Trailers": [("trailer_number", "varchar(50)"), ("storage", "bit"), ("status", "bit"), ("last_date", "smalldatetime"), ("LU", "bit")],
"check_yard": [("date", "smalldatetime"), ("company", "varchar(50)"), ("type", "varchar(50)"), ("unit_number", "varchar(50)"), ("status", "bit")],
"GN_History": [("company_name", "varchar(50)"), ("truck_number", "varchar(50)"), ("trailer_number", "varchar(50)"), ("fb_number", "varchar(50)"), ("datetime_event", "smalldatetime"), ("cargo", "bit"), ("status", "bit"), ("comment", "varchar(50)"), ("full_name", "varchar(50)")],
"Tenant_History": [("company_ID", "smallint"), ("truck_number", "varchar(50)"), ("trailer_number", "varchar(50)"), ("datetime_event", "smalldatetime"), ("status", "bit"), ("comment", "varchar(50)"), ("full_name", "varchar(50)")],
"Tenant_Trailers": [("company_ID", "smallint"), ("trailer_number", "varchar(50)"), ("storage", "bit"), ("status", "bit"), ("last_date", "smalldatetime")],
"Tenant_Trucks": [("company_ID", "smallint"), ("truck_number", "varchar(50)"), ("status", "bit"), ("last_date", "smalldatetime")],
"visitors": [("company_ID", "smallint"), ("plates", "varchar(50)"), ("driver_name", "varchar(50)"), ("car_model", "varchar(50)"), ("car_color", "varchar(50)"), ("expiration", "date"), ("private", "bit"), ("status", "bit"), ("last_date", "smalldatetime")],
"Company_List": [("company_ID", "smallint"), ("company_name", "varchar(50)"), ("designated", "smallint"), ("regular", "smallint"), ("trailer", "smallint"), ("truck", "smallint"), ("car", "smallint"), ("activity", "bit"), ("insurance", "date")],
"authentication": [("ID", "smallint identity"), ("login", "varchar(50)"), ("password", "varchar(MAX)"), ("full_name", "varchar(50)"), ("rights", "smallint"), ("activity", "bit")]
}

# DEBUGIN FUNCTION - Record Error in txt file
def debuger(val):
    try:
        with open(error_log, "r"):
            try:
                with open(error_log, "a") as file: file.write(str(datetime.now())+" "+str(val) + "\n")
            except Exception as e:
                error(f"An error occured: {e}")
    except FileNotFoundError:
        with open(error_log, "w") as file:
            try:
                file.write(str(val) + "\n")
            except:
                error(19)
                sys.exit()

# ERROR WINDOW POP-UP - UNDER CONSTRUCTION
def error(err_code):
    def message(winname, info): tk.messagebox.showerror(winname, info)
    if err_code == 1: message("Configuration", "Config file not found.\nDefault configuration was created.\nRestart Program.")
    elif err_code == 16: message("Configuration", "Settings file not found.\nDefault settings was created.\nRestart Program.")
    elif err_code == 2: message("Login", "Wrong Login or Password.")
    elif err_code == 3: message("Login","Login already exist.")
    elif err_code == 4: message("Login","Login do not exist.")
    elif err_code == 5: message("Input", "Entered value is not match!")
    elif err_code == 6: message("Input", "Company activity not defined")
    elif err_code == 7: message("Input", "Incorrect date. Must be YYYY-MM-DD")
    elif err_code == 10: message("Information", "Check Yard not generated.")
    elif err_code == 11: message("Information", "Unit not found.")
    elif err_code == 12: message("Input", "Unit already exist.")
    elif err_code == 13: message("Error", "Error was occurred. Check Error Log File.")
    elif err_code == 14: message("Input", "Entered Name is NOT IN database.")
    elif err_code == 15: message("Error", "Critical Impossible error!")
    elif err_code == 17: message("SQL", "Cannot connect to SQL Server...")
    elif err_code == 18: message("SQL", "SQL database not exist.\nNew database will be created.")
    elif err_code == 19: message("Error", "Cannot create a file.")
    elif err_code == 20: message("SQL", "Cannot create table/column in SQL.")
    elif err_code == 21: message("System", "Check will be initiated.\nIt may take some time.")
    elif err_code == 22: message("CLASS", "Error occurred in SQL_REQ class.")
    elif err_code == 23: message("OVERLIMIT", "Requesting list is too big\nLimit is 500 lines.\nTry specify range in filter.")
    else: message("Error", err_code)

# Function that test connection to SQL, check DB, tables, types of data, ini files and re-create if something is missing
def TESTER():
    global conf
    global sets
    global config_ini
    global SQLSERVERNAME
    global SQLUSERNAME
    global SQLPASSWORD
    global SQLDATABASE
    global SQLDRIVER

    #Function requesting login and password to connect SQL with admin right to create DB or tables.
    def admin_connect():
        while True:
            username = simpledialog.askstring("SQL Admin", "Username:")
            if username is not None and username.isprintable():
                password = simpledialog.askstring("SQL Admin", "Password:", show='*')
                if password is not None and password.isprintable(): return username, password
            choize = messagebox.askyesno("Data Mismatch", f"Empty field not allowed\nTry again?")
            if choize: admin_connect()
            else:
                connection.close()
                sys.exit()

    #CONFIG.INI
    # Testing config.ini and recreate if missing
    try:
        with open(config_ini, "r") as file: conf = {k: v for k, v in (s.split("=") for s in file.read().splitlines() if s!="")}
        with open(config_ini, "w") as file: file.write("\n".join([f"{key}={value}" for key, value in conf.items() if key and value]))
    except FileNotFoundError:
        error(1)
        with open(config_ini, "w") as file:
            try:
                file.write(config_string)
            except:
                error(19)
        sys.exit()
    except Exception as e:
        error(15)
        debuger(e)
        sys.exit()
    dfl_conf = {k:v for k, v in (s.split("=") for s in config_string.splitlines())}
    if len(conf)!=len(dfl_conf):
        for key in dfl_conf:
            if key not in conf:
                choize = messagebox.askyesno("Data Mismatch", f"Required configuration for {key} not found in {config_ini}.\nDo you want add default settings for this value?")
                if choize:
                    conf[key] = dfl_conf[key]
                    try:
                        with open(config_ini, "a") as file: file.write("\n"+key+"="+dfl_conf[key])
                    except Exception as e:
                        error(19)
                        debuger(e)
                        sys.exit()
                else:
                    error(15)
                    sys.exit()

    #SETTINGS.INI
    # Testing settings.ini and recreate if missing
    try:
        with open(settings_ini, "r") as file: sets = {k: v for k, v in (s.split("=") for s in file.read().splitlines() if s != "")}
        with open(settings_ini, "w") as file: file.write("\n".join([f"{key}={value}" for key, value in sets.items() if key and value]))
    except FileNotFoundError:
        error(16)
        with open(settings_ini, "w") as file:
            try:
                file.write(settings_string)
            except:
                error(19)
        sys.exit()
    except Exception as e:
        error(15)
        debuger(e)
        sys.exit()
    dfl_sets = {k: v for k, v in (s.split("=") for s in settings_string.splitlines())}
    if len(sets) != len(dfl_sets):
        for key in dfl_sets:
            if key not in sets:
                choize = messagebox.askyesno("Data Mismatch", f"Required configuration for {key} not found in {settings_ini}.\nDo you want add default settings for this value?")
                if choize:
                    sets[key] = dfl_sets[key]
                    try:
                        with open(settings_ini, "a") as file:
                            file.seek(0)
                            if file.tell() != 0 and file.read()[-1] != "\n":
                                file.write("\n")
                            file.write(key+"=" + dfl_sets[key])
                    except Exception as e:
                        error(19)
                        debuger(e)
                        sys.exit()
                else:
                    error(15)
                    sys.exit()

    # Return if Test was successful before
    try:
        with open(test_done, "r") as file:
            test_ok = file.read().split()
            if test_ok == "1": return
    except FileNotFoundError: error(21)

    # Defining vars for SQL connection
    SQLSERVERNAME = sets["SQL_Server_Name"]
    SQLUSERNAME = sets["SQL_Login"]
    SQLPASSWORD = sets["SQL_Password"]
    SQLDATABASE = sets["SQL_DB_Name"]
    SQLDRIVER = sets["SQL_Driver"]

    # Test SQL connection
    try:
        connection = pyodbc.connect(f"DRIVER={SQLDRIVER};Server={SQLSERVERNAME};UID={SQLUSERNAME};TrustedServerCertificate=1;Encrypt=No;PWD={SQLPASSWORD}")
        connection.autocommit = True
    except pyodbc.Error as e:
        error(17)
        debuger(e)
        connection.close()
        sys.exit()
    # Test if DB excists and creating new with tables if not
    try:
        cursor = connection.cursor()
        cursor.execute(f"SELECT name FROM sys.databases WHERE name = '{SQLDATABASE}'")
        result = cursor.fetchone()
        if not result:
            choize = messagebox.askyesno("Database Error", f"Database wasn't found or could not be connected.\nWould you like to create new?\n(SQL Admin Login Required)")
            if choize:
                while True:
                    login = admin_connect()
                    connection.close()
                    try:
                        admconnection = pyodbc.connect(f"DRIVER={SQLDRIVER};Server={SQLSERVERNAME};UID={login[0]};TrustedServerCertificate=1;Encrypt=No;PWD={login[1]}")
                        admconnection.autocommit=True
                    except Exception as e:
                        debuger(e)
                        choize = messagebox.askyesno("Connection Error", f"Cannot connect to Database\nCheck your Login and Password.\nTry again?")
                        if choize: continue
                        else:
                            admconnection.close()
                            sys.exit()
                    try:
                        admcursor = admconnection.cursor()
                        admcursor.execute(f"CREATE DATABASE {SQLDATABASE}")
                    except Exception as e:
                        debuger(e)
                        admconnection.close()
                        choize = messagebox.askyesno("Database Creation Error", f"Database could not be created.\nCheck your permission in SQL Server")
                        if choize: continue
                        else:
                            admconnection.close()
                            sys.exit()
                    try:
                        admcursor.execute(f"ALTER AUTHORIZATION ON DATABASE::{SQLDATABASE} TO {SQLUSERNAME}")
                    except pyodbc.Error as e:
                        error(20)
                        debuger(e)
                        admcursor.close()
                        admconnection.close()
                    try:
                        admcursor.execute(f"USE {SQLDATABASE}")
                        for table_name, columns in DataBaseStructure.items():
                            create_table_sql = f"CREATE TABLE [{table_name}] ("
                            for column, datatype in columns:
                                create_table_sql += f"[{column}] {datatype}, "
                            create_table_sql = create_table_sql.rstrip(", ")
                            create_table_sql += ");"
                            admcursor.execute(create_table_sql)
                        # Addin Default Admin Account for Parking Hawk
                        user = "admin"
                        password = "admin"
                        salt = bcrypt.gensalt()
                        encr_pass = bcrypt.hashpw(password.encode("utf-8"), salt)
                        name = "Default Admin"
                        rights = 1
                        activity = 1
                        try:
                            admcursor.execute(f"INSERT INTO dbo.authentication(login, password, full_name, rights, activity) values (?,?,?,?,?)", (user, encr_pass, name, rights, activity))
                        except pyodbc.Error as e:
                            error(13)
                            debuger(e)
                            admcursor.close()
                            admconnection.close()
                        break
                    except pyodbc.Error as e:
                        debuger(e)
                        admcursor.close()
                        admconnection.close()
            else:
                connection.close()
                sys.exit()
    except pyodbc.Error as e:
        error(15)
        debuger(e)
        cursor.close()
        connection.close()
        sys.exit()
    if "admconnection" in locals() and admconnection is not None:
        connection = admconnection
        cursor = admcursor

    # Test if Table exists and its type, request to create if not.
    cursor.execute(f"USE {SQLDATABASE}")
    for table_name, columns in DataBaseStructure.items():
        table_exist = cursor.execute(f"SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME='{table_name}'")
        table_exist = cursor.fetchone()
        if table_exist:
            table_columns = cursor.execute(f"EXEC sp_columns @table_name = N'{table_name}', @table_owner = 'dbo'")
            table_columns = cursor.fetchall()
            datatype_mapping = {"varchar": "varchar(50)", "text": "varchar(MAX)", "time": "time(7)"}
            existing_columns = {col.COLUMN_NAME: datatype_mapping.get(col.TYPE_NAME, col.TYPE_NAME) for col in table_columns}
            for column_name, data_type in columns:
                # column_name=str("dbo."+column_name)
                if column_name not in existing_columns:
                    #try add table
                    try:
                        #######
                        choize = messagebox.askyesno("Database Error", f"Required Table wasn't found in Database.\nWould you like to create new?\n(SQL Admin Login Required)")
                        if choize:
                            while True:
                                login = admin_connect()
                                try:
                                    admconnection = pyodbc.connect(f"DRIVER={SQLDRIVER};Server={SQLSERVERNAME};UID={login[0]};TrustedServerCertificate=1;Encrypt=No;PWD={login[1]}")
                                    admconnection.autocommit = True
                                except Exception as e:
                                    debuger(e)
                                    choize = messagebox.askyesno("Connection Error", f"Cannot connect to Database\nCheck your Login and Password.\nTry again?")
                                    if choize:
                                        continue
                                    else:
                                        admconnection.close()
                                        sys.exit()
                                try:
                                    admcursor = admconnection.cursor()
                                    admcursor.execute(f"ALTER TABLE [dbo.{table_name}] ADD {column_name} {data_type}")
                                except Exception as e:
                                    debuger(e)
                                    admconnection.close()
                                    choize = messagebox.askyesno("Table Creation Error", f"Table could not be created.\nCheck your permission in SQL Server")
                                    if choize:
                                        continue
                                    else:
                                        admconnection.close()
                                        sys.exit()
                    except Exception as e:
                        print(f"Error HERE!\n{e}")
                        error(15)
                        sys.exit()
                elif existing_columns[column_name] != data_type:
                    choize = messagebox.askyesno("Data Type Mismatch", f"Data type for column '{column_name}' in table '{table_name}' does not match.\nExpected data: {data_type}\nCurrent data:{existing_columns[column_name]}\nDo you want to replace it? Data might be lost.")
                    if choize:
                        #try to change column type
                        try:
                            cursor.execute(f"ALTER TABLE [dbo.{table_name}] ALTER COLUMN {column_name} {data_type}")
                        except:
                            error(15)
                            connection.close()
                            sys.exit()
                    else:
                        error(15)
                        connection.close()
                        sys.exit()
        else:
            cursor.execute(f"USE {SQLDATABASE}")
            create_gone_table_sql = f"CREATE TABLE [{table_name}] ("
            for column, datatype in columns: create_gone_table_sql += f"[{column}] {datatype}, "
            create_gone_table_sql = create_gone_table_sql.rstrip(", ")
            create_gone_table_sql += ");"
            cursor.execute(create_gone_table_sql)
    #Closing all possible connection with SQL
    if "admcursor" in locals() and admcursor is not None: admcursor.close()
    if "admconnection" in locals() and admconnection is not None: admconnection.close()
    if "cursor" in locals() and cursor is not None: cursor.close()
    if "connection" in locals() and connection is not None: connection.close()
    try:
        with open(test_done, "w") as file:
            file.write("1")
    except Exception as e:
        error(19)
        debuger(e)
        sys.close()

TESTER()

#SQL Requests Function
def SQL_REQ(query, vars, mode):
    def SQL_CLOSE():
        cursor.close()
        connection.close()
    try:
        connection = pyodbc.connect(f"DRIVER={SQLDRIVER};Server={SQLSERVERNAME};Database={SQLDATABASE};UID={SQLUSERNAME};TrustedServerCertificate=1;Encrypt=No;PWD={SQLPASSWORD}")
        connection.autocommit=True
        cursor = connection.cursor()
    except Exception as e:
        error(17)
        debuger(str(e)+" # ", query+" # ", str(vars)+" # ", str(mode))
        if os.path.exist(test_done): os.remove(test_done)
        SQL_CLOSE()
    try:
        if mode == "S_one":
            cursor.execute(query, vars)
            data = cursor.fetchone()
            SQL_CLOSE()
            return data
        elif mode == "S_all":
            cursor.execute(query, vars)
            data = cursor.fetchall()
            SQL_CLOSE()
            return data
        elif mode == "S_one_D":
            cursor.execute(query, vars)
            data = [cursor.fetchone(), cursor.description]
            SQL_CLOSE()
            return data
        elif mode == "S_all_D":
            cursor.execute(query, vars)
            data = [cursor.fetchall(), cursor.description]
            SQL_CLOSE()
            return data
        elif mode == "W":
            cursor.execute(query, vars)
        elif mode == "W_many":
            cursor.executemany(query, vars)
        # elif mode == "I_many":
        #     cursor.executemany(query, vars)
        # elif mode == "I_D":
        #     cursor.execute(query[0], query[1])
        # elif mode == "U_D":
        #     cursor.execute(query[0], query[1])
        # elif mode == "D":
        #     cursor.execute(query)
        SQL_CLOSE()
    except Exception as e:
        error(17)
        debuger(str(e)+" # "+query+" # "+str(vars)+" # "+str(mode))
        if os.path.exists(test_done): os.remove(test_done)
        SQL_CLOSE()


#READ SETTINGS FILE
def settings_file():
    path_ini = ("settings.ini")
    if os.path.exists(path_ini):
        sets_ini = open(path_ini)
        return_sets = {k: v for k, v in (s.split("=") for s in sets_ini.read().splitlines())}
        return return_sets
    else:
        error(2)
def settings_file_edit(key, value):
    settings = settings_file()
    settings[key] = value
    path_ini = "settings.ini"
    with open(path_ini, 'w') as sets_ini:
        for key, value in settings.items():
            sets_ini.write(f"{key}={value}\n")

# IN/OUT Sound Function
def beep(bool):
    if bool:
        winsound.Beep(440, 150)
        winsound.Beep(550, 150)
    else:
        winsound.Beep(550, 150)
        winsound.Beep(440, 150)
#Function for asking dialog to save file. Take 2 args File Path - str and extantion - str, return file pathin in forma C:/Folder/File.ext
def save_file_as(file_type, file_ext):
    filepath = asksaveasfilename(defaultextension=f".{file_ext}", filetypes=[(file_type, file_ext), ("All Files", "*.*")])
    if not filepath:
        return None  # Return None if the user cancels the operation
    return filepath

#LOGIN AND PASSWORD CLASS
class PasswordDatabase:
    def __init__(self):
        self.data = dict()
        row = SQL_REQ("SELECT login, password FROM dbo.authentication WHERE activity=1", (), "S_all")
        self.data = {k: v for k, v in row}
    def register(self, user, password, name, rights, activity):
        if user in self.data: return error(3)
        pwd_hash = self.hash_password(password)
        self.data[user] = pwd_hash
        val = (user, pwd_hash, name, rights, activity)
        SQL_REQ("INSERT INTO dbo.authentication(login, password, full_name, rights, activity) values (?,?,?,?,?)", val, "W")
        return True
    def hash_password(self, password):
        pwd_bytes = password.encode("utf-8")
        salt = bcrypt.gensalt()
        return bcrypt.hashpw(pwd_bytes, salt)
    def login(self, user, password):
        if user not in self.data: return False
        pwd_bytes = password.encode("utf-8")
        return bcrypt.checkpw(pwd_bytes, self.data[user].encode("utf-8"))
    def content(self):
        return self.data

class scroller(tk.Frame):
    def __init__(self, parent, **kwargs):
        try:
            super().__init__(parent, **kwargs)
            self.frame_deploy()
        except Exception as e:
            error(22)
            debuger(e)
    def frame_deploy(self):
        self.mainframe = tk.Frame(self, highlightthickness=0, bg=conf["window_bg"])
        self.mainframe.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.NW)
        self.canvas = tk.Canvas(self.mainframe, highlightthickness=0, bg=conf["window_bg"])
        self.frame = tk.Frame(self.canvas, bg=conf["window_bg"])
        self.scrollbar = ttk.Scrollbar(self.mainframe, orient=tk.VERTICAL, command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(fill=tk.BOTH, expand=0, side=tk.LEFT)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y, anchor=tk.NE)
        self.canvas.create_window((0, 0), window=self.frame, anchor=tk.NW)
        self.frame.bind("<Configure>", lambda event, canvas=self.canvas: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", self.update_scroll_region)
        self.frame.bind("<Configure>", self.update_scroll_region)
        self.frame.bind("<Enter>", self.enter_mousewheel, add="+")
        self.frame.bind("<Leave>", self.leave_mousewheel)
    def update_scroll_region(self, *event):
        self.canvas.update_idletasks()
        self.frame.update_idletasks()
        frame_width = self.frame.winfo_width()
        frame_height = self.frame.winfo_height()
        self.canvas.configure(scrollregion=(0,0, frame_width, frame_height))
        self.canvas.configure(width=frame_width, height=frame_height)
        if self.frame.winfo_height() <= self.canvas.winfo_height():
            self.scrollbar.pack_forget()
            self.canvas.configure(yscrollcommand=None)
            self.canvas.unbind("<Enter>")
            self.canvas.unbind_all("<MouseWheel>")
        else:
            self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            self.canvas.configure(yscrollcommand=self.scrollbar.set)
            self.canvas.bind("<Enter>", self.enter_mousewheel, add="+")
        self.canvas.configure(scrollregion=(0, 0, self.frame.winfo_width(), self.frame.winfo_height()))
    def top(self):self.canvas.yview_moveto(0.0)
    def refresh(self):
        try:
            self.canvas.update_idletasks()
            self.frame.update_idletasks()
            self.update_scroll_region()
            self.canvas.yview_moveto(0.0)
        except Exception as e:
            error(22)
            debuger(e)
    def delete(self):
        try:
            if self.frame.winfo_children():
                for widgets in self.frame.winfo_children(): widgets.destroy()
        except Exception as e:
            error(22)
            debuger(e)
    def on_mousewheel(self, event): self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    def enter_mousewheel(self, event): self.canvas.bind_all('<MouseWheel>', self.on_mousewheel, add="+")
    def leave_mousewheel(self, event): self.canvas.unbind_all('<MouseWheel>')
    # def grid_config(self, weight_vals):
    #     for i, weight in enumerate(weight_vals):
    #         self.frame.grid_columnconfigure(i, weight=weight)
    #         print(f"{self}.frame.grid_columnconfigure({i}, weight={weight})")


class filter_frame(tk.Frame):
    checkyard_marker = sets["chk_datetime"]
    checkyard_instances = []
    def __init__(self, parent, **kwargs):
        try:
            super().__init__(parent, **kwargs)
            self.frame_deploy()
        except Exception as e:
            error(22)
            debuger(e)
    # Deploying main frame for filter
    def frame_deploy(self):
        self.mainframe = tk.Frame(self, highlightthickness=3, relief=tk.RAISED, bg=conf["submenu_bg"], highlightbackground=conf["submenu_sel_bg"], width=conf["chk_filter_frame"])
        self.mainframe.pack_propagate(False)
        self.mainframe.pack(side=tk.LEFT, fill=tk.BOTH)
    ### Filters
    # Check Yard Filter (company-combobox, truck-checkbox, trailer-checkbox, storage-checkbox, time-checkbox+entry, generate-button, print-button)
    def checkyard(self, **kwargs):
        # adding instance in list of checkyard instances in class for bulk modification
        self.checkyard_instances.append(self)
        # initializing function responsible for activation
        self.company_func = kwargs.get("company_func")
        self.time_on_yard_func = kwargs.get("time_on_yard_func")
        self.check_generate = kwargs.get("check_generate")
        self.check_print = kwargs.get("check_print")
        self.truck_func = kwargs.get("truck_func")
        self.trailer_func = kwargs.get("trailer_func")
        self.storage_func = kwargs.get("storage_func")
        self.age_func = kwargs.get("age_func")
        # deployment of filters
        self.company()
        self.truck_box()
        self.trailer_box()
        self.storage_box()
        self.time_on_yard()
        self.generate_buttons()
    #Check Yard Filter for GN
    def checkyard_GN(self, **kwargs):
        # adding instance in list of checkyard instances in class for bulk modification
        self.checkyard_instances.append(self)
        # Top label
        self.GN_lb = tk.Label(self.mainframe, text="GNT:", foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["header_size"]), bg=conf["submenu_bg"])
        self.GN_lb.pack(side=tk.TOP, padx=5, pady=(10, 5), anchor=tk.W)
        # initializing function responsible for activation
        self.time_on_yard_func = kwargs.get("time_on_yard_func")
        self.truck_func = kwargs.get("truck_func")
        self.trailer_func = kwargs.get("trailer_func")
        self.fb_func = kwargs.get("fb_func")
        self.storage_func = kwargs.get("storage_func")
        self.cargo_func = kwargs.get("cargo_func")
        self.age_func = kwargs.get("age_func")
        self.check_generate = kwargs.get("check_generate")
        self.check_print = kwargs.get("check_print")
        # deployment of filters
        self.truck_box()
        self.trailer_box()
        self.fb_box()
        self.storage_box()
        self.cargo()
        self.time_on_yard()
        self.generate_buttons()
    # Check Filter for Vis
    def checkyard_Vis(self, **kwargs):
        # adding instance in list of checkyard instances in class for bulk modification
        self.checkyard_instances.append(self)
        # initializing function responsible for activation
        self.company_func = kwargs.get("company_func")
        self.time_on_yard_func = kwargs.get("time_on_yard_func")
        self.corp_func = kwargs.get("corp_func")
        self.private_func = kwargs.get("private_func")
        self.no_parking_func = kwargs.get("no_parking_func")
        self.expired_func = kwargs.get("expired_func")
        self.age_func = kwargs.get("age_func")
        self.check_generate = kwargs.get("check_generate")
        self.check_print = kwargs.get("check_print")
        # deployment of filters
        self.company()
        self.car_parking_box()
        self.time_on_yard()
        self.generate_buttons()

    # Tenant Statistics Filter
    def tenant_stat(self, **kwargs):
        # initializing function responsible for activation
        self.company_func = kwargs.get("company_func")
        self.time_on_yard_func = kwargs.get("time_on_yard_func")
        # self.check_generate = kwargs.get("check_generate")
        # self.check_print = kwargs.get("check_print")
        self.truck_func = kwargs.get("truck_func")
        self.trailer_func = kwargs.get("trailer_func")
        self.storage_func = kwargs.get("storage_func")
        self.age_func = kwargs.get("age_func")
        # deployment of filters
        self.company()
        self.truck_box()
        self.trailer_box()
        self.storage_box()
        self.time_on_yard()
        # buttons
    # Tenant History Filter
    def tenant_history(self, **kwargs):
        self.company_func = kwargs.get("comapny_func")
        self.truck_search_func = kwargs.get("truck_func")
        self.trailer_func = kwargs.get("trailer_func")
        self.scale_func = kwargs.get("scale_func")
        self.period_func = kwargs.get("period_func")
        self.truck_search_var = tk.StringVar()
        self.trailer_search_var = tk.StringVar()
        self.search_truck_checkbox = tk.BooleanVar()
        self.search_trailer_checkbox = tk.BooleanVar()
        self.company()
        self.date_period()
        self.date_scale()
        self.unit_search(label="Truck Search:", var=self.truck_search_var, checkbox=self.search_truck_checkbox, func=self.truck_search_func)
        self.unit_search(label="Trailer Search:", var=self.trailer_search_var, checkbox=self.search_trailer_checkbox, func=self.truck_search_func)
        self.history_buttons()

    def GN_histor(self, **kwargs):
        self.truck_search_func = kwargs.get("truck_func")
        self.trailer_func = kwargs.get("trailer_func")
        self.fb_func = kwargs.get("fb_func")
        self.scale_func = kwargs.get("scale_func")
        self.period_func = kwargs.get("period_func")
        self.truck_search_var = tk.StringVar()
        self.trailer_search_var = tk.StringVar()
        self.fb_search_var = tk.StringVar()
        self.search_truck_checkbox = tk.BooleanVar()
        self.search_trailer_checkbox = tk.BooleanVar()
        self.search_fb_checkbox = tk.BooleanVar()
        self.date_period()
        self.date_scale()
        self.unit_search(label="Truck Search:", var=self.truck_search_var, checkbox=self.search_truck_checkbox, func=self.truck_search_func)
        self.unit_search(label="Trailer Search:", var=self.trailer_search_var, checkbox=self.search_trailer_checkbox, func=self.truck_search_func)
        self.unit_search(label="Flatbed Search:", var=self.fb_search_var, checkbox=self.search_fb_checkbox, func=self.truck_search_func)
        self.history_buttons()

    def GN_city(self, **kwargs):
        self.truck_search_func = kwargs.get("truck_func")
        self.search_truck_checkbox = tk.BooleanVar()
        self.scale_func = kwargs.get("scale_func")
        self.period_func = kwargs.get("period_func")
        self.export_city_func = kwargs.get("export_city_func")
        self.truck_search_var = tk.StringVar()
        self.search_truck_checkbox = tk.BooleanVar()
        self.save_func = kwargs.get("save_func")
        self.date_period()
        self.date_scale()
        self.unit_search(label="Truck Search:", var=self.truck_search_var, checkbox=self.search_truck_checkbox, func=self.truck_search_func)
        self.save_button()

    def vis_stat(self, **kwargs):
        # adding instance in list of instances in class for bulk modification
        self.checkyard_instances.append(self)
        # initializing function responsible for activation
        self.company_func = kwargs.get("company_func")
        self.time_on_yard_func = kwargs.get("time_on_yard_func")
        self.corp_func = kwargs.get("corp_func")
        self.private_func = kwargs.get("private_func")
        self.no_parking_func = kwargs.get("no_parking_func")
        self.expired_func = kwargs.get("expired_func")
        self.age_func = kwargs.get("age_func")
        # deployment of filters
        self.company()
        self.car_parking_box()
        self.time_on_yard()


    def vis_history(self,**kwargs):
        self.company_func = kwargs.get("company_func")
        self.plate_search_func = kwargs.get("plate_func") #plate
        self.scale_func = kwargs.get("scale_func")
        self.period_func = kwargs.get("period_func")
        self.plate_search_var = tk.StringVar()
        self.search_plate_checkbox = tk.BooleanVar()
        self.company()
        self.date_period()
        self.date_scale()
        self.unit_search(label="Plates:", var=self.plate_search_var, checkbox=self.search_plate_checkbox, func=self.plate_search_func)
        self.history_buttons()

    def company(self):
        self.comp_list = ["All"] + units_lst("company")
        self.comp_lb = tk.Label(self.mainframe, text="Companies:", foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["header_size"]), bg=conf["submenu_bg"])
        self.comp_lb.pack(side=tk.TOP, padx=5, pady=(10, 5), anchor=tk.W)
        self.comp_box = ttk.Combobox(self.mainframe, values=self.comp_list, width=10, background=conf["submenu_sel_bg"], foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), state="readonly")
        self.comp_box.pack(fill=tk.X, side=tk.TOP, padx=5, pady=(0, 10))
        self.comp_box.current(0)
        self.comp_box.bind("<<ComboboxSelected>>", self.company_func)

    def truck_box(self):
        self.var_truck = tk.BooleanVar()
        self.truck_checkbutton = tk.Checkbutton(self.mainframe, text="Trucks", variable=self.var_truck, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], onvalue=True, offvalue=False)
        self.truck_checkbutton.pack(side=tk.TOP, padx=5, pady=(0, 10), anchor=tk.W)
        self.truck_checkbutton.select()
        self.var_truck.trace("w", self.truck_func)
    def trailer_box(self):
        self.var_trailer = tk.BooleanVar()
        self.trailer_checkbutton = tk.Checkbutton(self.mainframe, text="Trailers", variable=self.var_trailer, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], onvalue=True, offvalue=False)
        self.trailer_checkbutton.pack(side=tk.TOP, padx=5, pady=(0, 10), anchor=tk.W)
        self.trailer_checkbutton.select()
        self.var_trailer.trace("w", self.trailer_func)
    def fb_box(self):
        self.var_fb = tk.BooleanVar()
        self.fb_checkbutton = tk.Checkbutton(self.mainframe, text="Flatbed", variable=self.var_fb, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], onvalue=True, offvalue=False)
        self.fb_checkbutton.pack(side=tk.TOP, padx=5, pady=(0, 10), anchor=tk.W)
        self.fb_checkbutton.select()
        self.var_fb.trace("w", self.fb_func)
    def storage_box(self):
        self.var_storage = tk.BooleanVar()
        self.storage_checkbutton = tk.Checkbutton(self.mainframe, text="Storage", variable=self.var_storage, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], state=tk.NORMAL, onvalue=True, offvalue=False)
        self.storage_checkbutton.pack(side=tk.TOP, padx=5, pady=(10), anchor=tk.W)
        self.storage_checkbutton.select()
        self.var_storage.trace("w", self.storage_func)
    # Filter checkbuttons for Car Parking
    def car_parking_box(self):
        #corporate parking
        self.var_corp = tk.BooleanVar()
        self.corp_checkbutton = tk.Checkbutton(self.mainframe, text="Corporate", variable=self.var_corp, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], state=tk.NORMAL, onvalue=True, offvalue=False)
        self.corp_checkbutton.pack(side=tk.TOP, padx=5, pady=(10), anchor=tk.W)
        self.corp_checkbutton.select()
        # private parking
        self.var_private = tk.BooleanVar()
        self.private_checkbutton = tk.Checkbutton(self.mainframe, text="Private", variable=self.var_private, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], state=tk.NORMAL, onvalue=True, offvalue=False)
        self.private_checkbutton.pack(side=tk.TOP, padx=5, pady=(10), anchor=tk.W)
        self.private_checkbutton.select()
        # expired parking
        self.var_expired = tk.BooleanVar()
        self.expired_checkbutton = tk.Checkbutton(self.mainframe, text="Expired", variable=self.var_expired, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], state=tk.NORMAL, onvalue=True, offvalue=False)
        self.expired_checkbutton.pack(side=tk.TOP, padx=5, pady=(10), anchor=tk.W)
        # no parking
        self.var_no_parking = tk.BooleanVar()
        self.no_parking_checkbutton = tk.Checkbutton(self.mainframe, text="No Parking", variable=self.var_no_parking, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], state=tk.NORMAL, onvalue=True, offvalue=False)
        self.no_parking_checkbutton.pack(side=tk.TOP, padx=5, pady=(10), anchor=tk.W)
        self.no_parking_ID = self.var_no_parking.trace("w", self.no_parking_func)
        self.car_trace()
    def car_trace(self):
        self.corp_ID = self.var_corp.trace("w", self.corp_func)
        self.private_ID = self.var_private.trace("w", self.private_func)
        self.expired_ID = self.var_expired.trace("w", self.expired_func)
        self.no_parking_ID = self.var_no_parking.trace("w", self.no_parking_func)
    def car_untrace(self):
        self.var_corp.trace_remove("write", self.corp_ID)
        self.var_private.trace_remove("write", self.private_ID)
        self.var_expired.trace_remove("write", self.expired_ID)
        self.var_no_parking.trace_remove("write", self.no_parking_ID)

    def save_button(self):
        self.save_button_frame = tk.Frame(self.mainframe, highlightthickness=0, bg=conf["submenu_bg"], width=50)
        self.save_button_frame.pack(side=tk.BOTTOM, anchor=tk.W, fill=tk.X)
        self.save_bt = tk.Button(self.mainframe, text="SAVE", height=1, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["font_size"]), relief=tk.RAISED, command=self.save_func)
        self.save_bt.pack(side=tk.TOP, padx=(5, 0), fill=tk.X, expand=1)

    # Filter for limiting time on yard in days. With checkbox of activation and button to apply. Accept apply Funciton.
    def time_on_yard(self):
        self.aging_lb = tk.Label(self.mainframe, text="Time on Yard (days):", foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), bg=conf["submenu_bg"])
        self.aging_lb.pack(side=tk.TOP, padx=5, pady=(10, 5), anchor=tk.W)
        self.age_frame = tk.Frame(self.mainframe, highlightthickness=0, bg=conf["submenu_bg"])
        self.age_frame.pack(side=tk.TOP, anchor=tk.W)
        self.var_age = tk.BooleanVar()
        self.age_checkbutton = tk.Checkbutton(self.age_frame, variable=self.var_age, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], onvalue=True, offvalue=False)
        self.age_checkbutton.pack(fill=tk.X, side=tk.LEFT)
        self.var_age.trace("w", self.age_func)
        self.aging = None
        self.age_entry = tk.Entry(self.age_frame, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["entry_fg"], width=15, state=tk.DISABLED)
        self.age_entry.pack(side=tk.LEFT, fill=tk.BOTH)
        self.age_button_get = tk.Button(self.age_frame, text=u"\u23F5", bg=conf["submenu_sel_bg"], relief=tk.RAISED, command=self.time_on_yard_func)
        self.age_button_get.pack(side=tk.RIGHT, padx=(5, 0))
    # Check Yard Label for marker and Generate button
    def generate_buttons(self):
        self.gen_frame = tk.Frame(self.mainframe, highlightthickness=0, bg=conf["submenu_bg"], width=50)
        self.gen_frame.pack(side=tk.BOTTOM, anchor=tk.W)
        self.generate_lb = tk.Label(self.gen_frame, text="Check Yard Marker:", foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), bg=conf["submenu_bg"])
        self.generate_lb.pack(side=tk.TOP, padx=5, pady=(10, 5), anchor=tk.W)
        self.generate_marker = tk.Label(self.gen_frame, foreground=conf["status_fg"], text=filter_frame.checkyard_marker, font=(conf["submenu_font"], conf["notebook_tab_size"]), bg=conf["submenu_bg"])
        self.generate_marker.pack(side=tk.TOP, padx=5, pady=(10, 5), anchor=tk.CENTER)
        self.gen_button = tk.Button(self.gen_frame, text="GENERATE", height=1, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["font_size"]), relief=tk.RAISED, command=self.check_generate)
        self.gen_button.pack(side=tk.TOP, padx=(5, 0), fill=tk.X, expand=1)
        self.print_button = tk.Button(self.gen_frame, text="PRINT", height=1, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["font_size"]), relief=tk.RAISED, command=self.check_print)
        self.print_button.pack(side=tk.TOP, padx=(5, 0), fill=tk.X, expand=1)
    # Cargo Function for GN checkyard
    def cargo(self):
        self.cargo_lb = tk.Label(self.mainframe, text="Cargo:", foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), bg=conf["submenu_bg"])
        self.cargo_lb.pack(side=tk.TOP, padx=5, pady=(10, 5), anchor=tk.W)
        self.var_loaded = tk.BooleanVar()
        self.var_empty = tk.BooleanVar()
        self.loaded_checkbutton = tk.Checkbutton(self.mainframe, text="Loaded", variable=self.var_loaded, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], state=tk.NORMAL, onvalue=True, offvalue=False)
        self.loaded_checkbutton.pack(side=tk.TOP, padx=5, pady=(10), anchor=tk.W)
        self.loaded_checkbutton.select()
        self.var_loaded.trace("w", self.cargo_func)
        self.empty_checkbutton = tk.Checkbutton(self.mainframe, text="Empty", variable=self.var_empty, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], state=tk.NORMAL, onvalue=True, offvalue=False)
        self.empty_checkbutton.pack(side=tk.TOP, padx=5, pady=(10), anchor=tk.W)
        self.empty_checkbutton.select()
        self.var_empty.trace("w", self.cargo_func)
    # Date period filter ()
    def date_period(self, **kwargs):
        def Y_CHANGE(bul):
            if bul: self.year_label.config(text=str(int(self.year_label.cget("text")) + 1))
            elif not bul: self.year_label.config(text=str(int(self.year_label.cget("text"))-1))
            self.daylim = monthrange(int(self.year_label.cget("text")), int(self.month_label.cget("text")))
            if self.daylim[1] < int(self.day_label.cget("text")): self.day_label.config(text=str(self.daylim[1]))
            self.period_func()
        def M_CHANGE(bul):
            if bul:
                if int(self.month_label.cget("text")) == 12: return
                self.month_label.config(text=str(int(self.month_label.cget("text"))+1))
            elif not bul:
                if int(self.month_label.cget("text")) == 1: return
                self.month_label.config(text=str(int(self.month_label.cget("text"))-1))
            self.daylim = monthrange(int(self.year_label.cget("text")), int(self.month_label.cget("text")))
            if self.daylim[1] < int(self.day_label.cget("text")): self.day_label.config(text=str(self.daylim[1]))
            self.period_func()
        def D_CHANGE(bul):
            if bul:
                self.daylim = monthrange(int(self.year_label.cget("text")), int(self.month_label.cget("text")))
                if int(self.day_label.cget("text")) >= self.daylim[1]: return
                self.day_label.config(text=str(int(self.day_label.cget("text"))+1))
            elif not bul:
                if int(self.day_label.cget("text")) == 1: return
                self.day_label.config(text=str(int(self.day_label.cget("text"))-1))
            self.period_func()



        #setting up current date as default to display
        self.year = str(datetime.now().year)
        self.month = str(datetime.now().month)
        self.day = str(datetime.now().day)
        self.daylim = monthrange(int(self.year), int(self.month))
        #main frame for Period
        self.period_frame = tk.Frame(self.mainframe, highlightthickness=0, bg=conf["submenu_bg"])
        self.period_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 20))
        #Period Label
        self.period_lb = tk.Label(self.period_frame, text="Period:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
        self.period_lb.pack(fill=tk.X, side=tk.TOP, expand=1, pady=1)
        # year buttons
        self.year_frame = tk.Frame(self.period_frame, highlightthickness=0, bg=conf["window_bg"])
        self.year_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 1))
        self.year_L_button = tk.Button(self.year_frame, text="\N{LEFTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: Y_CHANGE(False))
        self.year_L_button.pack(fill=tk.X, side=tk.LEFT)
        self.year_label = tk.Label(self.year_frame, text=self.year, relief=tk.FLAT, bg=conf["widget_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
        self.year_label.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
        self.year_R_button = tk.Button(self.year_frame, text="\N{RIGHTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: Y_CHANGE(True))
        self.year_R_button.pack(fill=tk.X, side=tk.LEFT)
        # month buttons
        self.month_frame = tk.Frame(self.period_frame, highlightthickness=0, bg=conf["window_bg"])
        self.month_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 1))
        self.month_L_button = tk.Button(self.month_frame, text="\N{LEFTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: M_CHANGE(False))
        self.month_L_button.pack(fill=tk.X, side=tk.LEFT)
        self.month_label = tk.Label(self.month_frame, text=self.month, relief=tk.FLAT, bg=conf["widget_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
        self.month_label.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
        self.month_R_button = tk.Button(self.month_frame, text="\N{RIGHTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: M_CHANGE(True))
        self.month_R_button.pack(fill=tk.X, side=tk.LEFT)
        # day buttons
        self.day_frame = tk.Frame(self.period_frame, highlightthickness=0, bg=conf["window_bg"])
        self.day_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 1))
        self.day_L_button = tk.Button(self.day_frame, text="\N{LEFTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: D_CHANGE(False))
        self.day_L_button.pack(fill=tk.X, side=tk.LEFT)
        self.day_label = tk.Label(self.day_frame, text=self.day, relief=tk.FLAT, bg=conf["widget_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
        self.day_label.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
        self.day_R_button = tk.Button(self.day_frame, text="\N{RIGHTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: D_CHANGE(True))
        self.day_R_button.pack(fill=tk.X, side=tk.LEFT)
    def date_scale(self, **kwarg):
        self.chart_scale = "D"
        def Chart_Scale(func):
            self.scale_year_button.config(bg=conf["widget_bg"], fg=conf["on_parking"])
            self.scale_month_button.config(bg=conf["widget_bg"], fg=conf["on_parking"])
            self.scale_day_button.config(bg=conf["widget_bg"], fg=conf["on_parking"])
            if func == "Y":
                self.chart_scale = "Y"
                self.scale_year_button.config(bg=conf["widget_sel_bg"], fg=conf["expired_date"])
                self.scale_func() #################
            elif func == "M":
                self.chart_scale = "M"
                self.scale_month_button.config(bg=conf["widget_sel_bg"], fg=conf["expired_date"])
                self.scale_func() #################
            if func == "D":
                self.chart_scale = "D"
                self.scale_day_button.config(bg=conf["widget_sel_bg"], fg=conf["expired_date"])
                self.scale_func() #################

        # Filter by lentgh buttons
        self.scale_frame = tk.Frame(self.period_frame, highlightthickness=0, bg=conf["submenu_bg"])
        self.scale_frame.pack(side=tk.TOP, fill=tk.X)
        self.scale_lb = tk.Label(self.scale_frame, text="Scale:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
        self.scale_lb.pack(fill=tk.X, side=tk.TOP, expand=1, pady=1)
        self.scale_year_button = tk.Button(self.scale_frame, text="Year", bg=conf["widget_bg"], fg=conf["on_parking"], bd=0, highlightthickness=0, activebackground=conf["on_parking"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: Chart_Scale("Y"))
        self.scale_year_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
        self.scale_month_button = tk.Button(self.scale_frame, text="Month", bg=conf["widget_bg"], fg=conf["on_parking"], bd=0, highlightthickness=0, activebackground=conf["on_parking"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: Chart_Scale("M"))
        self.scale_month_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
        self.scale_day_button = tk.Button(self.scale_frame, text="Day", bg=conf["widget_bg"], fg=conf["on_parking"], bd=0, highlightthickness=0, activebackground=conf["on_parking"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: Chart_Scale("D"))
        self.scale_day_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
    def history_buttons(self):
        global security
        self.history_frame = tk.Frame(self.mainframe, highlightthickness=0, bg=conf["submenu_bg"], width=50)
        self.history_frame.pack(side=tk.BOTTOM, anchor=tk.W, fill=tk.X)
        self.arc_button = tk.Button(self.history_frame, text="ARCHIVE", height=1, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["font_size"]), relief=tk.RAISED, command=lambda:history_archive())
        if security[1]==1: self.arc_button.pack(side=tk.TOP, padx=(5, 0), fill=tk.X, expand=1)
        self.print_button = tk.Button(self.history_frame, text="PRINT", height=1, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["font_size"]), relief=tk.RAISED, command=lambda:history_print())
        self.print_button.pack(side=tk.TOP, padx=(5, 0), fill=tk.X, expand=1)
    #create entry for search. Require Lablel, tk variable for entry and function name for button
    def unit_search(self, **kwargs):
        label = kwargs.get("label")
        var = kwargs.get("var")
        checkbox = kwargs.get("checkbox")
        func = kwargs.get("func")
        def checkbox_trigger(*args, entry, button):
            if checkbox.get():
                entry.config(state=tk.NORMAL)
                button.config(state=tk.NORMAL)
            else:
                entry.config(state=tk.DISABLED)
                button.config(state=tk.DISABLED)
        self.search_lb = tk.Label(self.mainframe, text=label, foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), bg=conf["submenu_bg"])
        self.search_lb.pack(side=tk.TOP, padx=5, pady=(10, 5), anchor=tk.W)
        self.search_frame = tk.Frame(self.mainframe, highlightthickness=0, bg=conf["submenu_bg"])
        self.search_frame.pack(side=tk.TOP, anchor=tk.W)
        self.search_checkbox = tk.Checkbutton(self.search_frame, variable=checkbox, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], onvalue=True, offvalue=False)
        self.search_checkbox.pack(fill=tk.X, side=tk.LEFT)
        self.unit_search_entry = tk.Entry(self.search_frame, textvariable=var, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["entry_fg"], width=15, state=tk.DISABLED)
        self.unit_search_entry.pack(side=tk.LEFT, fill=tk.BOTH)
        self.unit_search_button = tk.Button(self.search_frame, text=u"\u23F5", bg=conf["submenu_sel_bg"], relief=tk.RAISED, command=func, state=tk.DISABLED)
        self.unit_search_button.pack(side=tk.RIGHT, padx=(5, 0))
        checkbox.trace("w", lambda *args, entry=self.unit_search_entry, button=self.unit_search_button: checkbox_trigger(entry=entry, button=button))




    # method delete everything from mainframe of class
    def delete(self):
        try:
            if self.mainframe.winfo_children():
                for widgets in self.mainframe.winfo_children(): widgets.destroy()
        except Exception as e:
            error(22)
            debuger(e)
    # method that accept new marker for check yard and reconfig markers labels
    def label_config(new_marker):
        filter_frame.checkyard_marker = new_marker
        for instance in filter_frame.checkyard_instances:
            instance.generate_marker.config(text=new_marker)

#Function retriving ID by name and otherway arround
def ID_NAME_company(name=None, ID=None):
    if name is not None: return SQL_REQ("SELECT company_ID FROM dbo.Company_List WHERE company_name=?", (name,), "S_one")[0]
    if ID is not None: return SQL_REQ("SELECT company_name FROM dbo.Company_List WHERE company_ID=?", (ID,), "S_one")[0]


# Check Insert Functions working with Class
def checkyard_insert(classobj, frame):
    C = classobj.comp_box.get()
    T = classobj.var_truck.get()
    t = classobj.var_trailer.get()
    S = classobj.var_storage.get()
    A = classobj.var_age.get()

    today = date.today()
    # tenant_his_scroll_frame.delete()

    if t:
        classobj.storage_checkbutton.config(state=tk.NORMAL)
    elif not classobj.var_trailer.get():
        classobj.storage_checkbutton.config(state=tk.DISABLED)
    if A:
        classobj.age_entry.configure(state=tk.NORMAL)
        a = classobj.age_entry.get().strip()
        if a != "":
            try:
                int(a)
            except:
                error(5)
                return
            aging = int(a)
        else:
            aging = 0
    else:
        classobj.age_entry.configure(state=tk.DISABLED)
        aging = 0
    if not T and not t: return
    data = get_onyard()["tenant"]

    frame.delete()

    if C == "All":
        l = list()
        for all_comp in data[0]: l.append(all_comp["company_ID"])
        for all_comp in data[1]: l.append(all_comp["company_ID"])
        allcompset = set(l)
        for all_comp in sorted(allcompset):
            c_frame = tk.Frame(frame.frame, highlightthickness=0, bg=conf["widget_sel_bg"])
            c_lb = tk.Label(c_frame, text=all_comp, bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=chk_data_size+1)
            c_lb.pack(side=tk.TOP, anchor=tk.NW, expand=0)
            checkT = False
            checkt = False
            if T:
                truck_label = tk.Label(c_frame, text="Trucks:", bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
                truck_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
                column_names_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
                column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
                unit_lb = tk.Label(column_names_fr, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
                unit_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                date_lb = tk.Label(column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"])
                date_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
                sum_lb = tk.Label(column_names_fr, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
                sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                checkT = False
                for all in data[0]:
                    delta_days = (today - all["last_date"].date()).days
                    if all["company_ID"] == all_comp and delta_days - aging >= 0:
                        checkT = True
                        rec_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
                        rec_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
                        T_lb = tk.Label(rec_fr, text=all["truck_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                        T_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                        T_time_lb = tk.Label(rec_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
                        T_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
                        T_sum_lb = tk.Label(rec_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                        T_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                if checkT:
                    truck_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
                    truck_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))

            if t:
                trailer_label = tk.Label(c_frame, text="Trailers:", bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
                trailer_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
                column_names_fr2 = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
                column_names_fr2.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
                unitT_lb = tk.Label(column_names_fr2, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
                unitT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                dateT_lb = tk.Label(column_names_fr2, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"])
                dateT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
                sumT_lb = tk.Label(column_names_fr2, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
                sumT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                checkt = False
                for all in data[1]:
                    delta_days = (today - all["last_date"].date()).days
                    if all["company_ID"] == all_comp and delta_days - aging >= 0:
                        if not S:
                            if all["storage"]: continue
                        checkt = True
                        recT_fr = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
                        recT_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
                        Tt_lb = tk.Label(recT_fr, text=all["trailer_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                        Tt_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                        Tt_time_lb = tk.Label(recT_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
                        Tt_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
                        delta_days = (today - all["last_date"].date()).days
                        Tt_sum_lb = tk.Label(recT_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                        Tt_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                        if all["storage"]:
                            Tt_lb.config(fg=conf["storage_fg"])
                            Tt_time_lb.config(fg=conf["storage_fg"])
                            Tt_sum_lb.config(fg=conf["storage_fg"])
                if checkt:
                    trailer_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
                    trailer_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
            if checkT or checkt:
                c_frame.pack(side=tk.TOP, fill=tk.X, expand=1, pady=(0, 2), padx=(2, 0))
    else:
        c1_frame = tk.Frame(frame.frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        c1_lb = tk.Label(c1_frame, text=C, bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=chk_data_size)
        c1_lb.pack(side=tk.TOP, anchor=tk.NW, expand=0)
        checkT = False
        checkt = False
        if T:
            C_truck_label = tk.Label(c1_frame, text="Trucks:", bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
            C_truck_frame = tk.Frame(c1_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
            C_column_names_fr = tk.Frame(C_truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
            C_column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
            C_unit_lb = tk.Label(C_column_names_fr, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
            C_unit_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
            C_date_lb = tk.Label(C_column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"])
            C_date_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
            C_sum_lb = tk.Label(C_column_names_fr, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
            C_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
            checkT = False
            for all in data[0]:
                if all["company_ID"] != C: continue
                delta_days = (today - all["last_date"].date()).days
                if delta_days - aging >= 0:
                    checkT = True
                    C_rec_fr = tk.Frame(C_truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
                    C_rec_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
                    C_T_lb = tk.Label(C_rec_fr, text=all["truck_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                    C_T_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                    C_T_time_lb = tk.Label(C_rec_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
                    C_T_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
                    C_T_sum_lb = tk.Label(C_rec_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                    C_T_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
            if checkT:
                C_truck_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
                C_truck_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
        if t:
            C_trailer_label = tk.Label(c1_frame, text="Trailers:", bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
            C_trailer_frame = tk.Frame(c1_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
            C_column_names_fr2 = tk.Frame(C_trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
            C_column_names_fr2.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
            C_unitT_lb = tk.Label(C_column_names_fr2, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
            C_unitT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
            C_dateT_lb = tk.Label(C_column_names_fr2, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"])
            C_dateT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
            C_sumT_lb = tk.Label(C_column_names_fr2, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
            C_sumT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
            checkt = False
            for all in data[1]:
                if all["company_ID"] != C: continue
                delta_days = (today - all["last_date"].date()).days
                if delta_days - aging >= 0:
                    if not S:
                        if all["storage"]: continue
                    checkt = True
                    C_recT_fr = tk.Frame(C_trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
                    C_recT_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
                    C_Tt_lb = tk.Label(C_recT_fr, text=all["trailer_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                    C_Tt_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                    C_Tt_time_lb = tk.Label(C_recT_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
                    C_Tt_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
                    delta_days = (today - all["last_date"].date()).days
                    C_Tt_sum_lb = tk.Label(C_recT_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                    C_Tt_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                    if all["storage"]:
                        C_Tt_lb.config(fg=conf["storage_fg"])
                        C_Tt_time_lb.config(fg=conf["storage_fg"])
                        C_Tt_sum_lb.config(fg=conf["storage_fg"])
            if checkt:
                C_trailer_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
                C_trailer_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
        if checkT or checkt:
            c1_frame.pack(side=tk.TOP, fill=tk.X, expand=1, pady=(0, 2), padx=(2, 0))

    frame.refresh()
    frame.top()
def checkyard_GN_insert(classobj, frame):



    ###############
    T = classobj.var_truck.get()
    t = classobj.var_trailer.get()
    S = classobj.var_storage.get()
    f = classobj.var_fb.get()
    L = classobj.var_loaded.get()
    U = classobj.var_empty.get()
    A = classobj.var_age.get()
    today = date.today()
    frame.delete()
    if not T and not t and not f: return
    elif not L and not U: return
    if t:
        classobj.storage_checkbutton.config(state=tk.NORMAL)
    elif not classobj.var_trailer.get():
        classobj.storage_checkbutton.config(state=tk.DISABLED)
    if A:
        classobj.age_entry.configure(state=tk.NORMAL)
        a = classobj.age_entry.get().strip()
        if a != "":
            try:
                int(a)
            except:
                error(5)
                return
            aging = int(a)
        else:
            aging = 0
    else:
        classobj.age_entry.configure(state=tk.DISABLED)
        aging = 0
    data = get_onyard()["GN"]
    #Clean scrollable frame
    frame.delete()

    c_frame = tk.Frame(frame.frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    c_frame.pack(side=tk.TOP, fill=tk.X, expand=1, pady=(0, 2), padx=(2, 0))
    if T:
        truck_label = tk.Label(c_frame, text="Trucks:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=chk_data_size)
        truck_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
        truck_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        truck_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
        column_names_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
        unit_lb = tk.Label(column_names_fr, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
        unit_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        cargo_lb = tk.Label(column_names_fr, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
        cargo_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        date_lb = tk.Label(column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"])
        date_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
        sum_lb = tk.Label(column_names_fr, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=30)
        sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        for all in data[0]:
            delta_days = (today - all["last_date"].date()).days
            if delta_days - aging >= 0:
                rec_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
                rec_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
                T_lb = tk.Label(rec_fr, text=all["truck_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                T_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                C_lb = tk.Label(rec_fr, text="", bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                C_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                T_time_lb = tk.Label(rec_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
                T_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
                T_sum_lb = tk.Label(rec_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=30)
                T_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))

    if t:
        trailer_label = tk.Label(c_frame, text="Trailers:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=chk_data_size)
        trailer_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
        trailer_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        trailer_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
        column_names_fr2 = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        column_names_fr2.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
        unitT_lb = tk.Label(column_names_fr2, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
        unitT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        cargo_t_lb = tk.Label(column_names_fr2, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
        cargo_t_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        dateT_lb = tk.Label(column_names_fr2, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"])
        dateT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
        sumT_lb = tk.Label(column_names_fr2, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=30)
        sumT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        for all in data[1]:
            delta_days = (today - all["last_date"].date()).days
            if delta_days - aging >= 0 and "trailer_number" in all:
                if not S:
                    if all["storage"]: continue
                if not L:
                    if all["LU"]: continue
                if not U:
                    if not all["LU"]: continue
                recT_fr = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
                recT_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
                Tt_lb = tk.Label(recT_fr, text=all["trailer_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                Tt_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                if all["LU"]:
                    cargo = "LOADED"
                else:
                    cargo = "EMPTY"
                tC_lb = tk.Label(recT_fr, text=cargo, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                tC_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                if all["LU"]:
                    tC_lb.config(fg=conf["func_button_fg"])
                else:
                    tC_lb.config(fg=conf["func_button_sel_fg"])
                Tt_time_lb = tk.Label(recT_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
                Tt_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
                delta_days = (today - all["last_date"].date()).days
                Tt_sum_lb = tk.Label(recT_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=30)
                Tt_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                if all["storage"]:
                    Tt_lb.config(fg=conf["storage_fg"])
                    Tt_time_lb.config(fg=conf["storage_fg"])
                    Tt_sum_lb.config(fg=conf["storage_fg"])
    if f:
        fb_label = tk.Label(c_frame, text="Flatbeds:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=chk_data_size)
        fb_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
        fb_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        fb_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
        column_names_fr3 = tk.Frame(fb_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        column_names_fr3.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
        unitf_lb = tk.Label(column_names_fr3, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
        unitf_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        cargo_f_lb = tk.Label(column_names_fr3, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
        cargo_f_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        datef_lb = tk.Label(column_names_fr3, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"])
        datef_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
        sumf_lb = tk.Label(column_names_fr3, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=30)
        sumf_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        for all in data[1]:
            delta_days = (today - all["last_date"].date()).days
            if delta_days - aging >= 0 and "fb_number" in all:
                if not S:
                    if all["storage"]: continue
                if not L:
                    if all["LU"]: continue
                if not U:
                    if not all["LU"]: continue
                recf_fr = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
                recf_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
                f_lb = tk.Label(recf_fr, text=all["fb_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                f_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                if all["LU"]:
                    cargo = "LOADED"
                else:
                    cargo = "EMPTY"
                fC_lb = tk.Label(recf_fr, text=cargo, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
                fC_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
                if all["LU"]:
                    fC_lb.config(fg=conf["func_button_fg"])
                else:
                    fC_lb.config(fg=conf["func_button_sel_fg"])
                f_time_lb = tk.Label(recf_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
                f_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
                delta_days = (today - all["last_date"].date()).days
                f_sum_lb = tk.Label(recf_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=30)
                f_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    frame.refresh()

def checkyard_vis_insert(classobj, frame):
    C = classobj.comp_box.get()
    c = classobj.var_corp.get()
    p = classobj.var_private.get()
    e = classobj.var_expired.get()
    NO = classobj.var_no_parking.get()
    A = classobj.var_age.get()
    today = datetime.now()

    def insert_cars(masta, data, color):
        if data["private"] is None:
            parking = ""
            expdate = ""
        else:
            parking = "YES"
            if data["private"]:
                expdate = data["expiration"]
            else:
                expdate = ""
        deltahours = int(delta_days.total_seconds() // 3600)
        if deltahours > 24:
            deltatime = f"{deltahours // 24} days {deltahours % 24} hours"
        else:
            deltatime = f"{deltahours} hours"
        rec_fr = tk.Frame(masta, highlightthickness=0, bg=conf["widget_sel_bg"])
        rec_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
        pl_lb = tk.Label(rec_fr, text=all["plates"], bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        pl_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        vend_lb = tk.Label(rec_fr, text=all["car_model"], bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        vend_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        dr_lb = tk.Label(rec_fr, text=all["driver_name"], bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        dr_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        car_time_lb = tk.Label(rec_fr, text=all["last_date"], bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color)
        car_time_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1), fill=tk.X, expand=1)
        park_lb = tk.Label(rec_fr, text=parking, bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        park_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        pr_lb = tk.Label(rec_fr, text=expdate, bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        pr_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        car_sum_lb = tk.Label(rec_fr, text=deltatime, bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        car_sum_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
    frame.delete()
    classobj.car_untrace()
    if e:
        classobj.private_checkbutton.select()
        p = True
        classobj.private_checkbutton.config(state=tk.DISABLED)
        classobj.corp_checkbutton.deselect()
        c = False
        classobj.corp_checkbutton.config(state=tk.DISABLED)
        classobj.no_parking_checkbutton.config(state=tk.DISABLED)
    else:
        state = classobj.no_parking_checkbutton["state"]
        if state == "disabled":
            classobj.no_parking_checkbutton.config(state=tk.NORMAL)
            classobj.private_checkbutton.config(state=tk.NORMAL)
            classobj.private_checkbutton.deselect()
            p = False
            classobj.corp_checkbutton.config(state=tk.NORMAL)
    if NO:
        classobj.private_checkbutton.deselect()
        p = False
        classobj.private_checkbutton.config(state=tk.DISABLED)
        classobj.corp_checkbutton.deselect()
        c = False
        classobj.corp_checkbutton.config(state=tk.DISABLED)
        classobj.expired_checkbutton.configure(state=tk.DISABLED)
    else:
        state2 = classobj.expired_checkbutton["state"]
        if state2 == "disabled":
            classobj.private_checkbutton.config(state=tk.NORMAL)
            classobj.corp_checkbutton.config(state=tk.NORMAL)
            classobj.expired_checkbutton.configure(state=tk.NORMAL)
    if A:
        classobj.age_entry.configure(state=tk.NORMAL)
        a = classobj.age_entry.get().strip()
        if a != "":
            try:
                int(a)
            except:
                error(5)
                return
            aging = int(a)
        else:
            aging = 0
    else:
        classobj.age_entry.configure(state=tk.NORMAL)
        classobj.age_entry.delete(0, tk.END)
        classobj.age_entry.configure(state=tk.DISABLED)
        aging = 0
    classobj.car_trace()

    data = get_onyard()["visitor"]
    if C == "All":
        l = list()
        for all_comp in data: l.append(all_comp["company_ID"])
        allcompset = set(l)
        for all_comp in sorted(allcompset):
            c_vis_frame = tk.Frame(frame.frame, highlightthickness=0, bg=conf["widget_sel_bg"])
            c_vis_lb = tk.Label(c_vis_frame, text=all_comp, bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=chk_data_size)
            c_vis_lb.pack(side=tk.TOP, anchor=tk.NW, expand=0)
            car_frame = tk.Frame(c_vis_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
            car_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
            column_names_fr = tk.Frame(car_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
            column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
            plate_lb = tk.Label(column_names_fr, text="plates:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            plate_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
            vendor_lb = tk.Label(column_names_fr, text="vendor:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            vendor_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
            driver_lb = tk.Label(column_names_fr, text="driver:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            driver_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
            date_lb = tk.Label(column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"])
            date_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1), fill=tk.X, expand=1)
            parking_lb = tk.Label(column_names_fr, text="on park:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            parking_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
            private_lb = tk.Label(column_names_fr, text="private:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            private_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
            sum_lb = tk.Label(column_names_fr, text="on yard / time", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            sum_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))

            checkE = False



            for all in data:
                delta_days = today - all["last_date"]
                if all["company_ID"] == all_comp and int(delta_days.total_seconds() // 86400) - aging >= 0:
                    if c and all["private"] is not None and not all["private"]:
                        checkE = True
                        insert_cars(car_frame, all, conf["submenu_fg"])
                    elif p and all["private"]:
                        if int((today.date() - all["expiration"]).days) > 0:
                            colorfg = conf["expired_date"]
                        else:
                            colorfg = conf["on_parking"]
                        if not e:
                            checkE = True
                            insert_cars(car_frame, all, colorfg)
                        else:
                            if int((today.date() - all["expiration"]).days) > 0:
                                checkE = True
                                insert_cars(car_frame, all, colorfg)
                    elif NO and all["private"] is None:
                        checkE = True
                        insert_cars(car_frame, all, conf["in_button_sel_fg"])
            if checkE: c_vis_frame.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
    else:
        if not any(d.get("company_ID") == C for d in data): return
        c_vis_frame = tk.Frame(frame.frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        c_vis_lb = tk.Label(c_vis_frame, text=C, bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=chk_data_size)
        c_vis_lb.pack(side=tk.TOP, anchor=tk.NW, expand=0)
        car_frame = tk.Frame(c_vis_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        car_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
        column_names_fr = tk.Frame(car_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
        plate_lb = tk.Label(column_names_fr, text="plates:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        plate_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        vendor_lb = tk.Label(column_names_fr, text="vendor:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        vendor_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        driver_lb = tk.Label(column_names_fr, text="driver:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        driver_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        date_lb = tk.Label(column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"])
        date_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1), fill=tk.X, expand=1)
        parking_lb = tk.Label(column_names_fr, text="parking:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        parking_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        private_lb = tk.Label(column_names_fr, text="private:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        private_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        sum_lb = tk.Label(column_names_fr, text="on yard / time", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        sum_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))

        checkE = False

        for all in data:
            delta_days = today - all["last_date"]
            if all["company_ID"] == C and int(delta_days.total_seconds() // 3600) - aging >= 0:
                if c and all["private"] is not None and not all["private"]:
                    checkE = True
                    insert_cars(car_frame, all, conf["submenu_fg"])
                elif p and all["private"]:
                    if int((today.date() - all["expiration"]).days) > 0:
                        colorfg = conf["expired_date"]
                    else:
                        colorfg = conf["on_parking"]
                    if not e:
                        checkE = True
                        insert_cars(car_frame, all, colorfg)
                    else:
                        if int((today.date() - all["expiration"]).days) > 0:
                            checkE = True
                            insert_cars(car_frame, all, colorfg)
                elif NO and all["private"] is None:
                    checkE = True
                    insert_cars(car_frame, all, conf["in_button_sel_fg"])
        if checkE: c_vis_frame.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
    frame.refresh()
    frame.top()




#Checkyard function generator. Clear previous check yard from SQL and create new. Generate EXCEL files and put Date marker in settings.ini
def checkyard_generate(classobj):
    SQL_REQ("DELETE FROM dbo.check_yard", (), "W")
    date = datetime.now().replace(microsecond=0)
    filter_frame.label_config(date)
    settings_file_edit("chk_datetime", date)
    complist = units_lst("company")

    # Tenant check

    wb = Workbook()
    ws = wb.active
    ws.title = "Tenant"
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5, header=0.3)
    ws.oddHeader.center.text = "Check Yard"
    ws.oddHeader.right.text = "&P of &N"
    ws.oddHeader.left.text = "&D"
    # ws.print_title_rows = "1:1"
    MaxX = 12
    # MaxY = 66
    X = 1
    Y = 1
    Z = 0
    for companyname in complist:
        truck = units_lst(companyname, "trucks")
        truck.update(units_lst(companyname, "trucks+"))
        trailer = {k: v[0] if isinstance(v, list) else v for k, v in units_lst(companyname, "trailers").items()}
        trailer.update(units_lst(companyname, "trailers+"))
        if not bool(truck) and not bool(trailer): continue
        ws.cell(row=Y, column=X).value = companyname
        ws.cell(row=Y, column=X).font = Font(name="Bahnschrift SemiBold SemiConden", size=12, b=True)
        for col in range(1, MaxX + 1):
            ws.cell(row=Y, column=col).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
            if col == 1:
                ws.cell(row=Y, column=col).border = Border(top=Side(border_style='medium'), left=Side(border_style='medium'))
            elif col == MaxX:
                ws.cell(row=Y, column=col).border = Border(top=Side(border_style='medium'), right=Side(border_style='medium'))
            else:
                ws.cell(row=Y, column=col).border = Border(top=Side(border_style='medium'))
        Y += 1
        if bool(truck):
            ws.cell(row=Y, column=X).value = "Trucks:"
            ws.cell(row=Y, column=X).font = Font(name="Bahnschrift SemiBold SemiConden", size=10, b=True)
            ws.cell(row=Y, column=X).border = Border(top=Side(border_style='medium'), left=Side(border_style='medium'), right=Side(border_style='thin'))

            Z = Y
            X += 1

            for trucknumber in sorted(truck.items(), key=lambda x: len(x[0])):
                unit_type = "truck"
                SQL_REQ("INSERT INTO dbo.check_yard (date, company, type, unit_number, status) VALUES (?,?,?,?,?)", (date, companyname, unit_type, trucknumber[0], trucknumber[1]), "W")
                ws.cell(row=Y, column=X).value = trucknumber[0]
                ws.cell(row=Y, column=X).font = Font(name="Bahnschrift SemiBold SemiConden", size=8, b=True)
                ws.cell(row=Y, column=X).alignment = Alignment(horizontal='center')
                if X == 2 and Y != Z: ws.cell(row=Y, column=X - 1).border = Border(left=Side(border_style='medium'))
                ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='thin'))
                if X == MaxX:
                    ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='medium'))
                    Y += 1
                    X = 2
                else:
                    X += 1
            if X != 2:
                while X != MaxX + 1:
                    ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='thin'))
                    if X == MaxX: ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'),
                                                                           right=Side(border_style='medium'))
                    X += 1
            else:
                Y -= 1
            for col in range(2, MaxX + 1):
                ws.cell(row=Z, column=col).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
                if col == MaxX: ws.cell(row=Z, column=col).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='thin'), right=Side(border_style='medium'),
                                                                           left=Side(border_style='thin'))

            X = 1
            Y += 1
        if bool(trailer):
            ws.cell(row=Y, column=X).value = "Trailers:"
            ws.cell(row=Y, column=X).font = Font(name="Bahnschrift SemiBold SemiConden", size=10, b=True)
            ws.cell(row=Y, column=X).border = Border(top=Side(border_style='medium'), left=Side(border_style='medium'), right=Side(border_style='thin'))
            Z = Y
            X += 1
            for trailernumber in sorted(trailer.items(), key=lambda x: len(x[0])):

                unit_type = "trailer"
                SQL_REQ("INSERT INTO dbo.check_yard (date, company, type, unit_number, status) VALUES (?,?,?,?,?)", (date, companyname, unit_type, trailernumber[0], trailernumber[1]), "W")
                if X == 2 and Y != Z: ws.cell(row=Y, column=X - 1).border = Border(left=Side(border_style='medium'))
                ws.cell(row=Y, column=X).value = trailernumber[0]
                ws.cell(row=Y, column=X).font = Font(name="Bahnschrift SemiBold SemiConden", size=8, b=True)
                ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='thin'))
                ws.cell(row=Y, column=X).alignment = Alignment(horizontal='center')
                if X == MaxX:
                    ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='medium'))
                    Y += 1
                    X = 2
                else:
                    X += 1
            if X != 2:
                while X != MaxX + 1:
                    ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='thin'))
                    if X == MaxX: ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'),
                                                                           right=Side(border_style='medium'))
                    X += 1
            else:
                Y -= 1
            for col in range(2, MaxX + 1):
                ws.cell(row=Z, column=col).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
                if col == MaxX: ws.cell(row=Z, column=col).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='thin'), right=Side(border_style='medium'),
                                                                           left=Side(border_style='thin'))
            X = 1
            Y += 1
    if Y - 1 != Z:
        ws.cell(row=Y - 1, column=1).border = Border(bottom=Side(border_style='thin'), left=Side(border_style='medium'))
    else:
        ws.cell(row=Y - 1, column=1).border = Border(bottom=Side(border_style='thin'), left=Side(border_style='medium'), top=Side(border_style='medium'))

    wS = wb.create_sheet(title="GN")
    wb.active = wS
    DEFAULT_FONT.name = "Bahnschrift SemiBold SemiConden"
    DEFAULT_FONT.sz = 10
    DEFAULT_FONT.b = True
    wS.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.3)
    wS.oddHeader.center.text = "Check Yard GNT"
    wS.oddHeader.right.text = "&P of &N"
    wS.oddHeader.left.text = "&D"

    # GN check

    GNtrucks = units_lst("GNtrucks")
    GNtrailers = {k: v[0] if isinstance(v, list) else v for k, v in units_lst("GNtrailers").items()}
    GNfb = {k: v[0] if isinstance(v, list) else v for k, v in units_lst("GNfb").items()}
    Csize = 11
    wS.column_dimensions["A"].width = Csize
    wS.column_dimensions["B"].width = Csize
    wS.column_dimensions["C"].width = Csize
    wS.column_dimensions["D"].width = Csize
    wS.column_dimensions["E"].width = Csize
    wS.column_dimensions["F"].width = Csize
    wS.column_dimensions["G"].width = Csize
    wS.column_dimensions["H"].width = Csize
    wS.column_dimensions["I"].width = Csize
    wS.cell(row=1, column=1).value = "GNT"
    wS.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    wS.cell(row=1, column=1).font = Font(name="Bahnschrift SemiBold SemiConden", size=20, b=True)
    wS.cell(row=2, column=1).value = "Trucks:"
    wS.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    wS.cell(row=2, column=1).font = Font(name="Bahnschrift SemiBold SemiConden", size=18, b=True)
    wS.cell(row=2, column=1).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'), right=Side(border_style='medium'), left=Side(border_style='medium'))
    wS.cell(row=2, column=1).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    wS.cell(row=2, column=4).value = "Trailers:"
    wS.cell(row=2, column=4).alignment = Alignment(horizontal='center')
    wS.cell(row=2, column=4).font = Font(name="Bahnschrift SemiBold SemiConden", size=18, b=True)
    wS.cell(row=2, column=4).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'), right=Side(border_style='medium'), left=Side(border_style='medium'))
    wS.cell(row=2, column=4).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    wS.cell(row=2, column=7).value = "Flatbeds:"
    wS.cell(row=2, column=7).alignment = Alignment(horizontal='center')
    wS.cell(row=2, column=7).font = Font(name="Bahnschrift SemiBold SemiConden", size=18, b=True)
    wS.cell(row=2, column=7).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'), right=Side(border_style='medium'), left=Side(border_style='medium'))
    wS.cell(row=2, column=7).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    wS.merge_cells('A1:I1')
    wS.merge_cells('D2:F2')
    wS.merge_cells('G2:I2')
    wS.merge_cells('A2:C2')

    row_coord1 = 3
    column_coord1 = 1
    for GNTtruck in sorted(GNtrucks.items(), key=lambda x: len(x[0])):
        type = "truck"
        SQL_REQ("INSERT INTO dbo.check_yard (date, company, type, unit_number, status) VALUES (?,?,?,?,?)", (date, "GN", type, GNTtruck[0], GNTtruck[1]), "W")
        wS.cell(row=row_coord1, column=column_coord1).value = GNTtruck[0]
        wS.cell(row=row_coord1, column=column_coord1).border = Border(bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
        wS.cell(row=row_coord1, column=column_coord1).font = Font(name="Bahnschrift SemiBold SemiConden", size=14, b=True)
        wS.cell(row=row_coord1, column=column_coord1).alignment = Alignment(horizontal='center')
        if column_coord1 == 3:
            column_coord1 = 1
            row_coord1 += 1
        else:
            column_coord1 += 1
    row_coord2 = 3
    column_coord2 = 4
    for GNTtrailer in sorted(GNtrailers.items(), key=lambda x: len(x[0])):
        type = "trailer"
        SQL_REQ("INSERT INTO dbo.check_yard (date, company, type, unit_number, status) VALUES (?,?,?,?,?)", (date, "GN", type, GNTtrailer[0], GNTtrailer[1]), "W")
        wS.cell(row=row_coord2, column=column_coord2).value = GNTtrailer[0]
        wS.cell(row=row_coord2, column=column_coord2).border = Border(bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
        wS.cell(row=row_coord2, column=column_coord2).font = Font(name="Bahnschrift SemiBold SemiConden", size=14, b=True)
        wS.cell(row=row_coord2, column=column_coord2).alignment = Alignment(horizontal='center')
        if column_coord2 == 6:
            column_coord2 = 4
            row_coord2 += 1
        else:
            column_coord2 += 1
    row_coord3 = 3
    column_coord3 = 7
    for GNTfb in sorted(GNfb.items(), key=lambda x: len(x[0])):
        type = "flatbed"
        SQL_REQ("INSERT INTO dbo.check_yard (date, company, type, unit_number, status) VALUES (?,?,?,?,?)", (date, "GN", type, GNTfb[0], GNTfb[1]), "W")
        wS.cell(row=row_coord3, column=column_coord3).value = GNTfb[0]
        wS.cell(row=row_coord3, column=column_coord3).border = Border(bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
        wS.cell(row=row_coord3, column=column_coord3).font = Font(name="Bahnschrift SemiBold SemiConden", size=14, b=True)
        wS.cell(row=row_coord3, column=column_coord3).alignment = Alignment(horizontal='center')
        if column_coord3 == 9:
            column_coord3 = 7
            row_coord3 += 1
        else:
            column_coord3 += 1
    minL = min(row_coord1, row_coord2, row_coord3)
    maxL = max(row_coord1, row_coord2, row_coord3)
    R = minL
    C = 1
    while True:
        wS.cell(row=R, column=C).border = Border(bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
        if C == 9:
            C = 1
            R += 1
        else:
            C += 1
        if R > maxL: break

    # Car parking check

    WS = wb.create_sheet(title="Car Parking")
    wb.active = WS
    WS.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.3)
    WS.oddHeader.center.text = "Car Parking Check"
    WS.oddHeader.right.text = "&P of &N"
    WS.oddHeader.left.text = "&D"

    carlist = list()
    for companyname in complist:
        main = cars(0, companyname, "main")
        main.extend(cars(0, companyname, "unreg"))
        if len(main) != 0:
            for item in main:

                if item["private"] is not None:
                    if item["private"]:
                        prv = item["expiration"]
                    elif not item["private"]:
                        car_amount = SQL_REQ("SELECT car FROM dbo.Company_List WHERE company_name=?", (companyname,), "S_one")
                        prv = car_amount[0]
                    carlist.append([item['plates'], item['car_model'], item["driver_name"], companyname, prv])
    carlist.sort(key=lambda x: x[0])
    Carsize = 20
    WS.column_dimensions["A"].width = Carsize
    WS.column_dimensions["B"].width = Carsize
    WS.column_dimensions["C"].width = Carsize + 3
    WS.column_dimensions["D"].width = Carsize
    WS.column_dimensions["E"].width = Carsize

    WS.cell(row=1, column=1).value = "Car Parking"
    WS.cell(row=1, column=1).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    WS.merge_cells("A1:E1")
    WS.cell(row=2, column=1).value = "Plate:"
    WS.cell(row=2, column=1).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=2, column=1).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
    WS.cell(row=2, column=1).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    WS.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    WS.cell(row=2, column=2).value = "Car:"
    WS.cell(row=2, column=2).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=2, column=2).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
    WS.cell(row=2, column=2).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    WS.cell(row=2, column=2).alignment = Alignment(horizontal='center')
    WS.cell(row=2, column=3).value = "Driver:"
    WS.cell(row=2, column=3).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=2, column=3).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
    WS.cell(row=2, column=3).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    WS.cell(row=2, column=3).alignment = Alignment(horizontal='center')
    WS.cell(row=2, column=4).value = "Company:"
    WS.cell(row=2, column=4).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=2, column=4).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
    WS.cell(row=2, column=4).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    WS.cell(row=2, column=4).alignment = Alignment(horizontal='center')
    WS.cell(row=2, column=5).value = "Spot/EXP.Date:"
    WS.cell(row=2, column=5).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=2, column=5).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
    WS.cell(row=2, column=5).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    WS.cell(row=2, column=5).alignment = Alignment(horizontal='center')
    row1 = 3
    colmn = 1
    for line in carlist:
        for each in line:
            if colmn == 5 and each is not None:
                a = each

                if not isinstance(a, int):
                    if date.date() > each:
                        WS.cell(row=row1, column=colmn).fill = PatternFill(fill_type="darkUp", start_color="FF0000")
            WS.cell(row=row1, column=colmn).value = each
            WS.cell(row=row1, column=colmn).font = Font(name="Bahnschrift SemiBold SemiConden", size=14, b=True)
            WS.cell(row=row1, column=colmn).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
            WS.cell(row=row1, column=colmn).alignment = Alignment(horizontal='center')
            colmn += 1
        row1 += 1
        colmn = 1
    wb.active = ws

    isExist = os.path.exists(sets["chk_path"])
    if not isExist:
        os.makedirs(sets["chk_path"])
    wb.save(sets["chk_path"] + "CheckYard " + date.strftime("%Y") + date.strftime("%m") + date.strftime("%d") + ".xlsx")


    pass

def checkyard_print():
    checkdatetime = settings_file()["chk_datetime"]
    if checkdatetime == "None": return
    check_date = datetime.strptime(checkdatetime, "%Y-%m-%d %H:%M:%S")
    filepath = sets["chk_path"].replace("\\\\", "/") + "CheckYard " + check_date.date().strftime("%Y%m%d") + ".xlsx"
    excel = win32.DispatchEx('Excel.Application')
    wb = excel.Workbooks.Open(filepath)
    sheets = wb.Sheets
    for sheet in sheets:
        sheet.PrintOut()
    wb.Close(False)
    excel.Quit()














#Login function Retrive log and pas, compare it with info in DB apply rights to access tabs in program.
def login_func(*args):
    global security
    global Log_Tracer
    login = login_name_entry.get()
    password = login_password_entry.get()
    login_class = PasswordDatabase()
    if login_class.login(login, password):
        login_frame.pack_forget()
        Menu_Bar.pack(fill=tk.BOTH, expand=1)
        security = SQL_REQ("SELECT full_name, rights FROM dbo.authentication WHERE login=?", (login,), "S_one")
        Log_Tracer = True
        Security_Name.configure(text=security[0])
        Security_Reset_Button.pack(side=tk.RIGHT, padx=(10, 50))
        Security_Name.pack(side=tk.RIGHT)
        for a in range(Menu_Bar.index(tk.END)): Menu_Bar.tab(a, state=tk.DISABLED)
        if security[1] == 1:
            for a in range(Menu_Bar.index(tk.END)):
                Menu_Bar.tab(a, state=tk.NORMAL)
                Menu_Bar.select(0)
        elif security[1] == 3:
            Menu_Bar.tab(0, state=tk.NORMAL)
            Menu_Bar.tab(1, state=tk.NORMAL)
            Menu_Bar.tab(3, state=tk.NORMAL)
            Menu_Bar.select(0)
        elif security[1] == 2:
            Menu_Bar.tab(3, state=tk.NORMAL)
            Menu_Bar.select(3)
    else:
        error(2)
        return
    login_button.unbind_all('<Return>')

#SHIFT CHANGE FUNCTION
def shift_change(event):
        Refresh("Tenant")
        Refresh("GN")
        Refresh("Visitor")
        Menu_Bar.pack_forget()
        login_password_entry.delete(0, tk.END)
        login_frame.pack(side=tk.TOP, pady=(300, 0))
        Security_Name.pack_forget()
        login_name_entry.focus_set()
        login_button.bind_all('<Return>', login_func)

def over_extract(M, Y, company):
    row = SQL_REQ("SELECT over_count, trucks_onyard, trailers_onyard, date, company_ID FROM dbo.OVERPARKING WHERE MONTH(date)=? AND YEAR(date)=? AND company_ID=? ORDER BY date", (str(M), str(Y), str(company)), "S_all_D")
    if not row[0]: return
    res = []
    for n in row[0]:
        over_list = {}
        index = 0
        for y in n:
            z = row[1][index]
            index += 1
            if y is not None:
                if z[0] == "date":
                    y = datetime.strptime(str(y), "%Y-%m-%d").date()
                over_list.update({z[0]: y})
            else:
                over_list.update({z[0]: 0})
        res.append(over_list)
    m_range = monthrange(Y, M)
    total_res = list()
    last = None
    for j in range(m_range[1]):
        for rec in res:
            if int(rec["date"].strftime('%d')) == j + 1:
                if rec["over_count"] == 0:
                    last = None
                    break
                else:
                    total_res.append(rec)
                    last = dict(rec)
                    break
        if last is not None and int(last["date"].strftime("%d")) != j + 1:
            last["date"] = last["date"].replace(day=j + 1)
            lst = dict(last)
            total_res.append(lst)
    if total_res is None or total_res == []: return
    for each in total_res:
        if each["trucks_onyard"] is not None:
            each["trucks_onyard"] = list(str(each["trucks_onyard"]).split("|"))
        if each["trailers_onyard"] is not None:
            each["trailers_onyard"] = list(str(each["trailers_onyard"]).split("|"))
        each["date"] = each["date"].strftime("%Y-%m-%d")
    return total_res

def chk_add_manual(*args):
    if settings_file()["chk_datetime"] == "None":
        error(10)
        return
    company = chk_c_entry.get()
    Type = chk_radio_var.get()
    unit_number = chk_t_entry.get()
    chk_t_entry.delete(0, tk.END)
    time = datetime.strptime(settings_file()["chk_datetime"],"%Y-%m-%d %H:%M:%S")
    status = True
    var = ((unit_number, Type, status, time, company))
    chk_set(var)
    UNTS(Current_chk_Company_obj, company, "CheckYard")
    chk_t_entry.focus_set()

def chk_del_manual(*args):
    if settings_file()["chk_datetime"] == "None":
        error(10)
        return
    company = chk_c_entry.get()
    Type = chk_radio_var.get()
    unit_number = chk_t_entry.get()
    chk_t_entry.delete(0, tk.END)
    time = datetime.strptime(settings_file()["chk_datetime"],"%Y-%m-%d %H:%M:%S")
    status = False
    var = ((unit_number, Type, status, time, company))
    chk_set(var, "DEL")
    UNTS(Current_chk_Company_obj, company, "CheckYard")
    chk_t_entry.focus_set()


# Function to appercase entry: accept obj key and entry name as value.
def UPPER_CASE(event, **kwargs):
    entry = kwargs["obj"].get()
    if entry:
        kwargs["obj"].delete(0, tk.END)
        kwargs["obj"].insert(0, entry.upper())

# OVERPARKING FUNCTION + statstic registering
# get: company, datetime, in_out
def OVERPARKING (event, func):
    global StatisticOVER
    def get_units(company_id):
        # making list of truck numbers on yard NULL if none
        truck_list_q = SQL_REQ("SELECT truck_number FROM dbo.Tenant_Trucks WHERE status=1 AND company_ID=?", (str(company_id),), "S_all")
        if truck_list_q:
            truck_list = []
            for x in truck_list_q: truck_list.append(x[0])
            l = str()
            for y in truck_list: l += (y + "|")
            over_trucks = l.strip("|")
        else:
            over_trucks = None
        truck_list_q = SQL_REQ("SELECT truck_number FROM dbo.Tenant_Trucks_UNREG WHERE status=1 AND company_ID=?", (str(company_id),), "S_all")
        if truck_list_q:
            truck_list = []
            for x in truck_list_q: truck_list.append(x[0])
            l = str()
            for y in truck_list: l += (y + "|")
            over_trucks_UNREG = l.strip("|")
        else:
            over_trucks_UNREG = None

        # making list of trailer numbers on yard NULL if none
        trailer_list_q = SQL_REQ("SELECT trailer_number FROM dbo.Tenant_Trailers WHERE status=1 AND company_ID=?", (str(company_id),), "S_all")
        if trailer_list_q:
            trailer_list = []
            for x in trailer_list_q: trailer_list.append(x[0])
            l = str()
            for y in trailer_list: l += (y + "|")
            over_trailers = l.strip("|")
        else:
            over_trailers = None
        trailer_list_q = SQL_REQ("SELECT trailer_number FROM dbo.Tenant_Trailers_UNREG WHERE status=1 AND company_ID=?", (str(company_id),), "S_all")
        if trailer_list_q:
            trailer_list = []
            for x in trailer_list_q: trailer_list.append(x[0])
            l = str()
            for y in trailer_list: l += (y + "|")
            over_trailers_UNREG = l.strip("|")
        else:
            over_trailers_UNREG = None
        if over_trucks is not None:
            if over_trucks_UNREG is not None:
                m = over_trucks+'|'+over_trucks_UNREG
            else: m = over_trucks
        else:
            if over_trucks_UNREG is not None:
                m = over_trucks_UNREG
            else: m = None
        if over_trailers is not None:
            if over_trailers_UNREG is not None:
                n = over_trailers + '|' + over_trailers_UNREG
            else:
                n = over_trailers
        else:
            if over_trailers_UNREG is not None:
                n = over_trailers_UNREG
            else:
                n = None
        list = [m, n]
        return (list)

    if func == "T":
        statistics_reg("T")
        if event["Company"] == "Euro Can": return  #return if Euro+Can (not applyable)

        # Retriving info how many spots is allowed and combine different plans in truckN and trailerN max value allowed
        row, col = SQL_REQ("SELECT regular, truck, trailer, designated, company_ID FROM dbo.Company_List WHERE company_name=? AND activity=1", (event["Company"],), "S_one_D")
        spot_list = {}
        if row:
            index = 0
            for y in row:
                z = col[index]
                index += 1
                if y is not None:
                    spot_list.update({z[0]: int(y)})
                else:
                    spot_list.update({z[0]: 0})
        else:
            error(15) #impossible error
            return
        truckN = spot_list["regular"] + spot_list["truck"] + spot_list["designated"]
        trailerN = spot_list["regular"] + spot_list["designated"] + spot_list["trailer"]

        #Retriving info how many trucks and trailers actually on yard including last event from perm and temp tables and form in TrucksONyard and TrailersOnyard var
        query = SQL_REQ("SELECT COUNT(*) FROM dbo.Tenant_Trucks WHERE company_ID=? AND status=1", (str(spot_list["company_ID"]),), "S_one")
        if query: Tval = int(query[0])
        else: Tval = 0
        query = SQL_REQ("SELECT COUNT(*) FROM dbo.Tenant_Trucks_UNREG WHERE company_ID=? AND status=1", (str(spot_list["company_ID"]),), "S_one")
        if query: Tval_UNREG = int(query[0])
        else: Tval_UNREG = 0
        query = SQL_REQ("SELECT COUNT(*) FROM dbo.Tenant_Trailers WHERE company_ID=? AND status=1", (str(spot_list["company_ID"]),), "S_one")
        if query: TRval = int(query[0])
        else: TRval = 0
        query = SQL_REQ("SELECT COUNT(*) FROM dbo.Tenant_Trailers_UNREG WHERE company_ID=? AND status=1", (str(spot_list["company_ID"]),), "S_one")
        if query: TRval_UNREG = int(query[0])
        else: TRval_UNREG = 0
        TrucksONyard = Tval + Tval_UNREG
        TrailersONyard = TRval + TRval_UNREG


        #get units numbers that on yard
        units = get_units(spot_list["company_ID"]) ##################################################################### redirect

        # calculate in var bigest overparking among trucks and trailers
        onYardOver = max(max(0, TrailersONyard - trailerN), max(0, TrucksONyard - truckN))  # removed abs()
        event_datetime = datetime.strptime(event["Date"] + " " + event["Time"], "%Y/%m/%d %H:%M:%S")
        #event_datetime = event_datetime - timedelta(hours=12) # to remove
        event["Time"] = event_datetime.strftime("%H:%M:%S")

        ######!!!MAIN!!!#######
        #check if there is over record TODAY
        row, col = SQL_REQ("SELECT * FROM dbo.OVERPARKING WHERE date=? AND company_ID=?", (event["Date"], str(spot_list["company_ID"])), "S_one_D")
        if row:
            last_record = {col[i][0]: y if y is not None else 0 for i, y in enumerate(row)}
            if last_record["over_count"] < onYardOver:
                #add current time in the list of over times
                if last_record["over_time"] !=0: over_time = last_record["over_time"]+"|"+str(event["Time"])
                else: over_time = str(event["Time"])
                #update over record
                SQL_REQ("UPDATE dbo.OVERPARKING SET over_count=?, trucks_onyard=?, trailers_onyard=?, last_time=?, over_time=?, last_over_count=?, last_trucks_onyard=?, last_trailers_onyard=? WHERE company_ID=? AND date=?",
                        (str(onYardOver), units[0], units[1], event["Time"], over_time, str(onYardOver), units[0], units[1], str(spot_list["company_ID"]), event["Date"]), "W")
            elif last_record["over_count"] > onYardOver:
                # # review previous overparking time markers and check if there is less "2h" difference remove latest marker and replace overparking with current
                if last_record["over_time"] != 0:
                    over_time_list = last_record["over_time"].split("|")
                    over_time_list.sort(reverse=True)
                    if over_time_list:
                        event_delta = int((event_datetime - datetime.strptime(f"{last_record['date']} {over_time_list[0]}", "%Y-%m-%d %H:%M:%S")).total_seconds()) // 3600
                        if event_delta < int(sets["Overparking_Timeout"]):
                            over_time_list.pop(0)
                            over_time_list_new = "|".join(over_time_list) if over_time_list else None
                            SQL_REQ("UPDATE dbo.OVERPARKING SET over_count=?, trucks_onyard=?, trailers_onyard=?, last_time=?, over_time=?, last_over_count=?, last_trucks_onyard=?, last_trailers_onyard=? WHERE company_ID=? AND date=?",
                                    (str(onYardOver), units[0], units[1], event["Time"], over_time_list_new, str(onYardOver), units[0], units[1], str(spot_list["company_ID"]), event["Date"]), "W")
                            return
                        SQL_REQ("UPDATE dbo.OVERPARKING SET last_time=?, last_over_count=?, last_trucks_onyard=?, last_trailers_onyard=? WHERE company_ID=? AND date=?",
                            (event["Time"], str(onYardOver), units[0], units[1], str(spot_list["company_ID"]), event["Date"]), "W")
            #when over the same - update last over, last trucks, last trailers
            else: SQL_REQ("UPDATE dbo.OVERPARKING SET last_time=?, last_over_count=?, last_trucks_onyard=?, last_trailers_onyard=? WHERE company_ID=? AND date=?",
                            (event["Time"], str(onYardOver), units[0], units[1], str(spot_list["company_ID"]), event["Date"]), "W")
        #if no overparking record for today
        else:  # NO overparking for today
                ###check if previous over was yesturday or later, if later - copy over to next day and reduce over for previous if its under over_time
            row, col = SQL_REQ("SELECT * FROM dbo.OVERPARKING WHERE date=(SELECT max(date) FROM dbo.OVERPARKING WHERE company_ID=?)", (str(spot_list["company_ID"]),), "S_one_D")
            if row:
                previous_record = {col[i][0]: y if y is not None else 0 for i, y in enumerate(row)}
                if previous_record["over_time"] != 0:
                    pr_time_list = previous_record["over_time"].split("|")
                    pr_time_list.sort()
                    trig = False # trigger need to run create previous+1 record if difference more than 1 day.
                    for time_str in pr_time_list:
                        midnight_time = datetime.strptime(str(previous_record["date"]) + " 00:00:00", "%Y-%m-%d %H:%M:%S")
                        previous_over_time = datetime.strptime(str(previous_record["date"]) + " " + str(time_str), "%Y-%m-%d %H:%M:%S")
                        delta_to_midnight = 24 + int((midnight_time - previous_over_time).total_seconds() // 3600)
                        ## block execute if time of previous over <= midnight
                        if delta_to_midnight <= int(sets["Overparking_Timeout"]):
                            pr_time_list.remove(time_str)
                            pr_time_new = "|".join(pr_time_list)
                            if pr_time_new == "": pr_time_new = None
                            new_over_count = int(previous_record["over_count"]) - 1
                            # creating overparking after previous over if difference over 1 day
                            days_delta = (datetime.strptime(event["Date"], "%Y/%m/%d").date() - previous_record["date"]).days
                            #rounding seconds to nearest minutes to fit smalldate type in history
                            if previous_over_time.second>=30: previous_over_time+=timedelta(minutes=1)
                            rounded_dt = previous_over_time.replace(second=0, microsecond=0)
                            unit_numbers_toremove = SQL_REQ("SELECT truck_number, trailer_number FROM dbo.Tenant_History WHERE datetime_event=?", (rounded_dt,), "S_all")
                            if unit_numbers_toremove:
                                new_trucks_onyard = previous_record["trucks_onyard"].replace("|"+unit_numbers_toremove[0][0], "").replace(unit_numbers_toremove[0][0], "")
                                new_trailer_onyard = previous_record["trailers_onyard"].replace("|"+unit_numbers_toremove[0][1], "").replace(unit_numbers_toremove[0][1], "")
                            SQL_REQ("UPDATE dbo.OVERPARKING SET over_time=?, over_count=?, trucks_onyard=?, trailers_onyard=?, last_over_count=?, last_trucks_onyard=?, last_trailers_onyard=? WHERE date=(SELECT max(date) FROM dbo.OVERPARKING WHERE company_ID=?)",
                                    (pr_time_new, new_over_count, str(new_trucks_onyard), str(new_trailer_onyard), new_over_count,  str(new_trucks_onyard), str(new_trailer_onyard), str(spot_list["company_ID"])), "W")
                            if days_delta > 1 and not trig:
                                next_date = previous_record["date"] + timedelta(days=1)

                                SQL_REQ("INSERT INTO dbo.OVERPARKING (date,company_ID,over_count,trucks_onyard,trailers_onyard,last_time,over_time, last_over_count, last_trucks_onyard, last_trailers_onyard) VALUES (?,?,?,?,?,?,?,?,?,?)",
                                        (next_date, str(spot_list["company_ID"]), previous_record["over_count"], previous_record["trucks_onyard"], previous_record["trailers_onyard"], None, None, previous_record["over_count"], previous_record["trucks_onyard"], previous_record["trailers_onyard"]), "W")
                                trig = True
                        ##block check if last over of previous record smaller than previous over and there is a gap in days - replace previous over on last over of the next day.
                        else:
                            if previous_record["last_over_count"] != "" and not trig:
                                # if previous over bigger than last over
                                if int(previous_record["over_count"])>int(previous_record["last_over_count"]):
                                    previous_day = datetime.strptime(str(previous_record["date"]), "%Y-%m-%d").date()
                                    print(f"previous_day {previous_day}")
                                    current_day = datetime.strptime(event["Date"], "%Y/%m/%d").date()
                                    print(f"current_day {current_day}")
                                    days_gap = current_day - previous_day
                                    # if gap between today and last record is more than 1 day - inserting last over as main over for day after last record
                                    if days_gap.days>1:
                                        next_day = previous_day + timedelta(days=1)
                                        print(next_day)
                                        for key in previous_record:
                                            print(key)
                                        SQL_REQ("INSERT INTO dbo.OVERPARKING (date, company_ID, over_count, trucks_onyard, trailers_onyard, last_time, over_time, last_over_count, last_trucks_onyard, last_trailers_onyard) VALUES (?,?,?,?,?,?,?,?,?,?)",
                                                (next_day, previous_record["company_ID"], previous_record["last_over_count"], previous_record["last_trucks_onyard"],
                                                 previous_record["last_trailers_onyard"], None, None, previous_record["last_over_count"], previous_record["last_trucks_onyard"],
                                                 previous_record["last_trailers_onyard"]), "W")
                                        trig = True


                        ##
                # IF previous over bigger make over record without marker
                if previous_record["over_count"] > onYardOver:
                    SQL_REQ("INSERT INTO dbo.OVERPARKING (date,company_ID,over_count,trucks_onyard,trailers_onyard,last_time, last_over_count, last_trucks_onyard, last_trailers_onyard) VALUES (?,?,?,?,?,?,?,?,?)",
                            (event["Date"], str(spot_list["company_ID"]), str(onYardOver), units[0], units[1], event["Time"], str(onYardOver), units[0], units[1]), "W")
                # IF previous over smaller - make over record with marker
                elif previous_record["over_count"] < onYardOver:
                    SQL_REQ("INSERT INTO dbo.OVERPARKING (date,company_ID,over_count,trucks_onyard,trailers_onyard,last_time,over_time,last_over_count, last_trucks_onyard, last_trailers_onyard) VALUES (?,?,?,?,?,?,?,?,?,?)",
                            (event["Date"], str(spot_list["company_ID"]), str(onYardOver), units[0], units[1], event["Time"], str(event["Time"]), str(onYardOver), units[0], units[1]), "W")
                # IF previous over same = make over record with marker IF time less than 2h.
                elif previous_record["over_count"] == onYardOver:
                    if event_datetime.hour < int(sets["Overparking_Timeout"]):
                        SQL_REQ("INSERT INTO dbo.OVERPARKING (date,company_ID,over_count,trucks_onyard,trailers_onyard,last_time,over_time, last_over_count, last_trucks_onyard, last_trailers_onyard) VALUES (?,?,?,?,?,?,?,?,?,?)",
                                (event["Date"], str(spot_list["company_ID"]), str(onYardOver), units[0], units[1], event["Time"], str(event["Time"]),  str(onYardOver), units[0], units[1]), "W")
                    else:
                        SQL_REQ("INSERT INTO dbo.OVERPARKING (date,company_ID,over_count,trucks_onyard,trailers_onyard,last_time, last_over_count, last_trucks_onyard, last_trailers_onyard) VALUES (?,?,?,?,?,?,?,?,?)",
                                (event["Date"], str(spot_list["company_ID"]), str(onYardOver), units[0], units[1], event["Time"],  str(onYardOver), units[0], units[1]), "W")
                return
                # create over if there is NO previous records but there is overparking now
            if onYardOver > 0: SQL_REQ("INSERT INTO dbo.OVERPARKING (date,company_ID,over_count,trucks_onyard,trailers_onyard,last_time,over_time, last_over_count, last_trucks_onyard, last_trailers_onyard) VALUES (?,?,?,?,?,?,?,?,?,?)",
                                       (event["Date"], str(spot_list["company_ID"]), str(onYardOver), units[0], units[1], event["Time"], str(event["Time"]), str(onYardOver), units[0], units[1]), "W")
            else: return
    elif func == "GN": #function for GN overparking if need in the future
        statistics_reg("GN")
    elif func == "V": #function for Visitors overparking if need in the future
        pass

#Function for registering statistics of the yard for Chart
def statistics_reg(func):
    global StatisticT
    global StatisticGN
    global Company_Var
    datenow = datetime.now().date()
    def stat_reg(table, amount):
        current = SQL_REQ("SELECT * FROM [dbo].[statistics] WHERE date=?", (str(datenow),), "S_one")
        if table == "tenant_amount": inx = 1
        elif table == "gn_amount": inx = 2
        else:
            error(15)
            return
        if current:
            if current[inx]:
                last_num = int(current[inx])
                if last_num >= amount: return
            SQL_REQ("UPDATE [dbo].[statistics] SET "+table+"=? WHERE date=?", (str(amount), str(datenow)), "W")
            return
        if amount > 0:
            list = [str(datenow), None, None]
            list[inx] = str(amount)
            SQL_REQ("INSERT INTO [dbo].[statistics] (date, tenant_amount, gn_amount) VALUES (?,?,?)", list, "W")



    if func == "T":
        reg = SQL_REQ("SELECT COUNT(*) FROM dbo.Tenant_Trailers WHERE status=1", (), "S_one")
        unreg = SQL_REQ("SELECT COUNT(*) FROM dbo.Tenant_Trailers_UNREG WHERE status=1", (), "S_one")
        amount = int(reg[0])+int(unreg[0])
        StatisticT.config(text=f"T: {amount}")
        stat_reg("tenant_amount", amount)
    elif func == "GN":
        tr = SQL_REQ("SELECT COUNT(*) FROM dbo.GN_Trailers WHERE status=1", (), "S_one")
        fb = SQL_REQ("SELECT COUNT(*) FROM dbo.GN_Flatbed WHERE status=1", (), "S_one")
        amount = int(tr[0])+int(fb[0])
        StatisticGN.config(text=f"GN: {amount}")
        stat_reg("gn_amount", amount)
    elif func == "O":
        if Company_Var is None: return
        statover = SQL_REQ("SELECT over_count FROM dbo.OVERPARKING AS ov INNER JOIN dbo.Company_List AS cl ON ov.company_ID=cl.company_ID WHERE ov.date=? AND cl.company_name=?", (str(datenow), Company_Var), "S_one")
        if statover:
            if int(statover[0])>0: StatisticOVER.config(text=f"O: {statover[0]}")
            else: StatisticOVER.config(text="")
        else: StatisticOVER.config(text="")

def to_Excel(date, company):

    def lot_extract(company):
        row, col = SQL_REQ("SELECT * FROM dbo.Company_List WHERE company_ID=?", (str(company),), "S_one_D")
        dict = {}
        if row:
            index = 0
            for y in row:
                z = col[index]
                index += 1
                if y is not None:
                    dict.update({z[0]: y})
                else:
                    dict.update({z[0]: 0})
        return dict
    over = over_extract(date.month, date.year, company)
    if over is None: return
    lot = lot_extract(company)
    wb = Workbook()
    ws = wb.active
    def_font = DEFAULT_FONT
    def_font.name = "Bahnschrift SemiBold SemiConden"
    def_font.sz = 8
    def_font.b = True
    ws.title = "Overparking"
    for cell in ws['A1:K1']:
        for all in cell:
            all.border = Border(bottom=Side(border_style="medium", color="000000"))
    for cell in "ABCDEFGHIJK":
        ws.column_dimensions[cell].width = 11
    ws.merge_cells('A1:C1')
    ws.merge_cells('D1:H1')
    ws.merge_cells('I1:K1')
    top_over = ws['A1']
    top_over.value = "Overparking For"
    top_company = ws['D1']
    top_company.value = lot["company_name"]
    top_date = ws['I1']
    top_date.value = date.strftime("%B, %Y")
    top_over.font = Font(name="Bahnschrift SemiBold SemiConden", size=14, b=True)
    top_company.font = Font(name="Bahnschrift SemiBold SemiConden", size=14, b=True, color="960404")
    top_company.alignment = Alignment(horizontal="center")
    top_date.font = Font(name="Bahnschrift SemiBold SemiConden", size=14, b=True, color="003399")
    top_date.alignment = Alignment(horizontal="right")
    ws.merge_cells('B2:C2')
    ws.merge_cells('E2:F2')
    lot_dis = ws["B2"]
    lot_dis_val = ws["D2"]
    lot_reg = ws["E2"]
    lot_reg_val = ws["G2"]
    lot_truck = ws["H2"]
    lot_truck_val = ws["I2"]
    lot_trailer = ws["J2"]
    lot_trailer_val = ws["K2"]
    lot_log = ws["A2"]
    font_lot = Font(name="Bahnschrift SemiBold SemiConden", color="494529", size=10)
    font_lot_val = Font(name="Bahnschrift SemiBold SemiConden", color="963634", size=10)
    lot_log.value = "Plan"
    lot_log.font = Font(name="Bahnschrift SemiBold SemiConden", size=10)
    lot_dis.value = "Designated:"
    lot_reg.value = "Truck+Trailer:"
    lot_truck.value = "Truck:"
    lot_trailer.value = "Trailer:"
    lot_dis.font = font_lot
    lot_reg.font = font_lot
    lot_truck.font = font_lot
    lot_trailer.font = font_lot
    lot_dis_val.value = lot["designated"]
    lot_reg_val.value = lot["regular"]
    lot_truck_val.value = lot["truck"]
    lot_trailer_val.value = lot["trailer"]
    lot_dis_val.font = font_lot_val
    lot_reg_val.font = font_lot_val
    lot_truck_val.font = font_lot_val
    lot_trailer_val.font = font_lot_val
    ws.print_title_rows = "1:3"
    Coord = 4
    over_total = 0
    th = Side(border_style='medium')
    tn = Side(border_style='thin')
    for one in over:
        len_truck = math.ceil(len(one["trucks_onyard"]) / 10)
        len_trailer = math.ceil(len(one["trailers_onyard"]) / 10)
        if 59 - ((Coord%59) + len_truck + len_trailer) < 0:
            pbreak = Break(id=(Coord-1))
            ws.row_breaks.append(pbreak)
        ws.cell(row=Coord, column=1).border = Border(left=th, top=th, bottom=th)
        for i in range(9): ws.cell(row=Coord, column=i + 2).border = Border(top=th, bottom=th)
        ws.cell(row=Coord, column=11).border = Border(right=th, top=th, bottom=th)
        ws.cell(row=Coord, column=1).value = "Date:"
        day = ws.cell(row=Coord, column=2)
        day.value = one["date"]
        ws.cell(row=Coord, column=10).value = "Over:"
        overcell = ws.cell(row=Coord, column=11)
        overcell.value = one["over_count"]
        over_total += int(one["over_count"])
        for cell in ws['A' + str(Coord) + ':K' + str(Coord)]:
            for all in cell:
                all.fill = PatternFill(start_color="B8CCE4", fill_type="solid")
        cc = 2
        Coord += 1
        ws.cell(row=Coord, column=1).value = "Trucks:"
        for unit in one["trucks_onyard"]:
            ws.cell(row=Coord, column=1).border = Border(left=tn)
            ws.cell(row=Coord, column=cc).value = unit
            ws.cell(row=Coord, column=cc).border = Border(left=tn, right=tn, top=tn, bottom=tn)
            if cc == 11:
                Coord += 1
                cc = 2
            else:
                cc += 1
        ws.cell(row=Coord, column=1).border = Border(left=tn, bottom=tn)
        if cc != 11:
            for i in range(12 - cc):
                ws.cell(row=Coord, column=i + cc).border = Border(left=tn, right=tn, top=tn, bottom=tn)
        Coord +=1
        ws.cell(row=Coord, column=1).value = "Trailers:"
        ws.cell(row=Coord, column=1).border = Border(top=tn)
        cc = 2
        for unit in one["trailers_onyard"]:
            ws.cell(row=Coord, column=1).border = Border(left=tn)
            ws.cell(row=Coord, column=cc).value = unit
            ws.cell(row=Coord, column=cc).border = Border(left=tn, right=tn, top=tn, bottom=tn)
            if cc == 11:
                Coord += 1
                cc = 2
            else:
                cc += 1
        ws.cell(row=Coord, column=1).border = Border(left=tn, bottom=tn)
        if cc != 11:
            for i in range(12 - cc):
                ws.cell(row=Coord, column=i + cc).border = Border(left=tn, right=tn, top=tn, bottom=tn)
        Coord += 1
    for i in range(11): ws.cell(row=Coord, column=i + 1).border = Border(top=th)
    tt = ws.cell(row=Coord, column=10)
    tt.value = "Total:"
    tt.font = font_lot
    ttr = ws.cell(row=Coord, column=11)
    ttr.value = over_total
    ttr.font = font_lot_val
    for cell in ws['A' + str(Coord) + ':K' + str(Coord)]:
        for all in cell:
            all.fill = PatternFill(start_color="C4D79B", fill_type="solid")
    isExist = os.path.exists(sets["SQL_path"] + date.strftime("%Y") + "\\" + date.strftime("%m"))
    if not isExist:
        os.makedirs(sets["SQL_path"] + date.strftime("%Y") + "\\" + date.strftime("%m"))
    wb.save(sets["SQL_path"] + date.strftime("%Y") + "\\" + date.strftime("%m") + "\\" + lot["company_name"] + ".xlsx")

def chk_update():
    if settings_file()["chk_datetime"] == "None":
        error(10)
        return
    check_list = SQL_REQ("SELECT * FROM dbo.check_yard", (), "S_all")
    for each in check_list:
        if each[1] == "GN":
            if each[2] == "truck":
                unit_read = SQL_REQ("SELECT truck_number, status, last_date FROM dbo.GN_Trucks WHERE truck_number=?", (each[3],), "S_one")
                if unit_read is not None:
                    if unit_read[1] != each[4]:
                        if unit_read[2] is None: unit_read[2] = datetime(2022, 1, 1, 1, 1, 1)
                        if unit_read[2]<each[0]:
                            comment = str(unit_read[1])+"-"+str(unit_read[2])
                            record = {
                                "Company": "GN",
                                "Truck": each[3],
                                "Trailer": None,
                                "Date": each[0].strftime("%Y/%m/%d"),
                                "Time": each[0].strftime("%H:%M:%S"),
                                "Status": each[4],
                                "Cargo": False,
                                "Type": None,
                                "Comment": comment
                            }
                            GN_Record(record)
            elif each[2] == "trailer":
                unit_read = SQL_REQ("SELECT trailer_number, status, last_date FROM dbo.GN_Trailers WHERE trailer_number=?", (each[3],), "S_one")
                if unit_read is not None:
                    if unit_read[1] != each[4]:
                        if unit_read[2] is None: unit_read[2] = datetime(2022, 1, 1, 1, 1, 1)
                        if unit_read[2] < each[0]:
                            comment = str(unit_read[1]) + "-" + str(unit_read[2])
                            record = {
                                "Company": "GN",
                                "Truck": None,
                                "Trailer": each[3],
                                "Date": each[0].strftime("%Y/%m/%d"),
                                "Time": each[0].strftime("%H:%M:%S"),
                                "Status": each[4],
                                "Cargo": False,
                                "Type": True,
                                "Comment": comment
                            }
                            GN_Record(record)
            elif each[2] == "flatbed":
                unit_read = SQL_REQ("SELECT fb_number, status, last_date FROM dbo.GN_Flatbed WHERE fb_number=?", (each[3],), "S_one")
                if unit_read is not None:
                    if unit_read[1] != each[4]:
                        if unit_read[2] is None: unit_read[2] = datetime(2022, 1, 1, 1, 1, 1)
                        if unit_read[2] < each[0]:
                            comment = str(unit_read[1]) + "-" + str(unit_read[2])
                            record = {
                                "Company": "GN",
                                "Truck": None,
                                "Trailer": each[3],
                                "Date": each[0].strftime("%Y/%m/%d"),
                                "Time": each[0].strftime("%H:%M:%S"),
                                "Status": each[4],
                                "Cargo": False,
                                "Type": False,
                                "Comment": comment
                            }
                            GN_Record(record)

        else:
            if each[2] == "truck":
                unit_read = SQL_REQ("SELECT truck_number, status, last_date FROM dbo.Tenant_Trucks AS tt INNER JOIN dbo.Company_List AS cl ON tt.company_ID=cl.company_ID WHERE tt.truck_number=? AND cl.company_name=?", (each[3], each[1]), "S_one")
                if unit_read is not None:
                    if unit_read[1] != each[4]:
                        if unit_read[2] is None: unit_read[2] = datetime(2022,1,1,1,1,1)
                        if unit_read[2]<each[0]:
                            comment = str(unit_read[1])+"-"+str(unit_read[2])
                        else:continue
                    else: continue
                else:
                    unit_read = SQL_REQ("SELECT truck_number, status, last_date FROM dbo.Tenant_Trucks_UNREG AS ttu INNER JOIN dbo.Company_List AS cl ON ttu.company_ID=cl.company_ID WHERE ttu.truck_number=? AND cl.company_name=?", (each[3], each[1]), "S_one")
                    if unit_read is not None:
                        if unit_read[1] != each[4]:
                            if unit_read[2] is None: unit_read[2] = datetime(2022, 1, 1, 1, 1, 1)
                            if unit_read[2] < each[0]:
                                comment = str(unit_read[1]) + "-" + str(unit_read[2])
                            else: continue
                        else: continue
                    else:
                        comment = "unregistered"
                        if each[1] == "Euro Can":
                            chk_eu_his = SQL_REQ("SELECT TOP 1 * FROM dbo.Tenant_History AS th INNER JOIN dbo.Company_List AS cl ON th.company_ID=cl.company_ID WHERE th.truck_number=? AND cl.company_name=? AND th.datetime_event>? ORDER BY th.datetime_event DESC", (each[3], each[1], each[0].strftime("%Y-%m-%d %H:%M:%S")), "S_one")
                            if chk_eu_his: continue
                record = {
                    "Company": each[1],
                    "Truck": each[3],
                    "Trailer": None,
                    "Date": each[0].strftime("%Y/%m/%d"),
                    "Time": each[0].strftime("%H:%M:%S"),
                    "Status": each[4],
                    "Comment": comment
                }
                Tenant_Record(record)
            elif each[2] == "trailer":
                unit_read = SQL_REQ("SELECT trailer_number, status, last_date FROM dbo.Tenant_Trailers AS TTr INNER JOIN dbo.Company_List AS cl ON TTr.company_ID=cl.company_ID WHERE TTr.trailer_number=? AND cl.company_name=?", (each[3], each[1]), "S_one")
                if unit_read is not None:
                    if unit_read[1] != each[4]:
                        if unit_read[2] is None: unit_read[2] = datetime(2022, 1, 1, 1, 1, 1)
                        if unit_read[2] < each[0]:
                            comment = str(unit_read[1]) + "-" + str(unit_read[2])
                        else: continue
                    else: continue
                else:
                    unit_read = SQL_REQ("SELECT trailer_number, status, last_date FROM dbo.Tenant_Trailers_UNREG AS TTru INNER JOIN dbo.Company_List AS cl ON TTru.company_ID=cl.company_ID WHERE TTru.trailer_number=? AND cl.company_name=?", (each[3], each[1]), "S_one")
                    if unit_read is not None:
                        if unit_read[1] != each[4]:
                            if unit_read[2] is None: unit_read[2] = datetime(2022, 1, 1, 1, 1, 1)
                            if unit_read[2] < each[0]:
                                comment = str(unit_read[1]) + "-" + str(unit_read[2])
                            else:
                                continue
                        else:
                            continue
                    else:
                        comment = "unregistered"
                        if each[1] == "Euro Can":
                            chk_eu_his = SQL_REQ("SELECT TOP 1 * FROM dbo.Tenant_History AS th INNER JOIN dbo.Company_List AS cl ON th.company_ID=cl.company_ID WHERE th.trailer_number=? AND cl.company_name=? AND th.datetime_event>? ORDER BY th.datetime_event DESC", (each[3], each[1], each[0].strftime("%Y-%m-%d %H:%M:%S")), "S_one")
                            if chk_eu_his: continue
                record = {
                    "Company": each[1],
                    "Truck": None,
                    "Trailer": each[3],
                    "Date": each[0].strftime("%Y/%m/%d"),
                    "Time": each[0].strftime("%H:%M:%S"),
                    "Status": each[4],
                    "Comment": comment
                }
                Tenant_Record(record)
    settings_file_edit("chk_datetime", "None")
    filter_frame.label_config("None")
    # tenant_chk_marker.config(text="None")
    # GN_chk_marker.config(text="None")
    # vis_chk_marker.config(text="None")




def check_generate(func):
    SQL_REQ("DELETE FROM dbo.check_yard", (), "W")
    date = datetime.now().replace(microsecond=0)
    #dateNow = date.strftime("%Y-%m-%d %H:%M:%S")

    tenant_chk_marker.config(text=date)  #### transfer date from here in SQL for checkyard
    GN_chk_marker.config(text=date)
    vis_chk_marker.config(text=date)
    settings_file_edit("chk_datetime", date)
    complist = units_lst("company")


    #Tenant check

    wb = Workbook()
    ws = wb.active
    ws.title = "Tenant"
    ws.page_margins=PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5, header=0.3)
    ws.oddHeader.center.text = "Check Yard"
    ws.oddHeader.right.text = "&P of &N"
    ws.oddHeader.left.text = "&D"
    # ws.print_title_rows = "1:1"
    MaxX = 12
    #MaxY = 66
    X = 1
    Y = 1
    Z = 0
    for companyname in complist:
        truck = units_lst(companyname, "trucks")
        truck.update(units_lst(companyname, "trucks+"))
        trailer = {k: v[0] if isinstance(v, list) else v for k, v in units_lst(companyname, "trailers").items()}
        trailer.update(units_lst(companyname, "trailers+"))
        if not bool(truck) and not bool(trailer): continue
        ws.cell(row=Y, column=X).value = companyname
        ws.cell(row=Y, column=X).font = Font(name="Bahnschrift SemiBold SemiConden", size=12, b=True)
        for col in range(1, MaxX+1):
            ws.cell(row=Y, column=col).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
            if col == 1: ws.cell(row=Y, column=col).border = Border(top=Side(border_style='medium'), left=Side(border_style='medium'))
            elif col == MaxX: ws.cell(row=Y, column=col).border = Border(top=Side(border_style='medium'), right=Side(border_style='medium'))
            else: ws.cell(row=Y, column=col).border = Border(top=Side(border_style='medium'))
        Y+=1
        if bool(truck):
            ws.cell(row=Y, column=X).value = "Trucks:"
            ws.cell(row=Y, column=X).font = Font(name="Bahnschrift SemiBold SemiConden", size=10, b=True)
            ws.cell(row=Y, column=X).border = Border(top=Side(border_style='medium'), left=Side(border_style='medium'), right=Side(border_style='thin'))

            Z = Y
            X+=1

            for trucknumber in sorted(truck.items(), key=lambda x:len(x[0])):
                unit_type="truck"
                SQL_REQ("INSERT INTO dbo.check_yard (date, company, type, unit_number, status) VALUES (?,?,?,?,?)", (date, companyname, unit_type, trucknumber[0], trucknumber[1]), "W")
                ws.cell(row=Y, column=X).value = trucknumber[0]
                ws.cell(row=Y, column=X).font = Font(name="Bahnschrift SemiBold SemiConden", size=8, b=True)
                ws.cell(row=Y, column=X).alignment = Alignment(horizontal='center')
                if X == 2 and Y != Z: ws.cell(row=Y, column=X-1).border = Border(left=Side(border_style='medium'))
                ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='thin'))
                if X == MaxX:
                    ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='medium'))
                    Y+=1
                    X=2
                else: X+=1
            if X!=2:
                while X != MaxX+1:
                    ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='thin'))
                    if X == MaxX: ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='medium'))
                    X+=1
            else: Y-=1
            for col in range(2, MaxX + 1):
                ws.cell(row=Z, column=col).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
                if col == MaxX: ws.cell(row=Z, column=col).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='thin'), right=Side(border_style='medium'), left=Side(border_style='thin'))

            X=1
            Y+=1
        if bool(trailer):
            ws.cell(row=Y, column=X).value = "Trailers:"
            ws.cell(row=Y, column=X).font = Font(name="Bahnschrift SemiBold SemiConden", size=10, b=True)
            ws.cell(row=Y, column=X).border = Border(top=Side(border_style='medium'), left=Side(border_style='medium'), right=Side(border_style='thin'))
            Z = Y
            X += 1
            for trailernumber in sorted(trailer.items(), key=lambda x:len(x[0])):

                unit_type="trailer"
                SQL_REQ("INSERT INTO dbo.check_yard (date, company, type, unit_number, status) VALUES (?,?,?,?,?)", (date, companyname, unit_type, trailernumber[0], trailernumber[1]), "W")
                if X == 2 and Y != Z: ws.cell(row=Y, column=X - 1).border = Border(left=Side(border_style='medium'))
                ws.cell(row=Y, column=X).value = trailernumber[0]
                ws.cell(row=Y, column=X).font = Font(name="Bahnschrift SemiBold SemiConden", size=8, b=True)
                ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='thin'))
                ws.cell(row=Y, column=X).alignment = Alignment(horizontal='center')
                if X == MaxX:
                    ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='medium'))
                    Y += 1
                    X = 2
                else:
                    X += 1
            if X!= 2:
                while X != MaxX+1:
                    ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='thin'))
                    if X == MaxX: ws.cell(row=Y, column=X).border = Border(left=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'), right=Side(border_style='medium'))
                    X += 1
            else: Y-=1
            for col in range(2, MaxX + 1):
                ws.cell(row=Z, column=col).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
                if col == MaxX: ws.cell(row=Z, column=col).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='thin'), right=Side(border_style='medium'), left=Side(border_style='thin'))
            X = 1
            Y += 1
    if Y-1!=Z: ws.cell(row=Y-1, column=1).border = Border(bottom=Side(border_style='thin'), left=Side(border_style='medium'))
    else: ws.cell(row=Y-1, column=1).border = Border(bottom=Side(border_style='thin'), left=Side(border_style='medium'), top=Side(border_style='medium'))

    wS = wb.create_sheet(title="GN")
    wb.active = wS
    DEFAULT_FONT.name = "Bahnschrift SemiBold SemiConden"
    DEFAULT_FONT.sz = 10
    DEFAULT_FONT.b = True
    wS.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.3)
    wS.oddHeader.center.text = "Check Yard GNT"
    wS.oddHeader.right.text = "&P of &N"
    wS.oddHeader.left.text = "&D"

    #GN check

    GNtrucks = units_lst("GNtrucks")
    GNtrailers = {k: v[0] if isinstance(v, list) else v for k, v in units_lst("GNtrailers").items()}
    GNfb = {k: v[0] if isinstance(v, list) else v for k, v in units_lst("GNfb").items()}
    Csize = 11
    wS.column_dimensions["A"].width = Csize
    wS.column_dimensions["B"].width = Csize
    wS.column_dimensions["C"].width = Csize
    wS.column_dimensions["D"].width = Csize
    wS.column_dimensions["E"].width = Csize
    wS.column_dimensions["F"].width = Csize
    wS.column_dimensions["G"].width = Csize
    wS.column_dimensions["H"].width = Csize
    wS.column_dimensions["I"].width = Csize
    wS.cell(row=1, column=1).value = "GNT"
    wS.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    wS.cell(row=1, column=1).font = Font(name="Bahnschrift SemiBold SemiConden", size=20, b=True)
    wS.cell(row=2, column=1).value = "Trucks:"
    wS.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    wS.cell(row=2, column=1).font = Font(name="Bahnschrift SemiBold SemiConden", size=18, b=True)
    wS.cell(row=2, column=1).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'), right=Side(border_style='medium'), left=Side(border_style='medium'))
    wS.cell(row=2, column=1).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    wS.cell(row=2, column=4).value = "Trailers:"
    wS.cell(row=2, column=4).alignment = Alignment(horizontal='center')
    wS.cell(row=2, column=4).font = Font(name="Bahnschrift SemiBold SemiConden", size=18, b=True)
    wS.cell(row=2, column=4).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'), right=Side(border_style='medium'), left=Side(border_style='medium'))
    wS.cell(row=2, column=4).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    wS.cell(row=2, column=7).value = "Flatbeds:"
    wS.cell(row=2, column=7).alignment = Alignment(horizontal='center')
    wS.cell(row=2, column=7).font = Font(name="Bahnschrift SemiBold SemiConden", size=18, b=True)
    wS.cell(row=2, column=7).border = Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'), right=Side(border_style='medium'), left=Side(border_style='medium'))
    wS.cell(row=2, column=7).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    wS.merge_cells('A1:I1')
    wS.merge_cells('D2:F2')
    wS.merge_cells('G2:I2')
    wS.merge_cells('A2:C2')

    row_coord1 = 3
    column_coord1 = 1
    for GNTtruck in sorted(GNtrucks.items(), key=lambda x:len(x[0])):
        type="truck"
        SQL_REQ("INSERT INTO dbo.check_yard (date, company, type, unit_number, status) VALUES (?,?,?,?,?)", (date, "GN", type, GNTtruck[0], GNTtruck[1]), "W")
        wS.cell(row=row_coord1, column=column_coord1).value = GNTtruck[0]
        wS.cell(row=row_coord1, column=column_coord1).border = Border(bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
        wS.cell(row=row_coord1, column=column_coord1).font = Font(name="Bahnschrift SemiBold SemiConden", size=14, b=True)
        wS.cell(row=row_coord1, column=column_coord1).alignment = Alignment(horizontal='center')
        if column_coord1 == 3:
            column_coord1 = 1
            row_coord1+=1
        else:
            column_coord1+=1
    row_coord2 = 3
    column_coord2 = 4
    for GNTtrailer in sorted(GNtrailers.items(), key=lambda x:len(x[0])):
        type="trailer"
        SQL_REQ("INSERT INTO dbo.check_yard (date, company, type, unit_number, status) VALUES (?,?,?,?,?)", (date, "GN", type, GNTtrailer[0], GNTtrailer[1]), "W")
        wS.cell(row=row_coord2, column=column_coord2).value = GNTtrailer[0]
        wS.cell(row=row_coord2, column=column_coord2).border = Border(bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
        wS.cell(row=row_coord2, column=column_coord2).font = Font(name="Bahnschrift SemiBold SemiConden", size=14, b=True)
        wS.cell(row=row_coord2, column=column_coord2).alignment = Alignment(horizontal='center')
        if column_coord2 == 6:
            column_coord2 = 4
            row_coord2+=1
        else:
            column_coord2+=1
    row_coord3 = 3
    column_coord3 = 7
    for GNTfb in sorted(GNfb.items(), key=lambda x:len(x[0])):
        type = "flatbed"
        SQL_REQ("INSERT INTO dbo.check_yard (date, company, type, unit_number, status) VALUES (?,?,?,?,?)", (date, "GN", type, GNTfb[0], GNTfb[1]), "W")
        wS.cell(row=row_coord3, column=column_coord3).value = GNTfb[0]
        wS.cell(row=row_coord3, column=column_coord3).border = Border(bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
        wS.cell(row=row_coord3, column=column_coord3).font = Font(name="Bahnschrift SemiBold SemiConden", size=14, b=True)
        wS.cell(row=row_coord3, column=column_coord3).alignment = Alignment(horizontal='center')
        if column_coord3 == 9:
            column_coord3 = 7
            row_coord3+=1
        else:
            column_coord3+=1
    minL = min(row_coord1, row_coord2, row_coord3)
    maxL = max(row_coord1, row_coord2, row_coord3)
    R = minL
    C = 1
    while True:
        wS.cell(row=R,  column=C).border = Border(bottom=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'))
        if C == 9:
            C = 1
            R+=1
        else:
            C+=1
        if R > maxL: break

    #Car parking check

    WS = wb.create_sheet(title="Car Parking")
    wb.active = WS
    WS.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.3)
    WS.oddHeader.center.text = "Car Parking Check"
    WS.oddHeader.right.text = "&P of &N"
    WS.oddHeader.left.text = "&D"

    carlist=list()
    for companyname in complist:
        main = cars(0, companyname, "main")
        main.extend(cars(0, companyname, "unreg"))
        if len(main) != 0:
            for item in main:

                if item["private"] is not None:
                    if item["private"]:
                        prv = item["expiration"]
                    elif not item["private"]:
                        car_amount = SQL_REQ("SELECT car FROM dbo.Company_List WHERE company_name=?", (companyname,), "S_one")
                        prv = car_amount[0]
                    carlist.append([item['plates'], item['car_model'], item["driver_name"], companyname, prv])
    carlist.sort(key=lambda x: x[0])
    Carsize=20
    WS.column_dimensions["A"].width = Carsize
    WS.column_dimensions["B"].width = Carsize
    WS.column_dimensions["C"].width = Carsize+3
    WS.column_dimensions["D"].width = Carsize
    WS.column_dimensions["E"].width = Carsize

    WS.cell(row=1, column=1).value = "Car Parking"
    WS.cell(row=1, column=1).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    WS.merge_cells("A1:E1")
    WS.cell(row=2, column=1).value = "Plate:"
    WS.cell(row=2, column=1).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=2, column=1).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
    WS.cell(row=2, column=1).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    WS.cell(row=2, column=1).alignment = Alignment(horizontal='center')
    WS.cell(row=2, column=2).value = "Car:"
    WS.cell(row=2, column=2).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=2, column=2).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
    WS.cell(row=2, column=2).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    WS.cell(row=2, column=2).alignment = Alignment(horizontal='center')
    WS.cell(row=2, column=3).value = "Driver:"
    WS.cell(row=2, column=3).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=2, column=3).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
    WS.cell(row=2, column=3).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    WS.cell(row=2, column=3).alignment = Alignment(horizontal='center')
    WS.cell(row=2, column=4).value = "Company:"
    WS.cell(row=2, column=4).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=2, column=4).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
    WS.cell(row=2, column=4).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    WS.cell(row=2, column=4).alignment = Alignment(horizontal='center')
    WS.cell(row=2, column=5).value = "Spot/EXP.Date:"
    WS.cell(row=2, column=5).font = Font(name="Bahnschrift SemiBold SemiConden", size=16, b=True)
    WS.cell(row=2, column=5).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
    WS.cell(row=2, column=5).fill = PatternFill(fill_type="solid", start_color="B8CCE4")
    WS.cell(row=2, column=5).alignment = Alignment(horizontal='center')
    row1 = 3
    colmn = 1
    for line in carlist:
        for each in line:
            if colmn==5 and each is not None:
                a = each

                if not isinstance(a, int):
                    if date.date()>each:
                        WS.cell(row=row1, column=colmn).fill = PatternFill(fill_type="darkUp", start_color="FF0000")
            WS.cell(row=row1, column=colmn).value = each
            WS.cell(row=row1, column=colmn).font = Font(name="Bahnschrift SemiBold SemiConden", size=14, b=True)
            WS.cell(row=row1, column=colmn).border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'), left=Side(border_style='thin'), bottom=Side(border_style='thin'))
            WS.cell(row=row1, column=colmn).alignment = Alignment(horizontal='center')
            colmn+=1
        row1+=1
        colmn=1
    wb.active=ws

    isExist = os.path.exists(sets["chk_path"])
    if not isExist:
        os.makedirs(sets["chk_path"])
    wb.save(sets["chk_path"] + "CheckYard " + date.strftime("%Y") + date.strftime("%m") + date.strftime("%d") + ".xlsx")





def check_print(trigger):
    checkdatetime = settings_file()["chk_datetime"]
    if checkdatetime == "None": return
    check_date = datetime.strptime(checkdatetime, "%Y-%m-%d %H:%M:%S")
    filepath = sets["chk_path"].replace("\\\\", "/")+"CheckYard "+check_date.date().strftime("%Y%m%d")+".xlsx"
    excel = win32.DispatchEx('Excel.Application')
    wb = excel.Workbooks.Open(filepath)
    sheets = wb.Sheets
    for sheet in sheets:
        sheet.PrintOut()
    wb.Close(False)
    excel.Quit()




def VIS_BUTTON(UNIT):
    global Current_Visitor_Unit
    global VIS_Company_Var
    if Current_Visitor_Unit != UNIT:
        if Current_Visitor_Unit is not None:
            for i in range(len(Current_Visitor_Unit)-1): Current_Visitor_Unit[i].configure(bg=conf["widget_bg"])
        Current_Visitor_Unit = UNIT
        UNIT[0].configure(bg=conf["widget_sel_fg"])
        VIS_Plates_Entry.delete(0, tk.END)
        VIS_Plates_Entry.insert(0, UNIT[5].get("plates"))
        VIS_Car_Entry.delete(0, tk.END)
        if UNIT[5].get("car_model") is not None:  VIS_Car_Entry.insert(0, UNIT[5].get("car_model"))
        VIS_Name_Entry.delete(0, tk.END)
        if UNIT[5].get("driver_name") is not None: VIS_Name_Entry.insert(0, UNIT[5].get("driver_name"))

def ADM_VIS_BUTTON(UNIT):
    global Current_Adm_Visitor_Unit
    global Adm_Vis_Company_Var
    global adm_vis_radio_var
    if Current_Adm_Visitor_Unit != UNIT:
        if Current_Adm_Visitor_Unit is not None:
            for i in range(len(Current_Adm_Visitor_Unit)-2): Current_Adm_Visitor_Unit[i].configure(bg=conf["widget_bg"])
        Current_Adm_Visitor_Unit = UNIT
        UNIT[0].configure(bg=conf["widget_sel_fg"])
        adm_Vis_t_entry.delete(0, tk.END)
        adm_Vis_t_entry.insert(0, UNIT[5].get("plates"))
        adm_Vis_car_entry.delete(0, tk.END)
        if UNIT[5].get("car_model") is not None:  adm_Vis_car_entry.insert(0, UNIT[5].get("car_model"))
        adm_Vis_n_entry.delete(0, tk.END)
        if UNIT[5].get("driver_name") is not None: adm_Vis_n_entry.insert(0, UNIT[5].get("driver_name"))
        adm_Vis_exp_entry.delete(0, tk.END)
        if UNIT[5].get("private") is True:
            adm_vis_radio_var.set("private")
            adm_vis_private()
            if UNIT[5].get("expiration") is not None: adm_Vis_exp_entry.insert(0, UNIT[5].get("expiration"))
        elif UNIT[5].get("private") is False:
            adm_vis_radio_var.set("com")
            adm_vis_com()
        elif UNIT[5].get("private") is None:
            adm_vis_radio_var.set("no")
            adm_vis_no()



def VIS_HOVER_ON(UNIT):
    for i in range(len(UNIT)-1):
        UNIT[i].configure(bg=conf["widget_sel_bg"])
def ADM_VIS_HOVER_ON(UNIT):
    for i in range(len(UNIT)-2):
        UNIT[i].configure(bg=conf["widget_sel_bg"])
def VIS_HOVER_OFF(UNIT):
    global Current_Visitor_Unit
    if Current_Visitor_Unit != UNIT:
        for i in range(len(UNIT)-1):
            UNIT[i].configure(bg=conf["widget_bg"])
def ADM_VIS_HOVER_OFF(UNIT):
    global Current_Adm_Visitor_Unit
    if Current_Adm_Visitor_Unit != UNIT:
        for i in range(len(UNIT)-2):
            UNIT[i].configure(bg=conf["widget_bg"])




#reset function
def Admin_VIS_RESET(*args):
    UNTS(Adm_Vis_Company_obj, Adm_Vis_Company_Var, "Admin_Vis_Co")





def _unit_vis_(master,x, func):
    if func == "main" or func == "UNREG":
        frame = tk.Frame(master, bg=conf["widget_bg"], highlightthickness=0)
        if func == "UNREG": frame.config(highlightthickness=1, highlightbackground=conf["UNREG_bg"])
        frame.pack(side=tk.TOP, pady=(1, 0), fill=tk.X)
        plates = tk.Label(frame, text=x.get("plates"), background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=20)
        plates.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, expand=1)
        car = tk.Label(frame, text=x.get("car_model"), background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=23)
        car.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, expand=1)
        name = tk.Label(frame, text=x.get("driver_name"), background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=23)
        name.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, expand=1)
        exp = tk.Label(frame, text=x.get("expiration"), background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=30)
        exp.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, expand=1)
        y = x.get("expiration")
        if y != None:
            exp_date = datetime.strptime(str(x.get("expiration")), "%Y-%m-%d")
            today = datetime.now()
            if today.date() > exp_date.date(): exp.configure(fg=conf["expired_date"])
        if x.get("status"):
            plates.configure(foreground=conf["on_parking"])
            car.configure(foreground=conf["on_parking"])
            name.configure(foreground=conf["on_parking"])
            exp.configure(foreground=conf["on_parking"])
        UNIT = [frame, plates, car, name, exp, x]
        frame.bind("<Button-1>", lambda z: VIS_BUTTON(UNIT))
        plates.bind("<Button-1>", lambda z: VIS_BUTTON(UNIT))
        car.bind("<Button-1>", lambda z: VIS_BUTTON(UNIT))
        name.bind("<Button-1>", lambda z: VIS_BUTTON(UNIT))
        exp.bind("<Button-1>", lambda z: VIS_BUTTON(UNIT))
        frame.bind("<Enter>", lambda zy: VIS_HOVER_ON(UNIT))
        frame.bind("<Leave>", lambda zyx: VIS_HOVER_OFF(UNIT))
        vis_car_canv.update()
        check_T_Vc_scroll_region()
        vis_car_canv.yview_moveto(0.0)
    elif func == "main_vis" or func == "UNREG_vis":
        frame = tk.Frame(master, bg=conf["widget_bg"], highlightthickness=0)
        table = False
        if func == "UNREG_vis":
            table = False
            frame.config(highlightthickness=1, highlightbackground=conf["UNREG_bg"])
        elif func == "main_vis": table = True
        frame.pack(side=tk.TOP, pady=(1, 0), fill=tk.X)
        plates = tk.Label(frame, text=x.get("plates"), background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=20)
        plates.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, expand=1)
        car = tk.Label(frame, text=x.get("car_model"), background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=23)
        car.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, expand=1)
        name = tk.Label(frame, text=x.get("driver_name"), background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=23)
        name.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, expand=1)
        exp = tk.Label(frame, text=x.get("expiration"), background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=30)
        exp.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, expand=1)
        y = x.get("expiration")
        if y != None:
            exp_date = datetime.strptime(str(x.get("expiration")), "%Y-%m-%d")
            today = datetime.now()
            if today.date() > exp_date.date(): exp.configure(fg=conf["expired_date"])
        if x.get("status"):
            plates.configure(foreground=conf["on_parking"])
            car.configure(foreground=conf["on_parking"])
            name.configure(foreground=conf["on_parking"])
            exp.configure(foreground=conf["on_parking"])
        UNIT = [frame, plates, car, name, exp, x, table]
        frame.bind("<Button-1>", lambda z: ADM_VIS_BUTTON(UNIT))
        plates.bind("<Button-1>", lambda z: ADM_VIS_BUTTON(UNIT))
        car.bind("<Button-1>", lambda z: ADM_VIS_BUTTON(UNIT))
        name.bind("<Button-1>", lambda z: ADM_VIS_BUTTON(UNIT))
        exp.bind("<Button-1>", lambda z: ADM_VIS_BUTTON(UNIT))
        frame.bind("<Enter>", lambda zy: ADM_VIS_HOVER_ON(UNIT))
        frame.bind("<Leave>", lambda zyx: ADM_VIS_HOVER_OFF(UNIT))



def cars(master, company, func):
    if func == "main" or func == "main_vis": table = "dbo.visitors"
    else: table = "dbo.visitors_UNREG"
    row, col = SQL_REQ(f"SELECT * FROM {table} INNER JOIN dbo.Company_List ON {table}.company_ID=dbo.Company_List.company_ID WHERE dbo.Company_List.company_name=?", (company,), "S_all_D")
    carlist = []
    for x in range(len(row)):
        a = {}
        index = 0
        for y in row[x]:
            z = col[index]
            index += 1
            a.update({z[0]: y})
        carlist.append(a)
    carlist = sorted(carlist, key=lambda x: x["plates"])
    if master == 0: return carlist
    else:
        for x in carlist:
            _unit_vis_(master, x, func)

# Registering tenant trucks and trailers if not in main table
def Tenant_Register_UNREG(comp_ID, unit_number, IN_OUT, event_date, func):
    if func == "truck":
        table = "dbo.Tenant_Trucks_UNREG"
        column = "truck_number"
    elif func == 'trailer':
        table = "dbo.Tenant_Trailers_UNREG"
        column = "trailer_number"
    if unit_number is None: return
    val = SQL_REQ(f"SELECT company_ID FROM {table} WHERE {column}=? AND company_ID=?", (unit_number, str(comp_ID)), "S_all")
    if IN_OUT: b = "1"
    else: b = "0"
    if not val:
        if comp_ID == "11" and b == "0": SQL_REQ(f"DELETE FROM {table} WHERE {column}=? AND company_ID=?", (unit_number, str(comp_ID)), "W")
        else: SQL_REQ(f"INSERT INTO {table} (company_ID, {column}, status, last_date) VALUES (?,?,?,?)", (comp_ID, unit_number, b, event_date), "W")
    else:
        if comp_ID == "11" and b == "0": SQL_REQ(f"DELETE FROM {table} WHERE {column}=? AND company_ID=?", (unit_number, str(comp_ID)), "W")
        else: SQL_REQ(f"UPDATE {table} SET status=?, last_date=? WHERE company_ID=? AND {column}=?", (b, event_date, str(comp_ID), unit_number), "W")

def Vis_Register_UNREG(comp_ID, plate, model, driver, event_date, IN_OUT):
    table = "dbo.visitors_UNREG"
    column = "plates"
    val = SQL_REQ(f"SELECT company_ID FROM {table} WHERE {column}=?", (plate,), "S_all")
    if IN_OUT:
        b = "1"
    else:
        b = "0"
    if not val:
        SQL_REQ(f"INSERT INTO {table} (company_ID, {column}, driver_name, car_model, status, last_date) VALUES (?,?,?,?,?,?)", (comp_ID, plate, driver, model, b, event_date), "W")
    else:
        SQL_REQ(f"UPDATE {table} SET status=?, last_date=? WHERE company_ID=? AND {column}=?", (b, event_date, str(comp_ID), plate), "W")

#REFRESH TENANT WINDOW FUNC
def Refresh(wnd):
    global Company_Var
    global Current_Company_obj
    global Truck_Var
    global Current_Truck_obj
    global Trailer_Var
    global Current_Trailer_obj
    global Tenant_Event
    global Tenant_Comment
    global GN_Other_Carrier_Var
    global GN_Truck_Var
    global GN_Trailer_Fb_Var
    global GN_Event
    global GN_Comment_Var
    global GN_Trigger_LU
    global GN_Trigger
    global GN_Trk_or_Fb_Var
    global VIS_Company_Var
    global Current_Visitor_Company_obj
    global Current_Visitor_Unit
    global Current_GN_Truck_obj
    global Current_GN_Trailer_Fb_obj
    global VIS_Plates_Var
    global VIS_Car_Var
    global VIS_Name_Var
    global VIS_Comment
    global chk_Company_Var
    global Current_chk_Company_obj
    global Adm_GN_Truck_Var
    global Adm_GN_Truck_obj
    global Adm_GN_Trailer_Var
    global Adm_GN_Trailer_obj
    global Adm_GN_Fb_Var
    global Adm_GN_Fb_obj
    global adm_GN_storage_var
    global adm_GN_LU_var
    global Adm_Company_obj
    global Adm_Company_Var
    global Adm_Unit_obj
    global Visitor_Company_Var

    if wnd == "Tenant":
        if Current_Company_obj is not None:
            # try to unhover selected company if choice in same window otherwise Current_Company_obj is true but widget no longer exist.
            try:
                Hover_Off(Current_Company_obj, None, Company_Var)
            except Exception as e:
                error("Error in Refresh T while Hover_Off: ", e)
        Company_Var = None
        Current_Company_obj = None
        if Current_Truck_obj is not None: Hover_Off(Current_Truck_obj, None, Truck_Var)
        Truck_Var = None
        Current_Truck_obj = None
        if Current_Trailer_obj is not None: Hover_Off(Current_Trailer_obj, None, Trailer_Var)
        Trailer_Var = None
        Current_Trailer_obj = None
        Company_Entry.config(state=tk.NORMAL)
        Company_Entry.delete(0, tk.END)
        Company_Entry.config(state=tk.DISABLED)
        Truck_Entry.delete(0, tk.END)
        Trailer_Entry.delete(0, tk.END)
        Comment_Entry.delete(0, tk.END)
        for widgets in second_truck_Frame.winfo_children(): widgets.destroy()
        for widgets in second_trailer_Frame.winfo_children(): widgets.destroy()
        comp_canv.update_idletasks()
        # truck_canv.update_idletasks()
        # check_T_C_scroll_region()
        # check_T_T_scroll_region()
        Menu_Bar_Parking.focus()
    if wnd == "GN":
        if Current_GN_Truck_obj is not None:
            try:
                Hover_Off(Current_GN_Truck_obj, None, GN_Truck_Var)
            except Exception as e:
                error("Error in Refresh GN while Hover_Off: ", e)
            Current_GN_Truck_obj = None
        GN_Truck_Var = None
        if Current_GN_Trailer_Fb_obj is not None:
            Hover_Off(Current_GN_Trailer_Fb_obj, None, GN_Trailer_Fb_Var)
            Current_GN_Trailer_Fb_obj = None
        GN_Trailer_Fb_Var = None
        GN_Trk_or_Fb_Var = None
        GN_Button(1)
        GN_Button(6)
        GN_Entry_Truck.delete(0, tk.END)
        GN_Entry_Trailer_Fb.delete(0, tk.END)
        GN_Comment_Entry.delete(0, tk.END)
        GN_Other_Entry.config(state=tk.NORMAL)
        GN_Other_Entry.delete(0, tk.END)
        GN_Other_Entry.config(state=tk.DISABLED)
        GN_Truck_SubFrame.delete()
        GN_Truck_SubFrame.refresh()
        GN_Trailer_SubFrame.delete()
        GN_Trailer_SubFrame.refresh()
        GN_Flatbed_SubFrame.delete()
        GN_Flatbed_SubFrame.refresh()
        # for widgets in GN_Truck_SubFrame.winfo_children(): widgets.destroy()
        # for widgets in GN_Trailer_SubFrame.winfo_children(): widgets.destroy()
        # for widgets in GN_Flatbed_SubFrame.winfo_children(): widgets.destroy()
        Implement(GN_Truck_SubFrame.frame, "GNtrucks", "GNtrucks", 4, Parking_GN_size)
        Implement(GN_Trailer_SubFrame.frame, "GNtrailers", "GNtrailers", 5, Parking_GN_size)
        Implement(GN_Flatbed_SubFrame.frame, "GNfb", "GNflatbeds", 6, Parking_GN_size)
        Menu_Bar_Parking.focus()
    if wnd == "Visitor":
        if Current_Visitor_Company_obj is not None:
            try:
                Current_Visitor_Company_obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])
            except: pass
        VIS_Company_Var = None
        Current_Visitor_Company_obj = None
        Current_Visitor_Unit = None
        VIS_Plates_Var = None
        VIS_Car_Var = None
        VIS_Name_Var = None
        VIS_Comment = None
        Visitor_Company_Var = None
        VIS_Company_Entry.delete(0, tk.END)
        VIS_Plates_Entry.delete(0, tk.END)
        VIS_Car_Entry.delete(0, tk.END)
        VIS_Name_Entry.delete(0, tk.END)
        VIS_Comment_Entry.delete(0, tk.END)
        for widgets in Vis_Second_Frame.winfo_children(): widgets.destroy()
        Menu_Bar_Parking.focus()
    if wnd == "CheckYard":
        if Current_chk_Company_obj is not None:
            try:
                Current_chk_Company_obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])
            except: pass
        Current_chk_Company_obj = None
        chk_Company_Var = None
    if wnd == "adm_GN":
        for widget in Admin_GN_T_Scroll.frame.winfo_children(): widget.destroy()
        for widget in Admin_GN_Tr_Scroll.frame.winfo_children(): widget.destroy()
        for widget in Admin_GN_Fb_Scroll.frame.winfo_children(): widget.destroy()
        Adm_GN_Truck_Var = None
        Adm_GN_Truck_obj = None
        Adm_GN_Trailer_Var = None
        Adm_GN_Trailer_obj = None
        Adm_GN_Fb_Var = None
        Adm_GN_Fb_obj = None
        adm_GN_storage_var.set(0)
        adm_GN_LU_var.set(0)
        adm_GN_t_entry.delete(0, tk.END)
        Implement(Admin_GN_T_Scroll.frame, "GNtrucks", "Admin_GN_Truck", 13, adm_gn_scrn_size)
        Implement(Admin_GN_Tr_Scroll.frame, "GNtrailers", "Admin_GN_Trailer", 14, adm_gn_scrn_size)
        Implement(Admin_GN_Fb_Scroll.frame, "GNfb", "Admin_GN_Flatbed", 15, adm_gn_scrn_size)
        adm_GN_t_entry.focus_set()
    if wnd == "adm_T":
        Admin_Tenant_Scroll.delete()
        Implement(Admin_Tenant_Scroll.frame, "company", "Admin_Units", 10, None)
        Adm_Company_obj = None  # Null company object to avoid error with configuration none existing Label
        Adm_Company_Var = None
        Admin_Tenant_T_Scroll.delete()
        Admin_Tenant_Tr_Scroll.delete()
        Admin_Tenant_Scroll.refresh()
        adm_c_entry.delete(0, tk.END)
        adm_t_entry.delete(0, tk.END)
        Adm_Unit_obj = None
        Adm_Truck_Var = None
        Adm_Trailer_Var = None
        adm_T_entry_focus()

def Tabs_Refresh(event):
    if H_Tenant_Main: H_Tenant_Main.pack_forget()
    if H_GN_Main: H_GN_Main.pack_forget()
    if H_VISITOR_Main: H_VISITOR_Main.pack_forget()
# def history_read(wnd):
#     if wnd == "Tenant":
#         path_ini = ("history.txt")
#         if os.path.exists(path_ini):
#             history = open(path_ini)
#             return_history = [v for v in (s.split("|") for s in history.read().splitlines())]
#             return return_history
#     if wnd == "GN":
#         path_ini = ("GN_history.txt")
#         if os.path.exists(path_ini):
#             history = open(path_ini)
#             return_history = [v for v in (s.split("|") for s in history.read().splitlines())]
#             return return_history
#     if wnd == "Visitor":
#         path_ini = ("history_vis.txt")
#         if os.path.exists(path_ini):
#             history = open(path_ini)
#             return_history = [v for v in (s.split("|") for s in history.read().splitlines())]
#             return return_history



def units_lst(query, func=None):
    # SQL QUERRIES
    # extract company data in format: full - list(name, id, list(Dis,Reg,Trl, Trk, car); insurance); D - list(name, id); None - list(name)
    if query == "company":
        if func == "D0": req = SQL_REQ("SELECT * FROM dbo.Company_List ORDER BY company_name", (), "S_all")
        else:  req = SQL_REQ("SELECT * FROM dbo.Company_List WHERE activity=1 ORDER BY company_name", (), "S_all")
        if func == "full":
            company_list = [[row[1], row[0], [row[2], row[3], row[4], row[5], row[6]], row[8]] for row in req]
        elif func == "D" or func == "D0": company_list = [[row[1], row[0]] for row in req]
        else:
            company_list = [row[1] for row in req]
        return company_list
    elif query == "GNtrucks":
            dict = {}
            row = SQL_REQ("SELECT truck_number, status FROM dbo.GN_Trucks ORDER BY len(truck_number), truck_number", (), "S_all")
            for x in row:
                dict.update({x[0]: x[1]})
            return dict
    elif query == "GNtrailers":
            dict = {}
            row = SQL_REQ("SELECT trailer_number, status, storage, LU FROM dbo.GN_Trailers ORDER BY len(trailer_number), trailer_number", (), "S_all")
            for x in row:
                dict.update({x[0]: list((x[1], x[2], x[3]))})
            return dict
    elif query == "GNfb":
            dict = {}
            row = SQL_REQ("SELECT fb_number, status, storage, LU FROM dbo.GN_Flatbed ORDER BY len(fb_number),fb_number", (), "S_all")
            for x in row:
                dict.update({x[0]: list((x[1], x[2], x[3]))})
            return dict
    else:
        if func == "trucks":
            dict = {}
            row = SQL_REQ("SELECT truck_number, status FROM dbo.Tenant_Trucks INNER JOIN dbo.Company_List ON dbo.Tenant_Trucks.company_ID=dbo.Company_List.company_ID WHERE dbo.Company_List.company_name=? ORDER BY len(truck_number), truck_number", (query,), "S_all")
            for x in row:
                dict.update({x[0]: x[1]})
            return dict
        elif func == "trailers":
            dict = {}
            row = SQL_REQ("SELECT trailer_number, status, storage FROM dbo.Tenant_Trailers INNER JOIN dbo.Company_List ON dbo.Tenant_Trailers.company_ID=dbo.Company_List.company_ID WHERE dbo.Company_List.company_name=? ORDER BY len(trailer_number), trailer_number", (query,), "S_all")
            for x in row:
                dict.update({x[0]: list((x[1], x[2]))})
            return dict
        elif func == "trucks+":
            dict = {}
            row = SQL_REQ("SELECT truck_number, status FROM dbo.Tenant_Trucks_UNREG INNER JOIN dbo.Company_List ON dbo.Tenant_Trucks_UNREG.company_ID=dbo.Company_List.company_ID WHERE dbo.Company_List.company_name=? ORDER BY len(truck_number), truck_number", (query,), "S_all")
            for x in row:
                dict.update({x[0]: x[1]})
            return dict
        elif func == "trailers+":
            dict = {}
            row = SQL_REQ("SELECT trailer_number, status FROM dbo.Tenant_Trailers_UNREG INNER JOIN dbo.Company_List ON dbo.Tenant_Trailers_UNREG.company_ID=dbo.Company_List.company_ID WHERE dbo.Company_List.company_name=? ORDER BY len(trailer_number), trailer_number", (query,), "S_all")
            for x in row:
                dict.update({x[0]: x[1]})
            return dict
        elif func == "check_yard_T_T" or func == "check_yard_T_Tr" or func == "check_yard_GN_T" or func == "check_yard_GN_Tr" or func == "check_yard_GN_Fb":
            dict = list()
            row = SQL_REQ("SELECT date, type, unit_number, status FROM dbo.check_yard WHERE company=? ORDER BY len(unit_number), unit_number", (query,), "S_all")
            for x in row:
                dict.append([x[2], x[1], x[3], x[0], query])
            return dict
        elif func == "adm_trucks":
            dict = list()
            row = SQL_REQ("SELECT truck_number, dbo.Company_List.company_ID FROM dbo.Tenant_Trucks INNER JOIN dbo.Company_List ON dbo.Tenant_Trucks.company_ID=dbo.Company_List.company_ID WHERE dbo.Company_List.company_name=? ORDER BY len(truck_number), truck_number", (query), "S_all")
            for x in row: dict.append([x[0], x[1], "REG"])
            return dict
        elif func == "adm_trailers":
            dict = list()
            row = SQL_REQ("SELECT trailer_number, dbo.Company_List.company_ID FROM dbo.Tenant_Trailers INNER JOIN dbo.Company_List ON dbo.Tenant_Trailers.company_ID=dbo.Company_List.company_ID WHERE dbo.Company_List.company_name=? ORDER BY len(trailer_number), trailer_number", (query), "S_all")
            for x in row: dict.append([x[0], x[1], "REG"])
            return dict
        elif func == "adm_trucks+":
            dict = list()
            row = SQL_REQ("SELECT truck_number, dbo.Company_List.company_ID FROM dbo.Tenant_Trucks_UNREG INNER JOIN dbo.Company_List ON dbo.Tenant_Trucks_UNREG.company_ID=dbo.Company_List.company_ID WHERE dbo.Company_List.company_name=? ORDER BY len(truck_number), truck_number", (query), "S_all")
            for x in row: dict.append([x[0], x[1], "UNREG"])
            return dict
        elif func == "adm_trailers+":
            dict = list()
            row = SQL_REQ("SELECT trailer_number, dbo.Company_List.company_ID FROM dbo.Tenant_Trailers_UNREG INNER JOIN dbo.Company_List ON dbo.Tenant_Trailers_UNREG.company_ID=dbo.Company_List.company_ID WHERE dbo.Company_List.company_name=? ORDER BY len(trailer_number), trailer_number", (query), "S_all")
            for x in row: dict.append([x[0], x[1], "UNREG"])
            return dict
        # funct for output data from Tenant SQL in format dict {Company_Name: {trucks:list(unit_number, date, status...), trailers:list(unit_number, date, status...)}, ... }
        elif func == "tenant_by_comp":
            company_list = units_lst("company", "D")
            company_list.sort()
            dict = {}
            for comp in company_list:
                trucks = SQL_REQ("SELECT * FROM dbo.Tenant_Trucks WHERE Company_ID=? ORDER BY truck_number", (comp[1],), "S_all")
                trucks_UNREG = SQL_REQ("SELECT * FROM dbo.Tenant_Trucks_UNREG WHERE Company_ID=? ORDER BY truck_number", (comp[1],), "S_all")
                all_trucks = [list(a) for a in trucks + trucks_UNREG]
                trailers = SQL_REQ("SELECT * FROM dbo.Tenant_Trailers WHERE Company_ID=? ORDER BY trailer_number", (comp[1],), "S_all")
                trailers_UNREG = SQL_REQ("SELECT * FROM dbo.Tenant_Trailers_UNREG WHERE Company_ID=? ORDER BY trailer_number", (comp[1],), "S_all")
                all_trailers = [list(a) for a in trailers + trailers_UNREG]
                dict = {comp[0]: [sorted(all_trucks, key=lambda x: x[1]), sorted(all_trailers, key=lambda x: x[1])]}
                return dict
            #
            #
            #
            #
            #
            #
            #
            # func to get data from sql in format dict company: dict truck: list(unit, date, status...), trailer: list(unit, date, status...)











#HOVER OFF FUNCTION - RETURN ORIGINAL BG COLOR IF NOT CLICKED
def Hover_Off(obj, var, current_select):
    if type(current_select) is not tuple:
        if var != current_select:
            try:
                obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])
            except Exception as e:
                error(["Error in Hover_Off ", e])
                debuger("obj.configure(bg=conf['widget_bg'], fg=conf['widget_fg']) where obj="+str(obj)+"; var="+str(var)+"; current_select="+str(current_select))
    else:
        if var is None:
            if type(current_select[1]) is not list:
                if current_select[1]: obj.configure(bg=conf["widget_bg"], fg=conf["on_parking"])
                else: obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])
            else:
                if current_select[1][1]: obj.configure(bg=conf["widget_bg"], fg=conf["storage_fg"])
                else:
                    if current_select[1][0]: obj.configure(bg=conf["widget_bg"], fg=conf["on_parking"])
                    else: obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])
        else:
            if var[0] != current_select[0]:
                if type(current_select[1]) is not list:
                    if current_select[1]:
                        obj.configure(bg=conf["widget_bg"], fg=conf["on_parking"])
                    else:
                        obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])
                else:
                    if current_select[1][1]:
                        obj.configure(bg=conf["widget_bg"], fg=conf["storage_fg"])
                    else:
                        if current_select[1][0]:
                            obj.configure(bg=conf["widget_bg"], fg=conf["on_parking"])
                        else:
                            obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])

def Hover_Off_Adm_Ten(obj, var, current_select):
    if var is None: obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])
    else:
        if var[0] != current_select[0]: obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])

def Hover_Off_Adm_GN(obj, var, current_select):
    if var is None: obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])
    else:
        if var[0] != current_select[0]: obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])

#HOVER ON FUN FOR GN CARRIER FEATURE BUTTONS
def GN_Button(trigger):
    global GN_Trigger
    global GN_Trigger_LU
    global GN_Other_Carrier_Var
    if trigger == 1:
        GN_Trigger = 1
        GN_Other_Carrier_Var = "GN"
        GN_Button_lb.configure(relief=tk.SUNKEN, fg=conf["func_button_sel_fg"], bg=conf["func_button_sel_bg"])
        GN_WTube_lb.configure(relief=tk.GROOVE, fg=conf["func_button_fg"], bg=conf["func_button_bg"])
        GN_Other_lb.configure(relief=tk.GROOVE, fg=conf["func_button_fg"], bg=conf["func_button_bg"])
        GN_Other_Entry.delete(0, tk.END)
        GN_Other_Entry.configure(state=tk.DISABLED)
    elif trigger == 2:
        GN_Trigger = 2
        GN_Other_Carrier_Var = "WT"
        GN_Button_lb.configure(relief=tk.GROOVE, fg=conf["func_button_fg"], bg=conf["func_button_bg"])
        GN_WTube_lb.configure(relief=tk.SUNKEN, fg=conf["func_button_sel_fg"], bg=conf["func_button_sel_bg"])
        GN_Other_lb.configure(relief=tk.GROOVE, fg=conf["func_button_fg"], bg=conf["func_button_bg"])
        GN_Other_Entry.delete(0, tk.END)
        GN_Other_Entry.configure(state=tk.DISABLED)
    elif trigger == 3:
        GN_Trigger = 3
        GN_Other_Carrier_Var = None
        GN_Button_lb.configure(relief=tk.GROOVE, fg=conf["func_button_fg"], bg=conf["func_button_bg"])
        GN_WTube_lb.configure(relief=tk.GROOVE, fg=conf["func_button_fg"], bg=conf["func_button_bg"])
        GN_Other_lb.configure(relief=tk.SUNKEN, fg=conf["func_button_sel_fg"], bg=conf["func_button_sel_bg"])
        GN_Other_Entry.configure(state=tk.NORMAL)
    if trigger == 4:
        GN_Trigger_LU = 4
        GN_Load_lb.configure(relief=tk.SUNKEN, fg=conf["func_button_sel_fg"], bg=conf["func_button_sel_bg"])
        GN_Unload_lb.configure(relief=tk.GROOVE, fg=conf["func_button_fg"], bg=conf["func_button_bg"])
    elif trigger == 5:
        GN_Trigger_LU = 5
        GN_Load_lb.configure(relief=tk.GROOVE, fg=conf["func_button_fg"], bg=conf["func_button_bg"])
        GN_Unload_lb.configure(relief=tk.SUNKEN, fg=conf["func_button_sel_fg"], bg=conf["func_button_sel_bg"])
    elif trigger == 6:
        GN_Trigger_LU = None
        GN_Load_lb.configure(relief=tk.GROOVE, fg=conf["func_button_fg"], bg=conf["func_button_bg"])
        GN_Unload_lb.configure(relief=tk.GROOVE, fg=conf["func_button_fg"], bg=conf["func_button_bg"])

#HOVER OFF FUNC FOR GN CARRIER FEATYRE BUTTONS
def GN_Hover_Off (obj, val, func):
    if func == "CARRIER":
        if GN_Trigger != val: obj.configure(fg=conf["func_button_fg"], bg=conf["func_button_bg"])
    if func == "LU":
        if GN_Trigger_LU != val:
            obj.configure(fg=conf["func_button_fg"], bg=conf["func_button_bg"])
        else:
            obj.configure(fg=conf["func_button_sel_fg"], bg=conf["func_button_sel_bg"])


#MOUSE CLICK FUNCTION - RETURN BG OF PREVIOUS SELECTION, SET CLICKED WIDGET IN THE VAR, INSERT DATA IN ENTRY
def UNTS(obj, var, func):
    global Company_Var
    global Current_Company_obj
    global Truck_Var
    global Current_Truck_obj
    global Trailer_Var
    global Current_Trailer_obj
    global GN_Truck_Var
    global Current_GN_Truck_obj
    global GN_Trailer_Fb_Var
    global Current_GN_Trailer_Fb_obj
    global GN_Trk_or_Fb_Var
    global Visitor_Company_Var
    global Current_Visitor_Company_obj
    global chk_Company_Var
    global Current_chk_Company_obj
    global Current_chk_T_T_obj
    global chk_T_T_Var
    global Current_chk_T_Tr_obj
    global chk_T_Tr_Var
    global Current_chk_GN_T_obj
    global chk_GN_T_Var
    global Current_chk_GN_Tr_Fb_obj
    global chk_GN_Tr_Fb_Var
    global Adm_Company_Var
    global Adm_Company_obj
    global Adm_Truck_Var
    global Adm_Unit_obj
    global Adm_Trailer_Var
    global adm_storage_var
    global Adm_GN_Truck_Var
    global Adm_GN_Truck_obj
    global Adm_GN_Trailer_Var
    global Adm_GN_Trailer_obj
    global Adm_GN_Fb_Var
    global Adm_GN_Fb_obj
    global Adm_Vis_Company_Var
    global Adm_Vis_Company_obj
    global Current_Adm_Visitor_Unit

    if func == "trucks":
        if Current_Truck_obj is not None and Current_Truck_obj != obj:
            Current_Truck_obj.config(bg=conf["widget_bg"])
            if Truck_Var[1] is True: Current_Truck_obj.config(fg=conf["on_parking"])
            else: Current_Truck_obj.config(fg=conf["widget_fg"])
        Current_Truck_obj = obj
        Truck_Var = var
        Truck_Entry.delete(0, tk.END)
        Truck_Entry.insert(0, var[0])
    elif func == "trailers":
        if Current_Trailer_obj is not None and Current_Trailer_obj != obj:
            Current_Trailer_obj.config(bg=conf["widget_bg"])
            if type(Trailer_Var[1]) is not list:
                if Trailer_Var[1]: Current_Trailer_obj.config(fg=conf["on_parking"])
                else: Current_Trailer_obj.config(fg=conf["widget_fg"])
            else:
                if Trailer_Var[1][1]: Current_Trailer_obj.config(fg=conf["storage_fg"])
                else:
                    if Trailer_Var[1][0]: Current_Trailer_obj.config(fg=conf["on_parking"])
                    else: Current_Trailer_obj.config(fg=conf["widget_fg"])
        Current_Trailer_obj = obj
        Trailer_Var = var
        Trailer_Entry.delete(0, tk.END)
        Trailer_Entry.insert(0, var[0])
    elif func == "trucks+":
        if Current_Truck_obj is not None and Current_Truck_obj != obj:
            Current_Truck_obj.config(bg=conf["widget_bg"])
            if Truck_Var[1] is True:
                Current_Truck_obj.config(fg=conf["on_parking"])
            else:
                Current_Truck_obj.config(fg=conf["widget_fg"])
        Current_Truck_obj = obj
        Truck_Var = var
        Truck_Entry.delete(0, tk.END)
        Truck_Entry.insert(0, var[0])
    elif func == "trailers+":
        if Current_Trailer_obj is not None and Current_Trailer_obj != obj:
            Current_Trailer_obj.config(bg=conf["widget_bg"])
            if Trailer_Var[1] is True:
                Current_Trailer_obj.config(fg=conf["on_parking"])
            else:
                Current_Trailer_obj.config(fg=conf["widget_fg"])
        Current_Trailer_obj = obj
        Trailer_Var = var
        Trailer_Entry.delete(0, tk.END)
        Trailer_Entry.insert(0, var[0])
    elif func == "companies":
        try:
            if Current_Company_obj is not None and Current_Company_obj != obj:
                Current_Company_obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])
        except Exception as e:
            error(["Error in UNTS: ", e])
            debuger("UNTS>> Current_Company_obj="+str(Current_Company_obj)+", obj="+str(obj))
        Refresh("Tenant")
        Current_Company_obj = obj
        Current_Company_obj.configure(bg=conf["widget_sel_bg"], fg=conf["widget_sel_fg"])
        Company_Var = var
        for widgets in second_truck_Frame.winfo_children(): widgets.destroy()
        for widgets in second_trailer_Frame.winfo_children(): widgets.destroy()
        Implement(second_truck_Frame, var, "trucks", 2, Parking_Tenant_size)
        Implement(second_truck_Frame, var, "trucks+", 2, Parking_Tenant_size)
        truck_canv.update_idletasks()
        check_T_T_scroll_region()
        truck_canv.yview_moveto(0.0)
        Implement(second_trailer_Frame, var, "trailers", 3, Parking_Tenant_size)
        Implement(second_trailer_Frame, var, "trailers+", 3, Parking_Tenant_size)
        trailer_canv.update_idletasks()
        check_T_Tr_scroll_region()
        trailer_canv.yview_moveto(0.0)

        Company_Entry.configure(state=tk.NORMAL)
        Company_Entry.delete(0, tk.END)
        Company_Entry.insert(0, var)
        Company_Entry.configure(state=tk.DISABLED)
        statistics_reg("O")
    elif func == "GNtrucks":
        if Current_GN_Truck_obj is not None and Current_GN_Truck_obj != obj:
            Current_GN_Truck_obj.config(bg=conf["widget_bg"])
            if GN_Truck_Var[1] is True: Current_GN_Truck_obj.config(fg=conf["on_parking"])
            else: Current_GN_Truck_obj.config(fg=conf["widget_fg"])
        Current_GN_Truck_obj = obj
        GN_Truck_Var = var
        GN_Entry_Truck.delete(0, tk.END)
        GN_Entry_Truck.insert(0, var[0])
    elif func == "GNtrailers":
        if Current_GN_Trailer_Fb_obj is not None and Current_GN_Trailer_Fb_obj != obj:
            Current_GN_Trailer_Fb_obj.config(bg=conf["widget_bg"])
            if type(GN_Trailer_Fb_Var[1]) is not list:
                if GN_Trailer_Fb_Var[1] is True: Current_GN_Trailer_Fb_obj.config(fg=conf["on_parking"])
                else: Current_GN_Trailer_Fb_obj.config(fg=conf["widget_fg"])
            else:
                if GN_Trailer_Fb_Var[1][1]: Current_GN_Trailer_Fb_obj.config(fg=conf["storage_fg"])
                else:
                    if GN_Trailer_Fb_Var[1][0] is True: Current_GN_Trailer_Fb_obj.config(fg=conf["on_parking"])
                    else: Current_GN_Trailer_Fb_obj.config(fg=conf["widget_fg"])
        Current_GN_Trailer_Fb_obj = obj
        GN_Trailer_Fb_Var = var
        GN_Trk_or_Fb_Var = True
        GN_Entry_Trailer_Fb.delete(0, tk.END)
        GN_Entry_Trailer_Fb.insert(0, var[0])
    elif func == "GNflatbeds":
        if Current_GN_Trailer_Fb_obj is not None and Current_GN_Trailer_Fb_obj != obj:
            Current_GN_Trailer_Fb_obj.config(bg=conf["widget_bg"])
            if type(GN_Trailer_Fb_Var[1]) is not list:
                if GN_Trailer_Fb_Var[1] is True:
                    Current_GN_Trailer_Fb_obj.config(fg=conf["on_parking"])
                else:
                    Current_GN_Trailer_Fb_obj.config(fg=conf["widget_fg"])
            else:
                if GN_Trailer_Fb_Var[1][1]:
                    Current_GN_Trailer_Fb_obj.config(fg=conf["storage_fg"])
                else:
                    if GN_Trailer_Fb_Var[1][0] is True:
                        Current_GN_Trailer_Fb_obj.config(fg=conf["on_parking"])
                    else:
                        Current_GN_Trailer_Fb_obj.config(fg=conf["widget_fg"])
        Current_GN_Trailer_Fb_obj = obj
        GN_Trailer_Fb_Var = var
        GN_Trk_or_Fb_Var = False
        GN_Entry_Trailer_Fb.delete(0, tk.END)
        GN_Entry_Trailer_Fb.insert(0, var[0])
    elif func == "Visitor":
        try:
            if Current_Visitor_Company_obj is not None and Current_Visitor_Company_obj != obj: Current_Visitor_Company_obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])
        except: debuger("UNTS>> Current_VisitorCompany_obj=" + str(Current_Visitor_Company_obj) + ", obj=" + str(obj))
        Refresh("Visitor")
        Current_Visitor_Company_obj = obj
        Visitor_Company_Var = var
        for widgets in Vis_Second_Frame.winfo_children(): widgets.destroy()
        cars(Vis_Second_Frame, var, "main")
        cars(Vis_Second_Frame, var, "UNREG")
        VIS_Company_Entry.delete(0, tk.END)
        VIS_Company_Entry.insert(0, var)
        vis_car_canv.update_idletasks()
        check_T_Vc_scroll_region()
        vis_car_canv.yview_moveto(0.0)
    elif func == "CheckYard":
        if Current_chk_Company_obj is not None and Current_chk_Company_obj != obj: Current_chk_Company_obj.configure(bg=conf["widget_bg"], fg=conf["widget_fg"])
        #Refresh("CheckYard")
        Current_chk_Company_obj = obj
        chk_Company_Var = var
        chk_c_entry.delete(0, tk.END)
        chk_c_entry.insert(0, var)
        if var is not None:
            if var != "G.N.Transport":
                chk_manual_entry_frame.pack(side=tk.TOP)
                chk_central_GN_frame.pack_forget()
                chk_central_tenant_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
                for widgets in second_chk_T_T_frame.winfo_children(): widgets.destroy()
                for widgets in second_chk_T_Tr_frame.winfo_children(): widgets.destroy()
                Implement(second_chk_T_T_frame, var, "check_yard_T_T", 9, edit_scrn_size_lb)
                Implement(second_chk_T_Tr_frame, var, "check_yard_T_Tr", 9, edit_scrn_size_lb)
            else:
                chk_manual_entry_frame.pack_forget()
                chk_central_tenant_frame.pack_forget()
                chk_central_GN_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
                for widgets in second_chk_GN_T_frame.winfo_children(): widgets.destroy()
                for widgets in second_chk_GN_Tr_frame.winfo_children(): widgets.destroy()
                for widgets in second_chk_GN_fb_frame.winfo_children(): widgets.destroy()
                Implement(second_chk_GN_T_frame, "GN", "check_yard_GN_T", 9, edit_scrn_size_lb)
                Implement(second_chk_GN_Tr_frame, "GN", "check_yard_GN_Tr", 9, edit_scrn_size_lb)
                Implement(second_chk_GN_fb_frame, "GN", "check_yard_GN_Fb", 9, edit_scrn_size_lb)
    elif func == "check_yard_T_T" or func == "check_yard_T_Tr" or func == "check_yard_GN_T" or func == "check_yard_GN_Tr" or func == "check_yard_GN_Fb":
        if settings_file()["chk_datetime"] == "None": return
        if var[2]:
            obj.config(fg=conf["widget_fg"])
            var[2] = False
            chk_set(var)
        else:
            obj.config(fg=conf["on_parking"])
            var[2] = True
            chk_set(var)
    elif func == "Admin_Units":
        if Adm_Company_obj is not None and Adm_Company_obj != obj: Adm_Company_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        #Refresh("Adm_Tenant")
        Adm_Company_obj = obj
        Adm_Company_Var = var
        if Admin_Tenant_T_Scroll.frame.winfo_children(): Admin_Tenant_T_Scroll.delete()
        if Admin_Tenant_Tr_Scroll.frame.winfo_children(): Admin_Tenant_Tr_Scroll.delete()
        adm_t_entry.delete(0, tk.END)
        adm_c_entry.config(state=tk.NORMAL)
        adm_c_entry.delete(0, tk.END)
        adm_c_entry.insert(0, var)
        adm_c_entry.config(state=tk.DISABLED)
        Adm_Truck_Var = None
        Adm_Trailer_Var = None
        Adm_Unit_obj = None
        Implement(Admin_Tenant_T_Scroll.frame, var, "adm_trucks", 11, 8)            #  Replace 8 on function - quantity of labels in row
        Implement(Admin_Tenant_T_Scroll.frame, var, "adm_trucks+", 11, 8)
        Implement(Admin_Tenant_Tr_Scroll.frame, var, "adm_trailers", 12, 8)         #
        Implement(Admin_Tenant_Tr_Scroll.frame, var, "adm_trailers+", 12, 8)
        adm_t_entry.focus_set()
        adm_storage_checkbox.pack_forget()

    elif func == "adm_trucks" or func == "adm_trucks+":
        if Adm_Unit_obj is not None and Adm_Unit_obj != obj: Adm_Unit_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        if Adm_Unit_obj == obj:
            Adm_Unit_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
            Adm_Unit_obj = None
            Adm_Truck_Var = None
            Adm_Trailer_Var = None
            adm_t_entry.delete(0, tk.END)
        else:
            Adm_Unit_obj = obj
            Adm_Truck_Var = var
            adm_t_entry.delete(0, tk.END)
            adm_t_entry.insert(0, Adm_Truck_Var[0])
        adm_radio_var.set("truck")
        adm_storage_checkbox.pack_forget()
    elif func == "adm_trailers" or func == "adm_trailers+":
        if Adm_Unit_obj is not None and Adm_Unit_obj != obj: Adm_Unit_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        if Adm_Unit_obj == obj:
            Adm_Unit_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
            Adm_Unit_obj = None
            Adm_Truck_Var = None
            Adm_Trailer_Var = None
            adm_t_entry.delete(0, tk.END)
            adm_radio_var.set("trailer")
            return
        else:
            Adm_Unit_obj = obj
            Adm_Trailer_Var = var
            adm_t_entry.delete(0, tk.END)
            adm_t_entry.insert(0, Adm_Trailer_Var[0])
        adm_radio_var.set("trailer")
        adm_storage_checkbox.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=15)
        val = SQL_REQ("SELECT company_ID FROM dbo.Company_List WHERE company_name=?", (Adm_Company_Var,), "S_one")
        if val: ID = str(val[0])
        if Adm_Trailer_Var[2] == "REG":
            check = SQL_REQ("SELECT storage FROM dbo.Tenant_Trailers WHERE company_ID=? AND trailer_number=?", (ID, Adm_Trailer_Var[0]), "S_one")
        elif Adm_Trailer_Var[2] == "UNREG":
            check = SQL_REQ("SELECT storage FROM dbo.Tenant_Trailers_UNREG WHERE company_ID=? AND trailer_number=?", (ID, Adm_Trailer_Var[0]), "S_one")
        if check[0] is True: st = 1
        else: st = 0
        adm_storage_var.set(st)

    elif func == "Admin_GN_Truck":
        if Adm_GN_Truck_obj is not None and Adm_GN_Truck_obj != obj: Adm_GN_Truck_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        if Adm_GN_Trailer_obj is not None: Adm_GN_Trailer_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        if Adm_GN_Fb_obj is not None: Adm_GN_Fb_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        if Adm_GN_Truck_obj == obj:
            Adm_GN_Truck_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
            Adm_GN_Truck_obj = None
            Adm_GN_Trailer_obj = None
            Adm_GN_Truck_Var = None
            Adm_GN_Trailer_Var = None
            Adm_GN_Fb_obj = None
            Adm_GN_Fb_Var = None
        else:
            Adm_GN_Truck_obj = obj
            Adm_GN_Truck_Var = var
            Adm_GN_Trailer_obj = None
            Adm_GN_Trailer_Var = None
            Adm_GN_Fb_obj = None
            Adm_GN_Fb_Var = None
        adm_GN_storage_checkbox.pack_forget()
        adm_GN_LU_checkbox.pack_forget()
        adm_GN_city_checkbox.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=15)
        city_check = SQL_REQ("SELECT city FROM dbo.GN_Trucks WHERE truck_number=?", (Adm_GN_Truck_Var[0],), "S_one")
        if city_check:
            adm_GN_city_var.set(city_check[0])

    elif func == "Admin_GN_Trailer":
        if Adm_GN_Trailer_obj is not None and Adm_GN_Trailer_obj != obj: Adm_GN_Trailer_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        if Adm_GN_Fb_obj is not None: Adm_GN_Fb_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        if Adm_GN_Truck_obj is not None: Adm_GN_Truck_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        if Adm_GN_Trailer_obj == obj:
            Adm_GN_Trailer_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
            Adm_GN_Truck_obj = None
            Adm_GN_Trailer_obj = None
            Adm_GN_Truck_Var = None
            Adm_GN_Trailer_Var = None
            Adm_GN_Fb_obj = None
            Adm_GN_Fb_Var = None
        else:
            Adm_GN_Trailer_obj = obj
            Adm_GN_Trailer_Var = var
            Adm_GN_Truck_obj = None
            Adm_GN_Truck_Var = None
            Adm_GN_Fb_obj = None
            Adm_GN_Fb_Var = None

        adm_GN_storage_checkbox.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=15)
        adm_GN_LU_checkbox.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=15)
        adm_GN_storage_var.set(Adm_GN_Trailer_Var[1][1])
        adm_GN_LU_var.set(Adm_GN_Trailer_Var[1][2])

    elif func == "Admin_GN_Flatbed":
        if Adm_GN_Fb_obj is not None and Adm_GN_Fb_obj != obj: Adm_GN_Fb_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        if Adm_GN_Truck_obj is not None: Adm_GN_Truck_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        if Adm_GN_Trailer_obj is not None: Adm_GN_Trailer_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        if Adm_GN_Fb_obj == obj:
            Adm_GN_Fb_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
            Adm_GN_Truck_obj = None
            Adm_GN_Trailer_obj = None
            Adm_GN_Truck_Var = None
            Adm_GN_Trailer_Var = None
            Adm_GN_Fb_obj = None
            Adm_GN_Fb_Var = None
        else:
            Adm_GN_Fb_obj = obj
            Adm_GN_Fb_Var = var
            Adm_GN_Truck_obj = None
            Adm_GN_Truck_Var = None
            Adm_GN_Trailer_obj = None
            Adm_GN_Trailer_Var = None

        adm_GN_storage_checkbox.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=15)
        adm_GN_LU_checkbox.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=15)
        adm_GN_storage_var.set(Adm_GN_Fb_Var[1][1])
        adm_GN_LU_var.set(Adm_GN_Fb_Var[1][2])

    elif func == "Admin_Vis_Co":
        if Adm_Vis_Company_obj is not None and Adm_Vis_Company_obj != obj: Adm_Vis_Company_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        Adm_Vis_Company_Var = var
        Adm_Vis_Company_obj = obj
        for widget in Admin_Vis_Car_Scroll.frame.winfo_children(): widget.destroy()
        cars(Admin_Vis_Car_Scroll.frame, var, "main_vis")
        cars(Admin_Vis_Car_Scroll.frame, var, "UNREG_vis")
        if var is not None:
            adm_Vis_c_entry.config(state=tk.NORMAL)
            adm_Vis_c_entry.delete(0, tk.END)
            adm_Vis_c_entry.insert(0, var)
            adm_Vis_c_entry.config(state=tk.DISABLED)
        adm_Vis_t_entry.delete(0, tk.END)
        adm_Vis_car_entry.delete(0, tk.END)
        adm_Vis_n_entry.delete(0, tk.END)
        Current_Adm_Visitor_Unit = None

def chk_set(var, *args):
    request = SQL_REQ("SELECT * FROM dbo.check_yard WHERE company=? AND unit_number=? AND type=?", (var[4], var[0], var[1]), "S_one")
    if request is not None:
        if len(args)>0:
            if args[0] == "DEL": SQL_REQ("DELETE FROM dbo.check_yard WHERE company=? AND unit_number=? AND type=?", (var[4], var[0], var[1]), "W")
        else: SQL_REQ("UPDATE dbo.check_yard SET status=? WHERE company=? AND unit_number=? AND type=?", (var[2], var[4], var[0], var[1]), "W")
    else:
        if len(args)>0:
            if args[0] == "DEL": error(11)
        else: SQL_REQ("INSERT INTO dbo.check_yard (date,company,type,unit_number,status) VALUES (?,?,?,?,?)", (var[3].strftime("%Y-%m-%d %H:%M:%S"), var[4], var[1], var[0], var[2]), "W")

###RECORD INFO IN TXT FILE - TEMP to replace on SQL
def Tenant_Record(record):

    global security
    val = SQL_REQ("SELECT company_ID FROM dbo.Company_List WHERE company_name=?", (record["Company"],), "S_all")
    if val:
        for v in val: _id = str(v[0])
    else: _id = "999"
    event_date = str(datetime.strptime(record["Date"]+" "+record["Time"],'%Y/%m/%d %H:%M:%S'))
    record_val = [_id, record["Truck"], record["Trailer"], event_date, record["Status"], record["Comment"], security[0]]
    SQL_REQ("INSERT INTO dbo.Tenant_History(company_ID, truck_number, trailer_number, datetime_event, status, comment, full_name) VALUES (?,?,?,?,?,?,?)", record_val, "W")
    Tsts = SQL_REQ("SELECT status FROM dbo.Tenant_Trucks WHERE company_ID=? AND truck_number=?", (_id, str(record["Truck"])),  "S_all")
    if Tsts:
        for w in Tsts: truck_status = str(w[0])
    else:
        truck_status = None
        Tenant_Register_UNREG(_id, record["Truck"], record["Status"], event_date, "truck")  ### function of adding unknown TRUCK in UNREG table
    if record["Trailer"] is not None:
        TRsts = SQL_REQ("SELECT status FROM dbo.Tenant_Trailers WHERE company_ID=? AND trailer_number=?", (_id, record["Trailer"]), "S_all")
        if TRsts:
            for w in TRsts: trailer_status = str(w[0])
        else:
            trailer_status = None
            Tenant_Register_UNREG(_id, record["Trailer"], record["Status"], event_date, "trailer")  ### function of adding unknown TRAILER in UNREG table
    else: trailer_status = None
    if record["Status"] is True:
        if truck_status is True: pass # function to investigate who fucked up
        if truck_status is not None: SQL_REQ("UPDATE dbo.Tenant_Trucks SET status=1, last_date=? WHERE company_ID=? AND truck_number=?", (event_date, _id, record["Truck"]), "W")
        if trailer_status is not None: SQL_REQ("UPDATE dbo.Tenant_Trailers SET status=1, last_date=? WHERE company_ID=? AND trailer_number=?", (event_date, _id, record["Trailer"]), "W")
    else:
        if truck_status is False: pass  # function to investigate who fucked up
        if truck_status is not None: SQL_REQ("UPDATE dbo.Tenant_Trucks SET status=0, last_date=?  WHERE company_ID=? AND truck_number=?", (event_date, _id, record["Truck"]), "W")
        if trailer_status is not None: SQL_REQ("UPDATE dbo.Tenant_Trailers SET status=0, last_date=? WHERE company_ID=? AND trailer_number=?", (event_date, _id, record["Trailer"]), "W")
    #ReCheckTruck = SQL_REQ("SELECT truck, status FROM dbo.Tenant_Truck")

    OVERPARKING(record, "T")


def GN_Record(record):
    global security
    if record["Type"]:
        trailer = record["Trailer"]
        fb = None
    elif record["Type"] is False:
        trailer = None
        fb = record["Trailer"]
    elif record["Type"] is None:
        trailer = None
        fb = None
    event_date = str(datetime.strptime(record["Date"] + " " + record["Time"],'%Y/%m/%d %H:%M:%S'))
    if record["Type"] is None and record["Trailer"] !="":
        record_val = [record["Company"], record["Truck"], record["Trailer"], fb, event_date, record["Cargo"], record["Status"], record["Comment"], security[0]]
    else: record_val = [record["Company"], record["Truck"], trailer, fb, event_date, record["Cargo"], record["Status"], record["Comment"], security[0]]
    SQL_REQ("INSERT INTO dbo.GN_History(company_name, truck_number, trailer_number, fb_number, datetime_event, cargo, status, comment, full_name) VALUES (?,?,?,?,?,?,?,?,?)", record_val, "W")
    if record["Company"] == "GN":
        GNT = SQL_REQ("SELECT status FROM dbo.GN_Trucks WHERE truck_number=?", (str(record["Truck"]),), "S_all")
        if GNT:
            for w in GNT: truck_status = w[0]
            if record["Status"]:
                if truck_status is not None:
                    if truck_status: pass   # function to investigate who fucked up
                SQL_REQ("UPDATE dbo.GN_Trucks SET status=1, last_date=? WHERE truck_number=?", (event_date, record["Truck"]), "W")
            else:
                if truck_status is not None:
                    if not truck_status: pass   # function to investigate who fucked up
                SQL_REQ("UPDATE dbo.GN_Trucks SET status=0, last_date=? WHERE truck_number=?", (event_date, record["Truck"]), "W")
        else:
            #
            # #<================
            # if record["Status"]:
            #     if truck_status is not None:
            #         if truck_status: pass   # function to investigate who fucked up
            #     SQL_REQ('UPDATE dbo.GN_Trucks SET status=1, last_date=\'' + event_date + '\' WHERE truck_number=\'' + record["Truck"] + '\'', "D")
            # else:
            #     if truck_status is not None:
            #         if not truck_status: pass   # function to investigate who fucked up
            #     SQL_REQ('UPDATE dbo.GN_Trucks SET status=0, last_date=\'' + event_date + '\' WHERE truck_number=\'' + record["Truck"] + '\'', "D")
            pass# <== NO TRUCK IN GN LIST - error

        if trailer is not None:
            GNTr = SQL_REQ("SELECT status FROM dbo.GN_Trailers WHERE trailer_number=?", (trailer,), "S_all")
            if GNTr:
                for x in GNTr: trailer_status = x[0]
                if record["Status"]:
                    if trailer_status is not None:
                        if trailer_status: pass # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Trailers SET status=1, last_date=?, LU=? WHERE trailer_number=?", (event_date, str(record["Cargo"]), trailer), "W")
                else:
                    if trailer_status is not None:
                        if not trailer_status: pass # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Trailers SET status=0, storage=0, last_date=?, LU=0 WHERE trailer_number=?", (event_date, trailer), "W")
            else: pass #<==== NO TRAILER IN GN LIST - error
        elif fb is not None:
            GNFb = SQL_REQ("SELECT status FROM dbo.GN_Flatbed WHERE fb_number=?", (fb,), "S_all")
            if GNFb:
                for x in GNFb: fb_status = x[0]
                if record["Status"]:
                    if fb_status is not None:
                            if fb_status: pass  # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Flatbed SET status=1, last_date=?, LU=? WHERE fb_number=?", (event_date, str(record["Cargo"]), fb), "W")
                else:
                    if fb_status is not None:
                        if not fb_status: pass  # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Flatbed SET status=0, storage=0, last_date=?, LU=0 WHERE fb_number=?", (event_date, fb), "W")
            else: pass  # <==== NO TRAILER IN GN LIST - error
        else: pass # <=== function for absent GN trailer
    elif record["Company"] == "WT":
        if trailer is not None:
            GNTr = SQL_REQ("SELECT status FROM dbo.GN_Trailers WHERE trailer_number=?", (trailer,), "S_all")
            if GNTr:
                for x in GNTr: trailer_status = x[0]
                if record["Status"]:
                    if trailer_status is not None:
                        if trailer_status: pass  # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Trailers SET status=1, last_date=?, LU=? WHERE trailer_number=?", (event_date, str(record["Cargo"]), trailer), "W")
                else:
                    if trailer_status is not None:
                        if not trailer_status: pass  # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Trailers SET status=0, storage=0, last_date=?, LU=0 WHERE trailer_number=?", (event_date, trailer), "W")
            else: pass  # <==== NO TRAILER IN GN LIST - error
        elif fb is not None:
            GNFb = SQL_REQ("SELECT status FROM dbo.GN_Flatbed WHERE fb_number=?", (fb,),"S_all")
            if GNFb:
                for x in GNFb: fb_status = x[0]
                if record["Status"]:
                    if fb_status is not None:
                        if fb_status: pass  # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Flatbed SET status=1, last_date=?, LU=? WHERE fb_number=?", (event_date, str(record["Cargo"]), fb), "W")
                else:
                    if fb_status is not None:
                        if not fb_status: pass  # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Flatbed SET status=0, storage=0, last_date=?, LU=0 WHERE fb_number=?", (event_date, fb), "W")
            else: pass  # <==== NO TRAILER IN GN LIST - error
        else: pass  # <=== function for absent GN trailer
    else:
        check_company = SQL_REQ("SELECT company_ID FROM dbo.Company_List WHERE company_name=?", (record["Company"],), "S_one")
        if check_company:
            c_ID = check_company[0]
        else:
            c_ID = None
            error(14)
            #return
        if record["Status"] is True:
            if c_ID is not None:
                check_truck = SQL_REQ("SELECT status FROM dbo.Tenant_Trucks WHERE truck_number=? AND company_ID=?", (record["Truck"], str(c_ID)), "S_one")
                if check_truck:
                    if check_truck[0]: pass # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.Tenant_Trucks SET status=1, last_date=? WHERE truck_number=? AND company_ID=?", (event_date, record["Truck"], str(c_ID)), "W")
                else:
                    check_truck = SQL_REQ("SELECT status FROM dbo.Tenant_Trucks_UNREG WHERE truck_number=? AND company_ID=?", (record["Truck"], str(c_ID)), "S_one")
                    if check_truck:
                        if check_truck[0]: pass  # function to invest who facuked up
                        SQL_REQ("UPDATE dbo.Tenant_Trucks_UNREG SET status=1, last_date=? WHERE truck_number=? AND company_ID=?", (event_date, record["Truck"], str(c_ID)), "W")
                    else:
                        Tenant_Register_UNREG(c_ID, record["Truck"], record["Status"], event_date, "truck")

            if trailer is not None:
                GNTr = SQL_REQ("SELECT status FROM dbo.GN_Trailers WHERE trailer_number=?", (trailer,), "S_one")
                if GNTr is not None:
                    trailer_status = GNTr[0]
                    if trailer_status: pass  # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Trailers SET status=1, last_date=?, LU=? WHERE trailer_number=?", (event_date, str(record["Cargo"]), trailer), "W")
            elif fb is not None:
                GNFb = SQL_REQ("SELECT status FROM dbo.GN_Flatbed WHERE fb_number=?", (fb,), "S_all")
                if GNFb is not None:
                    fb_status = GNFb[0]
                    if fb_status: pass  # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Flatbed SET status=1, last_date=?, LU=? WHERE fb_number=?", (event_date, str(record["Cargo"]), fb), "W")
        else:
            if c_ID is not None:
                check_truck = SQL_REQ("SELECT status FROM dbo.Tenant_Trucks WHERE truck_number=? AND company_ID=?", (record["Truck"], str(c_ID)), "S_one")
                if check_truck is not None:
                    if not check_truck[0]: pass # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.Tenant_Trucks SET status=0, last_date=? WHERE truck_number=? AND company_ID=?", (event_date, record["Truck"], str(c_ID)), "W")
                else:
                    check_truck = SQL_REQ("SELECT status FROM dbo.Tenant_Trucks_UNREG WHERE truck_number=? AND company_ID=?", (record["Truck"], str(c_ID)), "S_one")
                    if check_truck is not None:
                        if not check_truck[0]: pass  # function to invest who facuked up
                        SQL_REQ("UPDATE dbo.Tenant_Trucks_UNREG SET status=0, last_date=? WHERE truck_number=? AND company_ID=?", (event_date, record["Truck"], str(c_ID)), "W")
                    else:
                        Tenant_Register_UNREG(c_ID, record["Truck"], record["Status"], event_date, "truck")
            if trailer is not None:
                GNTr = SQL_REQ("SELECT status FROM dbo.GN_Trailers WHERE trailer_number=?", (trailer,), "S_one")
                if GNTr is not None:
                    trailer_status = GNTr[0]
                    if not trailer_status: pass  # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Trailers SET status=0, storage=0, last_date=?, LU=0 WHERE trailer_number=?", (event_date, trailer), "W")
            elif fb is not None:
                GNFb = SQL_REQ("SELECT status FROM dbo.GN_Flatbed WHERE fb_number=?", (fb,), "S_all")
                if GNFb is not None:
                    fb_status = GNFb[0]
                    if not fb_status: pass  # function to invest who facuked up
                    SQL_REQ("UPDATE dbo.GN_Flatbed SET status=0, storage=0, last_date=?, LU=0 WHERE fb_number=?", (event_date, fb), "W")

    OVERPARKING(record, "GN")

    # rec = open("GN_history.txt", "a")
    # str_line = str()
    # for key, value in record.items():
    #     str_line += str(value)
    #     str_line += "|"
    # str_line += "\n"
    # rec.write(str_line)
    # rec.close()

### RECORD INFO IN TXT FILE for VISITORS - TEMP to replace on SQL
def VIS_Record(record):
    global security
    event_date = str(datetime.strptime(record["Date"] + " " + record["Time"], '%Y/%m/%d %H:%M:%S'))
    val = SQL_REQ("SELECT company_ID FROM dbo.Company_List WHERE company_name=?", (record["Company"],), "S_one")
    if val is not None:
        _id = str(val[0])
        Vsts = SQL_REQ("SELECT status FROM dbo.visitors WHERE plates=?", (record["Plate"]), "S_one")
        if Vsts:
            V_status = str(Vsts[0])
        else:
            V_status = None
            Vis_Register_UNREG(_id, record["Plate"], record["Car"], record["Name"], event_date, record["Status"])
        if record["Status"] is True:
            if V_status is True: pass  # function to investigate who fucked up
            if V_status is not None:
                SQL_REQ("UPDATE dbo.visitors SET status=1, last_date=? WHERE plates=?", (event_date, record["Plate"]), "W")
        else:
            if V_status is False: pass  # function to investigate who fucked up
            if V_status is not None:
                SQL_REQ("UPDATE dbo.visitors SET status=0, last_date=NULL WHERE plates=?", (record["Plate"],), "W")
    else: _id = "999"

    record_val = [_id, record["Plate"], record["Car"], record["Name"], event_date, record["Status"], record["Comment"], security[0]]
    SQL_REQ("INSERT INTO dbo.visitors_history(company_ID, plates,  car_model, driver_name, datetime_event, status, comment, full_name) VALUES (?,?,?,?,?,?,?,?)", record_val, "W")

    OVERPARKING(record, "V")

#RECORD IN-OUT FUNCTION - TRUE-IN, FALSE-OUT. READ DATA FROM ENTRIES, REMOVE SPACES, BLOCK EMPTY COMPANY/TRUCK, RECORD DATA WITH DATE&TIME IN DICTIONARY, RUN REFRESH - TEMP to change on SQL
def Tenant_In_Out(bool):
    global Company_Var
    global Truck_Var
    global Trailer_Var
    global Tenant_Event
    if Company_Entry.get().strip() == "": return
    if Truck_Entry.get().strip() == "": return
    Company_Var = Company_Entry.get().strip()
    Truck_Var = (Truck_Entry.get().strip(), bool)
    if Trailer_Entry.get().strip() == "": Trailer = None
    else:
        Trailer_Var = (Trailer_Entry.get().strip(), bool)
        Trailer = Trailer_Var[0]
    if Comment_Entry.get().strip() != "": Tenant_Comment = Comment_Entry.get().strip()
    else: Tenant_Comment = None

    now = datetime.now()
    Tenant_Event = {
        "Company": Company_Var,
        "Truck": Truck_Var[0],
        "Trailer": Trailer,
        "Date": now.strftime("%Y/%m/%d"),
        "Time": now.strftime("%H:%M:%S"),
        "Status": bool,
        "Comment": Tenant_Comment
    }

    Last_Event_Company.configure(text=Company_Var)
    Last_Event_Truck.configure(text=Truck_Var[0])
    if Trailer is None: Last_Event_Trailer.configure(text="")
    else: Last_Event_Trailer.configure(text=Trailer)
    Last_Event_Date.configure(text=Tenant_Event["Date"])
    Last_Event_Time.configure(text=Tenant_Event["Time"])
    if Tenant_Event["Status"]:
        status = "IN"
    else:
        status = "OUT"
    Last_Event_Status.configure(text=status)
    try:
        Tenant_Record(Tenant_Event)
    except Exception as e:
        error(["Error in Tenant_Record: ", e])
        debuger(e)
    try:
        Refresh("Tenant")
    except Exception as e:
        error(["Error in Refresh: ", e])
        debuger(e)
    try:
        beep(bool)
    except Exception as e:
        error(["Error in Beep: ", e])
        debuger(e)

def GN_In_Out (bool):
    global GN_Other_Carrier_Var
    global GN_Truck_Var
    global GN_Trailer_Fb_Var
    global GN_Event
    global GN_Comment_Var
    global GN_Trigger_LU
    global GN_Trigger
    global GN_Trk_or_Fb_Var
    LU = None
    if GN_Entry_Truck.get().strip() == "": return
    GN_Truck_Var = (GN_Entry_Truck.get().strip(), bool)
    if GN_Entry_Trailer_Fb.get().strip() == "": Trailer = None
    else:
        GN_Trailer_Fb_Var = (GN_Entry_Trailer_Fb.get().strip(), bool)
        Trailer = GN_Trailer_Fb_Var[0]
        if GN_Trigger_LU == 4: LU = True
        elif GN_Trigger_LU == 5: LU = False
        if LU is None and bool is True: return
    if GN_Comment_Entry.get().strip() != "": GN_Comment_Var = GN_Comment_Entry.get().strip()
    else: GN_Comment_Var = ""
    if GN_Trigger == 3:
        if GN_Other_Entry.get().strip() == "": return
        else: GN_Other_Carrier_Var = GN_Other_Entry.get().strip()

    now = datetime.now()
    GN_Event = {
        "Company": GN_Other_Carrier_Var,
        "Truck": GN_Truck_Var[0],
        "Trailer": Trailer,
        "Date": now.strftime("%Y/%m/%d"),
        "Time": now.strftime("%H:%M:%S"),
        "Status": bool,
        "Cargo": LU,
        "Type": GN_Trk_or_Fb_Var,
        "Comment": GN_Comment_Var
    }

    GN_Last_Event_Carrier_lb.configure(text=GN_Other_Carrier_Var)
    GN_Last_Event_Truck_lb.configure(text=GN_Truck_Var[0])
    if Trailer is None: GN_Last_Event_Trailer_lb.configure(text="")
    else: GN_Last_Event_Trailer_lb.configure(text=Trailer)
    GN_Last_Event_Date_lb.configure(text=GN_Event["Date"])
    GN_Last_Event_Time_lb.configure(text=GN_Event["Time"])
    if GN_Event["Status"]: status = "IN"
    else: status = "OUT"
    GN_Last_Event_Status_lb.configure(text=status)
    GN_Record(GN_Event)
    Refresh("GN")
    beep(bool)

def VIS_IN_OUT(bool):
    global VIS_Company_Var
    global VIS_Company_Entry
    global VIS_Plates_Var
    global VIS_Plates_Entry
    global VIS_Car_Var
    global VIS_Car_Entry
    global VIS_Name_Var
    global VIS_Name_Entry
    global VIS_Comment_Entry
    global VIS_Comment
    global Current_Visitor_Unit
    if VIS_Company_Entry.get().strip() == "": return
    if VIS_Plates_Entry.get().strip() == "": return
    VIS_Company_Var = VIS_Company_Entry.get().strip()
    VIS_Plates_Var = VIS_Plates_Entry.get().strip()
    if VIS_Car_Entry.get().strip() != "": VIS_Car_Var = VIS_Car_Entry.get().strip()
    else: VIS_Car_Var = None
    if VIS_Name_Entry.get().strip() != "": VIS_Name_Var = VIS_Name_Entry.get().strip()
    else:  VIS_Name_Var = None
    if VIS_Comment_Entry.get().strip() != "": VIS_Comment = VIS_Comment_Entry.get().strip()
    else: VIS_Comment = None
    now = datetime.now()

    VIS_Event = {
        "Company": VIS_Company_Var,
        "Plate": VIS_Plates_Var,
        "Car": VIS_Car_Var,
        "Name": VIS_Name_Var,
        "Date": now.strftime("%Y/%m/%d"),
        "Time": now.strftime("%H:%M:%S"),
        "Status": bool,
        "Comment": VIS_Comment
    }

    VIS_Last_Event_Company.configure(text=VIS_Company_Var)
    VIS_Last_Event_Plates.configure(text=VIS_Plates_Var)
    if VIS_Car_Var is None: VIS_Last_Event_Car.configure(text="")
    else: VIS_Last_Event_Car.configure(text=VIS_Car_Var)
    VIS_Last_Event_Date.configure(text=VIS_Event["Date"])
    VIS_Last_Event_Time.configure(text=VIS_Event["Time"])
    if VIS_Event["Status"]:
        status = "IN"
    else:
        status = "OUT"
    VIS_Last_Event_Status.configure(text=status)
    VIS_Record(VIS_Event)
    Refresh("Visitor")
    # beep sound
    beep(bool)


#CREATE LABELS WIDGETS WITH FUNCTIONALITY FROM ITERABLE
def _unit_(masta, val, x, y, func, current_value):
    global Company_Var
    global Truck_Var
    global Trailer_Var
    global GN_Truck_Var
    global GN_Trailer_Fb_Var
    global Visitor_Company_Var
    global chk_T_T_Var
    global chk_T_Tr_Var
    global chk_GN_T_Var
    global chk_GN_Tr_Fb_Var
    global last_row
    if func == "companies" or func == "Visitor" or func == "CheckYard" or func == "Admin_Units" or func == "Admin_Vis_Co":
        button_var = tk.Label(masta, text=val, background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=22)
        button_var.pack(side=tk.TOP, pady=(1, 0), anchor=tk.N, fill=tk.X)
    elif func == "check_yard_T_T" or func == "check_yard_T_Tr" or func == "check_yard_GN_T" or func == "check_yard_GN_Tr" or func == "check_yard_GN_Fb":
        button_var = tk.Label(masta, text=val[0], background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.CENTER, width=9, highlightthickness=0)
        if val[2]: button_var.configure(fg=conf["on_parking"])
        button_var.grid(row=x, column=y, pady=(1, 0), padx=(0, 1), sticky=tk.NSEW)
    elif func == "adm_trucks" or func == "adm_trailers" or func == "adm_trucks+" or func == "adm_trailers+":
        button_var = tk.Label(masta, text=val[0], bg=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.CENTER, width=9, highlightthickness=0)
        button_var.grid(row=x, column=y, pady=(1, 0), padx=(0, 1), sticky=tk.NSEW)
        last_row[func] = x
        if func == "adm_trucks+" or func == "adm_trailers+": button_var.configure(highlightthickness=1, highlightbackground=conf["UNREG_bg"])
    else:
        if func == "trailers":
            button_var = tk.Label(masta, text=val[0], bg=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, width=9, highlightthickness=0)
            if val[1][0]: button_var.configure(fg=conf["on_parking"])
            if val[1][1]: button_var.configure(fg=conf["storage_fg"])
            button_var.grid(row=x, column=y, pady=(1, 0), padx=(0, 1), sticky=tk.NSEW)
        else:
            button_var = tk.Label(masta, text=val[0], bg=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, width=9, highlightthickness=0)
            if func != "Admin_GN_Truck" and func != "Admin_GN_Trailer" and func != "Admin_GN_Flatbed":
                if type(val[1]) is not list:
                    if val[1]: button_var.configure(fg=conf["on_parking"])
                else:
                    if val[1][0]: button_var.configure(fg=conf["on_parking"])
                    if val[1][1]: button_var.configure(fg=conf["storage_fg"])
            button_var.grid(row=x, column=y, pady=(1, 0), padx=(0, 1), sticky=tk.NSEW)
        if func == "trucks" or func == "trailers" or func == "adm_trucks" or func == "adm_trailers":
            last_row[func] = x
        if func == "trucks+" or func == "trailers+": button_var.configure(highlightthickness=1, highlightbackground=conf["UNREG_bg"])
    button_var.bind("<Button-1>", lambda x: UNTS(button_var, val, func))
    if func == "check_yard_T_T" or func == "check_yard_T_Tr" or func == "check_yard_GN_T" or func == "check_yard_GN_Tr" or func == "check_yard_GN_Fb":
        button_var.bind("<Enter>", lambda y: button_var.configure(bg=conf["widget_sel_bg"]))
    else: button_var.bind("<Enter>", lambda y: button_var.configure(bg=conf["widget_sel_bg"], fg=conf["widget_sel_fg"]))
    if current_value == 1: button_var.bind("<Leave>", lambda z: Hover_Off(button_var, Company_Var, val))
    if current_value == 2: button_var.bind("<Leave>", lambda z: Hover_Off(button_var, Truck_Var, val))
    if current_value == 3: button_var.bind("<Leave>", lambda z: Hover_Off(button_var, Trailer_Var, val))
    if current_value == 4: button_var.bind("<Leave>", lambda z: Hover_Off(button_var, GN_Truck_Var, val))
    if current_value == 5: button_var.bind("<Leave>", lambda z: Hover_Off(button_var, GN_Trailer_Fb_Var, val))
    if current_value == 6: button_var.bind("<Leave>", lambda z: Hover_Off(button_var, GN_Trailer_Fb_Var, val))
    if current_value == 7: button_var.bind("<Leave>", lambda z: Hover_Off(button_var, Visitor_Company_Var, val))
    if current_value == 8: button_var.bind("<Leave>", lambda z: Hover_Off(button_var, chk_Company_Var, val))
    if current_value == 9: button_var.bind("<Leave>", lambda z: button_var.configure(bg=conf["widget_bg"]))
    if current_value == 10: button_var.bind("<Leave>", lambda z: Hover_Off(button_var, Adm_Company_Var, val))
    if current_value == 11: button_var.bind("<Leave>", lambda z: Hover_Off_Adm_Ten(button_var, Adm_Truck_Var, val))
    if current_value == 12: button_var.bind("<Leave>", lambda z: Hover_Off_Adm_Ten(button_var, Adm_Trailer_Var, val))
    if current_value == 13: button_var.bind("<Leave>", lambda z: Hover_Off_Adm_GN(button_var, Adm_GN_Truck_Var, val))
    if current_value == 14: button_var.bind("<Leave>", lambda z: Hover_Off_Adm_GN(button_var, Adm_GN_Trailer_Var, val))
    if current_value == 15: button_var.bind("<Leave>", lambda z: Hover_Off_Adm_GN(button_var, Adm_GN_Fb_Var, val))
    if current_value == 16: button_var.bind("<Leave>", lambda z: Hover_Off(button_var, Adm_Vis_Company_Var, val))



# TENANT CLICKED FUNC - REMIND CLICKED POSITION, REMOVE PREVIOUS AND PACK TENANT FRAME
def Tenant(event):
    global Parking_Var
    Parking_Var = 1
    GN_lb.configure(bg=conf["submenu_bg"])
    Visitor_lb.configure(bg=conf["submenu_bg"])
    Refresh("GN")
    Refresh("Tenant")
    GN_Main.pack_forget()
    Refresh("Visitor")
    VISITOR_Main.pack_forget()
    Tenant_Main.pack(fill=tk.BOTH, expand=1, padx=5, pady=5)
    for widgets in second_comp_Frame.winfo_children(): widgets.destroy()
    Implement(second_comp_Frame, "company", "companies", 1, Parking_Tenant_size)

# GN CLICKED FUNC - REMIND CLICKED POSITION, REMOVE PREVIOUS AND PACK GN FRAME
def GN(event):
    global Parking_Var
    Parking_Var = 2
    Tenant_lb.configure(bg=conf["submenu_bg"])
    Visitor_lb.configure(bg=conf["submenu_bg"])
    Refresh("Tenant")
    Refresh("GN")
    Tenant_Main.pack_forget()
    Refresh("Visitor")
    VISITOR_Main.pack_forget()
    GN_Main.pack(fill=tk.BOTH, expand=1, padx=5, pady=5)
    Implement(GN_Truck_SubFrame.frame, "GNtrucks", "GNtrucks", 4, Parking_GN_size)
    Implement(GN_Trailer_SubFrame.frame, "GNtrailers", "GNtrailers", 5, Parking_GN_size)
    Implement(GN_Flatbed_SubFrame.frame, "GNfb", "GNflatbeds", 6, Parking_GN_size)

# VISITOR CLICKED - REMIND CLICKED POSITION, REMOVE PREVIOUS AND PACK VISITORS FRAME
def Visitor(event):
    global Parking_Var
    Parking_Var = 3
    Tenant_lb.configure(bg=conf["submenu_bg"])
    GN_lb.configure(bg=conf["submenu_bg"])
    Refresh("Tenant")
    Refresh("Visitor")
    Tenant_Main.pack_forget()
    Refresh("GN")
    GN_Main.pack_forget()
    VISITOR_Main.pack(fill=tk.BOTH, expand=1, padx=5, pady=5)
    for widgets in vis_second_comp_Frame.winfo_children(): widgets.destroy()
    Implement(vis_second_comp_Frame, "company", "Visitor", 7, Parking_Tenant_size)

#IMPLEMENTING FUNC - TAKE SIZE OF FRAME, CREATE COORDINATES AND IMPLEMENT DATA FROM ITERABLE BY COORDINATES.
def Implement(masta, query, func, current_value, counterSize):
    global last_row
    if func == "trailers+":
        if last_row["trailers"] is not None: counterX = last_row["trailers"]+1
        else: counterX = 1
    elif func == "trucks+":
        if last_row["trucks"] is not None: counterX = last_row["trucks"] + 1
        else: counterX = 1
    elif func == "adm_trucks+":
        if last_row["adm_trucks"] is not None: counterX = last_row["adm_trucks"]+1
        else: counterX = 1
    elif func == "adm_trailers+":
        if last_row["adm_trailers"] is not None: counterX = last_row["adm_trailers"]+1
        else: counterX = 1
    else: counterX = int(0)
    counterY = int(0)
    if func == "adm_trucks":
        dict = units_lst(query, "adm_trucks")
    elif func == "adm_trailers":
        dict = units_lst(query, "adm_trailers")
    else: dict = units_lst(query, func)
    last_row = {
        "trucks": None,
        "trailers": None,
        "adm_trucks": None,
        "adm_trailers": None
    }
    if func == "companies" or func == "Visitor" or func == "CheckYard" or func == "Admin_Units" or func == "Admin_Vis_Co":
        for item in dict: _unit_(masta, item, 1, 1, func, current_value)
    elif func == "check_yard_T_T":
        for item in dict:
            if item[1] == "truck" and item[4] != "GN":
                _unit_(masta, item, counterX, counterY, func, current_value)
                if counterY == counterSize - 1:
                    counterX += 1
                    counterY = 0
                else:
                    counterY += 1
    elif func == "check_yard_T_Tr":
        for item in dict:
            if item[1] == "trailer" and item[4] != "GN":
                _unit_(masta, item, counterX, counterY, func, current_value)
                if counterY == counterSize - 1:
                    counterX += 1
                    counterY = 0
                else:
                    counterY += 1
    elif func == "check_yard_GN_T":
        for item in dict:
            if item[1] == "truck" and item[4] == "GN":
                _unit_(masta, item, counterX, counterY, func, current_value)
                if counterY == counterSize - 1:
                    counterX += 1
                    counterY = 0
                else:
                    counterY += 1
    elif func == "check_yard_GN_Tr":
        for item in dict:
            if item[1] == "trailer" and item[4] == "GN":
                _unit_(masta, item, counterX, counterY, func, current_value)
                if counterY == counterSize - 1:
                    counterX += 1
                    counterY = 0
                else:
                    counterY += 1
    elif func == "check_yard_GN_Fb":
        for item in dict:
            if item[1] == "flatbed" and item[4] == "GN":
                _unit_(masta, item, counterX, counterY, func, current_value)
                if counterY == counterSize - 1:
                    counterX += 1
                    counterY = 0
                else:
                    counterY += 1
    elif func == "adm_trucks+" or func == "adm_trailers+":
        for item in dict:
            _unit_(masta, item, counterX, counterY, func, current_value)
            if counterY == counterSize - 1:
                counterX += 1
                counterY = 0
            else:
                counterY += 1
    elif func == "adm_trucks" or func == "adm_trailers":
        for item in dict:
            _unit_(masta, item, counterX, counterY, func, current_value)
            if counterY == counterSize-1:
                counterX += 1
                counterY = 0
            else:
                counterY += 1
    else:
        for item in dict.items():
            _unit_(masta, item, counterX, counterY, func, current_value)
            if counterY == counterSize-1:
                counterX += 1
                counterY = 0
            else:
                counterY += 1

#PARKING TOP MENU HOVER OFF FUNC - RETURN BG IF WIDGET NOT CLICKED
def Parking_Menu_Hover_Off(event):
    if Parking_Var != 1: Tenant_lb.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if Parking_Var != 2: GN_lb.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if Parking_Var != 3: Visitor_lb.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])

#GETTING HISTORY FROM SQL
def H_receive(tab):
    if tab == 1:
        row, col = SQL_REQ("SELECT * FROM dbo.Tenant_History WHERE day(datetime_event)=day(GETDATE()) AND MONTH(datetime_event) = MONTH(GETDATE()) AND YEAR(datetime_event) = YEAR(GETDATE()) ORDER BY datetime_event", (), "S_all_D")
        l = []
        for x in range(len(row)):
            a = {}
            index = 0
            for y in row[x]:
                z = col[index]
                index += 1
                a.update({z[0]: y})
            l.append(a)
        return l
    elif tab == 2:
        row, col = SQL_REQ("SELECT * FROM dbo.GN_History WHERE day(datetime_event)=day(GETDATE()) AND MONTH(datetime_event) = MONTH(GETDATE()) AND YEAR(datetime_event) = YEAR(GETDATE()) ORDER BY datetime_event", (), "S_all_D")
        l = []
        for x in range(len(row)):
            a = {}
            index = 0
            for y in row[x]:
                z = col[index]
                index += 1
                a.update({z[0]: y})
            l.append(a)
        return l
    elif tab == 3:
        row, col =SQL_REQ("SELECT * FROM dbo.visitors_history WHERE day(datetime_event)=day(GETDATE()) AND MONTH(datetime_event) = MONTH(GETDATE()) AND YEAR(datetime_event) = YEAR(GETDATE()) ORDER BY datetime_event", (), "S_all_D")
        l = []
        for x in range(len(row)):
            a = {}
            index = 0
            for y in row[x]:
                z = col[index]
                index += 1
                a.update({z[0]: y})
            l.append(a)
        return l

#INSERT HISTORY
def H_insert(masta, tab):
    if tab == 1:
        history_list = H_receive(tab)
        for i in range(len(history_list)):
            record = history_list[i]
            obj = SQL_REQ("SELECT company_name FROM dbo.Company_List WHERE dbo.Company_List.company_ID=?", (str(record['company_ID']),), "S_one")
            for x in obj: name=x
            frame = tk.Frame(masta, highlightthickness=0, bg=conf["window_bg"])
            frame.pack(side=tk.TOP, fill=tk.X, pady=(1, 0))
            cmpny = tk.Label(frame, text=name, bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=20)
            cmpny.pack(side=tk.LEFT, fill=tk.BOTH)
            truck = tk.Label(frame, text=record["truck_number"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            truck.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            trailer = tk.Label(frame, text=record["trailer_number"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            trailer.pack(side=tk.LEFT, fill=tk.BOTH)
            date = record["datetime_event"].date()
            time = record["datetime_event"].time()
            dt = tk.Label(frame, text=str(date), bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            dt.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            tm = tk.Label(frame, text=str(time), bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            tm.pack(side=tk.LEFT, fill=tk.BOTH)
            status = tk.Label(frame,  bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), anchor=tk.CENTER, highlightthickness=0, width=10)
            status.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            if record["status"]: status.configure(text="IN", fg=conf['in_button_fg'])
            else: status.configure(text="OUT", fg=conf["out_button_fg"])
            cmmnt = tk.Label(frame, text=record["comment"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=20)
            cmmnt.pack(side=tk.LEFT, fill=tk.BOTH)
            sec = tk.Label(frame, text=record["full_name"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=20)
            sec.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            edit_mark = tk.Label(frame, bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, width=10, highlightthickness=0)
            edit_mark.pack(side=tk.LEFT, fill=tk.BOTH)
        H_canv.update_idletasks()
        check_HT_scroll_region()
        H_canv.yview_moveto(0.0)

    elif tab == 2:
        history_list = H_receive(tab)
        for i in range(len(history_list)):
            record = history_list[i]
            if record["trailer_number"] is not None and record["fb_number"] is None: trl = record["trailer_number"]
            elif record["trailer_number"] is None and record["fb_number"] is not None: trl = "FB-"+record["fb_number"]
            else: trl = ""
            frame = tk.Frame(masta, highlightthickness=0, bg=conf["window_bg"])
            frame.pack(side=tk.TOP, fill=tk.X, pady=(1, 0))
            cmpny = tk.Label(frame, text=record["company_name"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=15)
            cmpny.pack(side=tk.LEFT, fill=tk.BOTH)
            truck = tk.Label(frame, text=record["truck_number"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            truck.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            trailer = tk.Label(frame, text=trl, bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            trailer.pack(side=tk.LEFT, fill=tk.BOTH)
            date = record["datetime_event"].date()
            time = record["datetime_event"].time()
            dt = tk.Label(frame, text=str(date), bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            dt.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            tm = tk.Label(frame, text=str(time), bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            tm.pack(side=tk.LEFT, fill=tk.BOTH)
            lu = tk.Label(frame, bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), anchor=tk.CENTER, highlightthickness=0, width=10)
            lu.pack(side=tk.LEFT, fill=tk.BOTH, padx=(1, 0))
            if record["cargo"]: lu.configure(text="LOADED", fg=conf['in_button_fg'])
            else: lu.configure(text="EMPTY", fg=conf['out_button_fg'])
            status = tk.Label(frame, bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), anchor=tk.CENTER, highlightthickness=0, width=10)
            status.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            if record["status"]: status.configure(text="IN", fg=conf['in_button_fg'])
            else: status.configure(text="OUT", fg=conf["out_button_fg"])
            cmmnt = tk.Label(frame, text=record["comment"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=20)
            cmmnt.pack(side=tk.LEFT, fill=tk.BOTH)
            sec = tk.Label(frame, text=record["full_name"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=15)
            sec.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            edit_mark = tk.Label(frame, bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, width=10, highlightthickness=0)
            edit_mark.pack(side=tk.LEFT, fill=tk.BOTH)
        HG_canv.update_idletasks()
        check_HG_scroll_region()
        HG_canv.yview_moveto(0.0)
    elif tab == 3:
        history_list = H_receive(tab)
        for i in range(len(history_list)):
            record = history_list[i]
            for x in SQL_REQ("SELECT company_name FROM dbo.Company_List WHERE dbo.Company_List.company_ID=?", (str( record['company_ID']),), "S_one"): name = x
            frame = tk.Frame(masta, highlightthickness=0, bg=conf["window_bg"])
            frame.pack(side=tk.TOP, fill=tk.X, pady=(1, 0))
            cmpny = tk.Label(frame, text=name, bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=15)
            cmpny.pack(side=tk.LEFT, fill=tk.BOTH)
            plate = tk.Label(frame, text=record["plates"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            plate.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            car = tk.Label(frame, text=record["car_model"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            car.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 1))
            driver = tk.Label(frame, text=record["driver_name"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=15)
            driver.pack(side=tk.LEFT, fill=tk.BOTH)
            date = record["datetime_event"].date()
            time = record["datetime_event"].time()
            dt = tk.Label(frame, text=str(date), bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            dt.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            tm = tk.Label(frame, text=str(time), bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=10)
            tm.pack(side=tk.LEFT, fill=tk.BOTH)
            status = tk.Label(frame, bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), anchor=tk.CENTER, highlightthickness=0, width=10)
            status.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            if record["status"]: status.configure(text="IN", fg=conf['in_button_fg'])
            else: status.configure(text="OUT", fg=conf["out_button_fg"])
            cmmnt = tk.Label(frame, text=record["comment"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=15)
            cmmnt.pack(side=tk.LEFT, fill=tk.BOTH)
            sec = tk.Label(frame, text=record["full_name"], bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, highlightthickness=0, width=15)
            sec.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            edit_mark = tk.Label(frame, bg=conf["widget_bg"], font=(conf["history_font"], conf["history_size"]), fg=conf["widget_fg"], anchor=tk.CENTER, width=10, highlightthickness=0)
            edit_mark.pack(side=tk.LEFT, fill=tk.BOTH)
        HV_canv.update_idletasks()
        check_HV_scroll_region()
        HV_canv.yview_moveto(0.0)





#@#@#@#@#    Future Function For QR creating     #@#@#@#@#
def to_QR(plates_number, company_name, driver_name, exp_date):
    import qrcode
    # Combine data into a single string
    code_data = f"{plates_number}\n{exp_date}\n{company_name}\n{driver_name}"
    # Choose a QR code version and error correction level
    qr = qrcode.QRCode(
        version=10,  # Adjust based on your data length
        error_correction=qrcode.constants.ERROR_CORRECT_H,  # High error correction
        box_size=10,
        border=4,
    )
    # Add data to the QR code
    qr.add_data(code_data)
    qr.make(fit=True)
    # Create an image from the QR code
    img = qr.make_image(fill_color="black", back_color="white")
    # check directory exist
    os.makedirs(sets['QR_path'], exist_ok=True)
    # Save the image or display it
    img.save(f"{sets['QR_path']}{company_name}-{plates_number}.png")
    img.show()
    return img
#@#@#@#@##@#@#@#@##@#@#@#@##@#@#@#@##@#@#@#@##@#@#@#@##@#@#

# qr_code = to_QR("ABCD123", "Company Name", "Vova Pupkin", "2023-04-08")
# qr_code.show()

########        ########        ########        ########        ########        ########        ########        ########

# MAIN WINDOW CONFIGURATIONS
root = tk.Tk()
root.title = ("Parking Hawk")
# Set the icon for the window
try:
    if getattr(sys, 'frozen', False):
        # Running in PyInstaller bundle
        root.iconbitmap(default=sys.executable)
    else:
        root.iconbitmap(default='icon.ico')
except:
    pass
root.attributes("-fullscreen", True)
root.configure(bg=conf["window_bg"])
stl = ttk.Style()           #to be configure
stl.theme_use("default")
screen_x = root.winfo_screenwidth()
screen_y = root.winfo_screenheight()


# CONFIGURE Notebook STYLE
stl.configure("TNotebook", background=conf["notebook_bg"])
stl.configure("TNotebook.Tab", background=conf["notebook_tab_unsel_bg"], foreground=conf["notebook_tab_unsel_fg"], font=(conf["notebook_tab_font"], conf["notebook_tab_size"]))
stl.map("TNotebook.Tab", background=[("selected", conf["notebook_tab_sel_bg"])], foreground=[("selected", conf["notebook_tab_sel_fg"])])
stl.configure("TEntry", fieldbackground=conf["entry_bg"], borderwidth=0)
stl.configure("Custom.TEntry", fieldbackground=conf["window_bg"])
stl.configure("CustomV.TEntry", fieldbackground=conf["widget_bg"])


# TOP WINDOW BAR AND FUNCTIONALITY BUTTONS
Top_Frame = tk.Frame(root, relief=tk.RAISED, bg=conf["window_bg"], borderwidth=2, highlightthickness=0)
Top_Frame.pack(side=tk.TOP, fill=tk.X)
Top_labe = tk.Label(Top_Frame, text="Parking Hawk 1.43", fg=conf["window_topbar_fg"], bg=conf["window_bg"], font=(conf["window_topbar_font"], conf["window_topbar_size"]))
Top_labe.pack(side=tk.LEFT)
StatisticT = tk.Label(Top_Frame, fg=conf["header_fg"], bg=conf["window_bg"], font=(conf["notebook_tab_font"], conf["notebook_tab_size"]))
StatisticT.pack(side=tk.LEFT, padx=(20,1))
StatisticGN = tk.Label(Top_Frame, fg=conf["on_parking"], bg=conf["window_bg"], font=(conf["notebook_tab_font"], conf["notebook_tab_size"]))
StatisticGN.pack(side=tk.LEFT)
StatisticOVER = tk.Label(Top_Frame, fg=conf["out_button_fg"], bg=conf["window_bg"], font=(conf["notebook_tab_font"], conf["notebook_tab_size"]))
StatisticOVER.pack(side=tk.LEFT, padx=1)
statistics_reg("T")
statistics_reg("GN")
Close_Button = tk.Button(Top_Frame, text="X", bg=conf["window_bg"], fg=conf["window_topbar_fg"], padx=15, pady=0, bd=0, command=root.destroy, highlightthickness=0, font=(conf["notebook_tab_font"], conf["notebook_tab_size"]), activebackground=conf["window_topbar_sel_bg"])
Close_Button.pack(side=tk.RIGHT, fill=tk.Y)
Min_Button = tk.Button(Top_Frame, text="_", bg=conf["window_bg"], fg=conf["window_topbar_fg"], padx=15, pady=0, bd=0, command=root.iconify, highlightthickness=0, font=(conf["notebook_tab_font"], conf["notebook_tab_size"]), activebackground=conf["window_topbar_sel_bg"])
Min_Button.pack(side=tk.RIGHT, fill=tk.Y)
Security_Reset_Button = tk.Label(Top_Frame, text="Log Out", bg=conf["window_bg"], fg=conf["window_topbar_fg"], font=(conf["window_topbar_font"], conf["window_topbar_size"]))
Security_Reset_Button.bind("<Button-1>", shift_change)
Security_Reset_Button.bind("<Enter>", lambda x: Security_Reset_Button.config(bg=conf["widget_sel_bg"]))
Security_Reset_Button.bind("<Leave>", lambda y: Security_Reset_Button.config(bg=conf["window_bg"]))
Security_Name = tk.Label(Top_Frame, bg=conf["window_bg"],  fg=conf["font_color"], font=(conf["font_name"], conf["font_size"]), relief=tk.SUNKEN)


# MENU BAR - FRAMES FOR EACH Notebook Tab
Menu_Bar = ttk.Notebook(root, style="TNotebook")
Menu_Bar_Parking = tk.Frame(Menu_Bar, bg=conf["window_bg"])
Menu_Bar_History = tk.Frame(Menu_Bar, bg=conf["window_bg"])
Menu_Bar_Overparking = tk.Frame(Menu_Bar, bg=conf["window_bg"])
Menu_Bar_GN = tk.Frame(Menu_Bar, bg=conf["window_bg"])
Menu_Bar_Check = tk.Frame(Menu_Bar, bg=conf["window_bg"])
Menu_Bar_Admin = tk.Frame(Menu_Bar, bg=conf["window_bg"])
Menu_Bar.add(Menu_Bar_Parking, text="Parking")
Menu_Bar.add(Menu_Bar_History, text="History")
Menu_Bar.add(Menu_Bar_Overparking, text="Overparking")
Menu_Bar.add(Menu_Bar_GN, text="GN")
Menu_Bar.add(Menu_Bar_Check, text="Check Yard")
Menu_Bar.add(Menu_Bar_Admin, text="Administration")
Menu_Bar.bind("<<NotebookTabChanged>>", Tabs_Refresh)


# LOGIN FRAME WITH BUTTONS
login_frame = tk.Frame(root, bg=conf["window_bg"], height=300, width=700, bd=2, relief=tk.GROOVE)
login_frame.pack_propagate(0)
login_header = tk.Label(login_frame, text="Please Log In", bg=conf["window_bg"], fg=conf["window_topbar_fg"], font=(conf["window_topbar_font"], conf["window_topbar_size"]))
login_header.pack(side=tk.TOP, pady=(15, 0))
login_name_frame = tk.Frame(login_frame, bg=conf["window_bg"])
login_name_frame.pack(side=tk.TOP, fill=tk.X, pady=(60, 0), padx=10)
login_name_lb = tk.Label(login_name_frame, text="Login:", bg=conf["window_bg"], fg=conf["widget_fg"], font=(conf["font_name"], conf["font_size"]))
login_name_lb.pack(side=tk.LEFT, fill=tk.Y)
login_name_entry = tk.Entry(login_name_frame, bg=conf["window_bg"], bd=1, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, textvariable=login_var, justify=tk.LEFT)
login_name_entry.pack(side=tk.RIGHT, padx=20)
login_password_frame = tk.Frame(login_frame, bg=conf["window_bg"])
login_password_frame.pack(side=tk.TOP, fill=tk.X, pady=(5, 0), padx=10)
login_password_lb = tk.Label(login_password_frame, text="Password:", bg=conf["window_bg"], fg=conf["widget_fg"], font=(conf["font_name"], conf["font_size"]))
login_password_lb.pack(side=tk.LEFT, fill=tk.Y)
login_password_entry = tk.Entry(login_password_frame, bg=conf["window_bg"], bd=1, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, textvariable=password_var, justify=tk.LEFT, show="*")
login_password_entry.pack(side=tk.RIGHT, padx=20)
login_button = tk.Button(login_frame, text="Login", bg=conf["window_bg"], fg=conf["window_topbar_fg"], font=(conf["window_topbar_font"], conf["window_topbar_size"]), command=login_func)
login_button.pack(side=tk.TOP, pady=10)
login_button.bind_all('<Return>', login_func)

#STARTING/RESTARTING LOGIN
if security==None:
    login_frame.pack(side=tk.TOP, pady=((screen_y-300)/2, 0)) #uncensore
    login_name_entry.focus_set()


# INSERING TABS in Notebook MENU BAR
# Menu_Bar.add(Menu_Bar_Parking, text="Parking")
# Menu_Bar.add(Menu_Bar_History, text="History")
# Menu_Bar.add(Menu_Bar_Overparking, text="Overparking")
# Menu_Bar.add(Menu_Bar_GN, text="GN")
# Menu_Bar.add(Menu_Bar_Check, text="Check Yard")
# Menu_Bar.add(Menu_Bar_Admin, text="Administration")

# PARKING TAB SUB MENU - TENANT/GN/VISITORS
Parking_Sub_Menu_Frame = tk.Frame(Menu_Bar_Parking, bg=conf["window_bg"])
Parking_Sub_Menu_Frame.pack(fill=tk.BOTH)

#CREATING TENANT BUTTON
Tenant_lb = tk.Label(Parking_Sub_Menu_Frame, text="Tenant", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
Tenant_lb.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
Tenant_lb.bind("<Button-1>", Tenant)
Tenant_lb.bind("<Enter>", lambda x: Tenant_lb.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
Tenant_lb.bind("<Leave>", Parking_Menu_Hover_Off)

#CREATING GN BUTTON
GN_lb = tk.Label(Parking_Sub_Menu_Frame, text="GN", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
GN_lb.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
GN_lb.bind("<Button-1>", GN)
GN_lb.bind("<Enter>", lambda x: GN_lb.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
GN_lb.bind("<Leave>", Parking_Menu_Hover_Off)
#CREATING VISITOR BUTTON
Visitor_lb = tk.Label(Parking_Sub_Menu_Frame, text="Visitor", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
Visitor_lb.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
Visitor_lb.bind("<Button-1>", Visitor)
Visitor_lb.bind("<Enter>", lambda x: Visitor_lb.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
Visitor_lb.bind("<Leave>", Parking_Menu_Hover_Off)

# PARKING SUBFRAME - UNDER SUM MENU BUTTONS
Parking_Main_Frame = tk.Frame(Menu_Bar_Parking, bg=conf["window_bg"])
Parking_Main_Frame.pack(fill=tk.BOTH, expand=1)

###########################################################################################
# PARKING TENANT FRAME
###########################################################################################
Tenant_Main = tk.Frame(Parking_Main_Frame, bg=conf["window_bg"], highlightthickness=0)


# COMPANY FRAME
Company_Frame = tk.Frame(Tenant_Main, bg=conf["window_bg"], highlightthickness=0, width=conf["p_t_company_w"])
Company_Frame.pack_propagate(0)
Company_Frame.pack(fill=tk.BOTH, side=tk.LEFT)

#LABEL FOR COMPANY
Company_Lb = tk.Label(Company_Frame, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
Company_Lb.pack(fill=tk.X, side=tk.TOP)
# PARKING FRAME FOR COMPANIES WITH SCROLL
sub_park_comp_frame = tk.Frame(Company_Frame, highlightthickness=0)
sub_park_comp_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
comp_canv = tk.Canvas(sub_park_comp_frame, highlightthickness=0, bg=conf["window_bg"])
second_comp_Frame = tk.Frame(comp_canv, bg=conf["window_bg"])
comp_scrl = ttk.Scrollbar(Company_Frame, orient=tk.VERTICAL, command=comp_canv.yview)
comp_canv.config(yscrollcommand=comp_scrl.set)
comp_scrl.pack(fill=tk.Y, side=tk.RIGHT)
comp_canv.pack(fill=tk.BOTH, side=tk.TOP, expand=1)
comp_canv.create_window((0, 0), window=second_comp_Frame, anchor=tk.NW)
second_comp_Frame.bind("<Configure>", lambda event, canvas=comp_canv: comp_canv.configure(scrollregion=comp_canv.bbox("all")))

def check_T_C_scroll_region(*event):
    if second_comp_Frame.winfo_height() <= comp_canv.winfo_height():
        comp_scrl.pack_forget()
        chk_canv.configure(yscrollcommand=None)
        second_comp_Frame.unbind("<Enter>")
        second_comp_Frame.unbind_all("<MouseWheel>")
    else:
        comp_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        comp_canv.configure(yscrollcommand=comp_scrl.set)
        second_comp_Frame.bind("<Enter>", _enter_mousewheel_tenant_comp, add="+")
comp_canv.bind("<Configure>", check_T_C_scroll_region)

#MOUSEWHEEL FUNCTION
#_mousewheel_(second_comp_Frame, comp_canv)
def _on_mousewheel(event): comp_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def _enter_mousewheel_tenant_comp(event): comp_canv.bind_all('<MouseWheel>', _on_mousewheel, add="+")
def _leave_mousewheel_tenant_comp(event): comp_canv.unbind_all('<MouseWheel>')
second_comp_Frame.bind("<Enter>", _enter_mousewheel_tenant_comp, add="+")
second_comp_Frame.bind("<Leave>", _leave_mousewheel_tenant_comp)

#TRUCKS FRAME
Trucks_Frame = tk.Frame(Tenant_Main, bg=conf["window_bg"], highlightthickness=0, width=screen_x-int(conf["p_t_company_w"])-15, height=conf["p_t_truck_h"])
Trucks_Frame.pack_propagate(0)
Trucks_Frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1, padx=(5, 0))
#LABEL FOR TRUCK
Truck_Lb = tk.Label(Trucks_Frame, text="TRUCK:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
Truck_Lb.pack(fill=tk.X, side=tk.TOP)
# SUBFRAME TRUCKS with SCROLL
sub_park_truck_frame = tk.Frame(Trucks_Frame, highlightthickness=0)
sub_park_truck_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
truck_canv = tk.Canvas(sub_park_truck_frame, bg=conf["window_bg"], highlightthickness=0)
second_truck_Frame = tk.Frame(truck_canv, bg=conf["window_bg"])
truck_scrl = ttk.Scrollbar(Trucks_Frame, orient=tk.VERTICAL, command=truck_canv.yview)
truck_canv.config(yscrollcommand=truck_scrl.set)
truck_scrl.pack(fill=tk.Y, side=tk.RIGHT)
truck_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
truck_canv.create_window((0, 0), window=second_truck_Frame, anchor=tk.NW)
truck_canv.bind("<Configure>", lambda event, canvas=truck_canv: truck_canv.configure(scrollregion=truck_canv.bbox("all")))

def check_T_T_scroll_region(*event):
    canvas_height = second_truck_Frame.winfo_height()
    truck_canv.configure(scrollregion=(0,0,0,canvas_height))
    if canvas_height <= truck_canv.winfo_height():
        truck_scrl.pack_forget()
        truck_canv.configure(yscrollcommand=None)
        second_truck_Frame.unbind("<Enter>")
        second_truck_Frame.unbind("<MouseWheel>")
    else:
        truck_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        truck_canv.configure(yscrollcommand=truck_scrl.set)
        second_truck_Frame.bind("<Enter>", _enter_mousewheel_tenant_truck, add="+")
truck_canv.bind("<Configure>", check_T_T_scroll_region)

#Mouse Func
def _on_mousewheel_tenant_truck(event):
    truck_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def _enter_mousewheel_tenant_truck(event): truck_canv.bind_all('<MouseWheel>', _on_mousewheel_tenant_truck, add="+")
def _leave_mousewheel_tenant_truck(event): truck_canv.unbind_all('<MouseWheel>')
second_truck_Frame.bind("<Enter>", _enter_mousewheel_tenant_truck, add="+")
second_truck_Frame.bind("<Leave>", _leave_mousewheel_tenant_truck)

#TRAILERS FRAME
Trailers_Frame = tk.Frame(Tenant_Main, bg=conf["window_bg"], highlightthickness=0)
Trailers_Frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1, padx=(5, 0), pady=(5, 0))
#LABEL FOR TRAILERS
Trailer_Lb = tk.Label(Trailers_Frame, text="TRAILER:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
Trailer_Lb.pack(fill=tk.X, side=tk.TOP)
# SUBFRAME TRAILERS with SCROLL
sub_park_trailer_frame = tk.Frame(Trailers_Frame, highlightthickness=0)
sub_park_trailer_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
trailer_canv = tk.Canvas(sub_park_trailer_frame, bg=conf["window_bg"], highlightthickness=0)
second_trailer_Frame = tk.Frame(trailer_canv, bg=conf["window_bg"])
trailer_scrl = ttk.Scrollbar(Trailers_Frame, orient=tk.VERTICAL, command=trailer_canv.yview)
trailer_canv.config(yscrollcommand=trailer_scrl.set)
trailer_scrl.pack(fill=tk.Y, side=tk.RIGHT, expand=0)
trailer_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
trailer_canv.create_window((0, 0), window=second_trailer_Frame, anchor=tk.NW)
trailer_canv.bind("<Configure>", lambda event, canvas=trailer_canv: trailer_canv.configure(scrollregion=trailer_canv.bbox("all")))
def check_T_Tr_scroll_region(*event):
    canvas_height = second_trailer_Frame.winfo_height()
    trailer_canv.configure(scrollregion=(0,0,0,canvas_height))
    if canvas_height <= trailer_canv.winfo_height():
        trailer_scrl.pack_forget()
        trailer_canv.configure(yscrollcommand=None)
        second_trailer_Frame.unbind("<Enter>")
        second_trailer_Frame.unbind("<MouseWheel>")
    else:
        trailer_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        trailer_canv.configure(yscrollcommand=trailer_scrl.set)
        second_trailer_Frame.bind("<Enter>", _enter_mousewheel_tenant_trailer, add="+")
trailer_canv.bind("<Configure>", check_T_Tr_scroll_region)

#Mouse func
def _on_mousewheel_tenant_trailer(event):
    trailer_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def _enter_mousewheel_tenant_trailer(event):
    trailer_canv.bind_all('<MouseWheel>', _on_mousewheel_tenant_trailer, add="+")
def _leave_mousewheel_tenant_trailer(event):
    trailer_canv.unbind_all('<MouseWheel>')
second_trailer_Frame.bind("<Enter>", _enter_mousewheel_tenant_trailer, add="+")
second_trailer_Frame.bind("<Leave>", _leave_mousewheel_tenant_trailer)

#IN-OUT FRAME
IN_OUT_Frame = tk.Frame(Tenant_Main, bg=conf["window_bg"], highlightthickness=0)
IN_OUT_Frame.pack(fill=tk.BOTH, side=tk.TOP, expand=0, padx=(5, 0), pady=(5, 0))
#IN-OUT SUB FRAME FOR MANUAL ENTRANCE AND BUTTONS
IN_OUT_Frame_Manual_TOP = tk.Frame(IN_OUT_Frame, bg=conf["window_bg"], highlightthickness=0, height=conf["p_t_manual_h"])
IN_OUT_Frame_Manual_TOP.pack_propagate(0)
IN_OUT_Frame_Manual_TOP.pack(fill=tk.BOTH, side=tk.TOP, expand=1)
IN_OUT_Frame_Manual_BOTTOM = tk.Frame(IN_OUT_Frame, bg=conf["window_bg"], highlightthickness=0, height=conf["p_t_manual_h"])
IN_OUT_Frame_Manual_BOTTOM.pack_propagate(0)
IN_OUT_Frame_Manual_BOTTOM.pack(fill=tk.BOTH, side=tk.TOP, expand=1, pady=(5, 0))
IN_OUT_Frame_Last_Event = tk.Frame(IN_OUT_Frame, bg=conf["window_bg"], highlightthickness=0, height=conf["p_t_last_event_h"])
IN_OUT_Frame_Last_Event.pack_propagate(0)
IN_OUT_Frame_Last_Event.pack(fill=tk.BOTH, side=tk.TOP, expand=0, pady=(5, 0))

#IN-OUT FRAME FOR COMPANY/TRUCK/TRAILER/BUTTONS with LABELS
IN_OUT_Company = tk.Frame(IN_OUT_Frame_Manual_TOP, bg=conf["window_bg"], highlightthickness=0)
IN_OUT_Company.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
IN_OUT_Company_Lb = tk.Label(IN_OUT_Company, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
IN_OUT_Company_Lb.pack(fill=tk.X, side=tk.TOP)
IN_OUT_Truck = tk.Frame(IN_OUT_Frame_Manual_TOP, bg=conf["window_bg"], highlightthickness=0)
IN_OUT_Truck.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, padx=5)
IN_OUT_Truck_Lb = tk.Label(IN_OUT_Truck, text="TRUCK:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
IN_OUT_Truck_Lb.pack(fill=tk.X, side=tk.TOP)
IN_OUT_Trailer = tk.Frame(IN_OUT_Frame_Manual_TOP, bg=conf["window_bg"], highlightthickness=0)
IN_OUT_Trailer.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
IN_OUT_Trailer_Lb = tk.Label(IN_OUT_Trailer, text="TRAILER:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
IN_OUT_Trailer_Lb.pack(fill=tk.X, side=tk.TOP)
IN_OUT_Frame_Comment = tk.Frame(IN_OUT_Frame_Manual_BOTTOM, bg=conf["window_bg"], highlightthickness=0)
IN_OUT_Frame_Comment.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
IN_OUT_Comment_lb = tk.Label(IN_OUT_Frame_Comment, text="COMMENT:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
IN_OUT_Comment_lb.pack(fill=tk.X, side=tk.TOP)

# ENTRY FOR MANUAL

Company_Entry = tk.Entry(IN_OUT_Company, bg=conf["entry_bg"], bd=0, font=(conf["entry_font"], conf["entry_size"]), fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=2, textvariable=Company_Var, justify=tk.CENTER, state=tk.DISABLED, disabledbackground=conf["entry_bg"], disabledforeground=conf["entry_fg"])
Company_Entry.pack(fill=tk.BOTH, expand=1)
Truck_Entry = tk.Entry(IN_OUT_Truck, bg=conf["entry_bg"], bd=0, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=2, textvariable=Truck_Var, justify=tk.CENTER)
Truck_Entry.pack(fill=tk.BOTH, expand=1)
Truck_Entry.bind("<KeyRelease>", lambda event: UPPER_CASE(event, obj=Truck_Entry))
Trailer_Entry = tk.Entry(IN_OUT_Trailer, bg=conf["entry_bg"], bd=0, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=2, textvariable=Trailer_Var, justify=tk.CENTER)
Trailer_Entry.pack(fill=tk.BOTH, expand=1)
Trailer_Entry.bind("<KeyRelease>", lambda event: UPPER_CASE(event, obj=Trailer_Entry))
Comment_Entry = tk.Entry(IN_OUT_Frame_Comment, bg=conf["entry_bg"], bd=0, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=2, textvariable=Tenant_Comment, justify=tk.CENTER)
Comment_Entry.pack(fill=tk.BOTH, expand=1)

# IN/OUT BUTTONS
in_button = tk.Button(IN_OUT_Frame_Manual_TOP, bg=conf["in_button_bg"], fg=conf["in_button_fg"], bd=0, highlightthickness=0, activebackground=conf["in_button_sel_bg"], activeforeground=conf["in_button_sel_fg"], width=conf["p_button_w"], text="IN", command=lambda: Tenant_In_Out(True))
in_button.pack_propagate(0)
in_button.pack(side=tk.RIGHT, fill=tk.BOTH, expand=0, padx=(5, 0))
in_button.bind("<Enter>", lambda y: in_button.configure(bg=conf["in_button_sel_bg"], fg=conf["in_button_sel_fg"]))
in_button.bind("<Leave>", lambda z: in_button.configure(bg=conf["in_button_bg"], fg=conf["in_button_fg"]))

out_button = tk.Button(IN_OUT_Frame_Manual_BOTTOM, bg=conf["out_button_bg"], fg=conf["out_button_fg"], bd=0, highlightthickness=0, activebackground=conf["out_button_sel_bg"], activeforeground=conf["out_button_sel_fg"], width=conf["p_button_w"], text="OUT", command=lambda: Tenant_In_Out(False))
out_button.pack_propagate(0)
out_button.pack(side=tk.RIGHT, fill=tk.BOTH, expand=0, padx=(5, 0))
out_button.bind("<Enter>", lambda y: out_button.configure(bg=conf["out_button_sel_bg"], fg=conf["out_button_sel_fg"]))
out_button.bind("<Leave>", lambda z: out_button.configure(bg=conf["out_button_bg"], fg=conf["out_button_fg"]))

# LAST EVENT FRAME WITH LABELS
Last_Event_Company_Frame = tk.Frame(IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
Last_Event_Company_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1, padx=(0,5))
Last_Event_Company = tk.Label(Last_Event_Company_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=20)
Last_Event_Company.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
Last_Event_Truck_Frame = tk.Frame(IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
Last_Event_Truck_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
Last_Event_Truck = tk.Label(Last_Event_Truck_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=9)
Last_Event_Truck.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
Last_Event_Trailer_Frame = tk.Frame(IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
Last_Event_Trailer_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
Last_Event_Trailer = tk.Label(Last_Event_Trailer_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=9)
Last_Event_Trailer.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
Last_Event_Date_Frame = tk.Frame(IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
Last_Event_Date_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
Last_Event_Date = tk.Label(Last_Event_Date_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=8)
Last_Event_Date.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
Last_Event_Time_Frame = tk.Frame(IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
Last_Event_Time_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
Last_Event_Time = tk.Label(Last_Event_Time_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=8)
Last_Event_Time.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
Last_Event_Status_Frame = tk.Frame(IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
Last_Event_Status_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
Last_Event_Status = tk.Label(Last_Event_Status_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=3)
Last_Event_Status.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)

#CALCULATE NUMBER OF WIDGED TO THE BORDER OF TRUCK FRAME

xyz = screen_x-int(conf["p_t_company_w"])-35

Parking_Tenant_size = xyz // 150





###########################################################################################
#     PARKING GN FRAME
###########################################################################################
GN_Main = tk.Frame(Parking_Main_Frame, bg=conf["window_bg"], highlightthickness=0)

#TRUCK FRAME AND LABEL
GN_Truck_Frame = tk.Frame(GN_Main, highlightthickness=0, bg=conf["window_bg"], height=conf["p_g_truck_h"])
GN_Truck_Frame.pack_propagate(0)
GN_Truck_Frame.pack(side=tk.TOP, fill=tk.BOTH, expand=0)
GN_Truck_Lb = tk.Label(GN_Truck_Frame, text="TRUCK:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
GN_Truck_Lb.pack(fill=tk.X, side=tk.TOP)
GN_Truck_SubFrame = scroller(GN_Truck_Frame)
GN_Truck_SubFrame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
#TRAILER FRAME AND LABEL
GN_Trailer_Frame = tk.Frame(GN_Main, highlightthickness=0, bg=conf["window_bg"], height=conf["p_g_trailer_h"])
GN_Trailer_Frame.pack_propagate(0)
GN_Trailer_Frame.pack(side=tk.TOP, fill=tk.BOTH, expand=0)
GN_Trailer_Lb = tk.Label(GN_Trailer_Frame, text="TRAILER:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
GN_Trailer_Lb.pack(fill=tk.X, side=tk.TOP)
GN_Trailer_SubFrame = scroller(GN_Trailer_Frame)
GN_Trailer_SubFrame.pack(side=tk.TOP, fill=tk.BOTH, expand=0)
#FLATBED FRAME AND LABEL
GN_Flatbed_Frame = tk.Frame(GN_Main, highlightthickness=0, bg=conf["window_bg"], height=conf["p_g_fb_h"])
#GN_Flatbed_Frame.pack_propagate(0)
GN_Flatbed_Frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
GN_Flatbed_Lb = tk.Label(GN_Flatbed_Frame, text="FLATBED:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
GN_Flatbed_Lb.pack(fill=tk.X, side=tk.TOP)
GN_Flatbed_SubFrame = scroller(GN_Flatbed_Frame)
GN_Flatbed_SubFrame.pack(side=tk.TOP, fill=tk.BOTH, expand=0)
#BOTTOM FRAME - FOR BUTTONS, MANUAL etc.
GN_Bottom_Frame = tk.Frame(GN_Main, highlightthickness=0, bg=conf["window_bg"], height=conf["p_g_bottom_h"])
GN_Bottom_Frame.pack_propagate(0)
GN_Bottom_Frame.pack(side=tk.TOP, fill=tk.BOTH, expand=0, pady=(5, 0))
#FEATURE MAIN FRAME
GN_Feature_Frame = tk.Frame(GN_Bottom_Frame, highlightthickness=0, bg=conf["window_bg"], width=conf["p_g_feature_w"])
GN_Feature_Frame.pack_propagate(0)
GN_Feature_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=0)
#FEATURE FRAME FOR CARRIER BUTTONS
GN_Feature_Company_Frame = tk.Frame(GN_Feature_Frame, highlightthickness=0, bg=conf["window_bg"], height=conf["p_g_feature_h"])
GN_Feature_Company_Frame.pack_propagate(0)
GN_Feature_Company_Frame.pack(side=tk.TOP, fill=tk.BOTH, expand=0)
#CREATING GN BUTTON
GN_Button_lb = tk.Label(GN_Feature_Company_Frame, text="GN", relief=tk.GROOVE, bg=conf["func_button_bg"], fg=conf["func_button_fg"], font=(conf["func_button_font"], conf["func_button_size"]))
GN_Button_lb.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
if GN_Trigger == 1: GN_Button_lb.configure(relief=tk.SUNKEN, fg=conf["func_button_sel_fg"], bg=conf["func_button_sel_bg"])
GN_Button_lb.bind("<Button-1>", lambda x: GN_Button(1))
GN_Button_lb.bind("<Enter>", lambda y: GN_Button_lb.configure(bg=conf["func_button_bg"], fg=conf["func_button_sel_fg"]))
GN_Button_lb.bind("<Leave>", lambda z: GN_Hover_Off(GN_Button_lb, 1, "CARRIER"))
#CREATING W-TUBES BUTTON
GN_WTube_lb = tk.Label(GN_Feature_Company_Frame, text="W-Tubes", relief=tk.GROOVE, bg=conf["func_button_bg"], fg=conf["func_button_fg"], font=(conf["func_button_font"], conf["func_button_size"]))
GN_WTube_lb.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
GN_WTube_lb.bind("<Button-1>", lambda x: GN_Button(2))
GN_WTube_lb.bind("<Enter>", lambda y: GN_WTube_lb.configure(bg=conf["func_button_sel_bg"], fg=conf["func_button_sel_fg"]))
GN_WTube_lb.bind("<Leave>", lambda z: GN_Hover_Off(GN_WTube_lb, 2, "CARRIER"))
#CREATING OTHER BUTTON
GN_Other_lb = tk.Label(GN_Feature_Company_Frame, text="OTHER", relief=tk.GROOVE, bg=conf["func_button_bg"], fg=conf["func_button_fg"], font=(conf["func_button_font"], conf["func_button_size"]))
GN_Other_lb.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
GN_Other_lb.bind("<Button-1>", lambda x: GN_Button(3))
GN_Other_lb.bind("<Enter>", lambda y: GN_Other_lb.configure(bg=conf["func_button_sel_bg"], fg=conf["func_button_sel_fg"]))
GN_Other_lb.bind("<Leave>", lambda z: GN_Hover_Off(GN_Other_lb, 3, "CARRIER"))
#CREATING OTHER ENTRY
GN_Feature_Other_Frame = tk.Frame(GN_Feature_Frame, highlightthickness=0, bg=conf["window_bg"])
GN_Feature_Other_Frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
GN_Other_Entry = AutocompleteEntry(GN_Feature_Other_Frame, background=conf["entry_bg"], font=(conf["entry_font"], conf["entry_size"]), cursor="shuttle", foreground=conf["entry_fg"], state=tk.DISABLED, textvariable=GN_Other_Carrier_Var, completevalues=units_lst("company"))
GN_Other_Entry.pack(fill=tk.BOTH, expand=1, side=tk.TOP)
#CREATING FRAME FOR LOAD UNLOAD BUTTONS
GN_Feature_LoadUnload_Frame = tk.Frame(GN_Feature_Frame, highlightthickness=0, bg=conf["window_bg"], height=conf["p_g_feature_h"])
GN_Feature_LoadUnload_Frame.pack_propagate(0)
GN_Feature_LoadUnload_Frame.pack(side=tk.TOP, fill=tk.BOTH, expand=0)
#CREATING LOAD BUTTON
GN_Load_lb = tk.Label(GN_Feature_LoadUnload_Frame, text="LOADED", relief=tk.GROOVE, bg=conf["func_button_bg"], fg=conf["func_button_fg"], font=(conf["func_button_font"], conf["func_button_size"]), width=1)
GN_Load_lb.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
GN_Load_lb.bind("<Button-1>", lambda x: GN_Button(4))
GN_Load_lb.bind("<Enter>", lambda y: GN_Load_lb.configure(bg=conf["func_button_sel_bg"], fg=conf["func_button_sel_fg"]))
GN_Load_lb.bind("<Leave>", lambda z: GN_Hover_Off(GN_Load_lb, 4, "LU"))
#CREATING UNLOAD BUTTON
GN_Unload_lb = tk.Label(GN_Feature_LoadUnload_Frame, text="EMPTY", relief=tk.GROOVE, bg=conf["func_button_bg"], fg=conf["func_button_fg"], font=(conf["func_button_font"], conf["func_button_size"]), width=1)
GN_Unload_lb.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
GN_Unload_lb.bind("<Button-1>", lambda x: GN_Button(5))
GN_Unload_lb.bind("<Enter>", lambda y: GN_Unload_lb.configure(bg=conf["func_button_sel_bg"], fg=conf["func_button_sel_fg"]))
GN_Unload_lb.bind("<Leave>", lambda z: GN_Hover_Off(GN_Unload_lb, 5, "LU"))
#CREATING ENTRY DATA FRAME - MAIN
GN_Entry_Main_Frame = tk.Frame(GN_Bottom_Frame, highlightthickness=0, bg=conf["window_bg"])
GN_Entry_Main_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1, padx=(5, 0))
#FRAME AND LABELS FOR ENTRIES
GN_Manual_Entry_TOP = tk.Frame(GN_Entry_Main_Frame, highlightthickness=0, bg=conf["window_bg"], height=conf["p_t_manual_h"])
GN_Manual_Entry_TOP.pack_propagate(0)
GN_Manual_Entry_TOP.pack(side=tk.TOP, fill=tk.BOTH, expand=0)
GN_Manual_Entry_BOTTOM = tk.Frame(GN_Entry_Main_Frame, highlightthickness=0, bg=conf["window_bg"], height=conf["p_t_manual_h"])
GN_Manual_Entry_BOTTOM.pack_propagate(0)
GN_Manual_Entry_BOTTOM.pack(side=tk.TOP, fill=tk.BOTH, expand=0, pady=(5, 0))
GN_Manual_Entry_Last_Event = tk.Frame(GN_Entry_Main_Frame, highlightthickness=0, bg=conf["window_bg"], height=conf["p_t_last_event_h"])
GN_Manual_Entry_Last_Event.pack_propagate(0)
GN_Manual_Entry_Last_Event.pack(side=tk.TOP, fill=tk.BOTH, expand=0, pady=(5, 0))
GN_Manual_Truck_Frame = tk.Frame(GN_Manual_Entry_TOP, highlightthickness=0, bg=conf["window_bg"], width=1)
GN_Manual_Truck_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1, padx=(0, 5))
GN_Manual_Truck_lb = tk.Label(GN_Manual_Truck_Frame, text="TRUCK:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
GN_Manual_Truck_lb.pack(fill=tk.X, expand=0, side=tk.TOP)
GN_Manual_Trailer_Frame = tk.Frame(GN_Manual_Entry_TOP, highlightthickness=0, bg=conf["window_bg"], width=1)
GN_Manual_Trailer_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1, padx=(0, 5))
GN_Manual_Trailer_lb = tk.Label(GN_Manual_Trailer_Frame, text="TRAILER/FLATBED:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
GN_Manual_Trailer_lb.pack(fill=tk.X, expand=0, side=tk.TOP)


# #CREATING GN ENTRIES
# def UPPER_CASE_GN_TRUCK(event):
#     v = GN_Entry_Truck.get()
#     if v:
#         GN_Entry_Truck.delete(0, tk.END)
#         GN_Entry_Truck.insert(0, v.upper())
# def UPPER_CASE_GN_TRAILER(event):
#     w = GN_Entry_Trailer_Fb.get()
#     if w:
#         GN_Entry_Trailer_Fb.delete(0, tk.END)
#         GN_Entry_Trailer_Fb.insert(0, w.upper())

GN_Entry_Truck = tk.Entry(GN_Manual_Truck_Frame, bg=conf["entry_bg"], bd=0, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=2, textvariable=GN_Truck_Var, justify=tk.CENTER)
GN_Entry_Truck.pack(fill=tk.BOTH, expand=1, side=tk.TOP)
GN_Entry_Truck.bind("<KeyRelease>", lambda event: UPPER_CASE(event, obj=GN_Entry_Truck))
GN_Entry_Trailer_Fb = tk.Entry(GN_Manual_Trailer_Frame, bg=conf["entry_bg"], bd=0, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=2, textvariable=GN_Trailer_Fb_Var, justify=tk.CENTER)
GN_Entry_Trailer_Fb.pack(fill=tk.BOTH, expand=1, side=tk.TOP)
GN_Entry_Trailer_Fb.bind("<KeyRelease>", lambda event: UPPER_CASE(event, obj=GN_Entry_Trailer_Fb))

#FRAME AND LABEL FOR COMMENTS
GN_Comment_Frame = tk.Frame(GN_Manual_Entry_BOTTOM, highlightthickness=0, bg=conf["window_bg"])
GN_Comment_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1, padx=(0, 5))
GN_Comment_lb = tk.Label(GN_Comment_Frame, text="COMMENT:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
GN_Comment_lb.pack(fill=tk.X, side=tk.TOP)
#CREATING COMMENT ENTRY
GN_Comment_Entry = tk.Entry(GN_Comment_Frame, bg=conf["entry_bg"], bd=0, font=(conf["entry_font"], conf["entry_size"]), cursor="shuttle", fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=2, textvariable=GN_Comment_Var, justify=tk.CENTER)
GN_Comment_Entry.pack(fill=tk.BOTH, expand=1, side=tk.TOP)

# GN IN/OUT BUTTONS
GN_in_button = tk.Button(GN_Manual_Entry_TOP, bg=conf["in_button_bg"], fg=conf["in_button_fg"], bd=0, highlightthickness=0, activebackground=conf["in_button_sel_bg"], activeforeground=conf["in_button_sel_fg"], width=conf["p_button_w"], text="IN", command=lambda: GN_In_Out(True))
GN_in_button.pack_propagate(0)
GN_in_button.pack(side=tk.RIGHT, fill=tk.BOTH, expand=0)
GN_in_button.bind("<Enter>", lambda y: GN_in_button.configure(bg=conf["in_button_sel_bg"], fg=conf["in_button_sel_fg"]))
GN_in_button.bind("<Leave>", lambda z: GN_in_button.configure(bg=conf["in_button_bg"], fg=conf["in_button_fg"]))
GN_out_button = tk.Button(GN_Manual_Entry_BOTTOM, bg=conf["out_button_bg"], fg=conf["out_button_fg"], bd=0, highlightthickness=0, activebackground=conf["out_button_sel_bg"], activeforeground=conf["out_button_sel_fg"], width=conf["p_button_w"], text="OUT", command=lambda: GN_In_Out(False))
GN_out_button.pack_propagate(0)
GN_out_button.pack(side=tk.RIGHT, fill=tk.BOTH, expand=0)
GN_out_button.bind("<Enter>", lambda y: GN_out_button.configure(bg=conf["out_button_sel_bg"], fg=conf["out_button_sel_fg"]))
GN_out_button.bind("<Leave>", lambda z: GN_out_button.configure(bg=conf["out_button_bg"], fg=conf["out_button_fg"]))

#FRAME FOR LAST EVENT
GN_Last_Event_Carrier_Frame = tk.Frame(GN_Manual_Entry_Last_Event, highlightthickness=0, bg=conf["window_bg"])
GN_Last_Event_Carrier_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
GN_Last_Event_Carrier_lb = tk.Label(GN_Last_Event_Carrier_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=20)
GN_Last_Event_Carrier_lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
GN_Last_Event_Truck_Frame = tk.Frame(GN_Manual_Entry_Last_Event, highlightthickness=0, bg=conf["window_bg"])
GN_Last_Event_Truck_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
GN_Last_Event_Truck_lb = tk.Label(GN_Last_Event_Truck_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=9)
GN_Last_Event_Truck_lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
GN_Last_Event_Trailer_Frame = tk.Frame(GN_Manual_Entry_Last_Event, highlightthickness=0, bg=conf["window_bg"])
GN_Last_Event_Trailer_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
GN_Last_Event_Trailer_lb = tk.Label(GN_Last_Event_Trailer_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=9)
GN_Last_Event_Trailer_lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
GN_Last_Event_Date_Frame = tk.Frame(GN_Manual_Entry_Last_Event, highlightthickness=0, bg=conf["window_bg"])
GN_Last_Event_Date_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
GN_Last_Event_Date_lb = tk.Label(GN_Last_Event_Date_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=8)
GN_Last_Event_Date_lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
GN_Last_Event_Time_Frame = tk.Frame(GN_Manual_Entry_Last_Event, highlightthickness=0, bg=conf["window_bg"])
GN_Last_Event_Time_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
GN_Last_Event_Time_lb = tk.Label(GN_Last_Event_Time_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=8)
GN_Last_Event_Time_lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
GN_Last_Event_Status_Frame = tk.Frame(GN_Manual_Entry_Last_Event, highlightthickness=0, bg=conf["window_bg"])
GN_Last_Event_Status_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
GN_Last_Event_Status_lb = tk.Label(GN_Last_Event_Status_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=3)
GN_Last_Event_Status_lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)

#SIZE OF WINDOW
Parking_GN_size = (screen_x // 150)




###########################################################################################
# PARKING VISITOR FRAME
###########################################################################################

VISITOR_Main = tk.Frame(Parking_Main_Frame, bg=conf["window_bg"], highlightthickness=0)

# VISITOR COMPANY FRAME
Visitor_Company_Frame = tk.Frame(VISITOR_Main, bg=conf["window_bg"], highlightthickness=0, width=conf["p_t_company_w"])
Visitor_Company_Frame.pack_propagate(0)
Visitor_Company_Frame.pack(fill=tk.BOTH, side=tk.LEFT)

#LABEL FOR COMPANY
Visitor_Company_Lb = tk.Label(Visitor_Company_Frame, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
Visitor_Company_Lb.pack(fill=tk.X, side=tk.TOP)
# PARKING FRAME FOR COMPANIES WITH SCROLL
sub_vis_comp_frame = tk.Frame(Visitor_Company_Frame, highlightthickness=0)
sub_vis_comp_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
vis_comp_canv = tk.Canvas(sub_vis_comp_frame, highlightthickness=0, bg=conf["window_bg"])
vis_second_comp_Frame = tk.Frame(vis_comp_canv, bg=conf["window_bg"])
vis_comp_scrl = ttk.Scrollbar(Visitor_Company_Frame, orient=tk.VERTICAL, command=vis_comp_canv.yview)
vis_comp_canv.config(yscrollcommand=vis_comp_scrl.set)
vis_comp_scrl.pack(fill=tk.Y, side=tk.RIGHT)
vis_comp_canv.pack(fill=tk.BOTH, side=tk.TOP, expand=1)
vis_comp_canv.create_window((0, 0), window=vis_second_comp_Frame, anchor=tk.NW)
vis_second_comp_Frame.bind("<Configure>", lambda event, canvas=vis_comp_canv: vis_comp_canv.configure(scrollregion=vis_comp_canv.bbox("all")))

def check_T_V_scroll_region(*event):
    canvas_height = vis_second_comp_Frame.winfo_height()
    vis_comp_canv.configure(scrollregion=(0,0,0,canvas_height))
    if canvas_height <= vis_comp_canv.winfo_height():
        vis_comp_scrl.pack_forget()
        vis_comp_canv.configure(yscrollcommand=None)
        vis_second_comp_Frame.unbind("<Enter>")
        vis_second_comp_Frame.unbind_all("<MouseWheel>")
    else:
        vis_comp_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        vis_comp_canv.configure(yscrollcommand=vis_comp_scrl.set)
        vis_second_comp_Frame.bind("<Enter>", _enter_mousewheel_vis_comp, add="+")
vis_comp_canv.bind("<Configure>", check_T_V_scroll_region)

#MOUSEWHEEL FUNCTION
def _on_mousewheel_vis_comp(event): vis_comp_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def _enter_mousewheel_vis_comp(event): vis_comp_canv.bind_all('<MouseWheel>', _on_mousewheel_vis_comp, add="+")
def _leave_mousewheel_vis_comp(event): vis_comp_canv.unbind_all('<MouseWheel>')
vis_second_comp_Frame.bind("<Enter>", _enter_mousewheel_vis_comp, add="+")
vis_second_comp_Frame.bind("<Leave>", _leave_mousewheel_vis_comp)

#CAR FRAME
Vis_Main_Frame = tk.Frame(VISITOR_Main, bg=conf["window_bg"], highlightthickness=0, width=screen_x-int(conf["p_t_company_w"])-15, height=conf["p_t_truck_h"])
Vis_Main_Frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1, padx=(5, 0))

#LABELS FOR VISITORS CARS
VIS_Label_Frame = tk.Frame(Vis_Main_Frame, bg=conf["window_bg"], highlightthickness=0)
VIS_Label_Frame.pack(fill=tk.X, side=tk.TOP)
VIS_Plates_Lb = tk.Label(VIS_Label_Frame, text="PLATE:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=8)
VIS_Plates_Lb.pack(fill=tk.X, side=tk.LEFT, expand=1)
VIS_Car_Lb = tk.Label(VIS_Label_Frame, text="CAR:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=12)
VIS_Car_Lb.pack(fill=tk.X, side=tk.LEFT, padx=1, expand=1)
VIS_Name_Lb= tk.Label(VIS_Label_Frame, text="NAME:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=12)
VIS_Name_Lb.pack(fill=tk.X, side=tk.LEFT, expand=1)
VIS_Exp_Lb= tk.Label(VIS_Label_Frame, text="EXPIRE:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
VIS_Exp_Lb.pack(fill=tk.X, side=tk.LEFT, padx=(1, 0),expand=1)
VIS_CarList_Frame = tk.Frame(Vis_Main_Frame, bg=conf["window_bg"], highlightthickness=0)
VIS_CarList_Frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1)
sub_vis_car_frame = tk.Frame(VIS_CarList_Frame, highlightthickness=0)
sub_vis_car_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.NW)
vis_car_canv = tk.Canvas(sub_vis_car_frame, highlightthickness=0, bg=conf["window_bg"])
Vis_Second_Frame = tk.Frame(vis_car_canv, bg=conf["window_bg"])
vis_car_scrl = ttk.Scrollbar(VIS_CarList_Frame, orient=tk.VERTICAL, command=vis_car_canv.yview)
vis_car_canv.config(yscrollcommand=vis_car_scrl.set)
vis_car_scrl.pack(fill=tk.Y, side=tk.RIGHT)
vis_car_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
vis_car_canv.create_window((0, 0), window=Vis_Second_Frame, anchor=tk.NW)
Vis_Second_Frame.bind("<Configure>", lambda event, canvas=vis_car_canv: vis_car_canv.configure(scrollregion=vis_car_canv.bbox("all")))

def check_T_Vc_scroll_region(*event):
    canvas_height = Vis_Second_Frame.winfo_height()
    vis_car_canv.configure(scrollregion=(0,0,0,canvas_height))
    if canvas_height <= vis_car_canv.winfo_height():
        vis_car_scrl.pack_forget()
        vis_car_canv.configure(yscrollcommand=None)
        Vis_Second_Frame.unbind("<Enter>")
        Vis_Second_Frame.unbind_all("<MouseWheel>")
    else:
        vis_car_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        vis_car_canv.configure(yscrollcommand=vis_car_scrl.set)
        Vis_Second_Frame.bind("<Enter>", _enter_mousewheel_vis_cars, add="+")
vis_car_canv.bind("<Configure>", check_T_Vc_scroll_region)

#MOUSEWHEEL FUNCTION
def _on_mousewheel_vis_cars(event): vis_car_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def _enter_mousewheel_vis_cars(event): vis_car_canv.bind_all('<MouseWheel>', _on_mousewheel_vis_cars, add="+")
def _leave_mousewheel_vis_cars(event): vis_car_canv.unbind_all('<MouseWheel>')
Vis_Second_Frame.bind("<Enter>", _enter_mousewheel_vis_cars, add="+")
Vis_Second_Frame.bind("<Leave>", _leave_mousewheel_vis_cars)

#IN-OUT VISITOR FRAME
VIS_IN_OUT_Frame = tk.Frame(VISITOR_Main, bg=conf["window_bg"], highlightthickness=0)
VIS_IN_OUT_Frame.pack(fill=tk.BOTH, side=tk.TOP, expand=0, padx=(5, 0), pady=(5, 0))
#IN-OUT VISITOR SUB FRAME FOR MANUAL ENTRANCE AND BUTTONS
VIS_IN_OUT_Frame_Manual_TOP = tk.Frame(VIS_IN_OUT_Frame, bg=conf["window_bg"], highlightthickness=0, height=conf["p_t_manual_h"])
VIS_IN_OUT_Frame_Manual_TOP.pack_propagate(0)
VIS_IN_OUT_Frame_Manual_TOP.pack(fill=tk.BOTH, side=tk.TOP, expand=1)
VIS_IN_OUT_Frame_Manual_BOTTOM = tk.Frame(VIS_IN_OUT_Frame, bg=conf["window_bg"], highlightthickness=0, height=conf["p_t_manual_h"])
VIS_IN_OUT_Frame_Manual_BOTTOM.pack_propagate(0)
VIS_IN_OUT_Frame_Manual_BOTTOM.pack(fill=tk.BOTH, side=tk.TOP, expand=1, pady=(5, 0))
VIS_IN_OUT_Frame_Last_Event = tk.Frame(VIS_IN_OUT_Frame, bg=conf["window_bg"], highlightthickness=0, height=conf["p_t_last_event_h"])
VIS_IN_OUT_Frame_Last_Event.pack_propagate(0)
VIS_IN_OUT_Frame_Last_Event.pack(fill=tk.BOTH, side=tk.TOP, expand=0, pady=(5, 0))

#IN-OUT FRAME FOR COMPANY/PLATES/CAR/DRIVER
VIS_IN_OUT_Company = tk.Frame(VIS_IN_OUT_Frame_Manual_TOP, bg=conf["window_bg"], highlightthickness=0)
VIS_IN_OUT_Company.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
VIS_IN_OUT_Company_Lb = tk.Label(VIS_IN_OUT_Company, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
VIS_IN_OUT_Company_Lb.pack(fill=tk.X, side=tk.TOP)
VIS_IN_OUT_Plates = tk.Frame(VIS_IN_OUT_Frame_Manual_TOP, bg=conf["window_bg"], highlightthickness=0)
VIS_IN_OUT_Plates.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, padx=5)
VIS_IN_OUT_Plates_Lb = tk.Label(VIS_IN_OUT_Plates, text="PLATES:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
VIS_IN_OUT_Plates_Lb.pack(fill=tk.X, side=tk.TOP)
VIS_IN_OUT_Car = tk.Frame(VIS_IN_OUT_Frame_Manual_TOP, bg=conf["window_bg"], highlightthickness=0)
VIS_IN_OUT_Car.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
VIS_IN_OUT_Car_Lb = tk.Label(VIS_IN_OUT_Car, text="CAR:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
VIS_IN_OUT_Car_Lb.pack(fill=tk.X, side=tk.TOP)
VIS_IN_OUT_Name = tk.Frame(VIS_IN_OUT_Frame_Manual_TOP, bg=conf["window_bg"], highlightthickness=0)
VIS_IN_OUT_Name.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, padx=(5, 0))
VIS_IN_OUT_Name_Lb = tk.Label(VIS_IN_OUT_Name, text="NAME:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
VIS_IN_OUT_Name_Lb.pack(fill=tk.X, side=tk.TOP)
VIS_IN_OUT_Frame_Comment = tk.Frame(VIS_IN_OUT_Frame_Manual_BOTTOM, bg=conf["window_bg"], highlightthickness=0)
VIS_IN_OUT_Frame_Comment.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
VIS_IN_OUT_Comment_lb = tk.Label(VIS_IN_OUT_Frame_Comment, text="COMMENT:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
VIS_IN_OUT_Comment_lb.pack(fill=tk.X, side=tk.TOP)

# # ENTRY FOR MANUAL
# def UPPER_CASE_PLATES(event):
#     v = VIS_Plates_Entry.get()
#     if v:
#         VIS_Plates_Entry.delete(0, tk.END)
#         VIS_Plates_Entry.insert(0, v.upper())

VIS_Company_Entry = tk.Entry(VIS_IN_OUT_Company, bg=conf["entry_bg"], bd=0, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=2, textvariable=VIS_Company_Var, justify=tk.CENTER, width=1)
VIS_Company_Entry.pack(fill=tk.BOTH, expand=0)
VIS_Plates_Entry = tk.Entry(VIS_IN_OUT_Plates, bg=conf["entry_bg"], bd=0, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=2, textvariable=VIS_Plates_Var, justify=tk.CENTER, width=1)
VIS_Plates_Entry.pack(fill=tk.BOTH, expand=0)
VIS_Plates_Entry.bind("<KeyRelease>", lambda event: UPPER_CASE(event, obj=VIS_Plates_Entry))

car_ven = SQL_REQ("SELECT Vendor FROM dbo.Car_Vendors ORDER BY Vendor", (), "S_all")
vendors = []
for x in car_ven: vendors.append(x[0])
VIS_Car_Entry = AutocompleteEntry(VIS_IN_OUT_Car, background=conf["entry_bg"], cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), foreground=conf["entry_fg"], completevalues=vendors, textvariable=VIS_Car_Var, width=1)
VIS_Car_Entry.pack(fill=tk.BOTH, expand=0)
VIS_Name_Entry = tk.Entry(VIS_IN_OUT_Name, bg=conf["entry_bg"], bd=0, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=2, textvariable=VIS_Name_Var, justify=tk.CENTER,width=1)
VIS_Name_Entry.pack(fill=tk.BOTH, expand=0)
VIS_Comment_Entry = tk.Entry(VIS_IN_OUT_Frame_Comment, bg=conf["entry_bg"], bd=0, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=2, textvariable=VIS_Comment, justify=tk.CENTER)
VIS_Comment_Entry.pack(fill=tk.BOTH, expand=0)

# VISITOR IN/OUT BUTTONS
vis_in_button = tk.Button(VIS_IN_OUT_Frame_Manual_TOP, bg=conf["in_button_bg"], fg=conf["in_button_fg"], bd=0, highlightthickness=0, activebackground=conf["in_button_sel_bg"], activeforeground=conf["in_button_sel_fg"], width=conf["p_button_w"], text="IN", command=lambda: VIS_IN_OUT(True))
vis_in_button.pack_propagate(0)
vis_in_button.pack(side=tk.RIGHT, fill=tk.BOTH, expand=0, padx=(5, 0))
vis_in_button.bind("<Enter>", lambda y: vis_in_button.configure(bg=conf["in_button_sel_bg"], fg=conf["in_button_sel_fg"]))
vis_in_button.bind("<Leave>", lambda z: vis_in_button.configure(bg=conf["in_button_bg"], fg=conf["in_button_fg"]))

vis_out_button = tk.Button(VIS_IN_OUT_Frame_Manual_BOTTOM, bg=conf["out_button_bg"], fg=conf["out_button_fg"], bd=0, highlightthickness=0, activebackground=conf["out_button_sel_bg"], activeforeground=conf["out_button_sel_fg"], width=conf["p_button_w"], text="OUT", command=lambda: VIS_IN_OUT(False))
vis_out_button.pack_propagate(0)
vis_out_button.pack(side=tk.RIGHT, fill=tk.BOTH, expand=0, padx=(5, 0))
vis_out_button.bind("<Enter>", lambda y: vis_out_button.configure(bg=conf["out_button_sel_bg"], fg=conf["out_button_sel_fg"]))
vis_out_button.bind("<Leave>", lambda z: vis_out_button.configure(bg=conf["out_button_bg"], fg=conf["out_button_fg"]))

# VISITOR LAST EVENT FRAME WITH LABELS
VIS_Last_Event_Company_Frame = tk.Frame(VIS_IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
VIS_Last_Event_Company_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1, padx=(0,5))
VIS_Last_Event_Company = tk.Label(VIS_Last_Event_Company_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=18)
VIS_Last_Event_Company.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
VIS_Last_Event_Plates_Frame = tk.Frame(VIS_IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
VIS_Last_Event_Plates_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
VIS_Last_Event_Plates = tk.Label(VIS_Last_Event_Plates_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=8)
VIS_Last_Event_Plates.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
VIS_Last_Event_Car_Frame = tk.Frame(VIS_IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
VIS_Last_Event_Car_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
VIS_Last_Event_Car = tk.Label(VIS_Last_Event_Car_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=10)
VIS_Last_Event_Car.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
VIS_Last_Event_Date_Frame = tk.Frame(VIS_IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
VIS_Last_Event_Date_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
VIS_Last_Event_Date = tk.Label(VIS_Last_Event_Date_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=10)
VIS_Last_Event_Date.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
VIS_Last_Event_Time_Frame = tk.Frame(VIS_IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
VIS_Last_Event_Time_Frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5), expand=1)
VIS_Last_Event_Time = tk.Label(VIS_Last_Event_Time_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=8)
VIS_Last_Event_Time.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
VIS_Last_Event_Status_Frame = tk.Frame(VIS_IN_OUT_Frame_Last_Event, highlightthickness=0, bg=conf["window_bg"])
VIS_Last_Event_Status_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
VIS_Last_Event_Status = tk.Label(VIS_Last_Event_Status_Frame, bg=conf["window_bg"], font=(conf["status_font"], conf["status_size"]), fg=conf["status_fg"], relief=tk.GROOVE, bd=3, width=3)
VIS_Last_Event_Status.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)


################################################################################################################################################################################
################################################################################################################################################################################
#                       HISTORY
################################################################################################################################################################################
################################################################################################################################################################################

# TENANT CLICKED FUNC - REMIND CLICKED POSITION, REMOVE PREVIOUS AND PACK TENANT FRAME
def H_Tenant(event):
    global H_Parking_Var
    H_Parking_Var = 1
    H_GN_lb.configure(bg=conf["submenu_bg"])
    H_Visitor_lb.configure(bg=conf["submenu_bg"])
#    Refresh("GN")
    H_GN_Main.pack_forget()
#    Refresh("Visitor")
    H_VISITOR_Main.pack_forget()
    H_Tenant_Main.pack(fill=tk.BOTH, expand=1, padx=5, pady=5, side=tk.TOP)
    for widgets in second_history_frame.winfo_children(): widgets.destroy()
    H_insert(second_history_frame, H_Parking_Var)



# GN CLICKED FUNC - REMIND CLICKED POSITION, REMOVE PREVIOUS AND PACK GN FRAME
def H_GN(event):
    global H_Parking_Var
    H_Parking_Var = 2
    H_Tenant_lb.configure(bg=conf["submenu_bg"])
    H_Visitor_lb.configure(bg=conf["submenu_bg"])
#    Refresh("Tenant")
    H_Tenant_Main.pack_forget()
 #   Refresh("Visitor")
    H_VISITOR_Main.pack_forget()
    H_GN_Main.pack(fill=tk.BOTH, expand=1, padx=5, pady=5)
    for widgets in second_history_GN_frame.winfo_children(): widgets.destroy()
    H_insert(second_history_GN_frame, H_Parking_Var)

# VISITOR CLICKED - REMIND CLICKED POSITION, REMOVE PREVIOUS AND PACK VISITORS FRAME
def H_Visitor(event):
    global H_Parking_Var
    H_Parking_Var = 3
    H_Tenant_lb.configure(bg=conf["submenu_bg"])
    H_GN_lb.configure(bg=conf["submenu_bg"])
 #   Refresh("Tenant")
    H_Tenant_Main.pack_forget()
 #   Refresh("GN")
    H_GN_Main.pack_forget()
    H_VISITOR_Main.pack(fill=tk.BOTH, expand=1, padx=5, pady=5)
    for widgets in second_history_VIS_frame.winfo_children(): widgets.destroy()
    H_insert(second_history_VIS_frame, H_Parking_Var)

def History_Menu_Hover_Off(event):
    if H_Parking_Var != 1: H_Tenant_lb.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if H_Parking_Var != 2: H_GN_lb.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if H_Parking_Var != 3: H_Visitor_lb.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])



#Menu_Bar_History
# HISTORY TAB SUB MENU - TENANT/GN/VISITORS
History_Sub_Menu_Frame = tk.Frame(Menu_Bar_History, bg=conf["window_bg"], highlightthickness=0)
History_Sub_Menu_Frame.pack(fill=tk.BOTH)
#CREATING TENANT BUTTON
H_Tenant_lb = tk.Label(History_Sub_Menu_Frame, text="Tenant", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
H_Tenant_lb.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
H_Tenant_lb.bind("<Button-1>", H_Tenant)
H_Tenant_lb.bind("<Enter>", lambda x: H_Tenant_lb.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
H_Tenant_lb.bind("<Leave>", History_Menu_Hover_Off)
#CREATING GN BUTTON
H_GN_lb = tk.Label(History_Sub_Menu_Frame, text="GN", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
H_GN_lb.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
H_GN_lb.bind("<Button-1>", H_GN)
H_GN_lb.bind("<Enter>", lambda x: H_GN_lb.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
H_GN_lb.bind("<Leave>", History_Menu_Hover_Off)
#CREATING VISITOR BUTTON
H_Visitor_lb = tk.Label(History_Sub_Menu_Frame, text="Visitor", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
H_Visitor_lb.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
H_Visitor_lb.bind("<Button-1>", H_Visitor)
H_Visitor_lb.bind("<Enter>", lambda x: H_Visitor_lb.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
H_Visitor_lb.bind("<Leave>", History_Menu_Hover_Off)

#HISTORY SUBFRAME - UNDER SUB MENU BUTTONS
History_Main_Frame = tk.Frame(Menu_Bar_History, bg=conf["window_bg"])
History_Main_Frame.pack(fill=tk.BOTH, expand=1)

##################################################
# HISTORY TENANT
##################################################
H_Tenant_Main = tk.Frame(History_Main_Frame, bg=conf["window_bg"], highlightthickness=0)

H_Header_Tenant_Frame = tk.Frame(H_Tenant_Main, bg=conf["window_bg"], highlightthickness=0)
H_Header_Tenant_Frame.pack(side=tk.TOP, anchor=tk.W, fill=tk.BOTH)
HT_Company_Lb = tk.Label(H_Header_Tenant_Frame, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
HT_Company_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HT_Truck_Lb = tk.Label(H_Header_Tenant_Frame, text="TRUCK:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HT_Truck_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HT_Trailer_Lb = tk.Label(H_Header_Tenant_Frame, text="TRAILER:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HT_Trailer_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HT_Date_Lb = tk.Label(H_Header_Tenant_Frame, text="DATE:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HT_Date_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HT_Time_Lb = tk.Label(H_Header_Tenant_Frame, text="TIME:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HT_Time_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HT_Status_Lb = tk.Label(H_Header_Tenant_Frame, text="STATUS:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HT_Status_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HT_Comment_Lb = tk.Label(H_Header_Tenant_Frame, text="COMMENT:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
HT_Comment_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HT_Security_Lb = tk.Label(H_Header_Tenant_Frame, text="Attendant:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
HT_Security_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HT_Edit_Lb = tk.Label(H_Header_Tenant_Frame, text="EDIT:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HT_Edit_Lb.pack(side=tk.LEFT, fill=tk.BOTH)

#SCROLLBAR FOR HISOTRY TENANT
history_frame = tk.Frame(H_Tenant_Main, bg=conf["window_bg"], highlightthickness=0)
history_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
sub_H_tenant_frame = tk.Frame(history_frame, highlightthickness=0)
sub_H_tenant_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
H_canv = tk.Canvas(sub_H_tenant_frame, bg=conf["window_bg"], highlightthickness=0)
second_history_frame = tk.Frame(H_canv, bg=conf["window_bg"])
H_scrl = ttk.Scrollbar(history_frame, orient=tk.VERTICAL, command=H_canv.yview)
H_canv.config(yscrollcommand=H_scrl.set)
H_scrl.pack(fill=tk.Y, side=tk.RIGHT)
H_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
H_canv.create_window((0, 0), window=second_history_frame, anchor=tk.NW)
H_canv.bind("<Configure>", lambda event, canvas=H_canv: H_canv.configure(scrollregion=H_canv.bbox("all")))

def check_HT_scroll_region(*event):
    canvas_height = second_history_frame.winfo_height()
    H_canv.configure(scrollregion=(0,0,0,canvas_height))
    if canvas_height <= H_canv.winfo_height():
        H_scrl.pack_forget()
        H_canv.configure(yscrollcommand=None)
        second_history_frame.unbind("<Enter>")
        second_history_frame.unbind_all("<MouseWheel>")
    else:
        H_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        H_canv.configure(yscrollcommand=H_scrl.set)
        second_history_frame.bind("<Enter>", H_enter_mousewheel_tenant_comp, add="+")
H_canv.bind("<Configure>", check_HT_scroll_region)


def H_on_mousewheel(event): H_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def H_enter_mousewheel_tenant_comp(event): H_canv.bind_all('<MouseWheel>', H_on_mousewheel, add="+")
def H_leave_mousewheel_tenant_comp(event): H_canv.unbind_all('<MouseWheel>')
second_history_frame.bind("<Enter>", H_enter_mousewheel_tenant_comp, add="+")
second_history_frame.bind("<Leave>", H_leave_mousewheel_tenant_comp)

#HISTORY GN
H_GN_Main = tk.Frame(History_Main_Frame, bg=conf["window_bg"], highlightthickness=0)
H_Header_GN_Frame = tk.Frame(H_GN_Main, bg=conf["window_bg"], highlightthickness=0)
H_Header_GN_Frame.pack(side=tk.TOP, anchor=tk.W, fill=tk.BOTH)
HG_Company_Lb = tk.Label(H_Header_GN_Frame, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
HG_Company_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HG_Truck_Lb = tk.Label(H_Header_GN_Frame, text="TRUCK:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HG_Truck_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HG_Trailer_Lb = tk.Label(H_Header_GN_Frame, text="TRAILER:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HG_Trailer_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HG_Date_Lb = tk.Label(H_Header_GN_Frame, text="DATE:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HG_Date_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HG_Time_Lb = tk.Label(H_Header_GN_Frame, text="TIME:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HG_Time_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HG_LU_Lb = tk.Label(H_Header_GN_Frame, text="LOADED:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HG_LU_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=(1, 0))
HG_Status_Lb = tk.Label(H_Header_GN_Frame, text="STATUS:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HG_Status_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HG_Comment_Lb = tk.Label(H_Header_GN_Frame, text="COMMENT:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
HG_Comment_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HG_Security_Lb = tk.Label(H_Header_GN_Frame, text="Attendant:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
HG_Security_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HG_Edit_Lb = tk.Label(H_Header_GN_Frame, text="EDIT:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HG_Edit_Lb.pack(side=tk.LEFT, fill=tk.BOTH)

#SCROLLBAR FOR HISOTRY GN
history_GN_frame = tk.Frame(H_GN_Main, bg=conf["window_bg"], highlightthickness=0)
history_GN_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
sub_H_GN_frame = tk.Frame(history_GN_frame, highlightthickness=0)
sub_H_GN_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
HG_canv = tk.Canvas(sub_H_GN_frame, bg=conf["window_bg"], highlightthickness=0)
second_history_GN_frame = tk.Frame(HG_canv, bg=conf["window_bg"])
HG_scrl = ttk.Scrollbar(history_GN_frame, orient=tk.VERTICAL, command=HG_canv.yview)
HG_canv.config(yscrollcommand=HG_scrl.set)
HG_scrl.pack(fill=tk.Y, side=tk.RIGHT)
HG_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
HG_canv.create_window((0, 0), window=second_history_GN_frame, anchor=tk.NW)
HG_canv.bind("<Configure>", lambda event, canvas=HG_canv: HG_canv.configure(scrollregion=HG_canv.bbox("all")))

def check_HG_scroll_region(*event):
    canvas_height = second_history_GN_frame.winfo_height()
    HG_canv.configure(scrollregion=(0,0,0,canvas_height))
    if canvas_height <= HG_canv.winfo_height():
        HG_scrl.pack_forget()
        HG_canv.configure(yscrollcommand=None)
        second_history_GN_frame.unbind("<Enter>")
        second_history_GN_frame.unbind_all("<MouseWheel>")
    else:
        HG_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        HG_canv.configure(yscrollcommand=HG_scrl.set)
        second_history_GN_frame.bind("<Enter>", HG_enter_mousewheel_tenant_comp, add="+")
HG_canv.bind("<Configure>", check_HG_scroll_region)

#mouse function
def HG_on_mousewheel(event): HG_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def HG_enter_mousewheel_tenant_comp(event): HG_canv.bind_all('<MouseWheel>', HG_on_mousewheel, add="+")
def HG_leave_mousewheel_tenant_comp(event): HG_canv.unbind_all('<MouseWheel>')
second_history_GN_frame.bind("<Enter>", HG_enter_mousewheel_tenant_comp, add="+")
second_history_GN_frame.bind("<Leave>", HG_leave_mousewheel_tenant_comp)

#HISTORY VISITOR
H_VISITOR_Main = tk.Frame(History_Main_Frame, bg=conf["window_bg"], highlightthickness=0)
H_Header_VIS_Frame = tk.Frame(H_VISITOR_Main, bg=conf["window_bg"], highlightthickness=0)
H_Header_VIS_Frame.pack(side=tk.TOP, anchor=tk.W, fill=tk.BOTH)
HV_Company_Lb = tk.Label(H_Header_VIS_Frame, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
HV_Company_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HV_Plate_Lb = tk.Label(H_Header_VIS_Frame, text="PLATE:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HV_Plate_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HV_CAR_Lb = tk.Label(H_Header_VIS_Frame, text="CAR:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HV_CAR_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HV_NAME_Lb = tk.Label(H_Header_VIS_Frame, text="DRIVER:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
HV_NAME_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=(1, 0))
HV_Date_Lb = tk.Label(H_Header_VIS_Frame, text="DATE:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HV_Date_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HV_Time_Lb = tk.Label(H_Header_VIS_Frame, text="TIME:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HV_Time_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HV_Status_Lb = tk.Label(H_Header_VIS_Frame, text="STATUS:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HV_Status_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HV_Comment_Lb = tk.Label(H_Header_VIS_Frame, text="COMMENT:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
HV_Comment_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
HV_Security_Lb = tk.Label(H_Header_VIS_Frame, text="Attendant:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
HV_Security_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
HV_Edit_Lb = tk.Label(H_Header_VIS_Frame, text="EDIT:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
HV_Edit_Lb.pack(side=tk.LEFT, fill=tk.BOTH)

#SCROLLBAR FOR HISOTRY GN
history_VIS_frame = tk.Frame(H_VISITOR_Main, bg=conf["window_bg"], highlightthickness=0)
history_VIS_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
sub_H_VIS_frame = tk.Frame(history_VIS_frame, highlightthickness=0)
sub_H_VIS_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
HV_canv = tk.Canvas(sub_H_VIS_frame, bg=conf["window_bg"], highlightthickness=0)
second_history_VIS_frame = tk.Frame(HV_canv, bg=conf["window_bg"])
HV_scrl = ttk.Scrollbar(history_VIS_frame, orient=tk.VERTICAL, command=HV_canv.yview)
HV_canv.config(yscrollcommand=HV_scrl.set)
HV_scrl.pack(fill=tk.Y, side=tk.RIGHT)
HV_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
HV_canv.create_window((0, 0), window=second_history_VIS_frame, anchor=tk.NW)
HV_canv.bind("<Configure>", lambda event, canvas=HV_canv: HV_canv.configure(scrollregion=HV_canv.bbox("all")))

def check_HV_scroll_region(*event):
    canvas_height = second_history_VIS_frame.winfo_height()
    HV_canv.configure(scrollregion=(0,0,0,canvas_height))
    if canvas_height <= HV_canv.winfo_height():
        HV_scrl.pack_forget()
        HV_canv.configure(yscrollcommand=None)
        second_history_VIS_frame.unbind("<Enter>")
        second_history_VIS_frame.unbind_all("<MouseWheel>")
    else:
        HV_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        HV_canv.configure(yscrollcommand=HV_scrl.set)
        second_history_VIS_frame.bind("<Enter>", HV_enter_mousewheel_tenant_comp, add="+")
HV_canv.bind("<Configure>", check_HV_scroll_region)

#mouse function
def HV_on_mousewheel(event): HV_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def HV_enter_mousewheel_tenant_comp(event): HV_canv.bind_all('<MouseWheel>', HV_on_mousewheel, add="+")
def HV_leave_mousewheel_tenant_comp(event): HV_canv.unbind_all('<MouseWheel>')
second_history_VIS_frame.bind("<Enter>", HV_enter_mousewheel_tenant_comp, add="+")
second_history_VIS_frame.bind("<Leave>", HV_leave_mousewheel_tenant_comp)

################################################################################################################################################################################
################################################################################################################################################################################
#                       OVERPARKING
################################################################################################################################################################################
################################################################################################################################################################################

Overparking_Submenu_Frame = tk.Frame(Menu_Bar_Overparking, highlightthickness=0, bg=conf["window_bg"])

Overparking_Submenu_Frame.pack(fill=tk.BOTH, expand=1)
Overparking_Submenu_Frame.pack_propagate(0)
combobox_fr = tk.Frame(Overparking_Submenu_Frame, highlightthickness=0, bg=conf["submenu_bg"])
combobox_fr.pack(fill=tk.X, side=tk.TOP)
Overparking_Preview_Fr = tk.Frame(Overparking_Submenu_Frame, highlightthickness=0, bg=conf["window_bg"])
#Overparking_Preview_Fr.pack_propagate(0)
Overparking_Preview_Fr.pack(fill=tk.BOTH, side=tk.TOP, expand=1)
over_preview_sc = scroller(Overparking_Preview_Fr)
over_preview_sc.pack(side=tk.TOP, fill=tk.BOTH, expand=0)

def over_preview(masta):
    global screen_x
    masta.delete()
    def insert_over(masta, line):
        if line is None or line == []: return
        OVER_Header_Frame = tk.Frame(masta.frame, bg=conf["window_bg"], highlightthickness=0)
        OVER_Header_Frame.pack(side=tk.TOP, anchor=tk.NW, fill=tk.BOTH, pady=5)
        for x in SQL_REQ("SELECT company_name FROM dbo.Company_list WHERE company_ID=? ORDER BY company_name", (str(line[0]["company_ID"]),), "S_one"): c_name = x
        OVER_Company_Lb = tk.Label(OVER_Header_Frame, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
        OVER_Company_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 1))
        OVER_Name_Lb = tk.Label(OVER_Header_Frame, text=c_name, bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
        OVER_Name_Lb.pack(side=tk.LEFT, fill=tk.BOTH)
        for each in line:
            OVER_Date_FR = tk.Frame(masta.frame, bg=conf["window_bg"], highlightthickness=0)
            OVER_Date_FR.pack(side=tk.TOP, fill=tk.BOTH, padx=6)
            OVER_Date_Lb = tk.Label(OVER_Date_FR, text="Date:", bg=conf["widget_sel_bg"], font=(conf["widget_font"], conf["notebook_tab_size"]), fg=conf["window_bg"], width=6)
            OVER_Date_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 1))
            OVER_Date = tk.Label(OVER_Date_FR, text=each["date"], bg=conf["widget_sel_bg"], font=(conf["widget_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=10)
            OVER_Date.pack(side=tk.LEFT, fill=tk.BOTH)
            OVER_Count_Lb = tk.Label(OVER_Date_FR, text="Over:", bg=conf["widget_sel_bg"], font=(conf["widget_font"], conf["notebook_tab_size"]), fg=conf["window_bg"], width=6)
            OVER_Count_Lb.pack(side=tk.LEFT, fill=tk.BOTH, padx=1)
            OVER_Count = tk.Label(OVER_Date_FR, text=each["over_count"], bg=conf["widget_sel_bg"], font=(conf["widget_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=3)
            OVER_Count.pack(side=tk.LEFT, fill=tk.BOTH)
            OVER_Trucklb_Fr = tk.Frame(masta.frame, bg=conf["window_bg"], highlightthickness=0)
            OVER_Trucklb_Fr.pack(side=tk.TOP, fill=tk.BOTH)
            OVER_Trucks_Lb = tk.Label(OVER_Trucklb_Fr, text="Trucks:", bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_fg"])
            OVER_Trucks_Lb.pack(side=tk.LEFT, fill=tk.BOTH, pady=(1, 0))
            OVER_Trucks_Fr = tk.Frame(masta.frame, bg=conf["window_bg"], highlightthickness=0)
            OVER_Trucks_Fr.pack(side=tk.TOP, fill=tk.Y, anchor=tk.W, padx=(5, 0))
            trucks_list = each["trucks_onyard"]
            counterSize = screen_x // 90
            counterX = 0
            counterY = 0
            for item in trucks_list:
                OVER_Truck_Item = tk.Label(OVER_Trucks_Fr, text=item, bg=conf["widget_bg"], font=(conf["widget_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], width=9)
                OVER_Truck_Item.grid(row=counterX, column=counterY, padx=1, pady=1, sticky=tk.NSEW)
                if counterY == counterSize-1:
                    counterX += 1
                    counterY = 0
                else:
                    counterY += 1
            OVER_Trailerlb_Fr = tk.Frame(masta.frame, bg=conf["window_bg"], highlightthickness=0)
            OVER_Trailerlb_Fr.pack(side=tk.TOP, fill=tk.BOTH, pady=(5, 0))
            OVER_Trailers_Lb = tk.Label(OVER_Trailerlb_Fr, text="Trailers:", bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_fg"])
            OVER_Trailers_Lb.pack(side=tk.LEFT, fill=tk.BOTH, pady=(1, 0))
            OVER_Trailers_Fr = tk.Frame(masta.frame, bg=conf["window_bg"], highlightthickness=0)
            OVER_Trailers_Fr.pack(side=tk.TOP, fill=tk.Y, anchor=tk.W, padx=5, pady=(0, 5))
            trailer_list = each["trailers_onyard"]
            counterSize = screen_x // 90
            counterX = 0
            counterY = 0
            for item in trailer_list:
                OVER_Trailer_Item = tk.Label(OVER_Trailers_Fr, text=item, bg=conf["widget_bg"], font=(conf["widget_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], width=9)
                OVER_Trailer_Item.grid(row=counterX, column=counterY, padx=1, pady=1, sticky=tk.NSEW)
                if counterY == counterSize - 1:
                    counterX += 1
                    counterY = 0
                else:
                    counterY += 1

    month = cb_month.get()
    year = cb_year.get()
    comp = cb_comp.get()
    date = datetime.strptime(year + "-" + month, "%Y-%B").date()

    if comp == "All":
        comp_list = SQL_REQ("SELECT company_ID FROM dbo.OVERPARKING WHERE month(date)=? and year(date)=?", (str(date.strftime("%m")), year), "S_all")
        if comp_list:
            comp_lst = set(list(x[0] for x in comp_list))
            for all in comp_lst:
                over_line = over_extract(int(date.strftime("%m")), int(year), all)
                insert_over(masta, over_line)
            else: return
        else: return
    else:
        comp_id = SQL_REQ("SELECT company_ID FROM dbo.Company_list WHERE company_name=? ORDER BY company_name", (comp,), "S_one")
        if comp_id:
            # c_ID = comp_id[0]
            over_line = over_extract(int(date.strftime("%m")), int(year), comp_id[0])
            insert_over(masta, over_line)
    masta.refresh()
    masta.top()
#Function of generating overparking and pass it to EXL generator
def generate():
    month = cb_month.get()
    year = cb_year.get()
    comp = cb_comp.get()
    date = datetime.strptime(year+"-"+month, "%Y-%B").date()
    if comp == "All":
        company_list = []
        row = SQL_REQ("SELECT company_ID FROM dbo.Company_list WHERE activity=1 ORDER BY company_name", (), "S_all")
        for x in row: company_list.append(x[0])
        for all in set(company_list):
            to_Excel(date, all)
    else:
        for x in SQL_REQ("SELECT company_ID FROM dbo.Company_list WHERE company_name=? ORDER BY company_name", (comp,), "S_one"): c_ID = x
        to_Excel(date, c_ID)
    #removing 12 month old over records
    try:
        SQL_REQ("DELETE FROM dbo.OVERPARKING WHERE [date]<DATEADD(year, -1, GETDATE())",(), "W")
        messagebox.showinfo("Information", f"Overparking have been created into folder: {sets['SQL_path']}\nOld Overparking were DELETED!")
    except Exception as e:
        error(f"Cannot delete records of old Overparking:\n{e}")
        debuger(e)

year_list = list(sets["Year_List"].split("|"))
month_list = list(sets["Month_List"].split("|"))

#initialisin list of all companies
comp_list = ["All"] + units_lst("company")

#comp_list.extend(units_lst("company"))
today = datetime.now()
cb_year = ttk.Combobox(combobox_fr, values=year_list, width=10, background=conf["submenu_bg"], foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
cb_year.current(int(today.strftime("%Y"))-2023)
cb_year.pack(fill=tk.BOTH, side=tk.LEFT, padx=5, pady=5)
cb_month = ttk.Combobox(combobox_fr, values=month_list, width=10, background=conf["submenu_bg"], foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))

if int(today.strftime("%m")) == 1:
    def_month = 11
    cb_year.current(int(today.strftime("%Y")) - 2024)
else: def_month = int(today.strftime("%m"))-2
cb_month.current(def_month)
cb_month.pack(fill=tk.BOTH, side=tk.LEFT, padx=5, pady=5)

cb_comp = ttk.Combobox(combobox_fr, values=comp_list, width=25, background=conf["submenu_bg"], foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
cb_comp.current(0)
cb_comp.pack(fill=tk.BOTH, side=tk.LEFT, padx=5, pady=5)



over_gen_button = tk.Button(combobox_fr, text="GENERATE", width=10, command=generate)
over_gen_button.pack(fill=tk.BOTH, side=tk.RIGHT, padx=5, pady=5)

over_gen_button = tk.Button(combobox_fr, text="PREVIEW", width=10, command=lambda: over_preview(over_preview_sc))
over_gen_button.pack(fill=tk.BOTH, side=tk.RIGHT, padx=5, pady=5)

################################################################################################################################################################################
################################################################################################################################################################################
##                      GN
################################################################################################################################################################################
################################################################################################################################################################################
def TEN_stat_insert(frame):
    today = date.today()
    data = units_lst("", "tenant_by_comp")
    c_frame = tk.Frame(frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    c_frame.pack(side=tk.TOP, fill=tk.X, expand=1, pady=(0, 2), padx=(2, 0))

    truck_label = tk.Label(c_frame, text="Trucks:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=GN_screen - 2)
    truck_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
    truck_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    truck_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
    column_names_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
    unit_lb = tk.Label(column_names_fr, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    unit_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    onyard_lb = tk.Label(column_names_fr, text="on yard:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    onyard_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    cargo_lb = tk.Label(column_names_fr, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    cargo_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    storage_lb = tk.Label(column_names_fr, text="storage:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=10)
    storage_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    date_lb = tk.Label(column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    date_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
    sum_lb = tk.Label(column_names_fr, text="on yard / days:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    for all in gn_truck:
        if all["last_date"] is not None and all["status"]:
            delta_days = (today - all["last_date"].date()).days
        else:
            delta_days = None
        if all["status"]:
            last_truck_time = all["last_date"]
        else:
            last_truck_time = ""
        rec_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        rec_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
        T_lb = tk.Label(rec_fr, text=all["truck_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        T_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["status"]:
            stateTruck = "on yard"
        else:
            stateTruck = ""
        st_lb = tk.Label(rec_fr, text=stateTruck, bg=conf["window_bg"], fg=conf["on_parking"], font=(conf["header_font"], conf["notebook_tab_size"]), width=15)
        st_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        C_lb = tk.Label(rec_fr, text="", bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=15)
        C_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        str_lb = tk.Label(rec_fr, text="", bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=10)
        str_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        T_time_lb = tk.Label(rec_fr, text=last_truck_time, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        T_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
        T_sum_lb = tk.Label(rec_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["status_fg"], width=20)
        T_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))

    trailer_label = tk.Label(c_frame, text="Trailers:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=GN_screen - 2)
    trailer_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
    trailer_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    trailer_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
    column_names_fr2 = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    column_names_fr2.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
    unitT_lb = tk.Label(column_names_fr2, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    unitT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    onyardT_lb = tk.Label(column_names_fr2, text="on yard:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    onyardT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    cargo_t_lb = tk.Label(column_names_fr2, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    cargo_t_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    st_t_lb = tk.Label(column_names_fr2, text="storage:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=10)
    st_t_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    dateT_lb = tk.Label(column_names_fr2, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    dateT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
    sumT_lb = tk.Label(column_names_fr2, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    sumT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    for all in gn_trailer:
        if all["last_date"] is not None and all["status"]:
            delta_days = (today - all["last_date"].date()).days
        else:
            delta_days = None
        if all["status"]:
            last_trailer_time = all["last_date"]
        else:
            last_trailer_time = ""
        recT_fr = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        recT_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
        Tt_lb = tk.Label(recT_fr, text=all["trailer_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        Tt_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["status"]:
            stateT = "on yard"
        else:
            stateT = ""
        onyard_T_lb = tk.Label(recT_fr, text=stateT, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["on_parking"], width=15)
        onyard_T_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["LU"]:
            cargo = "LOADED"
            fgTLU = conf["func_button_fg"]
        else:
            cargo = "EMPTY"
            fgTLU = conf["func_button_sel_fg"]
        tC_lb = tk.Label(recT_fr, text=cargo, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=fgTLU, width=15)
        tC_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["storage"]:
            strg = "storage"
        else:
            strg = ""
        strT_lb = tk.Label(recT_fr, text=strg, bg=conf["window_bg"], fg=conf["storage_fg"], font=(conf["header_font"], conf["notebook_tab_size"]), width=10)
        strT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        Tt_time_lb = tk.Label(recT_fr, text=last_trailer_time, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        Tt_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
        Tt_sum_lb = tk.Label(recT_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["status_fg"], width=20)
        Tt_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))

    fb_label = tk.Label(c_frame, text="Flatbeds:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=GN_screen - 2)
    fb_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
    fb_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    fb_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
    column_names_fr3 = tk.Frame(fb_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    column_names_fr3.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
    unitf_lb = tk.Label(column_names_fr3, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    unitf_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    onyardFb_lb = tk.Label(column_names_fr3, text="on yard:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    onyardFb_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    cargo_f_lb = tk.Label(column_names_fr3, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    cargo_f_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    stFb_f_lb = tk.Label(column_names_fr3, text="storage:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=10)
    stFb_f_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    datef_lb = tk.Label(column_names_fr3, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    datef_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
    sumf_lb = tk.Label(column_names_fr3, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    sumf_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    for all in gn_fb:
        if all["last_date"] is not None and all["status"]:
            delta_days = (today - all["last_date"].date()).days
        else:
            delta_days = None
        if all["status"]:
            last_fb_time = all["last_date"]
        else:
            last_fb_time = ""
        recf_fr = tk.Frame(fb_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        recf_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
        f_lb = tk.Label(recf_fr, text=all["fb_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        f_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["status"]:
            stateFb = "on yard"
        else:
            stateFb = ""
        onyard_fb_lb = tk.Label(recf_fr, text=stateFb, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["on_parking"], width=15)
        onyard_fb_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["LU"]:
            cargo = "LOADED"
            fgfb = conf["func_button_fg"]
        else:
            cargo = "EMPTY"
            fgfb = conf["func_button_sel_fg"]
        fC_lb = tk.Label(recf_fr, text=cargo, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=fgfb, width=15)
        fC_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["storage"]:
            strgFb = "storage"
        else:
            strgFb = ""
        strFb_lb = tk.Label(recf_fr, text=strgFb, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), width=10)
        if all["storage"]:
            strFb_lb.config(fg=conf["storage_fg"])
        else:
            strFb_lb.config(fg=conf["header_fg"])
        strFb_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        f_time_lb = tk.Label(recf_fr, text=last_fb_time, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        f_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
        f_sum_lb = tk.Label(recf_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["status_fg"], width=20)
        f_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))

def GN_stat_insert(frame):
    today = date.today()
    def extract(var):
        lst = list()
        for n in var[0]:
            u_dir = {}
            index = 0
            for y in n:
                z = var[1][index]
                index+=1
                if y is not None:
                    u_dir.update({z[0]: y})
                else: u_dir.update({z[0]: None})
            lst.append(u_dir)
        return lst

    gn_truck = extract(SQL_REQ("SELECT * FROM dbo.GN_Trucks", (), "S_all_D"))
    gn_trailer = extract(SQL_REQ("SELECT * FROM dbo.GN_Trailers", (), "S_all_D"))
    gn_fb = extract(SQL_REQ("SELECT * FROM dbo.GN_Flatbed", (), "S_all_D"))
    c_frame = tk.Frame(frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    c_frame.pack(side=tk.TOP, fill=tk.X, expand=1, pady=(0, 2), padx=(2, 0))
    truck_label = tk.Label(c_frame, text="Trucks:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=GN_screen-2)
    truck_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
    truck_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    truck_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
    column_names_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
    unit_lb = tk.Label(column_names_fr, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    unit_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    onyard_lb = tk.Label(column_names_fr, text="on yard:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    onyard_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    cargo_lb = tk.Label(column_names_fr, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    cargo_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    storage_lb = tk.Label(column_names_fr, text="storage:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=10)
    storage_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    date_lb = tk.Label(column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    date_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
    sum_lb = tk.Label(column_names_fr, text="on yard / days:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    for all in gn_truck:
        if all["last_date"] is not None and all["status"]: delta_days = (today - all["last_date"].date()).days
        else: delta_days = None
        if all["status"]: last_truck_time = all["last_date"]
        else: last_truck_time=""
        rec_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        rec_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
        T_lb = tk.Label(rec_fr, text=all["truck_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        T_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["status"]: stateTruck = "on yard"
        else: stateTruck = ""
        st_lb = tk.Label(rec_fr, text=stateTruck, bg=conf["window_bg"], fg=conf["on_parking"], font=(conf["header_font"], conf["notebook_tab_size"]), width=15)
        st_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        C_lb = tk.Label(rec_fr, text="", bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=15)
        C_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        str_lb = tk.Label(rec_fr, text="", bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=10)
        str_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        T_time_lb = tk.Label(rec_fr, text=last_truck_time, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        T_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
        T_sum_lb = tk.Label(rec_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["status_fg"], width=20)
        T_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))

    trailer_label = tk.Label(c_frame, text="Trailers:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=GN_screen-2)
    trailer_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
    trailer_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    trailer_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
    column_names_fr2 = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    column_names_fr2.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
    unitT_lb = tk.Label(column_names_fr2, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    unitT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    onyardT_lb = tk.Label(column_names_fr2, text="on yard:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    onyardT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    cargo_t_lb = tk.Label(column_names_fr2, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    cargo_t_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    st_t_lb = tk.Label(column_names_fr2, text="storage:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=10)
    st_t_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    dateT_lb = tk.Label(column_names_fr2, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    dateT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
    sumT_lb = tk.Label(column_names_fr2, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    sumT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    for all in gn_trailer:
        if all["last_date"] is not None and all["status"]: delta_days = (today - all["last_date"].date()).days
        else: delta_days = None
        if all["status"]:
            last_trailer_time = all["last_date"]
        else: last_trailer_time=""
        recT_fr = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        recT_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
        Tt_lb = tk.Label(recT_fr, text=all["trailer_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        Tt_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["status"]: stateT = "on yard"
        else: stateT = ""
        onyard_T_lb = tk.Label(recT_fr, text=stateT, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["on_parking"], width=15)
        onyard_T_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["LU"]:
            cargo = "LOADED"
            fgTLU = conf["func_button_fg"]
        else:
            cargo = "EMPTY"
            fgTLU = conf["func_button_sel_fg"]
        tC_lb = tk.Label(recT_fr, text=cargo, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=fgTLU, width=15)
        tC_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["storage"]:  strg = "storage"
        else:  strg = ""
        strT_lb = tk.Label(recT_fr, text=strg, bg=conf["window_bg"], fg=conf["storage_fg"], font=(conf["header_font"], conf["notebook_tab_size"]), width=10)
        strT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        Tt_time_lb = tk.Label(recT_fr, text=last_trailer_time, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        Tt_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
        Tt_sum_lb = tk.Label(recT_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["status_fg"], width=20)
        Tt_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))

    fb_label = tk.Label(c_frame, text="Flatbeds:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=GN_screen-2)
    fb_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
    fb_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    fb_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
    column_names_fr3 = tk.Frame(fb_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
    column_names_fr3.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
    unitf_lb = tk.Label(column_names_fr3, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    unitf_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    onyardFb_lb = tk.Label(column_names_fr3, text="on yard:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    onyardFb_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    cargo_f_lb = tk.Label(column_names_fr3, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=15)
    cargo_f_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    stFb_f_lb = tk.Label(column_names_fr3, text="storage:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=10)
    stFb_f_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    datef_lb = tk.Label(column_names_fr3, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    datef_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
    sumf_lb = tk.Label(column_names_fr3, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=20)
    sumf_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
    for all in gn_fb:
        if all["last_date"] is not None and all["status"]: delta_days = (today - all["last_date"].date()).days
        else: delta_days = None
        if all["status"]:
            last_fb_time = all["last_date"]
        else: last_fb_time=""
        recf_fr = tk.Frame(fb_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        recf_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
        f_lb = tk.Label(recf_fr, text=all["fb_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        f_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["status"]: stateFb = "on yard"
        else: stateFb = ""
        onyard_fb_lb = tk.Label(recf_fr, text=stateFb, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["on_parking"], width=15)
        onyard_fb_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["LU"]:
            cargo = "LOADED"
            fgfb = conf["func_button_fg"]
        else:
            cargo = "EMPTY"
            fgfb = conf["func_button_sel_fg"]
        fC_lb = tk.Label(recf_fr, text=cargo, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=fgfb, width=15)
        fC_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        if all["storage"]: strgFb = "storage"
        else: strgFb = ""
        strFb_lb = tk.Label(recf_fr, text=strgFb, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), width=10)
        if all["storage"]: strFb_lb.config(fg=conf["storage_fg"])
        else: strFb_lb.config(fg=conf["header_fg"])
        strFb_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
        f_time_lb = tk.Label(recf_fr, text=last_fb_time, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
        f_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
        f_sum_lb = tk.Label(recf_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["status_fg"], width=20)
        f_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))


def T_stat(*args):
    global GN_Menu_Var
    GN_Menu_Var = 3
    for all in GN_central_frame.winfo_children(): all.pack_forget()
    for all in data_Ten_stat.winfo_children(): all.pack_forget()
    ten_state_filter.delete()
    ten_state_scroller.delete()
    ten_state_filter.pack(side=tk.LEFT, fill=tk.Y)
    ten_state_filter.tenant_stat(
        company_func=lambda *args: checkyard_insert(ten_state_filter, ten_state_scroller),
        time_on_yard_func=lambda *args: checkyard_insert(ten_state_filter, ten_state_scroller),
        truck_func=lambda *args: checkyard_insert(ten_state_filter, ten_state_scroller),
        trailer_func=lambda *args: checkyard_insert(ten_state_filter, ten_state_scroller),
        storage_func=lambda *args: checkyard_insert(ten_state_filter, ten_state_scroller),
        age_func=lambda *args: checkyard_insert(ten_state_filter, ten_state_scroller),
        scale_func=lambda *args: checkyard_insert(ten_state_filter, ten_state_scroller),
        period_func=lambda *args: checkyard_insert(ten_state_filter, ten_state_scroller)
    )
    checkyard_insert(ten_state_filter, ten_state_scroller)
    data_Ten_stat.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    # Ten_state_scroll_fr.pack(side=tk.LEFT, fill=tk.BOTH)
    # Ten_his_fil
    ten_state_scroller.pack(side=tk.LEFT, fill=tk.BOTH)

    #TEN_stat_insert(Ten_state_date_sc_fr.frame)

    # ten_state_scroller.refresh
    # ten_state_scroller.delete


    #########
    # ten_state_filter = filter_frame(data_Ten_stat)
    # ten_state_scroller = scroller(data_Ten_stat)


def T_history(*args):
    global GN_Menu_Var
    global el_size
    GN_Menu_Var = 4
    for all in GN_central_frame.winfo_children(): all.pack_forget()
    for all in data_Ten_his.winfo_children(): all.pack_forget()
    tenant_his_filter.delete()
    tenant_his_scroll_frame.delete()
    tenant_his_filter.pack(side=tk.LEFT, fill=tk.Y)
    tenant_his_filter.tenant_history(
        company_func=lambda *args:history_insert("T"),
        truck_func=lambda *args:history_insert("T"),
        trailer_func=lambda *args:history_insert("T"),
        scale_func=lambda *args:history_insert("T"),
        period_func=lambda *args:history_insert("T")
    )
    #tenant_his_scroll_frame.pack(side=tk.LEFT, fill=tk.BOTH)
    data_Ten_his.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    #tenant_his_scroll_frame.refresh()

    T_signs_frame = tk.Frame(data_Ten_his, highlightthickness=0, bg=conf["window_bg"])
    T_signs_frame.pack(side=tk.TOP, fill=tk.X)
    T_signs_com_lb = tk.Label(T_signs_frame, text="Company:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    T_signs_com_lb.grid(row=0, column=0, sticky=tk.EW, pady=(1, 0))
    T_signs_T_lb = tk.Label(T_signs_frame, text="Truck:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    T_signs_T_lb.grid(row=0, column=1, sticky=tk.EW, pady=(1, 0))
    T_signs_Tr_lb = tk.Label(T_signs_frame, text="Trailer:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    T_signs_Tr_lb.grid(row=0, column=2, sticky=tk.EW, pady=(1, 0))
    T_signs_Datetime_lb = tk.Label(T_signs_frame, text="Datetime:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    T_signs_Datetime_lb.grid(row=0, column=3, sticky=tk.EW, pady=(1, 0))
    T_signs_Status_lb = tk.Label(T_signs_frame, text="Status:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    T_signs_Status_lb.grid(row=0, column=4, sticky=tk.EW, pady=(1, 0))
    T_signs_comm_lb = tk.Label(T_signs_frame, text="Comment:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    T_signs_comm_lb.grid(row=0, column=5, sticky=tk.EW, pady=(1, 0))
    T_signs_name_lb = tk.Label(T_signs_frame, text="Registered:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    T_signs_name_lb.grid(row=0, column=6, sticky=tk.EW, pady=(1, 0))
    T_signs_frame.grid_columnconfigure(0, weight=1)
    T_signs_frame.grid_columnconfigure(1, weight=1)
    T_signs_frame.grid_columnconfigure(2, weight=1)
    T_signs_frame.grid_columnconfigure(3, weight=4)
    T_signs_frame.grid_columnconfigure(4, weight=1)
    T_signs_frame.grid_columnconfigure(5, weight=2)
    T_signs_frame.grid_columnconfigure(6, weight=2)
    tenant_his_scroll_frame.pack(side=tk.LEFT, fill=tk.BOTH)
    tenant_his_scroll_frame.refresh()
    root.update_idletasks()
    el_size = [
        T_signs_com_lb.winfo_width(),
        T_signs_T_lb.winfo_width(),
        T_signs_Tr_lb.winfo_width(),
        T_signs_Datetime_lb.winfo_width(),
        T_signs_Status_lb.winfo_width(),
        T_signs_comm_lb.winfo_width(),
        T_signs_name_lb.winfo_width() - 25
    ]
    el_size = [x // 9 for x in el_size]


def GN_stat(*args):
    global GN_Menu_Var
    GN_Menu_Var = 1
    for all in GN_central_frame.winfo_children(): all.pack_forget()
    GN_state_data_sc_fr.delete()
    GN_stat_insert(GN_state_data_sc_fr.frame)
    GN_state_data_sc_fr.refresh()
    GN_state_data_sc_fr.top()
    data_GN_stat.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    data_Ten_stat.pack_forget()
    # GN_stat_canv.update_idletasks()
    # GN_stat_canv.yview_moveto(0)

def GN_history(*args):
    global GN_Menu_Var
    global el_size
    GN_Menu_Var = 2
    for all in GN_central_frame.winfo_children(): all.pack_forget()
    for all in data_GN_his.winfo_children(): all.pack_forget()
    GN_his_filter.delete()
    GN_his_scroll_frame.delete()
    GN_his_filter.pack(side=tk.LEFT, fill=tk.Y)
    GN_his_filter.GN_histor(
        truck_func=lambda *args: history_insert("GN"),
        trailer_func=lambda *args: history_insert("GN"),
        fb_func=lambda *args: history_insert("GN"),
        storage_func=lambda *args: history_insert("GN"),
        scale_func=lambda *args: history_insert("GN"),
        period_func=lambda *args: history_insert("GN")
    )
    data_GN_his.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    #GN_history_insert(second_GN_his_frame)
    GN_signs_frame = tk.Frame(data_GN_his, highlightthickness=0, bg=conf["window_bg"])
    GN_signs_frame.pack(side=tk.TOP, fill=tk.X)
    GN_signs_com_lb = tk.Label(GN_signs_frame, text="Company:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    GN_signs_com_lb.grid(row=0, column=0, sticky=tk.EW, pady=(1, 0))
    GN_signs_T_lb = tk.Label(GN_signs_frame, text="Truck:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    GN_signs_T_lb.grid(row=0, column=1, sticky=tk.EW, pady=(1, 0))
    GN_signs_Tr_lb = tk.Label(GN_signs_frame, text="Trailer:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    GN_signs_Tr_lb.grid(row=0, column=2, sticky=tk.EW, pady=(1, 0))
    GN_signs_FB_lb = tk.Label(GN_signs_frame, text="Flatbed:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    GN_signs_FB_lb.grid(row=0, column=3, sticky=tk.EW, pady=(1, 0))
    GN_signs_Datetime_lb = tk.Label(GN_signs_frame, text="Datetime:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    GN_signs_Datetime_lb.grid(row=0, column=4, sticky=tk.EW, pady=(1, 0))
    GN_signs_Cargo_lb = tk.Label(GN_signs_frame, text="Cargo:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    GN_signs_Cargo_lb.grid(row=0, column=5, sticky=tk.EW, pady=(1, 0))
    GN_signs_Status_lb = tk.Label(GN_signs_frame, text="Status:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    GN_signs_Status_lb.grid(row=0, column=6, sticky=tk.EW, pady=(1, 0))
    GN_signs_comm_lb = tk.Label(GN_signs_frame, text="Comment:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    GN_signs_comm_lb.grid(row=0, column=7, sticky=tk.EW, pady=(1, 0))
    GN_signs_name_lb = tk.Label(GN_signs_frame, text="Registered:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    GN_signs_name_lb.grid(row=0, column=8, sticky=tk.EW, pady=(1, 0))
    GN_signs_frame.grid_columnconfigure(0, weight=1)
    GN_signs_frame.grid_columnconfigure(1, weight=1)
    GN_signs_frame.grid_columnconfigure(2, weight=1)
    GN_signs_frame.grid_columnconfigure(3, weight=1)
    GN_signs_frame.grid_columnconfigure(4, weight=4)
    GN_signs_frame.grid_columnconfigure(5, weight=1)
    GN_signs_frame.grid_columnconfigure(6, weight=1)
    GN_signs_frame.grid_columnconfigure(7, weight=2)
    GN_signs_frame.grid_columnconfigure(8, weight=2)
    GN_his_scroll_frame.pack(side=tk.LEFT, fill=tk.BOTH)
    GN_his_scroll_frame.refresh()
    root.update_idletasks()
    el_size = [
        GN_signs_com_lb.winfo_width(),
        GN_signs_T_lb.winfo_width(),
        GN_signs_Tr_lb.winfo_width(),
        GN_signs_FB_lb.winfo_width(),
        GN_signs_Datetime_lb.winfo_width(),
        GN_signs_Cargo_lb.winfo_width(),
        GN_signs_Status_lb.winfo_width(),
        GN_signs_comm_lb.winfo_width(),
        GN_signs_name_lb.winfo_width()-25
    ]
    el_size = [x//9 for x in el_size]

def V_stat(*args):
    global GN_Menu_Var
    GN_Menu_Var = 8
    for all in GN_central_frame.winfo_children(): all.pack_forget()
    for all in data_Vis_stat.winfo_children(): all.pack_forget()
    Vis_state_filter.delete()
    Vis_state_scroller.delete()
    Vis_state_filter.pack(side=tk.LEFT, fill=tk.Y)
    Vis_state_filter.vis_stat(
        company_func=lambda *args: checkyard_vis_insert(Vis_state_filter, Vis_state_scroller),
        time_on_yard_func=lambda *args: checkyard_vis_insert(Vis_state_filter, Vis_state_scroller),
        age_func=lambda *args: checkyard_vis_insert(Vis_state_filter, Vis_state_scroller),
        corp_func=lambda *args: checkyard_vis_insert(Vis_state_filter, Vis_state_scroller),
        private_func=lambda *args: checkyard_vis_insert(Vis_state_filter, Vis_state_scroller),
        expired_func=lambda *args: checkyard_vis_insert(Vis_state_filter, Vis_state_scroller),
        no_parking_func=lambda *args: checkyard_vis_insert(Vis_state_filter, Vis_state_scroller)
    )
    data_Vis_stat.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    Vis_state_scroller.pack(side=tk.LEFT, fill=tk.BOTH)
    Vis_state_scroller.refresh()
    Vis_state_scroller.top()
    checkyard_vis_insert(Vis_state_filter, Vis_state_scroller)

def V_history(*args):
    global GN_Menu_Var
    global el_size
    GN_Menu_Var = 6
    for all in GN_central_frame.winfo_children(): all.pack_forget()
    for all in data_Vis_his.winfo_children(): all.pack_forget()
    Vis_his_filter.delete()
    Vis_his_scroll_frame.delete()
    Vis_his_filter.pack(side=tk.LEFT, fill=tk.Y)
    Vis_his_filter.vis_history(
        company_func=lambda *args: history_insert("V"),
        plate_func=lambda *args: history_insert("V"),
        scale_func=lambda *args: history_insert("V"),
        period_func=lambda *args: history_insert("V")
    )
    data_Vis_his.pack(side=tk.TOP, fill=tk.BOTH, expand=1)

    V_signs_frame = tk.Frame(data_Vis_his, highlightthickness=0, bg=conf["window_bg"])
    V_signs_frame.pack(side=tk.TOP, fill=tk.X)
    V_signs_com_lb = tk.Label(V_signs_frame, text="Company:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    V_signs_com_lb.grid(row=0, column=0, sticky=tk.EW, pady=(1, 0))
    V_signs_P_lb = tk.Label(V_signs_frame, text="Plates:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    V_signs_P_lb.grid(row=0, column=1, sticky=tk.EW, pady=(1, 0))
    V_signs_Car_lb = tk.Label(V_signs_frame, text="Car:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    V_signs_Car_lb.grid(row=0, column=2, sticky=tk.EW, pady=(1, 0))
    V_signs_driver_lb = tk.Label(V_signs_frame, text="Driver:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    V_signs_driver_lb.grid(row=0, column=3, sticky=tk.EW, pady=(1, 0))
    V_signs_Datetime_lb = tk.Label(V_signs_frame, text="Datetime:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    V_signs_Datetime_lb.grid(row=0, column=4, sticky=tk.EW, pady=(1, 0))
    V_signs_comm_lb = tk.Label(V_signs_frame, text="Comment:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    V_signs_comm_lb.grid(row=0, column=5, sticky=tk.EW, pady=(1, 0))
    V_signs_Status_lb = tk.Label(V_signs_frame, text="Status:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    V_signs_Status_lb.grid(row=0, column=6, sticky=tk.EW, pady=(1, 0))
    V_signs_name_lb = tk.Label(V_signs_frame, text="Registered:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    V_signs_name_lb.grid(row=0, column=7, sticky=tk.EW, pady=(1, 0))
    V_signs_frame.grid_columnconfigure(0, weight=1)
    V_signs_frame.grid_columnconfigure(1, weight=1)
    V_signs_frame.grid_columnconfigure(2, weight=1)
    V_signs_frame.grid_columnconfigure(3, weight=4)
    V_signs_frame.grid_columnconfigure(4, weight=1)
    V_signs_frame.grid_columnconfigure(5, weight=2)
    V_signs_frame.grid_columnconfigure(6, weight=2)
    V_signs_frame.grid_columnconfigure(7, weight=2)
    Vis_his_scroll_frame.pack(side=tk.LEFT, fill=tk.BOTH)
    Vis_his_scroll_frame.refresh()
    Vis_his_scroll_frame.top()
    root.update_idletasks()
    el_size = [
        V_signs_com_lb.winfo_width(),
        V_signs_P_lb.winfo_width(),
        V_signs_Car_lb.winfo_width(),
        V_signs_driver_lb.winfo_width(),
        V_signs_Datetime_lb.winfo_width(),
        V_signs_comm_lb.winfo_width(),
        V_signs_Status_lb.winfo_width(),
        V_signs_name_lb.winfo_width() - 25
    ]
    el_size = [x // 9 for x in el_size]

























def GN_city(*args):
    global GN_Menu_Var
    global el_size
    GN_Menu_Var = 7
    for all in GN_central_frame.winfo_children(): all.pack_forget()
    for all in data_city.winfo_children(): all.pack_forget()
    GN_city_filter.delete()
    GN_city_scroll_frame.delete()

###################
    GN_city_filter.pack(side=tk.LEFT, fill=tk.Y)
    GN_city_filter.GN_city(
        truck_func=lambda *args: city_insert(),
        scale_func=lambda *args: city_insert(),
        period_func=lambda *args: city_insert(),
        save_func=lambda *args: save_city_file()
    )
    data_city.pack(side=tk.TOP, fill=tk.BOTH, expand=1)


    city_signs_frame = tk.Frame(data_city, highlightthickness=0, bg=conf["window_bg"])
    city_signs_frame.pack(side=tk.TOP, fill=tk.X)
    city_signs_T_lb = tk.Label(city_signs_frame, text="Truck:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    city_signs_T_lb.grid(row=0, column=0, sticky=tk.EW, pady=(1, 0))
    city_signs_date_lb = tk.Label(city_signs_frame, text="Date:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    city_signs_date_lb.grid(row=0, column=1, sticky=tk.EW, pady=(1, 0))
    city_signs_timeOUT_lb = tk.Label(city_signs_frame, text="OUT:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    city_signs_timeOUT_lb.grid(row=0, column=2, sticky=tk.EW, pady=(1, 0))
    city_signs_timeIN_lb = tk.Label(city_signs_frame, text="IN:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    city_signs_timeIN_lb.grid(row=0, column=3, sticky=tk.EW, pady=(1, 0))
    city_signs_hours_lb = tk.Label(city_signs_frame, text="Hours:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
    city_signs_hours_lb.grid(row=0, column=4, sticky=tk.EW, pady=(1, 0))
    city_signs_frame.grid_columnconfigure(0, weight=2)
    city_signs_frame.grid_columnconfigure(1, weight=2)
    city_signs_frame.grid_columnconfigure(2, weight=2)
    city_signs_frame.grid_columnconfigure(3, weight=2)
    city_signs_frame.grid_columnconfigure(4, weight=1)
    GN_city_scroll_frame.pack(side=tk.LEFT, fill=tk.BOTH)
    GN_city_scroll_frame.refresh()
    root.update_idletasks()
    #recheck if needed el_size
    el_size = [
        city_signs_T_lb.winfo_width(),
        city_signs_date_lb.winfo_width(),
        city_signs_timeOUT_lb.winfo_width(),
        city_signs_timeIN_lb.winfo_width(),
        city_signs_hours_lb.winfo_width() - 25
    ]
    el_size = [x // 9 for x in el_size]
###################




def chart(*args):
    global GN_Menu_Var
    GN_Menu_Var = 5
    for all in GN_central_frame.winfo_children(): all.pack_forget()
    GN_Chart_Frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)

def GN_howeroff(*args):
    if GN_Menu_Var != 1: statistic_GN_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if GN_Menu_Var != 2: history_GN_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if GN_Menu_Var != 3: statistic_T_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if GN_Menu_Var != 4: history_T_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if GN_Menu_Var != 5: statistic_CH_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if GN_Menu_Var != 7: city_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if GN_Menu_Var != 8: statistic_V_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if GN_Menu_Var != 6: history_V_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])



#Defining number of days in Year and Month function
def update_GN_his(event, *args):
    return
    global GN_combo_days_list
    global GN_combo_T_list
    global GN_combo_filter_var
    global second_GN_his_frame
    def insert_records(list):
        for widgets in second_GN_his_frame.winfo_children(): widgets.destroy()
        for rec in list:
            record_frame = tk.Frame(second_GN_his_frame, bg=conf["window_bg"])
            record_frame.pack(side=tk.TOP, padx=5, pady=(0, 1), fill=tk.X, expand=1)
            company_var = tk.Label(record_frame, text=rec[0], background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=10)
            company_var.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X)
            truck_var = tk.Label(record_frame, text=rec[1], background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=10)
            truck_var.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, padx=1)
            trailer_var = tk.Label(record_frame, text=rec[2], background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=10)
            trailer_var.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X)
            fb_var = tk.Label(record_frame, text=rec[3], background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=10)
            fb_var.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, padx=1)
            if rec[5]:
                cargo = "LOADED"
                col = conf["func_button_fg"]
            else:
                cargo = "UNLOADED"
                col = conf["func_button_sel_fg"]
            cargo_var = tk.Label(record_frame, text=cargo, background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=col, anchor=tk.W, width=10)
            cargo_var.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X)
            if rec[6]:
                status = "IN"
                col2 = conf["in_button_fg"]
            else:
                status = "OUT"
                col2 = conf["out_button_fg"]
            status_var = tk.Label(record_frame, text=status, background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=col2, anchor=tk.W, width=10)
            status_var.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, padx=1)
            year = rec[4].year
            month = rec[4].month
            day = rec[4].day
            year_var = tk.Label(record_frame, text=year, background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=10)
            year_var.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X)
            month_var = tk.Label(record_frame, text=month, background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=10)
            month_var.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, padx=1)
            day_var = tk.Label(record_frame, text=day, background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=10)
            day_var.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X)
            comment_var = tk.Label(record_frame, text=rec[7], background=conf["widget_bg"], font=(conf["widget_font"], conf["widget_size"]), foreground=conf["widget_fg"], anchor=tk.W, width=20)
            comment_var.pack(side=tk.LEFT, anchor=tk.N, fill=tk.X, padx=(1, 0))

    args[0].selection_clear()
    value=args[0].get()
    if value == "ALL":
        if args[1] == 0:
            GN_combo_T_list = ["ALL"] + sorted(list({val[0] for val in SQL_REQ("SELECT truck_number FROM dbo.GN_History WHERE truck_number IS NOT NULL ORDER BY len(truck_number), truck_number", (), "S_all")}))
            GN_his_combobox_T.config(value=GN_combo_T_list)
            if GN_combo_filter_var is not None:
                if args[1] in GN_combo_filter_var: del GN_combo_filter_var[args[1]]
            args[0].current(0)
            args[0].selection_clear()
        elif args[1] == 41:
            if 41 in GN_combo_filter_var:
                del GN_combo_filter_var[41]
                GN_his_combobox_Y.current(0)
                GN_his_combobox_Y.selection_clear()
            if 42 in GN_combo_filter_var:
                GN_combo_days_list = ["ALL"]
                del GN_combo_filter_var[42]
                GN_his_combobox_M.current(0)
                GN_his_combobox_M.selection_clear()
                GN_his_combobox_D.config(value=GN_combo_days_list)
            if 43 in GN_combo_filter_var:
                GN_combo_days_list = ["ALL"]
                del GN_combo_filter_var[43]
                GN_his_combobox_D.current(0)
                GN_his_combobox_D.selection_clear()
                GN_his_combobox_D.config(value=GN_combo_days_list)
        elif args[1] == 42:
            if 42 in GN_combo_filter_var:
                GN_combo_days_list = ["ALL"]
                del GN_combo_filter_var[42]
                GN_his_combobox_M.current(0)
                GN_his_combobox_M.selection_clear()
                GN_his_combobox_D.config(value=GN_combo_days_list)
            if 43 in GN_combo_filter_var:
                del GN_combo_filter_var[43]
                GN_his_combobox_D.current(0)
                GN_his_combobox_D.selection_clear()
                GN_his_combobox_D.config(value=GN_combo_days_list)
        elif args[1] == 43:
            if 43 in GN_combo_filter_var:
                GN_combo_days_list = ["ALL"]
                del GN_combo_filter_var[43]
                GN_his_combobox_D.current(0)
                GN_his_combobox_D.selection_clear()
                GN_his_combobox_D.config(value=GN_combo_days_list)
        else:
            if args[1] in GN_combo_filter_var:
                del GN_combo_filter_var[args[1]]
                args[0].current(0)
                args[0].selection_clear()
    else:
        if args[1] == 43:
            if 42 not in GN_combo_filter_var or 41 not in GN_combo_filter_var:
                    args[0].current(0)
                    args[0].selection_clear()
                    return
        if args[1] == 42 and 41 not in GN_combo_filter_var:
            args[0].current(0)
            args[0].selection_clear()
            return
        if value == "LOADED" or value == "IN":
            GN_combo_filter_var.update({args[1]: True})
        elif value == "EMPTY" or value == "OUT":
            GN_combo_filter_var.update({args[1]: False})
        else:
            GN_combo_filter_var.update({args[1]: value})
        if args[1] == 0:
            GN_combo_T_list = ["ALL"] + sorted(list({val[0] for val in SQL_REQ("SELECT truck_number FROM dbo.GN_History WHERE company_name=\'"+value+"\' and truck_number IS NOT NULL ORDER BY len(truck_number), truck_number", "S_all")}))
            GN_his_combobox_T.config(value=GN_combo_T_list)
        if args[1] == 42:
            GN_combo_days_list = ["ALL"] + list(str(day) for day in range(1, monthrange(int(GN_combo_filter_var[41]), int(value))[1] + 1))
            GN_his_combobox_D.config(value=GN_combo_days_list)
    datapool = sorted(list(list(val) for val in SQL_REQ("SELECT * FROM dbo.GN_History ORDER BY datetime_event", (), "S_all")), key=lambda x:x[4])
    result_list = []
    if GN_combo_filter_var is not None:
        if len(GN_combo_filter_var) > 0:
            for line in datapool:
                if all((key in {41, 42, 43} or line[key] == value) for key, value in GN_combo_filter_var.items()):
                    if any(key in {41, 42, 43} and value != "ALL" for key, value in GN_combo_filter_var.items()):
                        date_filter = line[4].date()  # Extract the date part from the datetime object
                        if all((key in {41, 42, 43} and value != "ALL" and getattr(date_filter, {41: "year", 42: "month", 43: "day"}[key]) == int(value)) or key not in {41, 42, 43} for key, value in GN_combo_filter_var.items()):
                            result_list.append(line)
                    else:
                        result_list.append(line)
    else: result_list = list(datapool)
    insert_records(result_list)
    GN_his_canv.update_idletasks()

    second_GN_his_frame.update_idletasks()
    GN_his_scroll_region()
    GN_his_canv.yview_moveto(0)

#City insert in window function
def city_insert(*args):
    global City_Data
    masta = GN_city_scroll_frame
    filter_obj = GN_city_filter
    table_name = "dbo.GN_History"
    truck_table_name = "dbo.GN_Trucks"
    f_year = filter_obj.year_label.cget("text")
    f_month = filter_obj.month_label.cget("text")
    f_day = filter_obj.day_label.cget("text")
    his_scale = filter_obj.chart_scale
    truck_checkbox = filter_obj.search_truck_checkbox.get()
    if truck_checkbox: truck_unit = filter_obj.truck_search_var.get().strip() or None
    else: truck_unit = None
    where_list = [["YEAR(t.datetime_event)=?", f_year]]
    if his_scale == "M" or his_scale == "D": where_list.append(["MONTH(t.datetime_event)=?", f_month])
    if his_scale == "D": where_list.append(["DAY(t.datetime_event)=?", f_day])
    if truck_unit is not None: where_list.append(["t.truck_number=?", truck_unit])
    #getting list of city trucks city_trucks
    city_trucks = ", ".join(["'" + number[0] + "'" for number in SQL_REQ("SELECT truck_number FROM dbo.GN_Trucks WHERE city=1 ORDER BY len(truck_number), truck_number", (), "S_all")])
    masta.delete()
    condition_var, value_var = zip(*where_list)
    ### Final list with DATA for City
    ###
    city_data = SQL_REQ(
        f"""
        SELECT
            t.truck_number as truck,
            CONVERT(DATE, t.datetime_event) as date,
            MIN(CASE WHEN t.status = 0 THEN CONVERT(TIME, t.datetime_event) END) as out_time,
            MAX(CASE WHEN t.status = 1 THEN CONVERT(TIME, t.datetime_event) END) as in_time,
            DATEDIFF(MINUTE, MIN(CASE WHEN t.status = 0 THEN t.datetime_event END), MAX(CASE WHEN t.status = 1 THEN t.datetime_event END)) as time_difference
        FROM {table_name} t
        INNER JOIN {truck_table_name} tt ON t.truck_number = tt.truck_number
        WHERE {' AND '.join(condition_var)} AND DATENAME(dw, datetime_event) NOT IN ('Saturday', 'Sunday') AND tt.city = 1
        GROUP BY t.truck_number, CONVERT(DATE, datetime_event)
        """,
        value_var,
        "S_all"
    )
    X_ind = 0
    Y_ind = 1
    if city_data:
        City_Data = []
        for r in city_data:
            #converting data in lookable format
            new_record = [r[0], r[1].strftime("%Y-%m-%d"), r[2].strftime("%H:%M") if r[2] is not None else "", r[3].strftime("%H:%M") if r[3] is not None else "", r[4] if r[4] is not None else 0]#f"{divmod(r[4], 60)[0]}h {divmod(r[4], 60)[1]}m" if (r[4] is not None) and (r[4]>0) else "")
            City_Data.append(new_record)


    #####
        if len(City_Data) > 500:
            error(23)
            return
        City_Data.sort(key=lambda x: (x[0], x[1]))
        # Add total hours for each truck
        trucks_total_hours = {rec[0]: 0 for rec in City_Data}
        current_truck_val = ""
        temp_city_list = []
        for rec in City_Data:
            if current_truck_val != rec[0]:
                if current_truck_val != "": temp_city_list.append([current_truck_val,"","", "Total:", trucks_total_hours.get(current_truck_val)])
                current_truck_val = rec[0]
            if rec[4]: trucks_total_hours[rec[0]]+=rec[4]
            temp_city_list.append(rec)
        temp_city_list.append([current_truck_val,"","","Total:", trucks_total_hours.get(current_truck_val)])
        City_Data = [list(r[:4]) + ["{:d}:{:02d}".format(divmod(r[4], 60)[0], divmod(r[4], 60)[1])] if (r[4] is not None) and (r[4] > 0) else list(r[:4]) + [""] for r in temp_city_list]
        for rec_line in City_Data:
            rec_frame = tk.Frame(masta.frame, highlightthickness=0, bg=conf["window_bg"])
            rec_frame.pack(side=tk.TOP, fill=tk.X, expand=1)
            if rec_line[1] == "": fg_color = conf["entry_fg"]
            else: fg_color = conf["widget_fg"]
            for indx, element in enumerate(rec_line):
                rec_lb = tk.Label(rec_frame, text=element, bg=conf["widget_bg"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, width=el_size[indx], highlightbackground=conf["widget_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=fg_color, justify=tk.CENTER)
                rec_lb.grid(row=Y_ind, column=X_ind, sticky=tk.NSEW, padx=(0, 1), pady=(1, 0))
                X_ind += 1
            X_ind = 0
            Y_ind += 1
    else:
        tk.Label(masta.frame, text="NO DATA", bg=conf["window_bg"], fg=conf["chart_title"], font=(conf["header_font"], conf["header_size"])).grid(row=Y_ind, column=X_ind, sticky=tk.NSEW, padx=(0, 1), pady=(1, 0))

    masta.refresh()

#Function that trigger by SAVE button in City. Activate archivato function to safe current filtered data from City_Data to xlsx file by choice.
def save_city_file():
    if City_Data: archivator(City_Data, "", "city")
    else: return


    #
    # #################################################
    # #
    # def history_insert(func):
    #     global el_size
    #     global db_data
    #     query_string = ""
    #     if func == "T":
    #         masta = tenant_his_scroll_frame
    #         filter_obj = tenant_his_filter
    #         table_name = "dbo.Tenant_History"
    #         # getting data from company
    #         comp_name = filter_obj.comp_box.get()
    #         # getting date from period filter
    #         f_year = filter_obj.year_label.cget("text")
    #         f_month = filter_obj.month_label.cget("text")
    #         f_day = filter_obj.day_label.cget("text")
    #         filter_date = datetime.strptime(f"{f_year}-{f_month}-{f_day}", "%Y-%m-%d").date()  # ???????????
    #         his_scale = filter_obj.chart_scale
    #         # getting checkbox
    #         truck_checkbox = filter_obj.search_truck_checkbox.get()
    #         trailer_checkbox = filter_obj.search_trailer_checkbox.get()
    #         if truck_checkbox:
    #             truck_unit = filter_obj.truck_search_var.get().strip() or None
    #         else:
    #             truck_unit = None
    #         if trailer_checkbox:
    #             trailer_unit = filter_obj.trailer_search_var.get().strip() or None
    #         else:
    #             trailer_unit = None
    #         where_list = [["YEAR(datetime_event)=?", f_year]]
    #         if his_scale == "M" or his_scale == "D": where_list.append(["MONTH(datetime_event)=?", f_month])
    #         if his_scale == "D": where_list.append(["DAY(datetime_event)=?", f_day])
    #         if truck_unit is not None: where_list.append(["truck_number=?", truck_unit])
    #         if trailer_unit is not None: where_list.append(["trailer_number=?", trailer_unit])
    #         if comp_name != "All" and comp_name is not None:
    #             id_comp = ID_NAME_company(name=comp_name)
    #             where_list.append(["company_ID=?", id_comp])
    #
    #     elif func == "GN":
    #         masta = GN_his_scroll_frame
    #         filter_obj = GN_his_filter
    #         table_name = "dbo.GN_History"
    #         # getting date from period filter
    #         f_year = filter_obj.year_label.cget("text")
    #         f_month = filter_obj.month_label.cget("text")
    #         f_day = filter_obj.day_label.cget("text")
    #         filter_date = datetime.strptime(f"{f_year}-{f_month}-{f_day}", "%Y-%m-%d").date()  # ???????????
    #         his_scale = filter_obj.chart_scale
    #         # getting checkbox
    #         truck_checkbox = filter_obj.search_truck_checkbox.get()
    #         trailer_checkbox = filter_obj.search_trailer_checkbox.get()
    #         fb_checkbox = filter_obj.search_fb_checkbox.get()
    #         if truck_checkbox:
    #             truck_unit = filter_obj.truck_search_var.get().strip() or None
    #         else:
    #             truck_unit = None
    #         if trailer_checkbox:
    #             trailer_unit = filter_obj.trailer_search_var.get().strip() or None
    #         else:
    #             trailer_unit = None
    #         if fb_checkbox:
    #             fb_unit = filter_obj.fb_search_var.get().strip() or None
    #         else:
    #             fb_unit = None
    #         where_list = [["YEAR(datetime_event)=?", f_year]]
    #         if his_scale == "M" or his_scale == "D": where_list.append(["MONTH(datetime_event)=?", f_month])
    #         if his_scale == "D": where_list.append(["DAY(datetime_event)=?", f_day])
    #         if truck_unit is not None: where_list.append(["truck_number=?", truck_unit])
    #
    #     db_data = SQL_REQ(f"SELECT * FROM {table_name} WHERE {' AND '.join(condition_var)} ORDER BY datetime_event", value_var, "S_all")
    #     X_ind = 0
    #     Y_ind = 1
    #     if db_data:
    #         if len(db_data) > 500:
    #             error(23)
    #             return
    #         for rec_line in db_data:
    #             rec_frame = tk.Frame(masta.frame, highlightthickness=0, bg=conf["window_bg"])
    #             rec_frame.pack(side=tk.TOP, fill=tk.X, expand=1)
    #             for indx, element in enumerate(rec_line):
    #                 if indx == 0 and (func == "T" or func == "V"): element = ID_NAME_company(ID=element)
    #                 if func == "T" and indx == 4: element = "IN" if element else "OUT"
    #                 if func == "GN":
    #                     if indx == 6: element = "IN" if element else "OUT"
    #                     if indx == 5: element = "LOADED" if element else "" if element is None else "UNLOADED"
    #                 if func == "V" and indx == 6: element = "IN" if element else "OUT"
    #                 tk.Label(rec_frame, text=element, bg=conf["widget_bg"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, width=el_size[indx], highlightbackground=conf["widget_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.CENTER).grid(row=Y_ind, column=X_ind, sticky=tk.NSEW, padx=(0, 1), pady=(1, 0))
    #                 X_ind += 1
    #             X_ind = 0
    #             Y_ind += 1
    #     else:
    #         tk.Label(masta.frame, text="NO DATA", bg=conf["window_bg"], fg=conf["chart_title"], font=(conf["header_font"], conf["header_size"])).grid(row=Y_ind, column=X_ind, sticky=tk.NSEW,
    #                                                                                                                                                   padx=(0, 1), pady=(1, 0))
    #     masta.refresh()
    # ##############################

#Defining vars for comboboxed
GN_combo_C_list = ["ALL"] + sorted(list({val[0] for val in SQL_REQ("SELECT company_name FROM dbo.GN_History ORDER BY company_name", (), "S_all")}))
GN_combo_T_list = ["ALL"] + sorted(list({val[0] for val in SQL_REQ("SELECT truck_number FROM dbo.GN_History WHERE truck_number IS NOT NULL ORDER BY len(truck_number), truck_number", (), "S_all")}))
GN_combo_Tr_list = ["ALL"] + sorted(list({val[0] for val in SQL_REQ("SELECT trailer_number FROM dbo.GN_History WHERE trailer_number IS NOT NULL ORDER BY len(trailer_number), trailer_number", (), "S_all")}))
GN_combo_FB_list = ["ALL"] + sorted(list({val[0] for val in SQL_REQ("SELECT fb_number FROM dbo.GN_History WHERE fb_number IS NOT NULL ORDER BY len(fb_number), fb_number", (), "S_all")}))
GN_combo_year_list = ["ALL"] + sorted(list({val[0].year for val in SQL_REQ("SELECT datetime_event FROM dbo.GN_History ORDER BY datetime_event", (), "S_all")}))
GN_combo_month_list = ["ALL", "01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
GN_combo_cargo_list = ["ALL", "LOADED", "EMPTY"]
GN_combo_status_list = ["ALL", "IN", "OUT"]
GN_combo_days_list = ["ALL"]



GN_Main_Frame = tk.Frame(Menu_Bar_GN, highlightthickness=0, bg=conf["window_bg"])
GN_Main_Frame.pack(fill=tk.BOTH, expand=1)

GN_top_frame = tk.Frame(GN_Main_Frame, highlightthickness=0, bg=conf["window_bg"])
GN_top_frame.pack(side=tk.TOP, fill=tk.X)

statistic_T_button = tk.Label(GN_top_frame, text="T Status", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
statistic_T_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
statistic_T_button.bind("<Button-1>", T_stat)
statistic_T_button.bind("<Enter>", lambda x: statistic_T_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
statistic_T_button.bind("<Leave>", GN_howeroff)

history_T_button = tk.Label(GN_top_frame, text="T History", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
history_T_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
history_T_button.bind("<Button-1>", T_history)
history_T_button.bind("<Enter>", lambda x: history_T_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
history_T_button.bind("<Leave>", GN_howeroff)

statistic_GN_button = tk.Label(GN_top_frame, text="GN Status", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
statistic_GN_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
statistic_GN_button.bind("<Button-1>", GN_stat)
statistic_GN_button.bind("<Enter>", lambda x: statistic_GN_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
statistic_GN_button.bind("<Leave>", GN_howeroff)

history_GN_button = tk.Label(GN_top_frame, text="GN History", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
history_GN_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
history_GN_button.bind("<Button-1>", GN_history)
history_GN_button.bind("<Enter>", lambda x: history_GN_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
history_GN_button.bind("<Leave>", GN_howeroff)

statistic_V_button = tk.Label(GN_top_frame, text="V Status", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
statistic_V_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
statistic_V_button.bind("<Button-1>", V_stat)
statistic_V_button.bind("<Enter>", lambda x: statistic_V_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
statistic_V_button.bind("<Leave>", GN_howeroff)

history_V_button = tk.Label(GN_top_frame, text="V History", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
history_V_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
history_V_button.bind("<Button-1>", V_history)
history_V_button.bind("<Enter>", lambda x: history_V_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
history_V_button.bind("<Leave>", GN_howeroff)

city_button = tk.Label(GN_top_frame, text="City", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
city_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
city_button.bind("<Button-1>", GN_city)
city_button.bind("<Enter>", lambda x: city_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
city_button.bind("<Leave>", GN_howeroff)

statistic_CH_button = tk.Label(GN_top_frame, text="Chart", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
statistic_CH_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
statistic_CH_button.bind("<Button-1>", chart)
statistic_CH_button.bind("<Enter>", lambda x: statistic_CH_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
statistic_CH_button.bind("<Leave>", GN_howeroff)

GN_central_frame = tk.Frame(GN_Main_Frame, highlightthickness=0, bg=conf["window_bg"])
GN_central_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)

############################################################################################################################
#Frame with GN statistic information - packs by GN_stat

# Tenant stat
data_Ten_stat = tk.Frame(GN_central_frame, highlightthickness=0, bg=conf["window_bg"])
data_Ten_stat.pack_propagate(False)
ten_state_filter = filter_frame(data_Ten_stat)
ten_state_scroller = scroller(data_Ten_stat)

# Tenant history
data_Ten_his = tk.Frame(GN_central_frame, highlightthickness=0, bg=conf["window_bg"])
data_Ten_his.pack_propagate(False)
tenant_his_filter = filter_frame(data_Ten_his)
tenant_his_scroll_frame = scroller(data_Ten_his)

# GN stat
data_GN_stat = tk.Frame(GN_central_frame, highlightthickness=0, bg=conf["window_bg"])
data_GN_stat.pack_propagate(False)
GN_state_data_sc_fr = scroller(data_GN_stat)
GN_state_data_sc_fr.pack(side=tk.LEFT, fill=tk.X)

# GN history
data_GN_his = tk.Frame(GN_central_frame, highlightthickness=0, bg=conf["window_bg"])
data_GN_his.pack_propagate(False)
GN_his_filter = filter_frame(data_GN_his)
GN_his_scroll_frame = scroller(data_GN_his)

# V stat
data_Vis_stat = tk.Frame(GN_central_frame, highlightthickness=0, bg=conf["window_bg"])
data_Vis_stat.pack_propagate(False)
Vis_state_filter = filter_frame(data_Vis_stat)
Vis_state_scroller = scroller(data_Vis_stat)

# V history
data_Vis_his = tk.Frame(GN_central_frame, highlightthickness=0, bg=conf["window_bg"])
data_Vis_his.pack_propagate(False)
Vis_his_filter = filter_frame(data_Vis_his)
Vis_his_scroll_frame = scroller(data_Vis_his)


# City
data_city = tk.Frame(GN_central_frame, highlightthickness=0, bg=conf["window_bg"])
data_city.pack_propagate(False)
GN_city_filter = filter_frame(data_city)
GN_city_scroll_frame = scroller(data_city)





# History implement function
def history_insert(func):
    global el_size
    global db_data
    query_string = ""
    if func == "T":
        masta = tenant_his_scroll_frame
        filter_obj = tenant_his_filter
        table_name = "dbo.Tenant_History"
        #getting data from company
        comp_name = filter_obj.comp_box.get()
        # getting date from period filter
        f_year = filter_obj.year_label.cget("text")
        f_month = filter_obj.month_label.cget("text")
        f_day = filter_obj.day_label.cget("text")
        #filter_date = datetime.strptime(f"{f_year}-{f_month}-{f_day}", "%Y-%m-%d").date()  # ???????????
        his_scale = filter_obj.chart_scale
        # getting checkbox
        truck_checkbox = filter_obj.search_truck_checkbox.get()
        trailer_checkbox = filter_obj.search_trailer_checkbox.get()
        if truck_checkbox:
            truck_unit = filter_obj.truck_search_var.get().strip() or None
        else:
            truck_unit = None
        if trailer_checkbox:
            trailer_unit = filter_obj.trailer_search_var.get().strip() or None
        else:
            trailer_unit = None
        where_list = [["YEAR(datetime_event)=?", f_year]]
        if his_scale == "M" or his_scale == "D": where_list.append(["MONTH(datetime_event)=?", f_month])
        if his_scale == "D": where_list.append(["DAY(datetime_event)=?", f_day])
        if truck_unit is not None: where_list.append(["truck_number=?", truck_unit])
        if trailer_unit is not None: where_list.append(["trailer_number=?", trailer_unit])
        if comp_name != "All" and comp_name is not None:
            id_comp = ID_NAME_company(name=comp_name)
            where_list.append(["company_ID=?", id_comp])


    elif func == "GN":
        masta = GN_his_scroll_frame
        filter_obj = GN_his_filter
        table_name = "dbo.GN_History"
        # getting date from period filter
        f_year = filter_obj.year_label.cget("text")
        f_month = filter_obj.month_label.cget("text")
        f_day = filter_obj.day_label.cget("text")
        #filter_date = datetime.strptime(f"{f_year}-{f_month}-{f_day}", "%Y-%m-%d").date()  # ???????????
        his_scale = filter_obj.chart_scale
        # getting checkbox
        truck_checkbox = filter_obj.search_truck_checkbox.get()
        trailer_checkbox = filter_obj.search_trailer_checkbox.get()
        fb_checkbox = filter_obj.search_fb_checkbox.get()
        if truck_checkbox: truck_unit = filter_obj.truck_search_var.get().strip() or None
        else: truck_unit = None
        if trailer_checkbox: trailer_unit = filter_obj.trailer_search_var.get().strip() or None
        else:  trailer_unit = None
        if fb_checkbox: fb_unit = filter_obj.fb_search_var.get().strip() or None
        else: fb_unit = None
        where_list = [["YEAR(datetime_event)=?", f_year]]
        if his_scale == "M" or his_scale == "D": where_list.append(["MONTH(datetime_event)=?", f_month])
        if his_scale == "D": where_list.append(["DAY(datetime_event)=?", f_day])
        if truck_unit is not None: where_list.append(["truck_number=?", truck_unit])
        if trailer_unit is not None: where_list.append(["trailer_number=?", trailer_unit])
        if fb_unit is not None: where_list.append(["fb_number=?", fb_unit])
    elif func == "V":
        masta = Vis_his_scroll_frame
        filter_obj = Vis_his_filter
        table_name = "dbo.visitors_history"
        #############
        # getting data from company
        comp_name = filter_obj.comp_box.get()
        # getting date from period filter
        f_year = filter_obj.year_label.cget("text")
        f_month = filter_obj.month_label.cget("text")
        f_day = filter_obj.day_label.cget("text")
        #filter_date = datetime.strptime(f"{f_year}-{f_month}-{f_day}", "%Y-%m-%d").date()  # ???????????
        his_scale = filter_obj.chart_scale
        # getting checkbox
        plate_checkbox = filter_obj.search_plate_checkbox.get()
        if plate_checkbox:
            plate_unit = filter_obj.plate_search_var.get().strip() or None
        else:
            plate_unit = None
        where_list = [["YEAR(datetime_event)=?", f_year]]
        if his_scale == "M" or his_scale == "D": where_list.append(["MONTH(datetime_event)=?", f_month])
        if his_scale == "D": where_list.append(["DAY(datetime_event)=?", f_day])
        if plate_unit is not None: where_list.append(["plates=?", plate_unit])
        if comp_name != "All" and comp_name is not None:
            id_comp = ID_NAME_company(name=comp_name)
            where_list.append(["company_ID=?", id_comp])
        #####################

    masta.delete()
    condition_var, value_var = zip(*where_list)
    db_data = SQL_REQ(f"SELECT * FROM {table_name} WHERE {' AND '.join(condition_var)} ORDER BY datetime_event", value_var, "S_all")
    X_ind = 0
    Y_ind = 1
    if db_data:
        if len(db_data)>500:
            error(23)
            return
        for rec_line in db_data:
            rec_frame = tk.Frame(masta.frame, highlightthickness=0, bg=conf["window_bg"])
            rec_frame.pack(side=tk.TOP, fill=tk.X, expand=1)
            for indx, element in enumerate(rec_line):
                if indx==0 and (func=="T" or func=="V"): element = ID_NAME_company(ID=element)
                if func=="T" and indx==4: element = "IN" if element else "OUT"
                if func=="GN":
                    if indx==6: element = "IN" if element else "OUT"
                    if indx==5: element = "LOADED" if element else "" if element is None else "UNLOADED"
                if func=="V" and indx==6: element = "IN" if element else "OUT"
                tk.Label(rec_frame, text=element, bg=conf["widget_bg"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, width=el_size[indx], highlightbackground=conf["widget_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.CENTER).grid(row=Y_ind, column=X_ind, sticky=tk.NSEW, padx=(0, 1), pady=(1, 0))
                X_ind+=1
            X_ind=0
            Y_ind+=1
    else:
        tk.Label(masta.frame, text="NO DATA", bg=conf["window_bg"], fg=conf["chart_title"], font=(conf["header_font"], conf["header_size"])).grid(row=Y_ind, column=X_ind, sticky=tk.NSEW, padx=(0, 1), pady=(1, 0))
    masta.refresh()

#function take data as list, path check/create. Create exl files by path and delete all records that is in list from SQL DB.
def archivator(data, path, func):
    global GN_Menu_Var
    # creating folder if not exist
    try:
        if func != "city":
            isExist = os.path.exists(path)
            if not isExist: os.makedirs(path)
    except Exception as e:
        error(f"Cannot create path for archivator:\n{e}")
        debuger(e)
        return False
    # Create excel workbook and sheet
    wb = Workbook()
    ws = wb.active

    # creating excel file
    if func == "T":
        sheet_title = "Tenants"
        year_month = datetime.strptime(data[0][3], "%Y-%m-%d %H:%M:%S").strftime("%Y, %B")
        date_ind = [20, 10, 10, 20, 5, 15, 15, 10, 10]
        xl_title_row = ["Company", "Truck", "Trailer", "Date", "↑|↓", "Comment", "Name"]

    elif func == "GN":
        sheet_title = "GN"
        year_month = datetime.strptime(data[0][4], "%Y-%m-%d %H:%M:%S").strftime("%Y, %B")
        date_ind = [13, 9, 9, 6, 19, 11, 5, 13, 13]
        xl_title_row = ["Company", "Truck", "Trailer", "FB", "Date", "Cargo", "↑|↓", "Comment", "Name"]

    elif func == "V":
        sheet_title = "Visitors"
        year_month = datetime.strptime(data[0][4], "%Y-%m-%d %H:%M:%S").strftime("%Y, %B")
        date_ind = [15, 10, 10, 15, 20, 10, 5, 15, 10]
        xl_title_row = ["Company", "Plate", "Car", "Driver", "Date", "Comment", "↑|↓", "Name"]
    elif func == "print":
        sheet_title = "Custom"
        year_month = "Custom History"
        if GN_Menu_Var == 4:
            date_ind = [20, 10, 10, 20, 5, 15, 15, 10, 10]
            xl_title_row = ["Company", "Truck", "Trailer", "Date", "↑|↓", "Comment", "Name"]
        elif GN_Menu_Var == 2:
            date_ind = [13, 9, 9, 6, 19, 10, 5, 13, 13]
            xl_title_row = ["Company", "Truck", "Trailer", "FB", "Date", "Cargo", "↑|↓", "Comment", "Name"]
        elif GN_Menu_Var == 6:
            date_ind = [15, 10, 10, 15, 20, 10, 5, 15, 10]
            xl_title_row = ["Company", "Plate", "Car", "Driver", "Date", "Comment", "↑|↓", "Name"]
    elif func == "city":
        sheet_title = "City"
        year_month = "Custom"
        date_ind = [15, 12, 7, 7, 9]
        xl_title_row = ["Truck", "Date", "OUT", "IN", "Hours"]

    ws.title = sheet_title
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5, header=0.3)
    ws.oddHeader.center.text = sheet_title
    ws.oddHeader.right.text = "&P of &N"
    ws.oddHeader.left.text = year_month
    ws.append(xl_title_row)
    ws.freeze_panes = "A2"
    for cell in ws[1][0:ws.max_column]: cell.border = Border(top=Side(border_style='thin'), left=Side(border_style='thin'), right=Side(border_style='thin'), bottom=Side(border_style='thin'))
    for row_data in data: ws.append(row_data)
    for i, column_letter in enumerate(range(ord("A"), ord("I") + 1)):
        if i < len(date_ind):
            size = date_ind[i]
        else: break
        column_letter = chr(column_letter)
        ws.column_dimensions[column_letter].width = size

    #Applying formation for each func
    if func == "city":
        # Custom time format for OUT and IN columns
        time_format = NamedStyle(name='time_format', number_format='hh:mm')
        # Apply time format to OUT and IN columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=4):
            for cell in row: cell.style = time_format
        # ws['C'].style = time_format
        # ws['D'].style = time_format
        # Create a custom style for the "Hours" column
        hours_format = NamedStyle(name='hours_format', number_format='hh:mm')
        # Apply the custom style to the "Hours" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
            for cell in row: cell.style = hours_format
        # ws['E'].style = hours_format



    # excel over
    if func == "print": xl_file = f"{path}custom_history.xlsx"
    elif func == "city": xl_file = save_file_as("Excel File", "xlsx")
    else: xl_file = f"{path}{sheet_title} {year_month.replace(', ', '-')}.xlsx"
    #######
    try:
        filexist = os.path.exists(xl_file)
        if filexist:
            confirm = messagebox.askyesno("Replace confirmation", f"File\n{xl_file}\nalready exist.\nReplace it?")
            if not confirm: return None
    except Exception as e:
        error("Cannot access to folder/file for archive.")
        debuger(e)
        return False

    try:
        wb.save(xl_file)
        return True
    except Exception as e:
        error(19)
        debuger(e)
        return False


#archivation function fun=T for tenant GN for GN and V for visitors
def history_archive(*fun):
    global today
    T_list_of_years = SQL_REQ("SELECT DISTINCT YEAR(datetime_event) FROM dbo.Tenant_History", (), "S_all")
    GN_list_of_years = SQL_REQ("SELECT DISTINCT YEAR(datetime_event) FROM dbo.GN_History", (), "S_all")
    V_list_of_years = SQL_REQ("SELECT DISTINCT YEAR(datetime_event) FROM dbo.visitors_history", (), "S_all")
    T_last_years = [str(y[0]) for y in T_list_of_years if str(y[0]) != str(today.year)]
    GN_last_years = [str(y[0]) for y in GN_list_of_years if str(y[0]) != str(today.year)]
    V_last_years = [str(y[0]) for y in V_list_of_years if str(y[0]) != str(today.year)]
    #break out if no records found
    if len(T_last_years) == len(GN_last_years) == len(V_last_years) == 0:
        error("No previous years where found in History.\nData do not exist or have been archivated.")
        return

    #confirming continue archivation
    confirm = messagebox.askyesno("Archivation", f"Archivation will transfer old History in Excel files.\nWARNING!!! All history will be DELETED.\nFound:\nTenant: {'No Recors' if len(T_last_years) == 0 else T_last_years}\nGN: {'No Recors' if len(GN_last_years) == 0 else GN_last_years}\nVisitors: {'No Recors' if len(V_last_years) == 0 else V_last_years}\nContine?")
    if not confirm: return
    company_list = {id: name for name, id in units_lst("company", "D0")}
    if len(T_last_years) != 0:
        T_last_years.sort()
        for last_year in T_last_years:
            for M in range(1, 13):
                month = f"{M:02}"
                data_T = SQL_REQ(f"SELECT * FROM dbo.Tenant_History WHERE YEAR(datetime_event)=? AND MONTH(datetime_event)=? ORDER BY datetime_event", (last_year, month), "S_all")
                if data_T:
                    EXL_data_T = [[company_list[val] if i == 0 else val.strftime("%Y-%m-%d %H:%M:%S") if i == 3 else "" if i == 5 and val is None else "IN" if i == 4 and val else "OUT" if i == 4 and not val else val for i, val in enumerate(x)] for x in data_T]
                else: continue
                arch_func  = archivator(EXL_data_T, f"{sets['archive_path']}\\Tenant\\{last_year}\\", "T")
                if arch_func:
                    for data_line in data_T:
                        try:
                            if query_vars[2] is not None:
                                trailer_lv = "trailer_number=?"
                            else:
                                trailer_lv = "trailer_number is NULL"
                                query_vars.pop(2)
                            query_vars = [data_line[0], data_line[1], data_line[2], data_line[3].strftime("%Y-%m-%d %H:%M:%S")]
                            if query_vars[1] is not None:
                                truck_lv = "truck_number=?"
                            else:
                                truck_lv = "truck_number is NULL"
                                query_vars.pop(1)
                            SQL_REQ(f"DELETE FROM dbo.Tenant_History WHERE company_ID=? AND {truck_lv} AND {trailer_lv} AND datetime_event=?", (query_vars), "W")
                        except Exception as e:
                            error(f"Cannot delete line {query_vars} from SQL.")
                            debuger(e)
                            return
                elif arch_func is None: continue
                else: return

    if len(GN_last_years) != 0:
        GN_last_years.sort()
        for last_year in GN_last_years:
            for M in range(1, 13):
                month = f"{M:02}"
                data_GN = SQL_REQ(f"SELECT * FROM dbo.GN_History WHERE YEAR(datetime_event)=? AND MONTH(datetime_event)=? ORDER BY datetime_event", (last_year, month), "S_all")
                if data_GN:
                    EXL_data_GN = [[val.strftime("%Y-%m-%d %H:%M:%S") if i == 4 else "UNLOADED" if i == 5 and not val else "" if i == 7 and val is None else "LOADED" if i==5 and val else "IN" if i == 6 and val else "OUT" if i == 6 and not val else val for i, val in enumerate(x)] for x in data_GN]
                else: continue
                arch_func  = archivator(EXL_data_GN, f"{sets['archive_path']}\\GN\\{last_year}\\", "GN")
                if arch_func:
                    for data_line in data_GN:
                        try:
                            query_vars = [data_line[0], data_line[1], data_line[2], data_line[3], data_line[4].strftime("%Y-%m-%d %H:%M:%S")]
                            if query_vars[3] is not None:
                                fb_lv = "fb_number=?"
                            else:
                                fb_lv = "fb_number is NULL"
                                query_vars.pop(3)
                            if query_vars[2] is not None:
                                trailer_lv = "trailer_number=?"
                            else:
                                trailer_lv = "trailer_number is NULL"
                                query_vars.pop(2)
                            if query_vars[1] is not None:
                                truck_lv = "truck_number=?"
                            else:
                                truck_lv = "truck_number is NULL"
                                query_vars.pop(1)
                            SQL_REQ(f"DELETE FROM dbo.GN_History WHERE company_name=? AND {truck_lv} AND {trailer_lv} AND {fb_lv} AND datetime_event=?", (query_vars), "W")
                        except Exception as e:
                            error(f"Cannot delete line {query_vars} from SQL.")
                            debuger(e)
                            return
                elif arch_func is None: continue
                else: return
    if len(V_last_years) != 0:
        V_last_years.sort()
        for last_year in V_last_years:
            for M in range(1, 13):
                month = f"{M:02}"
                data_V = SQL_REQ(f"SELECT * FROM dbo.visitors_history WHERE YEAR(datetime_event)=? AND MONTH(datetime_event)=? ORDER BY datetime_event", (last_year, month), "S_all")
                if data_V:
                    EXL_data_V = [[company_list[val] if i == 0 else val.strftime("%Y-%m-%d %H:%M:%S") if i == 4 else "" if i == 5 and val is None else "IN" if i == 6 and val else "OUT" if i == 6 and not val else val for i, val in enumerate(x)] for x in data_V]
                else: continue
                arch_func  = archivator(EXL_data_V, f"{sets['archive_path']}\\Visitor\\{last_year}\\", "V")
                if arch_func:
                    for data_line in data_V:
                        try:
                            query_vars = [data_line[0], data_line[1], data_line[4].strftime("%Y-%m-%d %H:%M:%S")]
                            SQL_REQ(f"DELETE FROM dbo.visitors_history WHERE company_ID=? AND plates=? AND datetime_event=?", (query_vars), "W")
                        except Exception as e:
                            error(f"Cannot delete line {query_vars} from SQL.")
                            debuger(e)
                            return
                elif arch_func is None: continue
                else: return

    #
    # path for files
    # isExist = os.path.exists(sets["chk_path"])
    # if not isExist:
    #     os.makedirs(sets["chk_path"])
    # wb.save(sets["chk_path"] + "CheckYard " + date.strftime("%Y") + date.strftime("%m") + date.strftime("%d") + ".xlsx")
    #





def history_print():
    global db_data
    global GN_Menu_Var
    if db_data is None or db_data == "":
        error("No DATA to print")
        return
    company_list = {id: name for name, id in units_lst("company", "D0")}
    data = [[company_list[val] if i == 0 and GN_Menu_Var != 2 else val for i, val in enumerate(row)] for row in db_data]
    if GN_Menu_Var == 4:
        data = [[company_list[val] if i == 0 else val.strftime("%Y-%m-%d %H:%M:%S") if i == 3 else "" if i == 5 and val is None else "IN" if i == 4 and val else "OUT" if i == 4 and not val else val for i, val in enumerate(x)] for x in db_data]
    elif GN_Menu_Var == 2:
        data = [[val.strftime("%Y-%m-%d %H:%M:%S") if i == 4 else "UNLOADED" if i == 5 and not val else "" if i == 7 and val is None else "LOADED" if i == 5 and val else "IN" if i == 6 and val else "OUT" if i == 6 and not val else val for i, val in enumerate(x)] for x in db_data]
    elif GN_Menu_Var == 6:
        data = [[company_list[val] if i == 0 else val.strftime("%Y-%m-%d %H:%M:%S") if i == 4 else "" if i == 5 and val is None else "IN" if i == 6 and val else "OUT" if i == 6 and not val else val for i, val in enumerate(x)] for x in db_data]

    path = f"{sets['archive_path']}\\"
    excl_custom = archivator(data, path, "print")
    if excl_custom:
        filepath = path.replace("\\\\", "/")+"custom_history.xlsx"
        excel = win32.DispatchEx('Excel.Application')
        wb = excel.Workbooks.Open(filepath)
        sheets = wb.Sheets
        for sheet in sheets:
            sheet.PrintOut()
        wb.Close(False)
        excel.Quit()

    else:
        error("Cannot print custom file...\nCheck if file exist.")



#function for creating EXL files for archive. Takes list of events and path modefied for each history.




#history_archive("T")


# #Main frame for Header
# data_GN_his_label_fr = tk.Frame(data_GN_his, highlightthickness=0, bg=conf["window_bg"])
# data_GN_his_label_fr.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, anchor=tk.N)
#
# #Frame, label, combo for Company
# GN_his_comp_fr = tk.Frame(data_GN_his_label_fr, highlightthickness=0, bg=conf["window_bg"])
# GN_his_comp_fr.pack(side=tk.LEFT, padx=(0,1))
# GN_his_comp_lb = tk.Label(GN_his_comp_fr, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
# GN_his_comp_lb.pack(side=tk.TOP, fill=tk.BOTH)
# GN_his_combobox_C = ttk.Combobox(GN_his_comp_fr, values=GN_combo_C_list, background=conf["window_bg"], foreground=conf["submenu_fg"], font=(conf["widget_font"], conf["widget_size"]), state="readonly", width=10)
# GN_his_combobox_C.pack(fill=tk.BOTH, side=tk.TOP)
# GN_his_combobox_C.current(0)
# GN_his_combobox_C.bind("<<ComboboxSelected>>", lambda event: update_GN_his(event, GN_his_combobox_C, 0))
#
# #Frame, label, combo for Truck
# GN_his_T_fr = tk.Frame(data_GN_his_label_fr, highlightthickness=0, bg=conf["window_bg"])
# GN_his_T_fr.pack(side=tk.LEFT, padx=(0,1))
# GN_his_T_lb = tk.Label(GN_his_T_fr, text="TRUCK:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
# GN_his_T_lb.pack(side=tk.TOP, fill=tk.BOTH)
# GN_his_combobox_T = ttk.Combobox(GN_his_T_fr, values=GN_combo_T_list, background=conf["window_bg"], foreground=conf["submenu_fg"], font=(conf["widget_font"], conf["widget_size"]), state="readonly", width=10)
# GN_his_combobox_T.pack(fill=tk.BOTH, side=tk.TOP)
# GN_his_combobox_T.current(0)
# GN_his_combobox_T.bind("<<ComboboxSelected>>", lambda event: update_GN_his(event, GN_his_combobox_T, 1))
#
# #Frame, label, combo for Trailer
# GN_his_Tr_fr = tk.Frame(data_GN_his_label_fr, highlightthickness=0, bg=conf["window_bg"])
# GN_his_Tr_fr.pack(side=tk.LEFT, padx=(0,1))
# GN_his_Tr_lb = tk.Label(GN_his_Tr_fr, text="TRAILER:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
# GN_his_Tr_lb.pack(side=tk.TOP, fill=tk.BOTH)
# GN_his_combobox_Tr = ttk.Combobox(GN_his_Tr_fr, values=GN_combo_Tr_list, background=conf["window_bg"], foreground=conf["submenu_fg"], font=(conf["widget_font"], conf["widget_size"]), state="readonly", width=10)
# GN_his_combobox_Tr.pack(fill=tk.BOTH, side=tk.TOP)
# GN_his_combobox_Tr.current(0)
# GN_his_combobox_Tr.bind("<<ComboboxSelected>>", lambda event: update_GN_his(event, GN_his_combobox_Tr, 2))
#
# #Frame, label, combo for FB
# GN_his_fb_fr = tk.Frame(data_GN_his_label_fr, highlightthickness=0, bg=conf["window_bg"])
# GN_his_fb_fr.pack(side=tk.LEFT, padx=(0,1))
# GN_his_fb_lb = tk.Label(GN_his_fb_fr, text="FB:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
# GN_his_fb_lb.pack(side=tk.TOP, fill=tk.BOTH)
# GN_his_combobox_fb = ttk.Combobox(GN_his_fb_fr, values=GN_combo_FB_list, background=conf["window_bg"], foreground=conf["submenu_fg"], font=(conf["widget_font"], conf["widget_size"]), state="readonly", width=10)
# GN_his_combobox_fb.pack(fill=tk.BOTH, side=tk.TOP)
# GN_his_combobox_fb.current(0)
# GN_his_combobox_fb.bind("<<ComboboxSelected>>", lambda event: update_GN_his(event, GN_his_combobox_fb, 3))
#
# #Frame, label, combo for Cargo
# GN_his_LU_fr = tk.Frame(data_GN_his_label_fr, highlightthickness=0, bg=conf["window_bg"])
# GN_his_LU_fr.pack(side=tk.LEFT, padx=(0,1))
# GN_his_LU_lb = tk.Label(GN_his_LU_fr, text="CARGO:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
# GN_his_LU_lb.pack(side=tk.TOP, fill=tk.BOTH)
# GN_his_combobox_LU = ttk.Combobox(GN_his_LU_fr, values=GN_combo_cargo_list, background=conf["window_bg"], foreground=conf["submenu_fg"], font=(conf["widget_font"], conf["widget_size"]), state="readonly", width=10)
# GN_his_combobox_LU.pack(fill=tk.BOTH, side=tk.TOP)
# GN_his_combobox_LU.current(0)
# GN_his_combobox_LU.bind("<<ComboboxSelected>>", lambda event: update_GN_his(event, GN_his_combobox_LU, 5))
#
# #Frame, label, combo for Status
# GN_his_S_fr = tk.Frame(data_GN_his_label_fr, highlightthickness=0, bg=conf["window_bg"])
# GN_his_S_fr.pack(side=tk.LEFT, padx=(0,1))
# GN_his_S_lb = tk.Label(GN_his_S_fr, text="STATUS:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
# GN_his_S_lb.pack(side=tk.TOP, fill=tk.BOTH)
# GN_his_combobox_S = ttk.Combobox(GN_his_S_fr, values=GN_combo_status_list, background=conf["window_bg"], foreground=conf["submenu_fg"], font=(conf["widget_font"], conf["widget_size"]), state="readonly", width=10)
# GN_his_combobox_S.pack(fill=tk.BOTH, side=tk.TOP)
# GN_his_combobox_S.current(0)
# GN_his_combobox_S.bind("<<ComboboxSelected>>", lambda event: update_GN_his(event, GN_his_combobox_S, 6))
#
# #Frame, label, combo for Date
# GN_his_D_fr = tk.Frame(data_GN_his_label_fr, highlightthickness=0, bg=conf["window_bg"])
# GN_his_D_fr.pack(side=tk.LEFT, padx=(0,1))
# GN_his_D_lb = tk.Label(GN_his_D_fr, text="DATE:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=30)
# GN_his_D_lb.pack(side=tk.TOP, fill=tk.BOTH)
# GN_his_combo_fr = tk.Frame(GN_his_D_fr, highlightthickness=0, bg=conf["window_bg"])
# GN_his_combo_fr.pack(side=tk.TOP)
# GN_his_combobox_Y = ttk.Combobox(GN_his_combo_fr, values=GN_combo_year_list, background=conf["window_bg"], foreground=conf["submenu_fg"], font=(conf["widget_font"], conf["widget_size"]), state="readonly", width=10)
# GN_his_combobox_Y.pack(fill=tk.BOTH, side=tk.LEFT)
# GN_his_combobox_Y.current(0)
# GN_his_combobox_Y.bind("<<ComboboxSelected>>", lambda event: update_GN_his(event, GN_his_combobox_Y, 41))
# GN_his_combobox_M = ttk.Combobox(GN_his_combo_fr, values=GN_combo_month_list, background=conf["window_bg"], foreground=conf["submenu_fg"], font=(conf["widget_font"], conf["widget_size"]), state="readonly", width=10)
# GN_his_combobox_M.pack(fill=tk.BOTH, side=tk.LEFT)
# GN_his_combobox_M.current(0)
# GN_his_combobox_M.bind("<<ComboboxSelected>>", lambda event: update_GN_his(event, GN_his_combobox_M, 42))
# GN_his_combobox_D = ttk.Combobox(GN_his_combo_fr, values=GN_combo_days_list, background=conf["window_bg"], foreground=conf["submenu_fg"], font=(conf["widget_font"], conf["widget_size"]), state="readonly", width=10)
# GN_his_combobox_D.pack(fill=tk.BOTH, side=tk.LEFT)
# GN_his_combobox_D.current(0)
# GN_his_combobox_D.bind("<<ComboboxSelected>>", lambda event: update_GN_his(event, GN_his_combobox_D, 43))
#
# #Frame, label, combo for Comment
# GN_his_comm_fr = tk.Frame(data_GN_his_label_fr, highlightthickness=0, bg=conf["window_bg"])
# GN_his_comm_fr.pack(side=tk.LEFT)
# GN_his_comm_lb = tk.Label(GN_his_comm_fr, text="COMMENT:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
# GN_his_comm_lb.pack(side=tk.TOP, fill=tk.BOTH)
# GN_his_combobox_comm = ttk.Combobox(GN_his_comm_fr, state="disabled", background=conf["window_bg"], foreground=conf["submenu_fg"], font=(conf["widget_font"], conf["widget_size"]), width=10)
# GN_his_combobox_comm.pack(fill=tk.BOTH, side=tk.TOP)






#Comboboxes for rows





#
# # Scrollable frame for Data
# sub_GN_his_frame = tk.Frame(data_GN_his, highlightthickness=0)
# sub_GN_his_frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1)
#
#
# GN_his_canv = tk.Canvas(sub_GN_his_frame, bg=conf["window_bg"], highlightthickness=0)
# second_GN_his_frame = tk.Frame(GN_his_canv, bg=conf["window_bg"])
# GN_his_scrl = ttk.Scrollbar(sub_GN_his_frame, orient=tk.VERTICAL, command=GN_his_canv.yview)
# GN_his_canv.config(yscrollcommand=GN_his_scrl.set)
# GN_his_scrl.pack(fill=tk.Y, side=tk.RIGHT)
# GN_his_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
# GN_his_canv.create_window((0, 0), window=second_GN_his_frame, anchor=tk.NW)
# second_GN_his_frame.bind("<Configure>", lambda event, canvas=GN_his_canv: GN_his_canv.configure(scrollregion=GN_his_canv.bbox("all")))
# def GN_his_scroll_region(*event):
#     if second_GN_his_frame.winfo_height() <= GN_his_canv.winfo_height():
#         GN_his_scrl.pack_forget()
#         GN_his_canv.configure(yscrollcommand=None)
#         second_GN_his_frame.unbind("<Enter>")
#         second_GN_his_frame.unbind_all("<MouseWheel>")
#     else:
#         GN_his_scrl.pack(side=tk.RIGHT, fill=tk.Y)
#         GN_his_canv.configure(yscrollcommand=GN_his_scrl.set)
#         second_GN_his_frame.bind("<Enter>", GN_his_enter_mousewheel, add="+")
# GN_his_canv.bind("<Configure>", GN_his_scroll_region)
#
# def GN_his_on_mousewheel(event): GN_his_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
# def GN_his_enter_mousewheel(event): GN_his_canv.bind_all('<MouseWheel>', GN_his_on_mousewheel, add="+")
# def GN_his_leave_mousewheel(event): GN_his_canv.unbind_all('<MouseWheel>')
# second_GN_his_frame.bind("<Enter>", GN_his_enter_mousewheel, add="+")
# second_GN_his_frame.bind("<Leave>", GN_his_leave_mousewheel)
# GN_his_canv.config(bg=conf["window_bg"])
#

#############################################
#                   Chart                   #
#############################################
def Y_LEFT():
    GN_Chart_Year.config(text=str(int(GN_Chart_Year.cget("text"))-1))
    daylim = monthrange(int(GN_Chart_Year.cget("text")), int(GN_Chart_Month.cget("text")))
    if daylim[1] < int(GN_Chart_Day.cget("text")): GN_Chart_Day.config(text=str(daylim[1]))
    Chart_Draw()

def Y_RIGHT():
    GN_Chart_Year.config(text=str(int(GN_Chart_Year.cget("text"))+1))
    daylim = monthrange(int(GN_Chart_Year.cget("text")), int(GN_Chart_Month.cget("text")))
    if daylim[1] < int(GN_Chart_Day.cget("text")): GN_Chart_Day.config(text=str(daylim[1]))
    Chart_Draw()
def M_LEFT():
    if int(GN_Chart_Month.cget("text")) == 1: return
    GN_Chart_Month.config(text=str(int(GN_Chart_Month.cget("text"))-1))
    daylim = monthrange(int(GN_Chart_Year.cget("text")), int(GN_Chart_Month.cget("text")))
    if daylim[1] < int(GN_Chart_Day.cget("text")): GN_Chart_Day.config(text=str(daylim[1]))
    Chart_Draw()
def M_RIGHT():
    if int(GN_Chart_Month.cget("text")) ==12: return
    GN_Chart_Month.config(text=str(int(GN_Chart_Month.cget("text"))+1))
    daylim = monthrange(int(GN_Chart_Year.cget("text")), int(GN_Chart_Month.cget("text")))
    if daylim[1] < int(GN_Chart_Day.cget("text")): GN_Chart_Day.config(text=str(daylim[1]))
    Chart_Draw()
def D_LEFT():
    if int(GN_Chart_Day.cget("text")) == 1: return
    GN_Chart_Day.config(text=str(int(GN_Chart_Day.cget("text"))-1))
    Chart_Draw()
def D_RIGHT():
    daylim = monthrange(int(GN_Chart_Year.cget("text")), int(GN_Chart_Month.cget("text")))
    if int(GN_Chart_Day.cget("text")) >= daylim[1]: return
    GN_Chart_Day.config(text=str(int(GN_Chart_Day.cget("text"))+1))
    Chart_Draw()

def Chart_Scale(func):
    global chart_scale

    GN_Chart_Year.config(bg=conf["widget_bg"])
    GN_scale_year_button.config(fg=conf["on_parking"])
    GN_Chart_Month.config(bg=conf["widget_bg"])
    GN_scale_month_button.config(fg=conf["on_parking"])
    GN_Chart_Day.config(bg=conf["widget_bg"])
    GN_scale_day_button.config(fg=conf["on_parking"])

    if func == "Y":
        chart_scale = "Y"
        GN_Chart_Year.config(bg=conf["widget_sel_bg"])
        GN_scale_year_button.config(fg=conf["expired_date"])
        Chart_Draw()

    elif func == "M":
        chart_scale = "M"
        GN_Chart_Month.config(bg=conf["widget_sel_bg"])
        GN_scale_month_button.config(fg=conf["expired_date"])
        Chart_Draw()
    elif func == "D":
        chart_scale = "D"
        GN_Chart_Day.config(bg=conf["widget_sel_bg"])
        GN_scale_day_button.config(fg=conf["expired_date"])
        Chart_Draw()

def Chart_Draw():
    global chart_scale
    if chart_scale is None: return
    for widget in GN_Chart_Left_Frame.winfo_children(): widget.destroy()
    year = GN_Chart_Year.cget("text")
    month = GN_Chart_Month.cget("text")
    day = GN_Chart_Day.cget("text")
    ChartType = bool(GN_Chart_Type.get())
    if chart_scale == "Y": scale = f"YEAR(date)={year}"
    elif chart_scale == "M": scale = f"YEAR(date)={year} AND MONTH(date)={month}"
    elif chart_scale == "D": scale = f"YEAR(date)={year} AND MONTH(date)={month} AND DAY(date)={day}"
    query = f"SELECT * FROM dbo.[statistics] WHERE ({scale})"
    chart_data = SQL_REQ(query, (), "S_all")
    if chart_data:
        sorted_data = sorted(chart_data, key=lambda x: x[0])
        fig = Figure(figsize=(6, 3), dpi=100)
        ax = fig.add_subplot(111)
        ax.set_facecolor(conf["window_bg"])
        fig.patch.set_facecolor(conf["window_bg"])
        ###############
        x = []
        y = []
        for param in sorted_data:
            x.append(param[0]) #Date
            if ChartType:
                y.append(param[1]) #Tenant
            else: y.append(param[2]) #GN
        chart_lbl = "Tenant" if ChartType else "GN"
        chart_col = conf["chart_tenant"] if ChartType else conf["chart_gn"]


        # Plot the first line with a blue color
        #ax.scatter(x, y, label=chart_lbl, color=chart_col)
        if len(x)>1: ax.plot(x, y, label=chart_lbl, color=chart_col)
        else: ax.bar(x, y, label=chart_lbl, color=chart_col)

     #   ax.set_ylim(100,max(max(y1),max(y2))+10)
        # Customize labels, title, and legend
        ax.set_xlabel('Date')
        ax.set_ylabel('Units')
        title = ax.set_title('Chart of Parking Load')
        title.set_color(conf["chart_title"])
        ax.legend()

        #customise color
        ax.spines["bottom"].set_color(conf["chart_ax"])
        ax.spines["left"].set_color(conf["chart_ax"])
        ax.spines["right"].set_color(conf["chart_ax"])
        ax.spines["top"].set_color(conf["chart_ax"])
        ax.xaxis.label.set_color("yellow")
        ax.yaxis.label.set_color("yellow")
        ax.tick_params(axis='x', colors=conf["chart_ax"])
        ax.tick_params(axis='y', colors=conf["chart_ax"])
        #rotate date param
        ax.set_xticks(x)
        ax.set_xticklabels(ax.get_xticks(), rotation=90)
        date_format = mdates.DateFormatter("%Y-%m-%d")
        ax.xaxis.set_major_formatter(date_format)
        # Create a canvas to display the Matplotlib figure
        canvas = FigureCanvasTkAgg(fig, master=GN_Chart_Left_Frame)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill=tk.BOTH, expand=1)
        for d, v in zip(x,y): ax.text(d, v, str(v), ha="center", va="bottom", color="yellow")

    else:
        No_Data_lb = tk.Label(GN_Chart_Left_Frame, text="NO DATA", bg=conf["window_bg"], fg=conf["chart_title"], font=(conf["header_font"], conf["header_size"]))
        No_Data_lb.pack(side=tk.TOP, fill=tk.BOTH, anchor=tk.CENTER)

            #chart#

GN_Chart_Frame = tk.Frame(GN_central_frame, highlightthickness=0, bg=conf["window_bg"])
GN_Chart_Frame.pack_propagate(False)

#Left frame for Chart
GN_Chart_Left_Frame = tk.Frame(GN_Chart_Frame, highlightthickness=0, bg=conf["window_bg"])
GN_Chart_Left_Frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, padx=(0, 5))





#Right frame for Buttons
GN_Chart_Right_Frame = tk.Frame(GN_Chart_Frame, highlightthickness=3, relief=tk.RAISED, bg=conf["submenu_bg"], highlightbackground=conf["submenu_sel_bg"], width=conf["chk_filter_frame"])
GN_Chart_Right_Frame.pack_propagate(0)
GN_Chart_Right_Frame.pack(side=tk.RIGHT, fill=tk.BOTH)

GN_Chart_Sweetchers_Frame = tk.Frame(GN_Chart_Right_Frame, highlightthickness=0, bg=conf["submenu_bg"])
GN_Chart_Sweetchers_Frame.pack(side=tk.TOP, fill=tk.X, expand=0)
GN_Chart_Sweetchers_lb = tk.Label(GN_Chart_Sweetchers_Frame, text="Period:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
GN_Chart_Sweetchers_lb.pack(fill=tk.X, side=tk.TOP, expand=1, pady=1)

#year buttons
GN_Chart_Year_Frame = tk.Frame(GN_Chart_Sweetchers_Frame, highlightthickness=0, bg=conf["window_bg"])
GN_Chart_Year_Frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 1))
GN_Chart_Sw_Y_L_btn = tk.Button(GN_Chart_Year_Frame, text="\N{LEFTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=Y_LEFT)
GN_Chart_Sw_Y_L_btn.pack(fill=tk.X, side=tk.LEFT)
GN_Chart_Year = tk.Label(GN_Chart_Year_Frame, text=str(datetime.now().year), relief=tk.FLAT, bg=conf["widget_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
GN_Chart_Year.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
GN_Chart_Sw_Y_R_btn = tk.Button(GN_Chart_Year_Frame, text="\N{RIGHTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=Y_RIGHT)
GN_Chart_Sw_Y_R_btn.pack(fill=tk.X, side=tk.LEFT)
#month buttons
GN_Chart_Month_Frame = tk.Frame(GN_Chart_Sweetchers_Frame, highlightthickness=0, bg=conf["window_bg"])
GN_Chart_Month_Frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 1))
GN_Chart_Sw_M_L_btn = tk.Button(GN_Chart_Month_Frame, text="\N{LEFTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=M_LEFT)
GN_Chart_Sw_M_L_btn.pack(fill=tk.X, side=tk.LEFT)
GN_Chart_Month = tk.Label(GN_Chart_Month_Frame, text=str(datetime.now().month), relief=tk.FLAT, bg=conf["widget_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
GN_Chart_Month.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
GN_Chart_Sw_M_R_btn = tk.Button(GN_Chart_Month_Frame, text="\N{RIGHTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=M_RIGHT)
GN_Chart_Sw_M_R_btn.pack(fill=tk.X, side=tk.LEFT)
#day buttons
GN_Chart_Day_Frame = tk.Frame(GN_Chart_Sweetchers_Frame, highlightthickness=0, bg=conf["window_bg"])
GN_Chart_Day_Frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 1))
GN_Chart_Sw_D_L_btn = tk.Button(GN_Chart_Day_Frame, text="\N{LEFTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=D_LEFT)
GN_Chart_Sw_D_L_btn.pack(fill=tk.X, side=tk.LEFT)
GN_Chart_Day = tk.Label(GN_Chart_Day_Frame, text=str(datetime.now().day), relief=tk.FLAT, bg=conf["widget_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
GN_Chart_Day.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
GN_Chart_Sw_D_R_btn = tk.Button(GN_Chart_Day_Frame, text="\N{RIGHTWARDS ARROW}", bg=conf["submenu_bg"], fg=conf["submenu_fg"], bd=1, highlightthickness=1, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=D_RIGHT)
GN_Chart_Sw_D_R_btn.pack(fill=tk.X, side=tk.LEFT)

# Filter by lentgh buttons
GN_Chart_Scale_Frame = tk.Frame(GN_Chart_Right_Frame, highlightthickness=0, bg=conf["submenu_bg"])
GN_Chart_Scale_Frame.pack(side=tk.TOP, fill=tk.X, expand=0, pady = (20, 0))
GN_Chart_Scale_lb = tk.Label(GN_Chart_Scale_Frame, text="Scale:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
GN_Chart_Scale_lb.pack(fill=tk.X, side=tk.TOP, expand=1, pady=1)

GN_scale_year_button = tk.Button(GN_Chart_Scale_Frame, text="Year", bg=conf["widget_bg"], fg=conf["on_parking"], bd=0, highlightthickness=0, activebackground=conf["on_parking"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: Chart_Scale("Y"))
GN_scale_year_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
GN_scale_month_button = tk.Button(GN_Chart_Scale_Frame, text="Month", bg=conf["widget_bg"], fg=conf["on_parking"], bd=0, highlightthickness=0, activebackground=conf["on_parking"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: Chart_Scale("M"))
GN_scale_month_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
GN_scale_day_button = tk.Button(GN_Chart_Scale_Frame, text="Day", bg=conf["widget_bg"], fg=conf["on_parking"], bd=0, highlightthickness=0, activebackground=conf["on_parking"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=lambda: Chart_Scale("D"))
GN_scale_day_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
GN_Chart_Type = tk.BooleanVar()
GN_Chart_Tenant_R1 = tk.Radiobutton(GN_Chart_Scale_Frame, text="Tenant", bg=conf["submenu_bg"], fg=conf["submenu_fg"], highlightthickness=0, bd=0, font=(conf["submenu_font"], conf["notebook_tab_size"]), variable=GN_Chart_Type, command=Chart_Draw, value=True)
GN_Chart_Tenant_R1.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
GN_Chart_Tenant_R2 = tk.Radiobutton(GN_Chart_Scale_Frame, text="GN", bg=conf["submenu_bg"], fg=conf["submenu_fg"], highlightthickness=0, bd=0, font=(conf["submenu_font"], conf["notebook_tab_size"]), variable=GN_Chart_Type, command=Chart_Draw, value=False)
GN_Chart_Tenant_R2.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
GN_Chart_Tenant_R1.select()


##########################################
# #Button Frame for Adding/Delete Units
# adm_chart_buttons_frame = tk.Frame(GN_Chart_Right_Frame, highlightthickness=0, bg=conf["submenu_bg"])
# adm_chart_buttons_frame.pack(side=tk.TOP)
# adm_c_lb = tk.Label(adm_chart_buttons_frame, text="Company:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
# adm_c_lb.pack(fill=tk.X, side=tk.TOP)
# adm_c_entry = tk.Entry(adm_chart_buttons_frame, bg=conf["window_bg"], bd=1, cursor="shuttle", font=(conf["entry_font"], conf["notebook_tab_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, justify=tk.LEFT, disabledbackground=conf["window_bg"], disabledforeground=conf["widget_fg"])
# adm_c_entry.pack(fill=tk.X, side=tk.TOP)
# adm_t_lb = tk.Label(adm_chart_buttons_frame, text="UNIT:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
# adm_t_lb.pack(fill=tk.X, side=tk.TOP)
# adm_t_entry = tk.Entry(adm_chart_buttons_frame, bg=conf["window_bg"], bd=1, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, justify=tk.LEFT)
# adm_t_entry.pack(fill=tk.X, side=tk.TOP)
# adm_t_entry.bind("<Return>", adm_add)
# adm_radio_var = tk.StringVar()
# adm_truck_radio = tk.Radiobutton(adm_chart_buttons_frame, text="Truck", variable=adm_radio_var, value="truck", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
# adm_truck_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
# adm_trailer_radio = tk.Radiobutton(adm_chart_buttons_frame, text="Trailer", variable=adm_radio_var, value="trailer", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
# adm_trailer_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
# adm_radio_var.set("truck")
# adm_manual_add_button = tk.Button(adm_chart_buttons_frame, text="ADD", bg=conf["widget_bg"], fg=conf["in_button_fg"], bd=0, highlightthickness=0, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=adm_add)
# adm_manual_add_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
# adm_manual_remove_button = tk.Button(adm_chart_buttons_frame, text="REMOVE", bg=conf["widget_bg"], fg=conf["out_button_fg"], bd=0, highlightthickness=0, activebackground=conf["out_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["out_button_sel_fg"], command=adm_remove)
# adm_manual_remove_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
# adm_storage_var = tk.IntVar()
# adm_storage_checkbox = tk.Checkbutton(adm_chart_buttons_frame, text="Storage", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], variable=adm_storage_var, command=adm_storage_check)
# ###############################




################################################################################################################################################################################

GN_screen = screen_x//16

################################################################################################################################################################################
################################################################################################################################################################################
#                      Check yard
################################################################################################################################################################################
################################################################################################################################################################################
#screen size adjustment
s1ze = screen_x-int(conf["chk_filter_frame"])
chk_data_size = (s1ze//17)+4
def get_onyard():
    def extract(var):
        comp = units_lst("company", "D0")
        comp_dic = {}
        for x in comp: comp_dic.update({x[1]: x[0]})
        list = []
        if not var: return comp_dic
        comp_dic.update({999: "UNREGISTERED"})
        # for items in comp_dic.items(): print(items)
        for n in var[0]:
            u_dir = {}
            index = 0
            for y in n:
                z = var[1][index]
                index+=1
                if y is not None:
                    if z[0] == "company_ID":
                        y = comp_dic[y]
                    if z[0] == "fb_number": y = "FB"+y
                    u_dir.update({z[0]: y})
                else: u_dir.update({z[0]: None})
            list.append(u_dir)
        return list
    ten_truck = SQL_REQ("SELECT * FROM dbo.Tenant_Trucks WHERE status=1", (), "S_all_D")
    ten_truck_ur = SQL_REQ("SELECT * FROM dbo.Tenant_Trucks_UNREG WHERE status=1", (), "S_all_D")
    ten_trailer = SQL_REQ("SELECT * FROM dbo.Tenant_Trailers WHERE status=1", (), "S_all_D")
    ten_trailer_ur = SQL_REQ("SELECT * FROM dbo.Tenant_Trailers_UNREG WHERE status=1", (), "S_all_D")
    gn_truck = SQL_REQ("SELECT * FROM dbo.GN_Trucks Where status=1", (), "S_all_D")
    gn_trailer = SQL_REQ("SELECT * FROM dbo.GN_Trailers WHERE status=1", (), "S_all_D")
    gn_fb = SQL_REQ("SELECT * FROM dbo.GN_Flatbed WHERE status=1", (), "S_all_D")
    vis_cars = SQL_REQ("SELECT * FROM dbo.visitors WHERE status=1", (), "S_all_D")
    vis_cars_UNREG = SQL_REQ("SELECT * FROM dbo.visitors_UNREG WHERE status=1", (), "S_all_D")
    all_ten_trucks = extract(ten_truck)+extract(ten_truck_ur)
    all_ten_trailers = extract(ten_trailer)+extract(ten_trailer_ur)
    all_gn_trucks = extract(gn_truck)
    all_gn_trailers = extract(gn_trailer)+extract(gn_fb)
    all_vis_cars = extract(vis_cars)+extract(vis_cars_UNREG)
    res = {
        "tenant": [all_ten_trucks, all_ten_trailers],
        "GN": [all_gn_trucks, all_gn_trailers],
        "visitor": all_vis_cars
    }
    return res

Check_Main_Frame = tk.Frame(Menu_Bar_Check, highlightthickness=0, bg=conf["window_bg"])
Check_Main_Frame.pack(fill=tk.BOTH, expand=1)


def chk_tenant(*args):
    global Check_Menu_Var
    Check_Menu_Var = 1
    for all in chk_center_frame.winfo_children(): all.pack_forget()
    checkyard_Ten_Filter.delete()
    checkyard_Ten_Filter.pack(side=tk.LEFT, fill=tk.Y)
    checkyard_Ten_Filter.checkyard(
        company_func=lambda *args: checkyard_insert(checkyard_Ten_Filter, checkyard_Ten_sc_fr),
        time_on_yard_func=lambda *args: checkyard_insert(checkyard_Ten_Filter, checkyard_Ten_sc_fr),
        check_generate=lambda *args: checkyard_generate(checkyard_Ten_Filter),
        check_print=lambda *args: checkyard_print(),
        truck_func=lambda *args: checkyard_insert(checkyard_Ten_Filter, checkyard_Ten_sc_fr),
        trailer_func=lambda *args: checkyard_insert(checkyard_Ten_Filter, checkyard_Ten_sc_fr),
        storage_func=lambda *args: checkyard_insert(checkyard_Ten_Filter, checkyard_Ten_sc_fr),
        age_func=lambda *args: checkyard_insert(checkyard_Ten_Filter, checkyard_Ten_sc_fr)
    )
    checkyard_Ten_Main_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
    checkyard_Ten_sc_fr.pack(side=tk.LEFT, fill=tk.BOTH)
    checkyard_Ten_sc_fr.refresh()
    checkyard_Ten_sc_fr.top()
    checkyard_insert(checkyard_Ten_Filter, checkyard_Ten_sc_fr)

    #data_fr.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)

    # chk_insert(second_data_frame)
    # chk_canv.yview_moveto(0)


def chk_GN(*args):
    global Check_Menu_Var
    Check_Menu_Var = 2
    for all in chk_center_frame.winfo_children(): all.pack_forget()
    checkyard_GN_Filter.delete()
    checkyard_GN_Filter.pack(side=tk.LEFT, fill=tk.Y)
    checkyard_GN_Filter.checkyard_GN(
        time_on_yard_func=lambda *args: checkyard_GN_insert(checkyard_GN_Filter, checkyard_GN_sc_fr),
        check_generate=lambda *args: checkyard_generate(checkyard_GN_Filter),
        check_print=lambda *args: checkyard_print(),
        truck_func=lambda *args: checkyard_GN_insert(checkyard_GN_Filter, checkyard_GN_sc_fr),
        trailer_func=lambda *args: checkyard_GN_insert(checkyard_GN_Filter, checkyard_GN_sc_fr),
        storage_func=lambda *args: checkyard_GN_insert(checkyard_GN_Filter, checkyard_GN_sc_fr),
        age_func=lambda *args: checkyard_GN_insert(checkyard_GN_Filter, checkyard_GN_sc_fr),
        fb_func=lambda *args: checkyard_GN_insert(checkyard_GN_Filter, checkyard_GN_sc_fr),
        cargo_func=lambda *args: checkyard_GN_insert(checkyard_GN_Filter, checkyard_GN_sc_fr)
    )
    checkyard_GN_Main_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
    checkyard_GN_sc_fr.pack(side=tk.LEFT, fill=tk.BOTH)
    checkyard_GN_sc_fr.refresh()
    checkyard_GN_sc_fr.top()
    checkyard_GN_insert(checkyard_GN_Filter, checkyard_GN_sc_fr)

def chk_vis(*args):
    global Check_Menu_Var
    Check_Menu_Var = 3
    for all in chk_center_frame.winfo_children(): all.pack_forget()
    checkyard_Vis_Filter.delete()
    checkyard_Vis_Filter.pack(side=tk.LEFT, fill=tk.Y)

    checkyard_Vis_Filter.checkyard_Vis(
        company_func=lambda *args: checkyard_vis_insert(checkyard_Vis_Filter, checkyard_Vis_sc_fr),
        time_on_yard_func=lambda *args: checkyard_vis_insert(checkyard_Vis_Filter, checkyard_Vis_sc_fr),
        age_func=lambda *args: checkyard_vis_insert(checkyard_Vis_Filter, checkyard_Vis_sc_fr),
        check_generate=lambda *args: checkyard_generate(checkyard_Vis_Filter),
        check_print=lambda *args: checkyard_print(),
        corp_func=lambda *args: checkyard_vis_insert(checkyard_Vis_Filter, checkyard_Vis_sc_fr),
        private_func=lambda *args: checkyard_vis_insert(checkyard_Vis_Filter, checkyard_Vis_sc_fr),
        expired_func=lambda *args: checkyard_vis_insert(checkyard_Vis_Filter, checkyard_Vis_sc_fr),
        no_parking_func=lambda *args: checkyard_vis_insert(checkyard_Vis_Filter, checkyard_Vis_sc_fr),
    )
    checkyard_vis_Main_Frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
    checkyard_Vis_sc_fr.pack(side=tk.LEFT, fill=tk.BOTH)
    checkyard_Vis_sc_fr.refresh()
    checkyard_Vis_sc_fr.top()
    checkyard_vis_insert(checkyard_Vis_Filter, checkyard_Vis_sc_fr)


# def chk_gen(*args):
#     global Check_Menu_Var
#     Check_Menu_Var = 4
#     for all in chk_center_frame.winfo_children(): all.pack_forget()
#     filter_gen_fr.pack(side=tk.LEFT, fill=tk.BOTH)

def chk_edit(*args):
    global Check_Menu_Var
    global Current_chk_Company_obj
    Check_Menu_Var = 5
    for all in chk_center_frame.winfo_children(): all.pack_forget()
    chk_edit_mainframe.pack(side=tk.LEFT, fill=tk.BOTH, expand=1, padx=5, pady=5)
    Current_chk_Company_obj = None
    for widget in second_chk_comp_Frame.winfo_children(): widget.destroy()
    Implement(second_chk_comp_Frame, "company", "CheckYard", 8, edit_scrn_size_lb)
# def chk_truck(*args):
#     pass

def chk_howeroff(*args):
    if Check_Menu_Var != 1: chk_tenant_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if Check_Menu_Var != 2: chk_GN_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if Check_Menu_Var != 3: chk_vis_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
  #  if Check_Menu_Var != 4: chk_gen_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if Check_Menu_Var != 5: chk_edit_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
def adm_howeroff(*args):
    if Admin_Menu_Var !=0: adm_company_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if Admin_Menu_Var != 1: adm_tenant_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if Admin_Menu_Var != 2: adm_vis_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if Admin_Menu_Var != 3: adm_GN_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if Admin_Menu_Var != 4: adm_acc_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    if Admin_Menu_Var !=5: adm_ven_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])


chk_top_frame = tk.Frame(Check_Main_Frame, highlightthickness=0, bg=conf["window_bg"])
chk_top_frame.pack(side=tk.TOP, fill=tk.X)


chk_tenant_button = tk.Label(chk_top_frame, text="Tenant", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
chk_tenant_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
chk_tenant_button.bind("<Button-1>", chk_tenant)
chk_tenant_button.bind("<Enter>", lambda x: chk_tenant_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
chk_tenant_button.bind("<Leave>", chk_howeroff)

chk_GN_button = tk.Label(chk_top_frame, text="GN", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
chk_GN_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
chk_GN_button.bind("<Button-1>", chk_GN)
chk_GN_button.bind("<Enter>", lambda x: chk_GN_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
chk_GN_button.bind("<Leave>", chk_howeroff)

chk_vis_button = tk.Label(chk_top_frame, text="Visitors", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
chk_vis_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
chk_vis_button.bind("<Button-1>", chk_vis)
chk_vis_button.bind("<Enter>", lambda x: chk_vis_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
chk_vis_button.bind("<Leave>", chk_howeroff)

chk_edit_button = tk.Label(chk_top_frame, text="Edit", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
chk_edit_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
chk_edit_button.bind("<Button-1>", chk_edit)
chk_edit_button.bind("<Enter>", lambda x: chk_edit_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
chk_edit_button.bind("<Leave>", chk_howeroff)



chk_center_frame = tk.Frame(Check_Main_Frame, highlightthickness=0, bg=conf["window_bg"])
chk_center_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)

#####################
#  Tenant Chekcyard #
#####################

# def chk_insert(frame):
#     C = chk_comp_box.get()
#     T = chk_var_truck.get()
#     t = chk_var_trailer.get()
#     S = chk_var_storage.get()
#     A = chk_var_age.get()
#     today = date.today()
#     chk_canv.yview_moveto(0)
#     for all in frame.winfo_children(): all.destroy()
#     if t:
#         chk_chkbut_stor.config(state=tk.NORMAL)
#     elif not chk_var_trailer.get():
#         chk_chkbut_stor.config(state=tk.DISABLED)
#     if A:
#         chk_entry1.configure(state=tk.NORMAL)
#         a = chk_entry1.get().strip()
#         if a != "":
#             try:
#                 int(a)
#             except:
#                 error(5)
#                 return
#             aging = int(a)
#         else: aging = 0
#     else:
#         chk_entry1.configure(state=tk.DISABLED)
#         aging = 0
#     if not T and not t: return
#     data = get_onyard()["tenant"]
#
#     if C == "All":
#         l = list()
#         for all_comp in data[0]: l.append(all_comp["company_ID"])
#         for all_comp in data[1]: l.append(all_comp["company_ID"])
#         allcompset = set(l)
#         for all_comp in sorted(allcompset):
#             c_frame = tk.Frame(frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#             c_lb = tk.Label(c_frame, text=all_comp, bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=chk_data_size)
#             c_lb.pack(side=tk.TOP, anchor=tk.NW, expand=0)
#             checkT = False
#             checkt = False
#             if T:
#                 truck_label = tk.Label(c_frame, text="Trucks:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"])
#                 truck_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#                 column_names_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#                 column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
#                 unit_lb = tk.Label(column_names_fr, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=20)
#                 unit_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 date_lb = tk.Label(column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"])
#                 date_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#                 sum_lb = tk.Label(column_names_fr, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=20)
#                 sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 checkT = False
#                 for all in data[0]:
#                     delta_days = (today - all["last_date"].date()).days
#                     if all["company_ID"] == all_comp and delta_days-aging>=0:
#                         checkT = True
#                         rec_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#                         rec_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
#                         T_lb = tk.Label(rec_fr, text=all["truck_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
#                         T_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                         T_time_lb = tk.Label(rec_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
#                         T_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#                         T_sum_lb = tk.Label(rec_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
#                         T_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 if checkT:
#                     truck_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
#                     truck_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
#
#             if t:
#                 trailer_label = tk.Label(c_frame, text="Trailers:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"])
#                 trailer_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#                 column_names_fr2 = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#                 column_names_fr2.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
#                 unitT_lb = tk.Label(column_names_fr2, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=20)
#                 unitT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 dateT_lb = tk.Label(column_names_fr2, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"])
#                 dateT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#                 sumT_lb = tk.Label(column_names_fr2, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=20)
#                 sumT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 checkt = False
#                 for all in data[1]:
#                     delta_days = (today - all["last_date"].date()).days
#                     if all["company_ID"] == all_comp and delta_days-aging>=0:
#                         if not S:
#                             if all["storage"]: continue
#                         checkt=True
#                         recT_fr = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#                         recT_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
#                         Tt_lb = tk.Label(recT_fr, text=all["trailer_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
#                         Tt_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                         Tt_time_lb = tk.Label(recT_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
#                         Tt_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#                         delta_days = (today - all["last_date"].date()).days
#                         Tt_sum_lb = tk.Label(recT_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
#                         Tt_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                         if all["storage"]:
#                             Tt_lb.config(fg=conf["storage_fg"])
#                             Tt_time_lb.config(fg=conf["storage_fg"])
#                             Tt_sum_lb.config(fg=conf["storage_fg"])
#                 if checkt:
#                     trailer_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
#                     trailer_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
#             if checkT or checkt:
#                 c_frame.pack(side=tk.TOP, fill=tk.X, expand=1, pady=(0, 2), padx=(2, 0))
#     else:
#         c1_frame = tk.Frame(frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#         c1_lb = tk.Label(c1_frame, text=C, bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=chk_data_size)
#         c1_lb.pack(side=tk.TOP, anchor=tk.NW, expand=0)
#         checkT = False
#         checkt = False
#         if T:
#             C_truck_label = tk.Label(c1_frame, text="Trucks:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"])
#             C_truck_frame = tk.Frame(c1_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#             C_column_names_fr = tk.Frame(C_truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#             C_column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
#             C_unit_lb = tk.Label(C_column_names_fr, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=20)
#             C_unit_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#             C_date_lb = tk.Label(C_column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"])
#             C_date_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#             C_sum_lb = tk.Label(C_column_names_fr, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=20)
#             C_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#             checkT = False
#             for all in data[0]:
#                 if all["company_ID"] != C: continue
#                 delta_days = (today - all["last_date"].date()).days
#                 if delta_days - aging >= 0:
#                     checkT = True
#                     C_rec_fr = tk.Frame(C_truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#                     C_rec_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
#                     C_T_lb = tk.Label(C_rec_fr, text=all["truck_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
#                     C_T_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                     C_T_time_lb = tk.Label(C_rec_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
#                     C_T_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#                     C_T_sum_lb = tk.Label(C_rec_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
#                     C_T_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#             if checkT:
#                 C_truck_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
#                 C_truck_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
#         if t:
#             C_trailer_label = tk.Label(c1_frame, text="Trailers:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"])
#             C_trailer_frame = tk.Frame(c1_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#             C_column_names_fr2 = tk.Frame(C_trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#             C_column_names_fr2.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
#             C_unitT_lb = tk.Label(C_column_names_fr2, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=20)
#             C_unitT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#             C_dateT_lb = tk.Label(C_column_names_fr2, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"])
#             C_dateT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#             C_sumT_lb = tk.Label(C_column_names_fr2, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=20)
#             C_sumT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#             checkt = False
#             for all in data[1]:
#                 if all["company_ID"] != C: continue
#                 delta_days = (today - all["last_date"].date()).days
#                 if delta_days - aging >= 0:
#                     if not S:
#                         if all["storage"]: continue
#                     checkt = True
#                     C_recT_fr = tk.Frame(C_trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#                     C_recT_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
#                     C_Tt_lb = tk.Label(C_recT_fr, text=all["trailer_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
#                     C_Tt_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                     C_Tt_time_lb = tk.Label(C_recT_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
#                     C_Tt_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#                     delta_days = (today - all["last_date"].date()).days
#                     C_Tt_sum_lb = tk.Label(C_recT_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
#                     C_Tt_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                     if all["storage"]:
#                         C_Tt_lb.config(fg=conf["storage_fg"])
#                         C_Tt_time_lb.config(fg=conf["storage_fg"])
#                         C_Tt_sum_lb.config(fg=conf["storage_fg"])
#             if checkt:
#                 C_trailer_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
#                 C_trailer_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
#         if checkT or checkt:
#             c1_frame.pack(side=tk.TOP, fill=tk.X, expand=1, pady=(0, 2), padx=(2, 0))
#     chk_canv.update_idletasks()
#     check_scroll_region()
#
# def chk_GN_insert(frame):
#     T = chk_var_GN_truck.get()
#     t = chk_var_GN_trailer.get()
#     S = chk_var_GN_storage.get()
#     f = chk_var_GN_fb.get()
#     L = chk_var_GN_loaded.get()
#     U = chk_var_GN_unloaded.get()
#     A = chk_var_GN_age.get()
#     today = date.today()
#     chk_GN_canv.yview_moveto(0)
#     for all in frame.winfo_children(): all.destroy()
#     if not T and not t and not f: return
#     elif not L and not U: return
#     if t:
#         chk_chkbut_GN_stor.config(state=tk.NORMAL)
#     elif not chk_var_GN_trailer.get():
#         chk_chkbut_GN_stor.config(state=tk.DISABLED)
#     if A:
#         chk_entry2.configure(state=tk.NORMAL)
#         a = chk_entry2.get().strip()
#         if a != "":
#             try:
#                 int(a)
#             except:
#                 error(5)
#                 return
#             aging = int(a)
#         else:
#             aging = 0
#     else:
#         chk_entry2.configure(state=tk.DISABLED)
#         aging = 0
#     data = get_onyard()["GN"]
#
#
#     c_frame = tk.Frame(frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#     c_frame.pack(side=tk.TOP, fill=tk.X, expand=1, pady=(0, 2), padx=(2, 0))
#     if T:
#         truck_label = tk.Label(c_frame, text="Trucks:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=chk_data_size)
#         truck_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
#         truck_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#         truck_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
#         column_names_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#         column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
#         unit_lb = tk.Label(column_names_fr, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=15)
#         unit_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#         cargo_lb = tk.Label(column_names_fr, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=15)
#         cargo_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#         date_lb = tk.Label(column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"])
#         date_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#         sum_lb = tk.Label(column_names_fr, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=20)
#         sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#         for all in data[0]:
#             delta_days = (today - all["last_date"].date()).days
#             if delta_days - aging >= 0:
#                 rec_fr = tk.Frame(truck_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#                 rec_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
#                 T_lb = tk.Label(rec_fr, text=all["truck_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
#                 T_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 C_lb = tk.Label(rec_fr, text="", bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
#                 C_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 T_time_lb = tk.Label(rec_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
#                 T_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#                 T_sum_lb = tk.Label(rec_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
#                 T_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#
#
#     if t:
#         trailer_label = tk.Label(c_frame, text="Trailers:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=chk_data_size)
#         trailer_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
#         trailer_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#         trailer_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
#         column_names_fr2 = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#         column_names_fr2.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
#         unitT_lb = tk.Label(column_names_fr2, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=15)
#         unitT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#         cargo_t_lb = tk.Label(column_names_fr2, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=15)
#         cargo_t_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#         dateT_lb = tk.Label(column_names_fr2, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"])
#         dateT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#         sumT_lb = tk.Label(column_names_fr2, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=20)
#         sumT_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#         for all in data[1]:
#             delta_days = (today - all["last_date"].date()).days
#             if delta_days - aging >= 0 and "trailer_number" in all:
#                 if not S:
#                     if all["storage"]: continue
#                 if not L:
#                     if all["LU"]: continue
#                 if not U:
#                     if not all["LU"]: continue
#                 recT_fr = tk.Frame(trailer_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#                 recT_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
#                 Tt_lb = tk.Label(recT_fr, text=all["trailer_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
#                 Tt_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 if all["LU"]: cargo="LOADED"
#                 else: cargo="EMPTY"
#                 tC_lb = tk.Label(recT_fr, text=cargo, bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
#                 tC_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 if all["LU"]: tC_lb.config(fg=conf["func_button_fg"])
#                 else: tC_lb.config(fg=conf["func_button_sel_fg"])
#                 Tt_time_lb = tk.Label(recT_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
#                 Tt_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#                 delta_days = (today - all["last_date"].date()).days
#                 Tt_sum_lb = tk.Label(recT_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
#                 Tt_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 if all["storage"]:
#                     Tt_lb.config(fg=conf["storage_fg"])
#                     Tt_time_lb.config(fg=conf["storage_fg"])
#                     Tt_sum_lb.config(fg=conf["storage_fg"])
#     if f:
#         fb_label = tk.Label(c_frame, text="Flatbeds:", bg=conf["widget_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["submenu_fg"], width=chk_data_size)
#         fb_label.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
#         fb_frame = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#         fb_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
#         column_names_fr3 = tk.Frame(fb_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#         column_names_fr3.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
#         unitf_lb = tk.Label(column_names_fr3, text="unit number:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=15)
#         unitf_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#         cargo_f_lb = tk.Label(column_names_fr3, text="cargo:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=15)
#         cargo_f_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#         datef_lb = tk.Label(column_names_fr3, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"])
#         datef_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#         sumf_lb = tk.Label(column_names_fr3, text="on yard / days", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["widget_sel_fg"], width=20)
#         sumf_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#         for all in data[1]:
#             delta_days = (today - all["last_date"].date()).days
#             if delta_days - aging >= 0 and "fb_number" in all:
#                 if not S:
#                     if all["storage"]: continue
#                 if not L:
#                     if all["LU"]: continue
#                 if not U:
#                     if not all["LU"]: continue
#                 recf_fr = tk.Frame(c_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
#                 recf_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
#                 f_lb = tk.Label(recf_fr, text=all["fb_number"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
#                 f_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 if all["LU"]: cargo="LOADED"
#                 else: cargo="EMPTY"
#                 fC_lb = tk.Label(recf_fr, text=cargo, bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=15)
#                 fC_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#                 if all["LU"]: fC_lb.config(fg=conf["func_button_fg"])
#                 else: fC_lb.config(fg=conf["func_button_sel_fg"])
#                 f_time_lb = tk.Label(recf_fr, text=all["last_date"], bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
#                 f_time_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, fill=tk.X, expand=1, pady=(0, 1))
#                 delta_days = (today - all["last_date"].date()).days
#                 f_sum_lb = tk.Label(recf_fr, text=delta_days, bg=conf["window_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=20)
#                 f_sum_lb.pack(side=tk.LEFT, anchor=tk.NW, padx=1, pady=(0, 1))
#     chk_GN_canv.update_idletasks()
#     check_GN_scroll_region()


def chk_vis_insert(frame):
    global corp_id
    global pr_id
    global exp_id
    global age_id
    global nopark_id
    C = vis_chk_comp_box.get()
    c = chk_var_vis_corp.get()
    p = chk_var_vis_private.get()
    e = chk_var_vis_expired.get()
    NO = chk_var_vis_noparking.get()
    A = chk_var_vis_age.get()
    today = datetime.now()
    chk_vis_canv.yview_moveto(0)
    def insert_cars(masta, data, color):
        if data["private"] is None:
            parking=""
            expdate=""
        else:
            parking="YES"
            if data["private"]: expdate = data["expiration"]
            else: expdate = ""
        deltahours = int(delta_days.total_seconds()//3600)
        if deltahours > 24: deltatime = f"{deltahours//24} days {deltahours%24} hours"
        else: deltatime = f"{deltahours} hours"
        rec_fr = tk.Frame(masta, highlightthickness=0, bg=conf["widget_sel_bg"])
        rec_fr.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X)
        pl_lb = tk.Label(rec_fr, text=all["plates"], bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        pl_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        vend_lb = tk.Label(rec_fr, text=all["car_model"], bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        vend_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        dr_lb = tk.Label(rec_fr, text=all["driver_name"], bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        dr_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        car_time_lb = tk.Label(rec_fr, text=all["last_date"], bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color)
        car_time_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1), fill=tk.X, expand=1)
        park_lb = tk.Label(rec_fr, text=parking, bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        park_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        pr_lb = tk.Label(rec_fr, text=expdate, bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        pr_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        car_sum_lb = tk.Label(rec_fr, text=deltatime, bg=conf["widget_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=color, width=25)
        car_sum_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))

    for all in frame.winfo_children(): all.destroy()
    chk_var_vis_corp.trace_remove("write", corp_id)
    chk_var_vis_private.trace_remove("write", pr_id)
    chk_var_vis_expired.trace_remove("write", exp_id)
    chk_var_vis_age.trace_remove("write", age_id)
    chk_var_vis_noparking.trace_remove("write", nopark_id)
    if e:
        chk_chkbut_vis_private.select()
        p = True
        chk_chkbut_vis_private.config(state=tk.DISABLED)
        chk_chkbut_vis_corp.deselect()
        c = False
        chk_chkbut_vis_corp.config(state=tk.DISABLED)
        chk_chkbut_vis_noparking.config(state=tk.DISABLED)
    else:
        state = chk_chkbut_vis_noparking["state"]
        if state == "disabled":
            chk_chkbut_vis_noparking.config(state=tk.NORMAL)
            chk_chkbut_vis_private.config(state=tk.NORMAL)
            chk_chkbut_vis_private.deselect()
            p = False
            chk_chkbut_vis_corp.config(state=tk.NORMAL)
    if NO:
        chk_chkbut_vis_private.deselect()
        p = False
        chk_chkbut_vis_private.config(state=tk.DISABLED)
        chk_chkbut_vis_corp.deselect()
        c = False
        chk_chkbut_vis_corp.config(state=tk.DISABLED)
        chk_chkbut_vis_expired.configure(state=tk.DISABLED)
    else:
        state2 = chk_chkbut_vis_expired["state"]
        if state2 == "disabled":
            chk_chkbut_vis_private.config(state=tk.NORMAL)
            chk_chkbut_vis_corp.config(state=tk.NORMAL)
            chk_chkbut_vis_expired.configure(state=tk.NORMAL)
    if A:
        chk_vis_entry1.configure(state=tk.NORMAL)
        a = chk_vis_entry1.get().strip()
        if a !="":
            try:
                int(a)
            except:
                error(5)
                return
            aging = int(a)
        else: aging = 0
    else:
        chk_vis_entry1.configure(state=tk.DISABLED)
        aging = 0
    corp_id = chk_var_vis_corp.trace("w", lambda *args: chk_vis_insert(second_vis_data_frame))
    pr_id = chk_var_vis_private.trace("w", lambda *args: chk_vis_insert(second_vis_data_frame))
    exp_id = chk_var_vis_expired.trace("w", lambda *args: chk_vis_insert(second_vis_data_frame))
    age_id = chk_var_vis_age.trace("w", lambda *args: chk_vis_insert(second_vis_data_frame))
    nopark_id = chk_var_vis_noparking.trace("w", lambda *args: chk_vis_insert(second_vis_data_frame))
    # C = vis_chk_comp_box.get()
    # c = chk_var_vis_corp.get()
    # p = chk_var_vis_private.get()
    # e = chk_var_vis_expired.get()
    # NO = chk_var_vis_noparking.get()
    data = get_onyard()["visitor"]
    if C == "All":
        l = list()
        for all_comp in data: l.append(all_comp["company_ID"])
        allcompset = set(l)
        for all_comp in sorted(allcompset):
            c_vis_frame = tk.Frame(frame, highlightthickness=0, bg=conf["widget_sel_bg"])
            c_vis_lb = tk.Label(c_vis_frame, text=all_comp, bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=chk_data_size)
            c_vis_lb.pack(side=tk.TOP, anchor=tk.NW, expand=0)
            car_frame = tk.Frame(c_vis_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
            car_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
            column_names_fr = tk.Frame(car_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
            column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
            plate_lb = tk.Label(column_names_fr, text="plates:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            plate_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
            vendor_lb = tk.Label(column_names_fr, text="vendor:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            vendor_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
            driver_lb = tk.Label(column_names_fr, text="driver:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            driver_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
            date_lb = tk.Label(column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"])
            date_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1), fill=tk.X, expand=1)
            parking_lb = tk.Label(column_names_fr, text="on park:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            parking_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
            private_lb = tk.Label(column_names_fr, text="private:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            private_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
            sum_lb = tk.Label(column_names_fr, text="on yard / time", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
            sum_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))

            checkE = False

            for all in data:
                delta_days = today - all["last_date"]
                if all["company_ID"] == all_comp and int(delta_days.total_seconds()//3600) - aging >= 0:
                    if c and all["private"] is not None and not all["private"]:
                        checkE = True
                        insert_cars(car_frame, all, conf["submenu_fg"])
                    elif p and all["private"]:
                        if int((today.date()-all["expiration"]).days)>0: colorfg=conf["expired_date"]
                        else: colorfg=conf["on_parking"]
                        if not e:
                            checkE = True
                            insert_cars(car_frame, all, colorfg)
                        else:
                            if int((today.date()-all["expiration"]).days)>0:
                                checkE = True
                                insert_cars(car_frame, all, colorfg)
                    elif NO and all["private"] is None:
                        checkE = True
                        insert_cars(car_frame, all, conf["in_button_sel_fg"])
            if checkE: c_vis_frame.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)
    else:
        if not any(d.get("company_ID") == C for d in data): return
        c_vis_frame = tk.Frame(frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        c_vis_lb = tk.Label(c_vis_frame, text=C, bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=chk_data_size)
        c_vis_lb.pack(side=tk.TOP, anchor=tk.NW, expand=0)
        car_frame = tk.Frame(c_vis_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        car_frame.pack(side=tk.TOP, fill=tk.X, padx=3, pady=(0, 3))
        column_names_fr = tk.Frame(car_frame, highlightthickness=0, bg=conf["widget_sel_bg"])
        column_names_fr.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))
        plate_lb = tk.Label(column_names_fr, text="plates:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        plate_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        vendor_lb = tk.Label(column_names_fr, text="vendor:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        vendor_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        driver_lb = tk.Label(column_names_fr, text="driver:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        driver_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        date_lb = tk.Label(column_names_fr, text="on yard since:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"])
        date_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1), fill=tk.X, expand=1)
        parking_lb = tk.Label(column_names_fr, text="parking:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        parking_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        private_lb = tk.Label(column_names_fr, text="private:", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        private_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))
        sum_lb = tk.Label(column_names_fr, text="on yard / time", bg=conf["widget_sel_bg"], font=(conf["header_font"], conf["notebook_tab_size"]), fg=conf["widget_sel_fg"], width=25)
        sum_lb.pack(side=tk.LEFT, padx=1, pady=(0, 1))

        checkE = False

        for all in data:
            delta_days = today - all["last_date"]
            if all["company_ID"] == C and int(delta_days.total_seconds() // 3600) - aging >= 0:
                if c and all["private"] is not None and not all["private"]:
                    checkE = True
                    insert_cars(car_frame, all, conf["submenu_fg"])
                elif p and all["private"]:
                    if int((today.date() - all["expiration"]).days) > 0:
                        colorfg = conf["expired_date"]
                    else:
                        colorfg = conf["on_parking"]
                    if not e:
                        checkE = True
                        insert_cars(car_frame, all, colorfg)
                    else:
                        if int((today.date() - all["expiration"]).days) > 0:
                            checkE = True
                            insert_cars(car_frame, all, colorfg)
                elif NO and all["private"] is None:
                    checkE = True
                    insert_cars(car_frame, all, conf["in_button_sel_fg"])
        if checkE: c_vis_frame.pack(side=tk.TOP, fill=tk.X, expand=1, anchor=tk.NW, padx=3, pady=3)

    chk_vis_canv.update_idletasks()
    check_vis_scroll_region()



# Check yard Tenant window
checkyard_Ten_Main_Frame = tk.Frame(chk_center_frame, highlightthickness=0, relief=tk.RAISED, bg=conf["window_bg"], width=conf["chk_filter_frame"])
checkyard_Ten_Main_Frame.pack_propagate(False)
checkyard_Ten_Filter = filter_frame(checkyard_Ten_Main_Frame)
checkyard_Ten_sc_fr = scroller(checkyard_Ten_Main_Frame)


#####################
#        GN         #
#####################

# Check yard GN window
checkyard_GN_Main_Frame = tk.Frame(chk_center_frame, highlightthickness=0, relief=tk.RAISED, bg=conf["window_bg"], width=conf["chk_filter_frame"])
checkyard_GN_Main_Frame.pack_propagate(False)
checkyard_GN_Filter = filter_frame(checkyard_GN_Main_Frame)
checkyard_GN_sc_fr = scroller(checkyard_GN_Main_Frame)


#####################
#        VIS        #
#####################
checkyard_vis_Main_Frame = tk.Frame(chk_center_frame,  highlightthickness=0, relief=tk.RAISED, bg=conf["window_bg"], width=conf["chk_filter_frame"])
checkyard_vis_Main_Frame.pack_propagate(False)
checkyard_Vis_Filter = filter_frame(checkyard_vis_Main_Frame)
checkyard_Vis_sc_fr = scroller(checkyard_vis_Main_Frame)

#
# #
# chk_vis_lb = tk.Label(filter_vis_fr, text="Companies:", foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["header_size"]), bg=conf["submenu_bg"])
# chk_vis_lb.pack(side=tk.TOP, padx=5, pady=(10, 5), anchor=tk.W)
# #
# vis_chk_comp_box = ttk.Combobox(filter_vis_fr, values=comp_list, width=10, background=conf["submenu_sel_bg"], foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), state="readonly")
# vis_chk_comp_box.pack(fill=tk.X, side=tk.TOP, padx=5)
# vis_chk_comp_box.current(0)
# vis_chk_comp_box.bind("<<ComboboxSelected>>", lambda *args: chk_vis_insert(second_vis_data_frame))
# chk_var_vis_corp = tk.BooleanVar()
# chk_chkbut_vis_corp = tk.Checkbutton(filter_vis_fr, text="Corporate", variable=chk_var_vis_corp, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], onvalue=True, offvalue=False)
# chk_chkbut_vis_corp.pack(side=tk.TOP, padx=5, pady=(10, 0), anchor=tk.W)
# chk_chkbut_vis_corp.select()
# chk_var_vis_private = tk.BooleanVar()
# chk_chkbut_vis_private = tk.Checkbutton(filter_vis_fr, text="Private", variable=chk_var_vis_private, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], onvalue=True, offvalue=False)
# chk_chkbut_vis_private.pack(side=tk.TOP, padx=5, pady=(10, 0), anchor=tk.W)
# chk_chkbut_vis_private.select()
# chk_var_vis_expired = tk.BooleanVar()
# chk_chkbut_vis_expired = tk.Checkbutton(filter_vis_fr, text="Expired Only", variable=chk_var_vis_expired, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], state=tk.NORMAL, onvalue=True, offvalue=False)
# chk_chkbut_vis_expired.pack(side=tk.TOP, padx=5, pady=(10, 0), anchor=tk.W)
# chk_var_vis_noparking = tk.BooleanVar()
# chk_chkbut_vis_noparking = tk.Checkbutton(filter_vis_fr, text="No parking", variable=chk_var_vis_noparking, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], onvalue=True, offvalue=False)
# chk_chkbut_vis_noparking.pack(side=tk.TOP, padx=5, pady=(10, 0), anchor=tk.W)
#
# chk_v_age_lb = tk.Label(filter_vis_fr, text="Time on Yard (hours):", foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), bg=conf["submenu_bg"])
# chk_v_age_lb.pack(side=tk.TOP, padx=5, pady=(10, 5), anchor=tk.W)
#
# chk_vis_age_fr = tk.Frame(filter_vis_fr, highlightthickness=0, bg=conf["submenu_bg"])
# chk_vis_age_fr.pack(side=tk.TOP, anchor=tk.W)
# chk_var_vis_age = tk.BooleanVar()
# chk_vis_age_chkbox = tk.Checkbutton(chk_vis_age_fr, foreground=conf["submenu_fg"], bg=conf["submenu_bg"], onvalue=True, offvalue=False, variable=chk_var_vis_age)
# chk_vis_age_chkbox.pack(fill=tk.X, side=tk.LEFT)
# chk_vis_aging = None
# chk_vis_entry1 = tk.Entry(chk_vis_age_fr, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["entry_fg"], width=15, state=tk.DISABLED)
# chk_vis_entry1.pack(side=tk.LEFT, fill=tk.BOTH)
# chk_vis_button_get = tk.Button(chk_vis_age_fr, text=u"\u23F5", bg=conf["submenu_sel_bg"], relief=tk.RAISED, command=lambda *args: chk_vis_insert(second_vis_data_frame))
# chk_vis_button_get.pack(side=tk.RIGHT, padx=(5, 0))
#
# vis_gen_frame = tk.Frame(filter_vis_fr, highlightthickness=0, bg=conf["submenu_bg"], width=50)
# vis_gen_frame.pack(side=tk.BOTTOM, anchor=tk.W)
# vis_chk_marker_lb = tk.Label(vis_gen_frame, text="Check Yard Marker:", foreground=conf["submenu_fg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), bg=conf["submenu_bg"])
# vis_chk_marker_lb.pack(side=tk.TOP, padx=5, pady=(10, 5), anchor=tk.W)
# vis_chk_marker = tk.Label(vis_gen_frame, foreground=conf["status_fg"], text=sets["chk_datetime"], font=(conf["submenu_font"], conf["notebook_tab_size"]), bg=conf["submenu_bg"])
# vis_chk_marker.pack(side=tk.TOP, padx=5, pady=(10, 5), anchor=tk.CENTER)
# vis_gen_button = tk.Button(vis_gen_frame, text="GENERATE", height=1, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["font_size"]), relief=tk.RAISED, command = lambda: check_generate(3))
# vis_gen_button.pack(side=tk.TOP, padx=(5, 0), fill=tk.X, expand=1)
# vis_print_button = tk.Button(vis_gen_frame, text="PRINT", height=1, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["font_size"]), relief=tk.RAISED, command = lambda: check_print(3))
# vis_print_button.pack(side=tk.TOP, padx=(5, 0), fill=tk.X, expand=1)
# #
# data_vis_fr = tk.Frame(chk_center_frame, highlightthickness=0, bg=conf["window_bg"])
# data_vis_fr.pack_propagate(False)
# #
# sub_chk_vis_frame = tk.Frame(data_vis_fr, highlightthickness=0)
# sub_chk_vis_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
# #
# # # Scrolable frame for Data
# chk_vis_canv = tk.Canvas(sub_chk_vis_frame, bg=conf["window_bg"], highlightthickness=0)
# second_vis_data_frame = tk.Frame(chk_vis_canv, bg=conf["window_bg"])
# chk_vis_scrl = ttk.Scrollbar(sub_chk_vis_frame, orient=tk.VERTICAL, command=chk_vis_canv.yview)
# chk_vis_canv.config(yscrollcommand=chk_vis_scrl.set)
# chk_vis_scrl.pack(fill=tk.Y, side=tk.RIGHT)
# chk_vis_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
# chk_vis_canv.create_window((0, 0), window=second_vis_data_frame, anchor=tk.NW)
# second_vis_data_frame.bind("<Configure>", lambda event, canvas=chk_vis_canv: chk_vis_canv.configure(scrollregion=chk_vis_canv.bbox("all")))
#
# def check_vis_scroll_region(*event):
#     if second_vis_data_frame.winfo_height() <= chk_vis_canv.winfo_height():
#         chk_vis_scrl.pack_forget()
#         chk_vis_canv.configure(yscrollcommand=None)
#         second_vis_data_frame.unbind("<Enter>")
#         chk_vis_canv.unbind_all('<MouseWheel>')
#     else:
#         chk_vis_scrl.pack(side=tk.RIGHT, fill=tk.Y)
#         chk_vis_canv.configure(yscrollcommand=chk_vis_scrl.set)
#         second_vis_data_frame.bind("<Enter>", chk_vis_enter_mousewheel_tenant_comp, add="+")
# chk_vis_canv.bind("<Configure>", check_vis_scroll_region)
#
# def chk_vis_on_mousewheel(event): chk_vis_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
# def chk_vis_enter_mousewheel_tenant_comp(event): chk_vis_canv.bind_all('<MouseWheel>', chk_vis_on_mousewheel, add="+")
# def chk_vis_leave_mousewheel_tenant_comp(event): chk_vis_canv.unbind_all('<MouseWheel>')
# second_vis_data_frame.bind("<Enter>", chk_vis_enter_mousewheel_tenant_comp, add="+")
# second_vis_data_frame.bind("<Leave>", chk_vis_leave_mousewheel_tenant_comp)
#
# corp_id = chk_var_vis_corp.trace("w", lambda *args: chk_vis_insert(second_vis_data_frame))
# pr_id = chk_var_vis_private.trace("w", lambda *args: chk_vis_insert(second_vis_data_frame))
# exp_id = chk_var_vis_expired.trace("w", lambda *args: chk_vis_insert(second_vis_data_frame))
# age_id = chk_var_vis_age.trace("w", lambda *args: chk_vis_insert(second_vis_data_frame))
# nopark_id = chk_var_vis_noparking.trace("w", lambda *args: chk_vis_insert(second_vis_data_frame))

#####################
#       EDIT        #
#####################

#Main Frame in Edit => packs with Edit button
chk_edit_mainframe = tk.Frame(chk_center_frame, highlightthickness=0, bg=conf["window_bg"])

#LEFT frame for Company list and GN
chk_listframe = tk.Frame(chk_edit_mainframe, bg=conf["window_bg"], highlightthickness=0, width=conf["p_t_company_w"])
chk_listframe.pack_propagate(0)
chk_listframe.pack(fill=tk.BOTH, side=tk.LEFT)



#Company list Frame
chk_Company_Frame = tk.Frame(chk_listframe, bg=conf["window_bg"], highlightthickness=0)
# chk_Company_Frame.pack_propagate(0)
chk_Company_Frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1)

#LABEL FOR COMPANY
chk_Company_Lb = tk.Label(chk_Company_Frame, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
chk_Company_Lb.pack(fill=tk.X, side=tk.TOP)

#SCROLLABLE FRAME FOR COMPANIES WITH MOUSE CONTROL
chk_comp_frame = tk.Frame(chk_Company_Frame, highlightthickness=0)
chk_comp_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
chk_comp_canv = tk.Canvas(chk_comp_frame, highlightthickness=0, bg=conf["window_bg"])
second_chk_comp_Frame = tk.Frame(chk_comp_canv, bg=conf["window_bg"])
chk_comp_scrl = ttk.Scrollbar(chk_Company_Frame, orient=tk.VERTICAL, command=chk_comp_canv.yview)
chk_comp_canv.config(yscrollcommand=chk_comp_scrl.set)
chk_comp_scrl.pack(fill=tk.Y, side=tk.RIGHT)
chk_comp_canv.pack(fill=tk.BOTH, side=tk.TOP, expand=1)
chk_comp_canv.create_window((0, 0), window=second_chk_comp_Frame, anchor=tk.NW)
second_chk_comp_Frame.bind("<Configure>", lambda event, canvas=chk_comp_canv: chk_comp_canv.configure(scrollregion=chk_comp_canv.bbox("all")))

def check_chk_comp_scroll_region(*event):
    if second_chk_comp_Frame.winfo_height() <= chk_comp_canv.winfo_height():
        chk_comp_scrl.pack_forget()
        chk_canv.configure(yscrollcommand=None)
        second_chk_comp_Frame.unbind("<Enter>")
        second_chk_comp_Frame.unbind_all("<MouseWheel>")
    else:
        chk_comp_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        chk_comp_canv.configure(yscrollcommand=chk_comp_scrl.set)
        second_chk_comp_Frame.bind("<Enter>", _enter_mousewheel_chk_comp, add="+")
chk_comp_canv.bind("<Configure>", check_chk_comp_scroll_region)
#MOUSEWHEEL FUNCTION
def _on_chk_comp_mousewheel(event): chk_comp_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def _enter_mousewheel_chk_comp(event): chk_comp_canv.bind_all('<MouseWheel>', _on_chk_comp_mousewheel, add="+")
def _leave_mousewheel_chk_comp(event): chk_comp_canv.unbind_all('<MouseWheel>')
second_chk_comp_Frame.bind("<Enter>", _enter_mousewheel_chk_comp, add="+")
second_chk_comp_Frame.bind("<Leave>", _leave_mousewheel_chk_comp)

#GN label
chk_GN_Lb = tk.Label(chk_listframe, text="GNT:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
chk_GN_Lb.pack(fill=tk.X, side=tk.TOP, pady=(5, 0))
#GNT Button
_unit_(chk_listframe, "G.N.Transport", 1, 1, "CheckYard", 8)

#Central Frame for Units
chk_central_frame = tk.Frame(chk_edit_mainframe, bg=conf["window_bg"], highlightthickness=0)
chk_central_frame.pack_propagate(0)
chk_central_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, padx=5)



#creating frames for Tenants
chk_central_tenant_frame = tk.Frame(chk_central_frame, bg=conf["window_bg"], highlightthickness=0)

#Truck Label
chk_T_T_Lb = tk.Label(chk_central_tenant_frame, text="TRUCK:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
chk_T_T_Lb.pack(fill=tk.X, side=tk.TOP)
#Trucks Frame
chk_central_tenant_truck_frame = tk.Frame(chk_central_tenant_frame, bg=conf["window_bg"], highlightthickness=0, height=350)
chk_central_tenant_truck_frame.propagate(0)
chk_central_tenant_truck_frame.pack(fill=tk.BOTH, side=tk.TOP)

#scrollable frame for Trucks
chk_T_T_canv = tk.Canvas(chk_central_tenant_truck_frame, bg=conf["window_bg"], highlightthickness=0)
second_chk_T_T_frame = tk.Frame(chk_T_T_canv, bg=conf["window_bg"])
chk_T_T_scrl = ttk.Scrollbar(chk_central_tenant_truck_frame, orient=tk.VERTICAL, command=chk_T_T_canv.yview)
chk_T_T_canv.config(yscrollcommand=chk_T_T_scrl.set)
chk_T_T_scrl.pack(fill=tk.Y, side=tk.RIGHT)
chk_T_T_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
chk_T_T_canv.create_window((0, 0), window=second_chk_T_T_frame, anchor=tk.NW)
second_chk_T_T_frame.bind("<Configure>", lambda event, canvas=chk_T_T_canv: chk_T_T_canv.configure(scrollregion=chk_T_T_canv.bbox("all")))
def check_chk_T_T_scroll_region(*event):
    if second_chk_T_T_frame.winfo_height() <= chk_T_T_canv.winfo_height():
        chk_T_T_scrl.pack_forget()
        chk_T_T_canv.configure(yscrollcommand=None)
        second_chk_T_T_frame.unbind("<Enter>")
        second_chk_T_T_frame.unbind_all("<MouseWheel>")
    else:
        chk_T_T_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        chk_T_T_canv.configure(yscrollcommand=chk_T_T_scrl.set)
        second_chk_T_T_frame.bind("<Enter>", chk_enter_mousewheel_truck, add="+")
chk_T_T_canv.bind("<Configure>", check_chk_T_T_scroll_region)
def chk_on_mousewheel_truck(event): chk_T_T_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def chk_enter_mousewheel_truck(event): chk_T_T_canv.bind_all('<MouseWheel>', chk_on_mousewheel_truck, add="+")
def chk_leave_mousewheel_truck(event): chk_T_T_canv.unbind_all('<MouseWheel>')
second_chk_T_T_frame.bind("<Enter>", chk_enter_mousewheel_truck, add="+")
second_chk_T_T_frame.bind("<Leave>", chk_leave_mousewheel_truck)

#Trailer Label
chk_T_Tr_Lb = tk.Label(chk_central_tenant_frame, text="TRAILER:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
chk_T_Tr_Lb.pack(fill=tk.X, side=tk.TOP)
#Trailer Frame
chk_central_tenant_trailer_frame = tk.Frame(chk_central_tenant_frame, bg=conf["window_bg"], highlightthickness=0)
chk_central_tenant_trailer_frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1)

#scrollable frame for Trailers
chk_T_Tr_canv = tk.Canvas(chk_central_tenant_trailer_frame, bg=conf["window_bg"], highlightthickness=0)
second_chk_T_Tr_frame = tk.Frame(chk_T_Tr_canv, bg=conf["window_bg"])
chk_T_Tr_scrl = ttk.Scrollbar(chk_central_tenant_trailer_frame, orient=tk.VERTICAL, command=chk_T_Tr_canv.yview)
chk_T_Tr_canv.config(yscrollcommand=chk_T_Tr_scrl.set)
chk_T_Tr_scrl.pack(fill=tk.Y, side=tk.RIGHT)
chk_T_Tr_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
chk_T_Tr_canv.create_window((0, 0), window=second_chk_T_Tr_frame, anchor=tk.NW)
second_chk_T_Tr_frame.bind("<Configure>", lambda event, canvas=chk_T_Tr_canv: chk_T_Tr_canv.configure(scrollregion=chk_T_Tr_canv.bbox("all")))
def check_chk_T_Tr_scroll_region(*event):
    if second_chk_T_Tr_frame.winfo_height() <= chk_T_Tr_canv.winfo_height():
        chk_T_Tr_scrl.pack_forget()
        chk_T_Tr_canv.configure(yscrollcommand=None)
        second_chk_T_Tr_frame.unbind("<Enter>")
        second_chk_T_Tr_frame.unbind_all("<MouseWheel>")
    else:
        chk_T_Tr_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        chk_T_Tr_canv.configure(yscrollcommand=chk_T_Tr_scrl.set)
        second_chk_T_Tr_frame.bind("<Enter>", chk_enter_mousewheel_trailer, add="+")
chk_T_Tr_canv.bind("<Configure>", check_chk_T_Tr_scroll_region)
def chk_on_mousewheel_trailer(event): chk_T_Tr_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def chk_enter_mousewheel_trailer(event): chk_T_Tr_canv.bind_all('<MouseWheel>', chk_on_mousewheel_trailer, add="+")
def chk_leave_mousewheel_trailer(event): chk_T_Tr_canv.unbind_all('<MouseWheel>')
second_chk_T_Tr_frame.bind("<Enter>", chk_enter_mousewheel_trailer, add="+")
second_chk_T_Tr_frame.bind("<Leave>", chk_leave_mousewheel_trailer)

#creating frames for GN
chk_central_GN_frame = tk.Frame(chk_central_frame, bg=conf["window_bg"], highlightthickness=0)


#GN Truck Label
chk_GN_T_Lb = tk.Label(chk_central_GN_frame, text="TRUCK:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
chk_GN_T_Lb.pack(fill=tk.X, side=tk.TOP)
#GN Truck Frame
chk_central_GN_truck_frame = tk.Frame(chk_central_GN_frame, bg=conf["window_bg"], highlightthickness=0, height=200)
chk_central_GN_truck_frame.propagate(0)
chk_central_GN_truck_frame.pack(fill=tk.BOTH, side=tk.TOP, expand=0)

#scrollable frame for GN Trucks
chk_GN_T_canv = tk.Canvas(chk_central_GN_truck_frame, bg=conf["window_bg"], highlightthickness=0)
second_chk_GN_T_frame = tk.Frame(chk_GN_T_canv, bg=conf["window_bg"])
chk_GN_T_scrl = ttk.Scrollbar(chk_central_GN_truck_frame, orient=tk.VERTICAL, command=chk_GN_T_canv.yview)
chk_GN_T_canv.config(yscrollcommand=chk_GN_T_scrl.set)
chk_GN_T_scrl.pack(fill=tk.Y, side=tk.RIGHT)
chk_GN_T_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
chk_GN_T_canv.create_window((0, 0), window=second_chk_GN_T_frame, anchor=tk.NW)
second_chk_GN_T_frame.bind("<Configure>", lambda event, canvas=chk_GN_T_canv: chk_GN_T_canv.configure(scrollregion=chk_GN_T_canv.bbox("all")))
def check_chk_GN_T_scroll_region(*event):
    if second_chk_GN_T_frame.winfo_height() <= chk_GN_T_canv.winfo_height():
        chk_GN_T_scrl.pack_forget()
        chk_GN_T_canv.configure(yscrollcommand=None)
        second_chk_GN_T_frame.unbind("<Enter>")
        second_chk_GN_T_frame.unbind_all("<MouseWheel>")
    else:
        chk_GN_T_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        chk_GN_T_canv.configure(yscrollcommand=chk_GN_T_scrl.set)
        second_chk_GN_T_frame.bind("<Enter>", chk_enter_mousewheel_GN_T, add="+")
chk_GN_T_canv.bind("<Configure>", check_chk_GN_T_scroll_region)
def chk_on_mousewheel_GN_T(event): chk_GN_T_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def chk_enter_mousewheel_GN_T(event): chk_GN_T_canv.bind_all('<MouseWheel>', chk_on_mousewheel_GN_T, add="+")
def chk_leave_mousewheel_GN_T(event): chk_GN_T_canv.unbind_all('<MouseWheel>')
second_chk_GN_T_frame.bind("<Enter>", chk_enter_mousewheel_GN_T, add="+")
second_chk_GN_T_frame.bind("<Leave>", chk_leave_mousewheel_GN_T)


#GN Trailer Label
chk_GN_Tr_Lb = tk.Label(chk_central_GN_frame, text="TRAILER:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
chk_GN_Tr_Lb.pack(fill=tk.X, side=tk.TOP)
#GN Trailer Frame
chk_central_GN_trailer_frame = tk.Frame(chk_central_GN_frame, bg=conf["entry_bg"], highlightthickness=0)
chk_central_GN_trailer_frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1)

#scrollable frame for GN Trailers
chk_GN_Tr_canv = tk.Canvas(chk_central_GN_trailer_frame, bg=conf["window_bg"], highlightthickness=0)
second_chk_GN_Tr_frame = tk.Frame(chk_GN_Tr_canv, bg=conf["window_bg"])
chk_GN_Tr_scrl = ttk.Scrollbar(chk_central_GN_trailer_frame, orient=tk.VERTICAL, command=chk_GN_Tr_canv.yview)
chk_GN_Tr_canv.config(yscrollcommand=chk_GN_Tr_scrl.set)
chk_GN_Tr_scrl.pack(fill=tk.Y, side=tk.RIGHT)
chk_GN_Tr_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
chk_GN_Tr_canv.create_window((0, 0), window=second_chk_GN_Tr_frame, anchor=tk.NW)
second_chk_GN_Tr_frame.bind("<Configure>", lambda event, canvas=chk_GN_Tr_canv: chk_GN_Tr_canv.configure(scrollregion=chk_GN_Tr_canv.bbox("all")))
def check_chk_GN_Tr_scroll_region(*event):
    if second_chk_GN_Tr_frame.winfo_height() <= chk_GN_Tr_canv.winfo_height():
        chk_GN_Tr_scrl.pack_forget()
        chk_GN_Tr_canv.configure(yscrollcommand=None)
        second_chk_GN_Tr_frame.unbind("<Enter>")
        second_chk_GN_Tr_frame.unbind_all("<MouseWheel>")
    else:
        chk_GN_Tr_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        chk_GN_Tr_canv.configure(yscrollcommand=chk_GN_Tr_scrl.set)
        second_chk_GN_Tr_frame.bind("<Enter>", chk_enter_mousewheel_GN_Tr, add="+")
chk_GN_Tr_canv.bind("<Configure>", check_chk_GN_Tr_scroll_region)
def chk_on_mousewheel_GN_Tr(event): chk_GN_Tr_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def chk_enter_mousewheel_GN_Tr(event): chk_GN_Tr_canv.bind_all('<MouseWheel>', chk_on_mousewheel_GN_Tr, add="+")
def chk_leave_mousewheel_GN_Tr(event): chk_GN_Tr_canv.unbind_all('<MouseWheel>')
second_chk_GN_Tr_frame.bind("<Enter>", chk_enter_mousewheel_GN_Tr, add="+")
second_chk_GN_Tr_frame.bind("<Leave>", chk_leave_mousewheel_GN_Tr)

#GN FB Label
chk_GN_fb_Lb = tk.Label(chk_central_GN_frame, text="FLATBED:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
chk_GN_fb_Lb.pack(fill=tk.X, side=tk.TOP)
#GN FB Frame
chk_central_GN_fb_frame = tk.Frame(chk_central_GN_frame, bg=conf["window_bg"], highlightthickness=0, height=300)
chk_central_GN_fb_frame.propagate(0)
chk_central_GN_fb_frame.pack(fill=tk.BOTH, side=tk.TOP)

#scrollable frame for GN FB
chk_GN_fb_canv = tk.Canvas(chk_central_GN_fb_frame, bg=conf["window_bg"], highlightthickness=0)
second_chk_GN_fb_frame = tk.Frame(chk_GN_fb_canv, bg=conf["window_bg"])
chk_GN_fb_scrl = ttk.Scrollbar(chk_central_GN_fb_frame, orient=tk.VERTICAL, command=chk_GN_fb_canv.yview)
chk_GN_fb_canv.config(yscrollcommand=chk_GN_fb_scrl.set)
chk_GN_fb_scrl.pack(fill=tk.Y, side=tk.RIGHT)
chk_GN_fb_canv.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
chk_GN_fb_canv.create_window((0, 0), window=second_chk_GN_fb_frame, anchor=tk.NW)
second_chk_GN_fb_frame.bind("<Configure>", lambda event, canvas=chk_GN_fb_canv: chk_GN_fb_canv.configure(scrollregion=chk_GN_fb_canv.bbox("all")))
def check_chk_GN_fb_scroll_region(*event):
    if second_chk_GN_fb_frame.winfo_height() <= chk_GN_fb_canv.winfo_height():
        chk_GN_fb_scrl.pack_forget()
        chk_GN_fb_canv.configure(yscrollcommand=None)
        second_chk_GN_fb_frame.unbind("<Enter>")
        second_chk_GN_fb_frame.unbind_all("<MouseWheel>")
    else:
        chk_GN_fb_scrl.pack(side=tk.RIGHT, fill=tk.Y)
        chk_GN_fb_canv.configure(yscrollcommand=chk_GN_fb_scrl.set)
        second_chk_GN_fb_frame.bind("<Enter>", chk_enter_mousewheel_GN_fb, add="+")
chk_GN_fb_canv.bind("<Configure>", check_chk_GN_fb_scroll_region)
def chk_on_mousewheel_GN_fb(event): chk_GN_fb_canv.yview_scroll(int(-1 * (event.delta / 120)), "units")
def chk_enter_mousewheel_GN_fb(event): chk_GN_fb_canv.bind_all('<MouseWheel>', chk_on_mousewheel_GN_fb, add="+")
def chk_leave_mousewheel_GN_fb(event): chk_GN_fb_canv.unbind_all('<MouseWheel>')
second_chk_GN_fb_frame.bind("<Enter>", chk_enter_mousewheel_GN_fb, add="+")
second_chk_GN_fb_frame.bind("<Leave>", chk_leave_mousewheel_GN_fb)

#creating frames for Car Parking
chk_central_vis_frame = tk.Frame(chk_central_frame, bg=conf["window_bg"], highlightthickness=0)
#chk_central_vis_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)




#Right Frame for buttons
chk_right_frame = tk.Frame(chk_edit_mainframe, highlightthickness=3, relief=tk.RAISED, bg=conf["submenu_bg"], highlightbackground=conf["submenu_sel_bg"], width=conf["chk_filter_frame"])
chk_right_frame.pack_propagate(0)
chk_right_frame.pack(fill=tk.BOTH, side=tk.LEFT)

chk_manual_entry_frame = tk.Frame(chk_right_frame, highlightthickness=0, bg=conf["submenu_bg"])
chk_c_lb = tk.Label(chk_manual_entry_frame, text="Company:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
chk_c_lb.pack(fill=tk.X, side=tk.TOP)
chk_c_entry = tk.Entry(chk_manual_entry_frame, bg=conf["window_bg"], bd=1, cursor="shuttle", font=(conf["entry_font"], conf["notebook_tab_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, justify=tk.LEFT)
chk_c_entry.pack(fill=tk.X, side=tk.TOP)
chk_t_lb = tk.Label(chk_manual_entry_frame, text="UNIT:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
chk_t_lb.pack(fill=tk.X, side=tk.TOP)
chk_t_entry = tk.Entry(chk_manual_entry_frame, bg=conf["window_bg"], bd=1, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, justify=tk.LEFT)
chk_t_entry.pack(fill=tk.X, side=tk.TOP)
chk_t_entry.bind("<Return>", chk_add_manual)
chk_radio_var = tk.StringVar()
chk_truck_radio = tk.Radiobutton(chk_manual_entry_frame, text="Truck", variable=chk_radio_var, value="truck", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
chk_truck_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
chk_trailer_radio = tk.Radiobutton(chk_manual_entry_frame, text="Trailer", variable=chk_radio_var, value="trailer", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
chk_trailer_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
chk_radio_var.set("truck")
chk_manual_add_button = tk.Button(chk_manual_entry_frame, text="ADD", bg=conf["widget_bg"], fg=conf["in_button_fg"], bd=0, highlightthickness=0, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=chk_add_manual)
chk_manual_add_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
chk_manual_remove_button = tk.Button(chk_manual_entry_frame, text="REMOVE", bg=conf["widget_bg"], fg=conf["out_button_fg"], bd=0, highlightthickness=0, activebackground=conf["out_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["out_button_sel_fg"], command=chk_del_manual)
chk_manual_remove_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
chk_set_button = tk.Button(chk_right_frame, text="SET", bg=conf["widget_bg"], fg=conf["out_button_fg"], bd=0, highlightthickness=0, activebackground=conf["out_button_sel_bg"], activeforeground=conf["out_button_sel_fg"], font=(conf["submenu_font"], conf["header_size"], "bold"), command=chk_update)
chk_set_button.pack(fill=tk.X, side=tk.BOTTOM, padx=5, pady=5, ipady=10)



#unit_read = SQL_REQ("SELECT truck_number, status FROM dbo.GN_Trucks WHERE truck_number=\'928\'", (), "S_one")



edit_scrn_size = screen_x-int(conf["p_t_company_w"])-int(conf["chk_filter_frame"])
edit_scrn_size_lb = edit_scrn_size//150

######################
#     Admin Menu     #
######################
def adm_comp(*args):
    global Admin_Menu_Var
    global Admin_Company_Entries

    Admin_Menu_Var = 0
    adm_GN_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_acc_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_tenant_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_vis_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_ven_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    Admin_Company_Frame.pack(fill=tk.BOTH, expand=1, padx=5, pady=5)
    Admin_Tenant_Frame.pack_forget()
    Admin_GN_Frame.pack_forget()
    Admin_Account_Frame.pack_forget()
    Admin_Visitor_Frame.pack_forget()
    Admin_Ven_Frame.pack_forget()
    Admin_Company_Entries = admin_company_insert(Admin_Company_Scroll.frame)
    Admin_Company_Scroll.refresh()


def adm_ten(*args):
    global Admin_Menu_Var
    global Adm_Company_obj
    global Adm_Company_Var
    Admin_Menu_Var = 1
    adm_GN_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_acc_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_company_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_vis_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_ven_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    Admin_Tenant_Frame.pack(fill=tk.BOTH, expand=1, padx=(5, 0), pady=5)
    Admin_Company_Frame.pack_forget()
    Admin_GN_Frame.pack_forget()
    Admin_Account_Frame.pack_forget()
    Admin_Visitor_Frame.pack_forget()
    Admin_Ven_Frame.pack_forget()
    Refresh("adm_T")

def adm_vis(*args):
    global Admin_Menu_Var
    global Adm_Vis_Company_Var
    global Adm_Vis_Company_obj
    Admin_Menu_Var = 2
    adm_tenant_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_acc_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_company_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_ven_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_GN_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    Admin_Visitor_Frame.pack(fill=tk.BOTH, expand=1, padx=(5, 0), pady=5)
    Admin_GN_Frame.pack_forget()
    Admin_Company_Frame.pack_forget()
    Admin_Tenant_Frame.pack_forget()
    Admin_Account_Frame.pack_forget()
    Admin_Ven_Frame.pack_forget()
    Admin_Vis_Scroll.delete()
    Implement(Admin_Vis_Scroll.frame, "company", "Admin_Vis_Co", 16, None)
    Admin_Vis_Scroll.refresh()
    Adm_Vis_Company_obj = None
    Adm_Vis_Company_Var = None
    Admin_VIS_RESET()


def adm_GN(*args):
    global Admin_Menu_Var
    Admin_Menu_Var = 3
    adm_tenant_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_acc_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_company_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_vis_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_ven_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    Admin_GN_Frame.pack(fill=tk.BOTH, expand=1, padx=(5, 0), pady=5)
    Admin_Company_Frame.pack_forget()
    Admin_Tenant_Frame.pack_forget()
    Admin_Account_Frame.pack_forget()
    Admin_Visitor_Frame.pack_forget()
    Admin_Ven_Frame.pack_forget()
    Refresh("adm_GN")


def adm_acc(*args):
    global Admin_Menu_Var
    Admin_Menu_Var = 4
    adm_GN_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_tenant_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_company_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_vis_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_ven_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    Admin_Account_Frame.pack(fill=tk.BOTH, expand=1, padx=5, pady=5)
    Admin_Company_Frame.pack_forget()
    Admin_Tenant_Frame.pack_forget()
    Admin_GN_Frame.pack_forget()
    Admin_Visitor_Frame.pack_forget()
    Admin_Ven_Frame.pack_forget()

def adm_ven(*args):
    global Admin_Menu_Var
    Admin_Menu_Var = 5
    adm_GN_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_tenant_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_company_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_vis_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    adm_acc_button.configure(bg=conf["submenu_bg"], fg=conf["submenu_fg"])
    Admin_Account_Frame.pack_forget()
    Admin_Company_Frame.pack_forget()
    Admin_Tenant_Frame.pack_forget()
    Admin_GN_Frame.pack_forget()
    Admin_Visitor_Frame.pack_forget()
    Admin_Ven_Frame.pack(fill=tk.BOTH, expand=1, padx=5, pady=5)
    # Function to create list Vendors
    Adm_Ven_Insert(Admin_Ven_List_Frame)

Admin_Sub_Menu_Frame = tk.Frame(Menu_Bar_Admin, bg=conf["window_bg"])
Admin_Sub_Menu_Frame.pack(fill=tk.BOTH, expand=1)

adm_top_frame = tk.Frame(Admin_Sub_Menu_Frame, highlightthickness=0, bg=conf["window_bg"])
adm_top_frame.pack(side=tk.TOP, fill=tk.X)

adm_company_button = tk.Label(adm_top_frame, text="Company", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
adm_company_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
adm_company_button.bind("<Button-1>", adm_comp)
adm_company_button.bind("<Enter>", lambda x: adm_company_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
adm_company_button.bind("<Leave>", adm_howeroff)

adm_tenant_button = tk.Label(adm_top_frame, text="Tenant", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"], font=(conf["submenu_font"], conf["submenu_size"]))
adm_tenant_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
adm_tenant_button.bind("<Button-1>", adm_ten)
adm_tenant_button.bind("<Enter>", lambda x: adm_tenant_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
adm_tenant_button.bind("<Leave>", adm_howeroff)

adm_vis_button = tk.Label(adm_top_frame, text="Visitors", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_vis_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
adm_vis_button.bind("<Button-1>", adm_vis)
adm_vis_button.bind("<Enter>", lambda x: adm_vis_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
adm_vis_button.bind("<Leave>", adm_howeroff)

adm_GN_button = tk.Label(adm_top_frame, text="GN", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_GN_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
adm_GN_button.bind("<Button-1>", adm_GN)
adm_GN_button.bind("<Enter>", lambda x: adm_GN_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
adm_GN_button.bind("<Leave>", adm_howeroff)

adm_acc_button = tk.Label(adm_top_frame, text="Accounts", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_acc_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
adm_acc_button.bind("<Button-1>", adm_acc)
adm_acc_button.bind("<Enter>", lambda x: adm_acc_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
adm_acc_button.bind("<Leave>", adm_howeroff)

adm_ven_button = tk.Label(adm_top_frame, text="Vendors", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_ven_button.pack(fill=tk.BOTH, side=tk.LEFT, expand=1)
adm_ven_button.bind("<Button-1>", adm_ven)
adm_ven_button.bind("<Enter>", lambda x: adm_ven_button.configure(bg=conf["submenu_sel_bg"], fg=conf["submenu_sel_fg"]))
adm_ven_button.bind("<Leave>", adm_howeroff)

adm_main_frame = tk.Frame(Admin_Sub_Menu_Frame, highlightthickness=0, bg=conf["window_bg"])
adm_main_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=1)


#################################################################################
# Admin Company
#################################################################################
def admin_company_insert(*args):
    global today
    global Admin_Company_Quantity_Var
    global Admin_Company_Quantity_D_Var
    global Admin_Company_Quantity_R_Var
    global Admin_Company_Quantity_T_Var
    global Admin_Company_Quantity_Tr_Var
    global Admin_Company_Quantity_V_Var
    #entry creator in 2D by provided list
    def create_entry_grid(masta, data):
        rows = len(data)
        cols = len(data[0])
        # Create a 2D list to store the Entry widgets
        entries = [[None for _ in range(cols)] for _ in range(rows)]
        #inserting data in entries
        for i in range(rows):
            if data[i][7]: color = conf["widget_fg"]
            elif not data[i][7]: color = conf["widget_sel_bg"]
            else: error(6)
            for j in range(cols):
                entry = tk.Entry(masta, bg=conf["widget_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=color, justify=tk.CENTER, disabledbackground=conf["widget_bg"], disabledforeground=color)
                if j == 0: entry.config(width=10)
                elif (2<=j<=7): entry.config(width=20)
                elif j == 8: entry.config(width=29)
                elif j == 1: entry.config(width=adm_comp_name_lb.winfo_reqwidth())
                if data[i][j] is None or data[i][j] == "":
                    value = ""
                else:
                    if j == 8 and color == conf["widget_fg"]:
                        if int((data[i][j]-today.date()).days)<0: entry.config(fg=conf["expired_date"])
                        elif int((data[i][j]-today.date()).days)<30: entry.config(fg=conf["on_parking"])
                    value = data[i][j]
                entry.insert(0, value)
                entry.grid(row=i, column=j, sticky=tk.E, padx=(0, 1), pady=(1, 0))
                entry.update_idletasks()
                if j == 0: entry.config(state=tk.DISABLED)
                else: entry.config(state=tk.NORMAL)
                entries[i][j] = entry
        return entries

    Admin_Company_Quantity_Var = 0
    Admin_Company_Quantity_D_Var = 0
    Admin_Company_Quantity_R_Var = 0
    Admin_Company_Quantity_T_Var = 0
    Admin_Company_Quantity_Tr_Var = 0
    Admin_Company_Quantity_V_Var = 0
    if len(args)>1:
        adm_comp_list = args[1]
    else:
        raw = SQL_REQ("SELECT * FROM dbo.Company_list ORDER BY company_name", (), "S_all")
        adm_comp_list = list()
        for line in raw: adm_comp_list.append(list(line))
    for line in adm_comp_list:
        if line[7]:
            Admin_Company_Quantity_Var+=1
            if line[2] is not None and line[2] != "": Admin_Company_Quantity_D_Var += int(line[2])
            if line[3] is not None and line[3] != "": Admin_Company_Quantity_R_Var += int(line[3])
            if line[5] is not None and line[5] != "": Admin_Company_Quantity_T_Var += int(line[5])
            if line[4] is not None and line[4] != "": Admin_Company_Quantity_Tr_Var += int(line[4])
            if line[6] is not None and line[6] != "": Admin_Company_Quantity_V_Var += int(line[6])

    Admin_Company_Info_num_lb.config(text=Admin_Company_Quantity_Var)
    Admin_Designated_num_Info_num_lb.config(text=Admin_Company_Quantity_D_Var)
    Admin_Regular_num_Info_num_lb.config(text=Admin_Company_Quantity_R_Var)
    Admin_Truck_num_Info_num_lb.config(text=Admin_Company_Quantity_T_Var)
    Admin_Trailer_num_Info_num_lb.config(text=Admin_Company_Quantity_Tr_Var)
    Admin_Vehicle_num_Info_num_lb.config(text=Admin_Company_Quantity_V_Var)
    return create_entry_grid(args[0], adm_comp_list)
def admin_company_register(*args):
    raw = SQL_REQ("SELECT * FROM dbo.Company_list ORDER BY company_name", (), "S_all")
    comp_list = list()
    for line in raw: comp_list.append(list(line))
    new_record = [
        Admin_Company_Add_Company_Entry_name.get().strip(),
        Admin_Company_Add_Company_Entry_des.get().strip(),
        Admin_Company_Add_Company_Entry_reg.get().strip(),
        Admin_Company_Add_Company_Entry_tr.get().strip(),
        Admin_Company_Add_Company_Entry_t.get().strip(),
        Admin_Company_Add_Company_Entry_v.get().strip(),
        Admin_Company_Add_Company_Entry_a.get().strip(),
        Admin_Company_Add_Company_Entry_i.get().strip()
    ]

    for i in range(len(new_record)):
        if new_record[i] == "": new_record[i] = None
        if 1<= i <= 5:
            if new_record[i] !="" and new_record[i] is not None:
                try:
                    int(new_record[i])
                except:
                    error(5)
                    return
        if i == 6:
            if new_record[i] is None or new_record[i] == "":
                error(6)
                return
        if i ==7:
            try:
                datetime.strptime(new_record[i], "%Y-%m-%d")
            except:
                error(7)
                return

    ID = []
    NEW_ID = None
#check if company name exist and popup error
    for line in comp_list:
        if comp_list[1] == new_record:
            error(8)
            return
        ID.append(int(line[0]))
#check first available ID
    for i in range(1, 1000):
        if i not in ID:
            NEW_ID = i
            break
    new_record.insert(0, str(NEW_ID))
    SQL_REQ("INSERT INTO dbo.Company_List(company_ID, company_name, designated, regular, trailer, truck, car, activity, insurance) VALUES (?,?,?,?,?,?,?,?,?)", new_record, "W")
    admin_company_cancel()
    admin_company_insert(Admin_Company_Scroll.frame)

def admin_company_commit_changes(entries):
    global Admin_Company_Entries
    modified_list = []
    for row in entries:
        #modified_row = [entry.get().strip() for entry in row]
        modified_row = []
        for entry in row:
            value = entry.get().strip()
            modified_row.append(value)
        if modified_row[8] is not None and modified_row[8] != "":
            try:
                modified_row[8] = datetime.strptime(modified_row[8], "%Y-%m-%d").date()
            except:
                error(7)
                return
        if modified_row[7] == "1": modified_row[7] = True
        elif modified_row[7] == "0": modified_row[7] = False
        else:
            error(6)
            return
        for i in range(len(modified_row)):
            if 2<= i <= 6:
                if modified_row[i] is not None and modified_row[i] != "":
                    try: int(modified_row[i])
                    except:
                        error(5)
                        return

        modified_list.append(modified_row)
    for widgets in Admin_Company_Scroll.frame.winfo_children(): widgets.destroy()
    # update SQL

    for line in modified_list:
        line = [value if value != "" else None for value in line]
        SQL_REQ("UPDATE dbo.Company_List SET company_name=?, designated=?, regular=?, trailer=?, truck=?, car=?, activity=?, insurance=? WHERE company_ID=?", (line[1], line[2], line[3], line[4], line[5], line[6], line[7], line[8] , line[0]), "W")
    #############

    Admin_Company_Entries = admin_company_insert(Admin_Company_Scroll.frame, modified_list)



def admin_company_add():
    Admin_Info_Frame.pack_forget()
    Admin_Company_Add_Button.pack_forget()
    Admin_Company_Cancel_Button.pack(side=tk.LEFT, fill=tk.Y, padx=2, pady=2)
    Admin_Company_Save_Button.config(command=admin_company_register, fg=conf["widget_sel_fg"])
    Admin_Company_Add_Entry_fr.pack(fill=tk.X, side=tk.LEFT, expand=1)


def admin_company_cancel():
    Admin_Company_Cancel_Button.pack_forget()
    Admin_Company_Add_Entry_fr.pack_forget()
    Admin_Company_Save_Button.config(command=lambda:admin_company_commit_changes(Admin_Company_Entries), fg=conf["submenu_fg"])
    Admin_Company_Add_Button.pack(side=tk.LEFT, fill=tk.Y, padx=2, pady=2)
    Admin_Info_Frame.pack(fill=tk.X, side=tk.LEFT, expand=1)


Admin_Company_Frame = tk.Frame(adm_main_frame, bg=conf["window_bg"], highlightthickness=0)
Admin_Company_Frame.pack_propagate(0)

Admin_Company_Table_Frame = tk.Frame(Admin_Company_Frame, bg=conf["window_bg"], highlightthickness=0)
Admin_Company_Table_Frame.pack(fill=tk.X, side=tk.TOP, expand=0)
#Signs in Table
adm_comp_id_lb = tk.Label(Admin_Company_Table_Frame, text="ID", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=10)
adm_comp_id_lb.grid(row=0, column=0, sticky=tk.EW)
adm_comp_name_lb = tk.Label(Admin_Company_Table_Frame, text="Name", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
adm_comp_name_lb.grid(row=0, column=1, sticky=tk.EW)
adm_comp_des_lb = tk.Label(Admin_Company_Table_Frame, text="Designated", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
adm_comp_des_lb.grid(row=0, column=2, sticky=tk.EW)
adm_comp_reg_lb = tk.Label(Admin_Company_Table_Frame, text="Regular", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=19)
adm_comp_reg_lb.grid(row=0, column=3, sticky=tk.EW)
adm_comp_trk_lb = tk.Label(Admin_Company_Table_Frame, text="Trailer", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
adm_comp_trk_lb.grid(row=0, column=4, sticky=tk.EW)
adm_comp_trl_lb = tk.Label(Admin_Company_Table_Frame, text="Truck", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
adm_comp_trl_lb.grid(row=0, column=5, sticky=tk.EW)
adm_comp_car_lb = tk.Label(Admin_Company_Table_Frame, text="Car", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
adm_comp_car_lb.grid(row=0, column=6, sticky=tk.EW)
adm_comp_car_lb = tk.Label(Admin_Company_Table_Frame, text="Active", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=18)
adm_comp_car_lb.grid(row=0, column=7, sticky=tk.EW)
adm_comp_ins_lb = tk.Label(Admin_Company_Table_Frame, text="Insurance", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=32)
adm_comp_ins_lb.grid(row=0, column=8, sticky=tk.EW)

# Admin_Company_Table_Frame.grid_columnconfigure(0, weight=1)

Admin_Company_Table_Frame.grid_columnconfigure(1, weight=1)

# Admin_Company_Table_Frame.grid_columnconfigure(2, weight=2)
# Admin_Company_Table_Frame.grid_columnconfigure(3, weight=2)
# Admin_Company_Table_Frame.grid_columnconfigure(4, weight=2)
# Admin_Company_Table_Frame.grid_columnconfigure(5, weight=2)
# Admin_Company_Table_Frame.grid_columnconfigure(6, weight=2)
# Admin_Company_Table_Frame.grid_columnconfigure(7, weight=4)
# Admin_Company_Table_Frame.grid_columnconfigure(8, weight=2)

Admin_Company_Buttons_Frame = tk.Frame(Admin_Company_Frame, highlightthickness=3, relief=tk.RAISED, bg=conf["submenu_bg"], highlightbackground=conf["submenu_sel_bg"], height=conf["admin_company_button_height"])
Admin_Company_Buttons_Frame.pack(fill=tk.X, side=tk.BOTTOM)
Admin_Company_Scroll = scroller(Admin_Company_Frame)
Admin_Company_Scroll.pack(fill=tk.BOTH, side=tk.TOP, expand=1, anchor=tk.N)
Admin_Company_Refresh_Button = tk.Button(Admin_Company_Buttons_Frame, text="REFRESH", command=lambda:admin_company_insert(Admin_Company_Scroll.frame), width=20, fg=conf["submenu_fg"])
Admin_Company_Refresh_Button.pack(side=tk.LEFT, fill=tk.Y,  padx=2, pady=2)
Admin_Company_Save_Button = tk.Button(Admin_Company_Buttons_Frame, text="SAVE", command=lambda:admin_company_commit_changes(Admin_Company_Entries), width=20, fg=conf["submenu_fg"])
Admin_Company_Save_Button.pack(side=tk.LEFT, fill=tk.Y,  padx=2, pady=2)
Admin_Company_Add_Button = tk.Button(Admin_Company_Buttons_Frame, text="ADD", command=admin_company_add, width=20, fg=conf["submenu_fg"])
Admin_Company_Add_Button.pack(side=tk.LEFT, fill=tk.Y,  padx=2, pady=2)
Admin_Company_Cancel_Button = tk.Button(Admin_Company_Buttons_Frame, text="Cancel", command=admin_company_cancel, width=20, fg=conf["submenu_fg"])

#Frame and widgets for Info panel
Admin_Info_Frame = tk.Frame(Admin_Company_Buttons_Frame, highlightthickness=0, bg=conf["submenu_bg"])
Admin_Info_Frame.pack(fill=tk.X, side=tk.LEFT, expand=1)
Admin_Company_Info_lb = tk.Label(Admin_Info_Frame, text="Companies:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=20)
Admin_Company_Info_lb.pack(side=tk.LEFT)
Admin_Company_Info_num_lb = tk.Label(Admin_Info_Frame, text=Admin_Company_Quantity_Var, bg=conf["window_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], width=5)
Admin_Company_Info_num_lb.pack(side=tk.LEFT, anchor=tk.W)
Admin_Designated_Info_lb = tk.Label(Admin_Info_Frame, text="Designated:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=20)
Admin_Designated_Info_lb.pack(side=tk.LEFT)
Admin_Designated_num_Info_num_lb = tk.Label(Admin_Info_Frame, text=Admin_Company_Quantity_D_Var, bg=conf["window_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], width=5)
Admin_Designated_num_Info_num_lb.pack(side=tk.LEFT, anchor=tk.W)
Admin_Regular_Info_lb = tk.Label(Admin_Info_Frame, text="Regulars:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=20)
Admin_Regular_Info_lb.pack(side=tk.LEFT)
Admin_Regular_num_Info_num_lb = tk.Label(Admin_Info_Frame, text=Admin_Company_Quantity_R_Var, bg=conf["window_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], width=5)
Admin_Regular_num_Info_num_lb.pack(side=tk.LEFT, anchor=tk.W)
Admin_Trailer_Info_lb = tk.Label(Admin_Info_Frame, text="Trailers:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=20)
Admin_Trailer_Info_lb.pack(side=tk.LEFT)
Admin_Trailer_num_Info_num_lb = tk.Label(Admin_Info_Frame, text=Admin_Company_Quantity_Tr_Var, bg=conf["window_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], width=5)
Admin_Trailer_num_Info_num_lb.pack(side=tk.LEFT, anchor=tk.W)
Admin_Truck_Info_lb = tk.Label(Admin_Info_Frame, text="Trucks:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=20)
Admin_Truck_Info_lb.pack(side=tk.LEFT)
Admin_Truck_num_Info_num_lb = tk.Label(Admin_Info_Frame, text=Admin_Company_Quantity_T_Var, bg=conf["window_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], width=5)
Admin_Truck_num_Info_num_lb.pack(side=tk.LEFT, anchor=tk.W)
Admin_Vehicle_Info_lb = tk.Label(Admin_Info_Frame, text="Vehichles:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=20)
Admin_Vehicle_Info_lb.pack(side=tk.LEFT)
Admin_Vehicle_num_Info_num_lb = tk.Label(Admin_Info_Frame, text=Admin_Company_Quantity_V_Var, bg=conf["window_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], width=5)
Admin_Vehicle_num_Info_num_lb.pack(side=tk.LEFT, anchor=tk.W)
#Frame and widgets for Entry Add Company panel
Admin_Company_Add_Entry_fr = tk.Frame(Admin_Company_Buttons_Frame, highlightthickness=0, bg=conf["submenu_bg"])
Admin_Company_Add_Company_Entry_name_lb = tk.Label(Admin_Company_Add_Entry_fr, text="Name:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Company_Add_Company_Entry_name_lb.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_name = tk.Entry(Admin_Company_Add_Entry_fr, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=20)
Admin_Company_Add_Company_Entry_name.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_des_lb = tk.Label(Admin_Company_Add_Entry_fr, text="Designated:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Company_Add_Company_Entry_des_lb.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_des = tk.Entry(Admin_Company_Add_Entry_fr, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=3)
Admin_Company_Add_Company_Entry_des.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_reg_lb = tk.Label(Admin_Company_Add_Entry_fr, text="Regular:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Company_Add_Company_Entry_reg_lb.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_reg = tk.Entry(Admin_Company_Add_Entry_fr, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=3)
Admin_Company_Add_Company_Entry_reg.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_tr_lb = tk.Label(Admin_Company_Add_Entry_fr, text="Trailer:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Company_Add_Company_Entry_tr_lb.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_tr = tk.Entry(Admin_Company_Add_Entry_fr, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=3)
Admin_Company_Add_Company_Entry_tr.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_t_lb = tk.Label(Admin_Company_Add_Entry_fr, text="Truck:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Company_Add_Company_Entry_t_lb.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_t = tk.Entry(Admin_Company_Add_Entry_fr, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=3)
Admin_Company_Add_Company_Entry_t.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_v_lb = tk.Label(Admin_Company_Add_Entry_fr, text="Vehicle:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Company_Add_Company_Entry_v_lb.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_v = tk.Entry(Admin_Company_Add_Entry_fr, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=3)
Admin_Company_Add_Company_Entry_v.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_a_lb = tk.Label(Admin_Company_Add_Entry_fr, text="Active:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Company_Add_Company_Entry_a_lb.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_a = tk.Entry(Admin_Company_Add_Entry_fr, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=3)
Admin_Company_Add_Company_Entry_a.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_i_lb = tk.Label(Admin_Company_Add_Entry_fr, text="Insurance:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Company_Add_Company_Entry_i_lb.pack(side=tk.LEFT)
Admin_Company_Add_Company_Entry_i = tk.Entry(Admin_Company_Add_Entry_fr, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=10)
Admin_Company_Add_Company_Entry_i.pack(side=tk.LEFT)




#################################################################################
# Admin Tenant
#################################################################################
Admin_Tenant_Frame = tk.Frame(adm_main_frame, bg=conf["window_bg"], highlightthickness=0)
Admin_Tenant_Frame.pack_propagate(0)

# Admin Tenant COMPANY FRAME
adm_C_fr = tk.Frame(Admin_Tenant_Frame, bg=conf["window_bg"], highlightthickness=0, width=conf["p_t_company_w"])
adm_C_fr.pack_propagate(0)
adm_C_fr.pack(fill=tk.BOTH, side=tk.LEFT)

#LABEL FOR COMPANY
adm_C_Lb = tk.Label(adm_C_fr, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
adm_C_Lb.pack(fill=tk.X, side=tk.TOP)

#Scroll Frame for Companies
Admin_Tenant_Scroll = scroller(adm_C_fr)
Admin_Tenant_Scroll.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
Admin_Tenant_Scroll.refresh()

#Central frame
adm_central_frame = tk.Frame(Admin_Tenant_Frame, highlightthickness=0, bg=conf["window_bg"])
adm_central_frame.pack_propagate(0)
adm_central_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, padx=5)

#Right Frame for buttons
adm_right_frame = tk.Frame(Admin_Tenant_Frame, highlightthickness=3, relief=tk.RAISED, bg=conf["submenu_bg"], highlightbackground=conf["submenu_sel_bg"], width=conf["chk_filter_frame"])
adm_right_frame.pack_propagate(0)
adm_right_frame.pack(fill=tk.BOTH, side=tk.LEFT)

#TRUCKS FRAME
adm_Tenant_Trucks_Frame = tk.Frame(adm_central_frame, bg=conf["window_bg"], highlightthickness=0)
adm_Tenant_Trucks_Frame.pack_propagate(0)
adm_Tenant_Trucks_Frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1)

#LABEL FOR TRUCK
adm_Tenant_Truck_Lb = tk.Label(adm_Tenant_Trucks_Frame, text="TRUCK:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
adm_Tenant_Truck_Lb.pack(fill=tk.X, side=tk.TOP)

#Scroll Frame for Trucks
Admin_Tenant_T_Scroll = scroller(adm_Tenant_Trucks_Frame)
Admin_Tenant_T_Scroll.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
Admin_Tenant_T_Scroll.refresh()

#TRAILERS FRAME
adm_Tenant_Trailers_Frame = tk.Frame(adm_central_frame, bg=conf["window_bg"], highlightthickness=0)
adm_Tenant_Trailers_Frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1, pady=(5, 0))

#LABEL FOR TRAILERS
adm_Tenant_Trailer_Lb = tk.Label(adm_Tenant_Trailers_Frame, text="TRAILER:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
adm_Tenant_Trailer_Lb.pack(fill=tk.X, side=tk.TOP)

#Scroll Frame for Trailers
Admin_Tenant_Tr_Scroll = scroller(adm_Tenant_Trailers_Frame)
Admin_Tenant_Tr_Scroll.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
Admin_Tenant_Tr_Scroll.refresh()



def adm_add(*args):
    if Adm_Company_Var is None: return
    opt = adm_radio_var.get()
    unit = adm_t_entry.get().strip()
    if unit is None or unit == "": return
    var = SQL_REQ("SELECT company_ID FROM dbo.Company_List WHERE company_name=?", (Adm_Company_Var,), "S_one")
    if var: ID = str(var[0])
    if opt == "truck":
        table = "dbo.Tenant_Trucks"
        unreg_table = "dbo.Tenant_Trucks_UNREG"
        column = "truck_number"
        unit_sel = Adm_Truck_Var
    elif opt == "trailer":
        table = "dbo.Tenant_Trailers"
        unreg_table = "dbo.Tenant_Trailers_UNREG"
        column = "trailer_number"
        unit_sel = Adm_Trailer_Var
    if Adm_Unit_obj is None:
        check = SQL_REQ(f"SELECT * FROM {unreg_table} WHERE company_ID=? AND {column}=?", (ID, unit), "S_one")
        if check:
            recheck = SQL_REQ(f"SELECT * FROM {table} WHERE company_ID=? AND {column}=?", (ID, unit), "S_one")
            if not recheck:
                SQL_REQ(f"INSERT INTO {table} SELECT * FROM {unreg_table} WHERE company_ID=? AND {column}=?", (ID, unit), "W")
                SQL_REQ(f"DELETE FROM {unreg_table} WHERE company_ID=? AND {column}=?", (ID, unit), "W")
            else: error(12)
        else:
            recheck = SQL_REQ(f"SELECT * FROM {table} WHERE company_ID=? AND {column}=?", (ID, unit), "S_one")
            if not recheck:
                query = f"INSERT INTO {table} (company_ID, {column}) VALUES (?, ?)"
                values = (ID, unit)
                SQL_REQ(query, values, "W")   #fix insert select*
            else: error(12)
    else:
        if unit == unit_sel[0]:
            check = SQL_REQ(f"SELECT * FROM {unreg_table} WHERE company_ID=? AND {column}=?", (ID, unit), "S_one")
            if check:
                recheck = SQL_REQ(f"SELECT * FROM {table} WHERE company_ID=? AND {column}=?", (ID, unit), "S_one")
                if not recheck:
                    SQL_REQ(f"INSERT INTO {table} SELECT * FROM {unreg_table} WHERE company_ID=? AND {column}=?", (ID, unit), "W")
                    SQL_REQ(f"DELETE FROM {unreg_table} WHERE company_ID=? AND {column}=?", (ID, unit), "W")
                else:
                    SQL_REQ(f"DELETE FROM {unreg_table} WHERE company_ID=? AND {column}=?", (ID, unit), "W")
            else:
                recheck = SQL_REQ(f"SELECT * FROM {table} WHERE company_ID=? AND {column}=?", (ID, unit), "S_one")
                if not recheck:
                    error(15)
                    return
                else:
                    return
        else:
            check = SQL_REQ(f"SELECT * FROM {unreg_table} WHERE company_ID=? AND {column}=?", (ID, unit), "S_one")
            if check:
                error(12)
                return
            recheck = SQL_REQ(f"SELECT * FROM {table} WHERE company_ID=? AND {column}=?", (ID, unit), "S_one")
            if recheck:
                error(12)
                return
            SQL_REQ(f"UPDATE {table} SET {column}=? WHERE company_ID=? AND {column}=?", (unit, ID, unit_sel[0]), "W")

    adm_t_entry.delete(0,tk.END)
    adm_t_entry.focus_set()
    UNTS(Adm_Company_obj, Adm_Company_Var, "Admin_Units")

def adm_remove(*args):
    global Adm_Truck_Var
    global Adm_Trailer_Var
    global adm_radio_var
    if Adm_Company_Var is None: return
    if Adm_Truck_Var is None and Adm_Trailer_Var is None: return
    val = SQL_REQ("SELECT company_ID FROM dbo.Company_List WHERE company_name=?", (Adm_Company_Var), "S_one")
    if val: ID =  str(val[0])
    if Adm_Truck_Var is not None:
        if Adm_Truck_Var[2] == "REG":
            check = SQL_REQ("SELECT * FROM dbo.Tenant_Trucks WHERE company_ID=? AND truck_number=?", (ID, Adm_Truck_Var[0]), "S_one")
            if check: SQL_REQ("DELETE FROM dbo.Tenant_Trucks WHERE company_ID=? AND truck_number=?", (ID, Adm_Truck_Var[0]), "W")
        elif Adm_Truck_Var[2] == "UNREG":
            check = SQL_REQ("SELECT * FROM dbo.Tenant_Trucks_UNREG WHERE company_ID=? AND truck_number=?", (ID, Adm_Truck_Var[0]), "S_one")
            if check: SQL_REQ("DELETE FROM dbo.Tenant_Trucks_UNREG WHERE company_ID=? AND truck_number=?", (ID, Adm_Truck_Var[0]), "W")
    elif Adm_Trailer_Var is not None:
        if Adm_Trailer_Var[2] == "REG":
            check = SQL_REQ("SELECT * FROM dbo.Tenant_Trailers WHERE company_ID=? AND trailer_number=?", (ID, Adm_Trailer_Var[0]), "S_one")
            if check: SQL_REQ("DELETE FROM dbo.Tenant_Trailers WHERE company_ID=? AND trailer_number=?", (ID, Adm_Trailer_Var[0]), "W")
        elif Adm_Trailer_Var[2] == "UNREG":
            check = SQL_REQ("SELECT * FROM dbo.Tenant_Trailers_UNREG WHERE company_ID=? AND trailer_number=?", (ID, Adm_Trailer_Var[0]), "S_one")
            if check: SQL_REQ("DELETE FROM dbo.Tenant_Trailers_UNREG WHERE company_ID=? AND trailer_number=?", (ID, Adm_Trailer_Var[0]), "W")
    adm_t_entry.focus_set()
    UNTS(Adm_Company_obj, Adm_Company_Var, "Admin_Units")
def adm_T_entry_focus(*args):
    global Adm_Truck_Var
    global Adm_Trailer_Var
    global Adm_Unit_obj
    if Adm_Unit_obj is not None:
        Adm_Unit_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        Adm_Trailer_Var = None
        Adm_Truck_Var = None
    adm_t_entry.delete(0, tk.END)
    adm_radio_var.set("truck")

def adm_storage_check():
    global adm_storage_var
    if Adm_Company_Var is None: return
    if Adm_Trailer_Var is None: return
    str_status = adm_storage_var.get()
    val = SQL_REQ("SELECT company_ID FROM dbo.Company_List WHERE company_name=?", (Adm_Company_Var,), "S_one")
    if val: ID = str(val[0])
    if Adm_Trailer_Var[2] == "REG":
        check = SQL_REQ("SELECT * FROM dbo.Tenant_Trailers WHERE company_ID=? AND trailer_number=?", (ID, Adm_Trailer_Var[0]), "S_one")
        if check: SQL_REQ("UPDATE dbo.Tenant_Trailers SET storage=? WHERE company_ID=? AND trailer_number=?", (str(str_status), ID, Adm_Trailer_Var[0]), "W")
    elif Adm_Trailer_Var[2] == "UNREG":
        check = SQL_REQ("SELECT * FROM dbo.Tenant_Trailers_UNREG WHERE company_ID=? AND trailer_number=?", (ID, Adm_Trailer_Var[0]), "S_one")
        if check: SQL_REQ("UPDATE dbo.Tenant_Trailers_UNREG SET storage=? WHERE company_ID=? AND trailer_number=?", (str(str_status), ID, Adm_Trailer_Var[0]), "W")

#Button Frame for Adding/Delete Units
adm_manual_entry_frame = tk.Frame(adm_right_frame, highlightthickness=0, bg=conf["submenu_bg"])
adm_manual_entry_frame.pack(side=tk.TOP)
adm_c_lb = tk.Label(adm_manual_entry_frame, text="Company:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_c_lb.pack(fill=tk.X, side=tk.TOP)
adm_c_entry = tk.Entry(adm_manual_entry_frame, bg=conf["window_bg"], bd=1, cursor="shuttle", font=(conf["entry_font"], conf["notebook_tab_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, justify=tk.LEFT, disabledbackground=conf["window_bg"], disabledforeground=conf["widget_fg"])
adm_c_entry.pack(fill=tk.X, side=tk.TOP)
adm_t_lb = tk.Label(adm_manual_entry_frame, text="UNIT:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_t_lb.pack(fill=tk.X, side=tk.TOP)
adm_t_entry = tk.Entry(adm_manual_entry_frame, bg=conf["window_bg"], bd=1, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, justify=tk.LEFT)
adm_t_entry.pack(fill=tk.X, side=tk.TOP)
adm_t_entry.bind("<Return>", adm_add)
adm_t_entry.bind("<Delete>", adm_remove)
# adm_t_entry.bind("<FocusIn>", adm_T_entry_focus)
# adm_t_entry.bind("<Button-1>", adm_T_entry_focus)

adm_radio_var = tk.StringVar()
adm_truck_radio = tk.Radiobutton(adm_manual_entry_frame, text="Truck", variable=adm_radio_var, value="truck", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
adm_truck_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
adm_trailer_radio = tk.Radiobutton(adm_manual_entry_frame, text="Trailer", variable=adm_radio_var, value="trailer", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
adm_trailer_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
adm_radio_var.set("truck")
adm_manual_add_button = tk.Button(adm_manual_entry_frame, text="ADD", bg=conf["widget_bg"], fg=conf["in_button_fg"], bd=0, highlightthickness=0, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=adm_add)
adm_manual_add_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
adm_manual_add_button.bind("<Return>", adm_add)
adm_manual_remove_button = tk.Button(adm_manual_entry_frame, text="REMOVE", bg=conf["widget_bg"], fg=conf["out_button_fg"], bd=0, highlightthickness=0, activebackground=conf["out_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["out_button_sel_fg"], command=adm_remove)
adm_manual_remove_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
adm_manual_remove_button.bind("<Delete>", adm_remove)
adm_storage_var = tk.IntVar()
adm_storage_checkbox = tk.Checkbutton(adm_manual_entry_frame, text="Storage", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], variable=adm_storage_var, command=adm_storage_check)


Implement(Admin_Tenant_Scroll.frame, "company", "Admin_Units", 10, None)
Admin_Tenant_Scroll.refresh()


#################################################################################
# Admin VIS
#################################################################################
def adm_Vis_add(*args):
    global adm_vis_radio_var
    global Adm_Vis_Company_Var
    print(Adm_Vis_Company_Var)
    if Adm_Vis_Company_Var is None: return
    plate = adm_Vis_t_entry.get().strip() if adm_Vis_t_entry.get().strip() != "" else None
    print(plate)
    if plate is None: return
    car = adm_Vis_car_entry.get().strip() if adm_Vis_car_entry.get().strip() != "" else None
    print(car)
    name = adm_Vis_n_entry.get().strip() if adm_Vis_n_entry.get().strip() != "" else None
    print(name)
    ID = ID_NAME_company(name=Adm_Vis_Company_Var) #str(SQL_REQ("SELECT company_ID FROM dbo.Company_List WHERE company_name=?", (Adm_Vis_Company_Var,), "S_one")[0])
    print(ID)
    prv = adm_vis_radio_var.get()
    last_date = None
    status = None
    if prv == "private":
        if adm_Vis_exp_entry.get() != "":
            try:
                exp = datetime.strptime(adm_Vis_exp_entry.get().strip(), "%Y-%m-%d").date()
                private_var = True
            except:
                error(7)
                return
        else:
            error(7)
            return
    else:
        exp=0
        if prv == "com": private_var = False
        else: private_var = None

    rec = [ID, plate, name, car, exp, private_var, last_date, status]
    print(rec)
    if rec[4] == 0: rec[4] = None
    if Current_Adm_Visitor_Unit is not None:
        check = SQL_REQ("SELECT * FROM dbo.visitors_UNREG WHERE plates=?", (Current_Adm_Visitor_Unit[5].get("plates")), "S_one_D")
        if check:
            SQL_REQ("DELETE FROM dbo.visitors_UNREG WHERE plates=?", (Current_Adm_Visitor_Unit[5].get("plates")), "W")
            #taking last_date and status from old record
            rec[6] = Current_Adm_Visitor_Unit[5].get("last_date")
            rec[7] = Current_Adm_Visitor_Unit[5].get("status")
        check = SQL_REQ("SELECT * FROM dbo.visitors WHERE plates=?", (Current_Adm_Visitor_Unit[5].get("plates")), "S_one_D")
        if check:
            SQL_REQ("DELETE FROM dbo.visitors WHERE plates=?", (Current_Adm_Visitor_Unit[5].get("plates")), "W")
            # taking last_date and status from old record
            rec[6] = Current_Adm_Visitor_Unit[5].get("last_date")
            rec[7] = Current_Adm_Visitor_Unit[5].get("status")
    SQL_REQ("INSERT INTO dbo.visitors (company_ID, plates, driver_name, car_model, expiration, private, last_date, status) VALUES (?,?,?,?,?,?,?,?)", rec, "W")
    Admin_VIS_RESET()

def adm_Vis_remove(*args):
    plate = adm_Vis_t_entry.get().strip()
    if plate == "": return
    if Current_Adm_Visitor_Unit[6] is None:
        error(15)
        return
    elif Current_Adm_Visitor_Unit[6] is False:
        table = "dbo.visitors_UNREG"
    else:
        table = "dbo.visitors"
    SQL_REQ(f"DELETE FROM {table} WHERE plates=?", (Current_Adm_Visitor_Unit[5].get("plates"),), "W")
    Admin_VIS_RESET()

def adm_vis_no():
    adm_Vis_exp_entry.config(state=tk.NORMAL)
    adm_Vis_exp_entry.delete(0, tk.END)
    adm_Vis_exp_entry.config(state=tk.DISABLED)
def adm_vis_com():
    adm_Vis_exp_entry.config(state=tk.NORMAL)
    adm_Vis_exp_entry.delete(0, tk.END)
    adm_Vis_exp_entry.config(state=tk.DISABLED)
def adm_vis_private():
    adm_Vis_exp_entry.config(state=tk.NORMAL)


Admin_Visitor_Frame = tk.Frame(adm_main_frame, bg=conf["window_bg"], highlightthickness=0)
Admin_Visitor_Frame.pack_propagate(0)

# Admin Tenant COMPANY FRAME
adm_vis_C_fr = tk.Frame(Admin_Visitor_Frame, bg=conf["window_bg"], highlightthickness=0, width=conf["p_t_company_w"])
adm_vis_C_fr.pack_propagate(0)
adm_vis_C_fr.pack(fill=tk.BOTH, side=tk.LEFT)

#LABEL FOR COMPANY
adm_vis_C_Lb = tk.Label(adm_vis_C_fr, text="COMPANY:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
adm_vis_C_Lb.pack(fill=tk.X, side=tk.TOP)

#Scroll Frame for Companies
Admin_Vis_Scroll = scroller(adm_vis_C_fr)
Admin_Vis_Scroll.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
Admin_Vis_Scroll.refresh()

#Central frame
adm_Vis_central_frame = tk.Frame(Admin_Visitor_Frame, highlightthickness=0, bg=conf["window_bg"])
adm_Vis_central_frame.pack_propagate(0)
adm_Vis_central_frame.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, padx=5)

#frame for sings
adm_sign_central_frame = tk.Frame(adm_Vis_central_frame, highlightthickness=0, bg=conf["window_bg"])
adm_sign_central_frame.pack(fill=tk.X, side=tk.TOP, anchor=tk.N)
#frame for list
adm_vis_carlist_central_frame = tk.Frame(adm_Vis_central_frame, highlightthickness=0, bg=conf["window_bg"])
adm_vis_carlist_central_frame.pack(fill=tk.X, side=tk.TOP, expand=1, anchor=tk.N)

#Scroll Frame for Car list
Admin_Vis_Car_Scroll = scroller(adm_vis_carlist_central_frame)
Admin_Vis_Car_Scroll.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
Admin_Vis_Car_Scroll.refresh()



#Right Frame for buttons
adm_Vis_right_frame = tk.Frame(Admin_Visitor_Frame, highlightthickness=3, relief=tk.RAISED, bg=conf["submenu_bg"], highlightbackground=conf["submenu_sel_bg"], width=conf["chk_filter_frame"])
adm_Vis_right_frame.pack_propagate(0)
adm_Vis_right_frame.pack(fill=tk.BOTH, side=tk.LEFT)

#Button Frame for Adding/Delete Units
adm_Vis_manual_entry_frame = tk.Frame(adm_Vis_right_frame, highlightthickness=0, bg=conf["submenu_bg"])
adm_Vis_manual_entry_frame.pack(side=tk.TOP, fill=tk.Y, expand=1)
adm_Vis_c_lb = tk.Label(adm_Vis_manual_entry_frame, text="Company:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_Vis_c_lb.pack(fill=tk.X, side=tk.TOP)
adm_Vis_c_entry = tk.Entry(adm_Vis_manual_entry_frame, state=tk.DISABLED, bg=conf["window_bg"], bd=1, font=(conf["entry_font"], conf["notebook_tab_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, justify=tk.LEFT, disabledbackground=conf["window_bg"], disabledforeground=conf["widget_fg"])
adm_Vis_c_entry.pack(fill=tk.X, side=tk.TOP)
adm_Vis_t_lb = tk.Label(adm_Vis_manual_entry_frame, text="PLATES:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_Vis_t_lb.pack(fill=tk.X, side=tk.TOP)
adm_Vis_t_entry = tk.Entry(adm_Vis_manual_entry_frame, bg=conf["window_bg"], bd=1, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, justify=tk.LEFT)
adm_Vis_t_entry.pack(fill=tk.X, side=tk.TOP)
adm_Vis_t_entry.bind("<KeyRelease>", lambda event: UPPER_CASE(event, obj=adm_Vis_t_entry))
#adm_Vis_t_entry.bind("<Tab>", adm_Vis_add)


adm_Vis_car_lb = tk.Label(adm_Vis_manual_entry_frame, text="CAR:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_Vis_car_lb.pack(fill=tk.X, side=tk.TOP)

adm_Vis_car_entry = AutocompleteEntry(adm_Vis_manual_entry_frame, completevalues=vendors, textvariable=VIS_Car_Var, style="Custom.TEntry", background=conf["window_bg"], cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), foreground=conf["entry_fg"], justify=tk.LEFT)
adm_Vis_car_entry.pack(fill=tk.X, side=tk.TOP)

adm_Vis_n_lb = tk.Label(adm_Vis_manual_entry_frame, text="NAME:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_Vis_n_lb.pack(fill=tk.X, side=tk.TOP)
adm_Vis_n_entry = tk.Entry(adm_Vis_manual_entry_frame, bg=conf["window_bg"], bd=1, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, justify=tk.LEFT)
adm_Vis_n_entry.pack(fill=tk.X, side=tk.TOP)

adm_vis_radio_var = tk.StringVar()
adm_vis_N_radio = tk.Radiobutton(adm_Vis_manual_entry_frame, text="No parking", command=adm_vis_no, variable=adm_vis_radio_var, value="no", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
adm_vis_N_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
adm_vis_com_radio = tk.Radiobutton(adm_Vis_manual_entry_frame, text="Commercial", command=adm_vis_com, variable=adm_vis_radio_var, value="com", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
adm_vis_com_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
adm_vis_private_radio = tk.Radiobutton(adm_Vis_manual_entry_frame, text="Private", command=adm_vis_private, variable=adm_vis_radio_var, value="private", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
adm_vis_private_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
adm_vis_radio_var.set("no")

adm_Vis_exp_lb = tk.Label(adm_Vis_manual_entry_frame, text="Exp/Date:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_Vis_exp_lb.pack(fill=tk.X, side=tk.TOP)
adm_Vis_exp_entry = tk.Entry(adm_Vis_manual_entry_frame, bg=conf["window_bg"], state=tk.DISABLED, bd=1, cursor="shuttle", disabledbackground=conf["submenu_bg"], font=(conf["entry_font"], conf["entry_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, justify=tk.LEFT)
adm_Vis_exp_entry.pack(fill=tk.X, side=tk.TOP)

adm_Vis_exp_details = tk.Label(adm_Vis_manual_entry_frame, text="YYYY-MM-DD", bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["notebook_tab_size"]))
adm_Vis_exp_details.pack(fill=tk.X, side=tk.TOP, anchor=tk.CENTER)

adm_Vis_manual_remove_button = tk.Button(adm_Vis_manual_entry_frame, text="RESET", bg=conf["widget_bg"], fg=conf["header_fg"], bd=0, highlightthickness=0, activebackground=conf["header_fg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["out_button_sel_fg"], command=Admin_VIS_RESET)
adm_Vis_manual_remove_button.pack(fill=tk.X, side=tk.BOTTOM, padx=5, pady=5, ipady=10)
adm_Vis_manual_remove_button = tk.Button(adm_Vis_manual_entry_frame, text="REMOVE", bg=conf["widget_bg"], fg=conf["out_button_fg"], bd=0, highlightthickness=0, activebackground=conf["out_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["out_button_sel_fg"], command=adm_Vis_remove)
adm_Vis_manual_remove_button.pack(fill=tk.X, side=tk.BOTTOM, padx=5, pady=5, ipady=10)
adm_Vis_manual_add_button = tk.Button(adm_Vis_manual_entry_frame, text="ADD", bg=conf["widget_bg"], fg=conf["in_button_fg"], bd=0, highlightthickness=0, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=adm_Vis_add)
adm_Vis_manual_add_button.pack(fill=tk.X, side=tk.BOTTOM, padx=5, pady=5, ipady=10)

#Labels for central list
adm_VIS_Plates_Lb = tk.Label(adm_sign_central_frame, text="PLATE:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=8)
adm_VIS_Plates_Lb.pack(fill=tk.X, side=tk.LEFT, expand=1)
adm_VIS_Car_Lb = tk.Label(adm_sign_central_frame, text="CAR:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=12)
adm_VIS_Car_Lb.pack(fill=tk.X, side=tk.LEFT, padx=1, expand=1)
adm_VIS_Name_Lb= tk.Label(adm_sign_central_frame, text="NAME:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=12)
adm_VIS_Name_Lb.pack(fill=tk.X, side=tk.LEFT, expand=1)
adm_VIS_Exp_Lb= tk.Label(adm_sign_central_frame, text="EXPIRE:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"], width=10)
adm_VIS_Exp_Lb.pack(fill=tk.X, side=tk.LEFT, padx=(1, 0),expand=1)

Implement(Admin_Vis_Scroll.frame, "company", "Admin_Vis_Co", 16, None)



#################################################################################
# Admin GN
#################################################################################
Admin_GN_Frame = tk.Frame(adm_main_frame, bg=conf["window_bg"], highlightthickness=0)
Admin_GN_Frame.pack_propagate(0)

#############@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

# Admin GN Frame for data
adm_GN_LEFT_fr = tk.Frame(Admin_GN_Frame, bg=conf["window_bg"], highlightthickness=0)
adm_GN_LEFT_fr.pack_propagate(0)
adm_GN_LEFT_fr.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, padx=(0, 5))


#Right GN Frame for buttons
adm_GN_right_fr = tk.Frame(Admin_GN_Frame, highlightthickness=3, relief=tk.RAISED, bg=conf["submenu_bg"], highlightbackground=conf["submenu_sel_bg"], width=conf["chk_filter_frame"])
adm_GN_right_fr.pack_propagate(0)
adm_GN_right_fr.pack(fill=tk.BOTH, side=tk.RIGHT)

#grid configuration for T, Tr, Fb
adm_GN_LEFT_fr.rowconfigure(0, weight=1)
adm_GN_LEFT_fr.rowconfigure(1, weight=2)
adm_GN_LEFT_fr.rowconfigure(2, weight=2)
adm_GN_LEFT_fr.columnconfigure(0, weight=1)

#frames for T, Tr, Fb
adm_GN_T_fr = tk.Frame(adm_GN_LEFT_fr, bg=conf["window_bg"], highlightthickness=0)
adm_GN_T_fr.grid(row=0, column=0, sticky=tk.NSEW, pady=(0, 5))
adm_GN_Tr_fr = tk.Frame(adm_GN_LEFT_fr, bg=conf["window_bg"], highlightthickness=0)
adm_GN_Tr_fr.grid(row=1, column=0, sticky=tk.NSEW, pady=(0, 5))
adm_GN_Fb_fr = tk.Frame(adm_GN_LEFT_fr, bg=conf["window_bg"], highlightthickness=0)
adm_GN_Fb_fr.grid(row=2, column=0, sticky=tk.NSEW)

#LABEL FOR T, Tr, Fb
adm_GN_T_Lb = tk.Label(adm_GN_T_fr, text="TRUCKS:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
adm_GN_T_Lb.pack(fill=tk.X, side=tk.TOP)
adm_GN_Tr_Lb = tk.Label(adm_GN_Tr_fr, text="TRAILERS:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
adm_GN_Tr_Lb.pack(fill=tk.X, side=tk.TOP)
adm_GN_Fb_Lb = tk.Label(adm_GN_Fb_fr, text="FB:", bg=conf["header_bg"], font=(conf["header_font"], conf["header_size"]), fg=conf["header_fg"])
adm_GN_Fb_Lb.pack(fill=tk.X, side=tk.TOP)

#Scroll Frame for Trucks
Admin_GN_T_Scroll = scroller(adm_GN_T_fr)
Admin_GN_T_Scroll.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
Admin_GN_T_Scroll.refresh()

#Scroll Frame for Trailers
Admin_GN_Tr_Scroll = scroller(adm_GN_Tr_fr)
Admin_GN_Tr_Scroll.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
Admin_GN_Tr_Scroll.refresh()

#Scroll Frame for Fb
Admin_GN_Fb_Scroll = scroller(adm_GN_Fb_fr)
Admin_GN_Fb_Scroll.pack(fill=tk.BOTH, side=tk.LEFT, expand=1, anchor=tk.N)
Admin_GN_Fb_Scroll.refresh()

def adm_GN_entry_focus(*args):
    global Adm_GN_Truck_Var
    global Adm_GN_Truck_obj
    global Adm_GN_Trailer_Var
    global Adm_GN_Trailer_obj
    global Adm_GN_Fb_Var
    global Adm_GN_Fb_obj

    if Adm_GN_Truck_obj is not None:
        Adm_GN_Truck_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        Adm_GN_Truck_Var = None
        Adm_GN_Truck_obj = None
    if Adm_GN_Trailer_obj is not None:
        Adm_GN_Trailer_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        Adm_GN_Trailer_Var = None
        Adm_GN_Trailer_obj = None
    if Adm_GN_Fb_obj is not None:
        Adm_GN_Fb_obj.config(bg=conf["widget_bg"], fg=conf["widget_fg"])
        Adm_GN_Fb_Var = None
        Adm_GN_Fb_obj = None
    adm_GN_storage_var.set(0)
    adm_GN_LU_var.set(0)

def adm_GN_add(*args):
    opt = adm_GN_radio_var.get()
    unit = adm_GN_t_entry.get().strip()
    if unit is None or unit == "": return
    if opt == "truck":
        table = "dbo.GN_Trucks"
        column = "truck_number"
    elif opt == "trailer":
        table = "dbo.GN_Trailers"
        column = "trailer_number"
    elif opt == "flatbed":
        table = "dbo.GN_Flatbed"
        column = "fb_number"
    check = SQL_REQ(f"SELECT * FROM {table} WHERE {column}=?", (unit,), "S_one")
    if check:
        error(12)
        return
    else:
        if opt == "trailer" or opt == "flatbed":
            st = adm_GN_storage_var.get()
            lu = adm_GN_LU_var.get()
            query = f"INSERT INTO {table} ({column}, status, storage, LU) VALUES (?, ?, ?, ?)"
            values = (unit, "0", st, lu)
        else:
            query = f"INSERT INTO {table} ({column}, status) VALUES (?, ?)"
            values = (unit, "0")
        SQL_REQ(query, values, "W")
    adm_GN_t_entry.delete(0, tk.END)
    adm_GN_t_entry.focus_set()
    Refresh("adm_GN")

#
def adm_GN_remove(*args):
    if Adm_GN_Truck_Var is None and Adm_GN_Trailer_Var is None and Adm_GN_Fb_Var is None: return
    if Adm_GN_Truck_Var is not None:
        check = SQL_REQ("SELECT * FROM dbo.GN_Trucks WHERE truck_number=?", (Adm_GN_Truck_Var[0],), "S_one")
        if check: SQL_REQ("DELETE FROM dbo.GN_Trucks WHERE truck_number=?", (Adm_GN_Truck_Var[0],), "W")
    elif Adm_GN_Trailer_Var is not None:
        check = SQL_REQ("SELECT * FROM dbo.GN_Trailers WHERE trailer_number=?", (Adm_GN_Trailer_Var[0],), "S_one")
        if check: SQL_REQ("DELETE FROM dbo.GN_Trailers WHERE trailer_number=?", (Adm_GN_Trailer_Var[0],), "W")
    elif Adm_GN_Fb_Var is not None:
        check = SQL_REQ("SELECT * FROM dbo.GN_Flatbed WHERE fb_number=?", (Adm_GN_Fb_Var[0],), "S_one")
        if check: SQL_REQ("DELETE FROM dbo.GN_Flatbed WHERE fb_number=?", (Adm_GN_Fb_Var[0],), "W")
    else:
        error(11)
    adm_GN_t_entry.focus_set()
    Refresh("adm_GN")

def adm_GN_storage_check(*args):
    global adm_GN_storage_var
    str_status = adm_GN_storage_var.get()
    if Adm_GN_Trailer_Var is not None:
        SQL_REQ("UPDATE dbo.GN_Trailers SET storage=? WHERE trailer_number=?", (str(str_status), Adm_GN_Trailer_Var[0]), "W")
        Adm_GN_Trailer_Var[1][1] = str_status
    elif Adm_GN_Fb_Var is not None:
        SQL_REQ("UPDATE dbo.GN_Flatbed SET storage=? WHERE fb_number=?", (str(str_status), Adm_GN_Fb_Var[0]), "W")
        Adm_GN_Fb_Var[1][1] = str_status
    else: return


def adm_GN_LU_check(*args):
    global adm_GN_LU_var
    str_status = adm_GN_LU_var.get()
    if Adm_GN_Trailer_Var is not None:
        SQL_REQ("UPDATE dbo.GN_Trailers SET LU=? WHERE trailer_number=?", (str(str_status), Adm_GN_Trailer_Var[0]), "W")
        Adm_GN_Trailer_Var[1][2] = str_status
    elif Adm_GN_Fb_Var is not None:
        SQL_REQ("UPDATE dbo.GN_Flatbed SET LU=? WHERE fb_number=?", (str(str_status), Adm_GN_Fb_Var[0]), "W")
        Adm_GN_Fb_Var[1][2] = str_status
    else:
        return
def adm_GN_city_check(*args):
    global adm_GN_city_var
    city_status = adm_GN_city_var.get()
    if Adm_GN_Truck_Var is not None:
        SQL_REQ("UPDATE dbo.GN_Trucks SET city=? WHERE truck_number=?", (str(city_status), Adm_GN_Truck_Var[0]), "W")
    else: return


def adm_GN_radio_T(*args):
    adm_GN_storage_checkbox.pack_forget()
    adm_GN_LU_checkbox.pack_forget()

def adm_GN_radio_Tr(*args):
    adm_GN_storage_checkbox.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=15)
    adm_GN_LU_checkbox.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=15)

# #Button Frame for Adding/Delete Units
adm_GN_manual_entry_frame = tk.Frame(adm_GN_right_fr, highlightthickness=0, bg=conf["submenu_bg"])
adm_GN_manual_entry_frame.pack(side=tk.TOP)

adm_GN_t_lb = tk.Label(adm_GN_manual_entry_frame, text="UNIT:", relief=tk.GROOVE, bg=conf["submenu_bg"], fg=conf["submenu_fg"],font=(conf["submenu_font"], conf["submenu_size"]))
adm_GN_t_lb.pack(fill=tk.X, side=tk.TOP)
adm_GN_t_entry = tk.Entry(adm_GN_manual_entry_frame, bg=conf["window_bg"], bd=1, cursor="shuttle", font=(conf["entry_font"], conf["entry_size"]), relief=tk.SUNKEN, fg=conf["entry_fg"], highlightbackground=conf["entry_unsel_frame"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, justify=tk.LEFT)
adm_GN_t_entry.pack(fill=tk.X, side=tk.TOP)
adm_GN_t_entry.bind("<Return>", adm_GN_add)
adm_GN_t_entry.bind("<FocusIn>", adm_GN_entry_focus)
adm_GN_t_entry.bind("<Button-1>", adm_GN_entry_focus)
adm_GN_radio_var = tk.StringVar()
adm_GN_truck_radio = tk.Radiobutton(adm_GN_manual_entry_frame, text="Truck", command=adm_GN_radio_T, variable=adm_GN_radio_var, value="truck", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
adm_GN_truck_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
adm_GN_trailer_radio = tk.Radiobutton(adm_GN_manual_entry_frame, text="Trailer", command=adm_GN_radio_Tr, variable=adm_GN_radio_var, value="trailer", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
adm_GN_trailer_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
adm_GN_fb_radio = tk.Radiobutton(adm_GN_manual_entry_frame, text="Flatbed", command=adm_GN_radio_Tr, variable=adm_GN_radio_var, value="flatbed", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
adm_GN_fb_radio.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=10)
adm_GN_radio_var.set("truck")
adm_GN_manual_add_button = tk.Button(adm_GN_manual_entry_frame, text="ADD", bg=conf["widget_bg"], fg=conf["in_button_fg"], bd=0, highlightthickness=0, activebackground=conf["in_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["in_button_sel_fg"], command=adm_GN_add)
adm_GN_manual_add_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
adm_GN_manual_remove_button = tk.Button(adm_GN_manual_entry_frame, text="REMOVE", bg=conf["widget_bg"], fg=conf["out_button_fg"], bd=0, highlightthickness=0, activebackground=conf["out_button_sel_bg"], font=(conf["submenu_font"], conf["header_size"]), activeforeground=conf["out_button_sel_fg"], command=adm_GN_remove)
adm_GN_manual_remove_button.pack(fill=tk.X, side=tk.TOP, padx=5, pady=5, ipady=10)
adm_GN_storage_var = tk.IntVar()
adm_GN_storage_checkbox = tk.Checkbutton(adm_GN_manual_entry_frame, text="Storage", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], variable=adm_GN_storage_var, command=adm_GN_storage_check)
adm_GN_LU_var = tk.IntVar()
adm_GN_LU_checkbox = tk.Checkbutton(adm_GN_manual_entry_frame, text="Loaded", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], variable=adm_GN_LU_var, command=adm_GN_LU_check)
adm_GN_city_var = tk.IntVar()
adm_GN_city_checkbox = tk.Checkbutton(adm_GN_manual_entry_frame, text="City", bg=conf["submenu_bg"], font=(conf["submenu_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], variable=adm_GN_city_var, command=adm_GN_city_check)

adm_gn_scrn_size = 10

# Implement(Admin_GN_T_Scroll.frame, "GNtrucks", "Admin_GN_Truck", 13, adm_gn_scrn_size)
# Implement(Admin_GN_Tr_Scroll.frame, "GNtrailers", "Admin_GN_Trailer", 14, adm_gn_scrn_size)
# Implement(Admin_GN_Fb_Scroll.frame, "GNfb", "Admin_GN_Flatbed", 15, adm_gn_scrn_size)

##############@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

#################################################################################
# Admin Accounts
#################################################################################
Admin_Account_Frame = tk.Frame(adm_main_frame, bg=conf["window_bg"], highlightthickness=0)
Admin_Account_Frame.pack_propagate(0)

#getting list of accounts with info
account_list = sorted((list(val) for val in SQL_REQ("SELECT * FROM dbo.authentication", (), "S_all")), key=lambda x: x[0])

##############################################################################################################################################################################################
def admin_company_insert(*args):
    global today
    global Admin_Company_Quantity_Var
    global Admin_Company_Quantity_D_Var
    global Admin_Company_Quantity_R_Var
    global Admin_Company_Quantity_T_Var
    global Admin_Company_Quantity_Tr_Var
    global Admin_Company_Quantity_V_Var
    #entry creator in 2D by provided list
    def create_entry_grid(masta, data):
        rows = len(data)
        cols = len(data[0])
        # Create a 2D list to store the Entry widgets
        entries = [[None for _ in range(cols)] for _ in range(rows)]
        #inserting data in entries
        for i in range(rows):
            if data[i][7]: color = conf["widget_fg"]
            elif not data[i][7]: color = conf["widget_sel_bg"]
            else: error(6)
            for j in range(cols):
                entry = tk.Entry(masta, bg=conf["widget_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=color, justify=tk.CENTER, disabledbackground=conf["widget_bg"], disabledforeground=color)
                if j == 0: entry.config(width=10)
                elif (2<=j<=7): entry.config(width=20)
                elif j == 8: entry.config(width=29)
                elif j == 1: entry.config(width=adm_comp_name_lb.winfo_reqwidth())
                if data[i][j] is None or data[i][j] == "":
                    value = ""
                else:
                    if j == 8 and color == conf["widget_fg"]:
                        if int((data[i][j]-today.date()).days)<0: entry.config(fg=conf["expired_date"])
                        elif int((data[i][j]-today.date()).days)<30: entry.config(fg=conf["on_parking"])
                    value = data[i][j]
                entry.insert(0, value)
                entry.grid(row=i, column=j, sticky=tk.E, padx=(0, 1), pady=(1, 0))
                entry.update_idletasks()
                if j == 0: entry.config(state=tk.DISABLED)
                else: entry.config(state=tk.NORMAL)
                entries[i][j] = entry
        return entries

    Admin_Company_Quantity_Var = 0
    Admin_Company_Quantity_D_Var = 0
    Admin_Company_Quantity_R_Var = 0
    Admin_Company_Quantity_T_Var = 0
    Admin_Company_Quantity_Tr_Var = 0
    Admin_Company_Quantity_V_Var = 0
    if len(args)>1:
        adm_comp_list = args[1]
    else:
        raw = SQL_REQ("SELECT * FROM dbo.Company_list ORDER BY company_name", (), "S_all")
        adm_comp_list = list()
        for line in raw: adm_comp_list.append(list(line))
    for line in adm_comp_list:
        if line[7]:
            Admin_Company_Quantity_Var+=1
            if line[2] is not None and line[2] != "": Admin_Company_Quantity_D_Var += int(line[2])
            if line[3] is not None and line[3] != "": Admin_Company_Quantity_R_Var += int(line[3])
            if line[5] is not None and line[5] != "": Admin_Company_Quantity_T_Var += int(line[5])
            if line[4] is not None and line[4] != "": Admin_Company_Quantity_Tr_Var += int(line[4])
            if line[6] is not None and line[6] != "": Admin_Company_Quantity_V_Var += int(line[6])

    Admin_Company_Info_num_lb.config(text=Admin_Company_Quantity_Var)
    Admin_Designated_num_Info_num_lb.config(text=Admin_Company_Quantity_D_Var)
    Admin_Regular_num_Info_num_lb.config(text=Admin_Company_Quantity_R_Var)
    Admin_Truck_num_Info_num_lb.config(text=Admin_Company_Quantity_T_Var)
    Admin_Trailer_num_Info_num_lb.config(text=Admin_Company_Quantity_Tr_Var)
    Admin_Vehicle_num_Info_num_lb.config(text=Admin_Company_Quantity_V_Var)
    return create_entry_grid(args[0], adm_comp_list)
def admin_company_register(*args):
    raw = SQL_REQ("SELECT * FROM dbo.Company_list ORDER BY company_name", (), "S_all")
    comp_list = list()
    for line in raw: comp_list.append(list(line))
    new_record = [
        Admin_Company_Add_Company_Entry_name.get().strip(),
        Admin_Company_Add_Company_Entry_des.get().strip(),
        Admin_Company_Add_Company_Entry_reg.get().strip(),
        Admin_Company_Add_Company_Entry_tr.get().strip(),
        Admin_Company_Add_Company_Entry_t.get().strip(),
        Admin_Company_Add_Company_Entry_v.get().strip(),
        Admin_Company_Add_Company_Entry_a.get().strip(),
        Admin_Company_Add_Company_Entry_i.get().strip()
    ]

    for i in range(len(new_record)):
        if new_record[i] == "": new_record[i] = None
        if 1<= i <= 5:
            if new_record[i] !="" and new_record[i] is not None:
                try:
                    int(new_record[i])
                except:
                    error(5)
                    return
        if i == 6:
            if new_record[i] is None or new_record[i] == "":
                error(6)
                return
        if i ==7:
            try:
                datetime.strptime(new_record[i], "%Y-%m-%d")
            except:
                error(7)
                return

    ID = []
    NEW_ID = None
#check if company name exist and popup error
    for line in comp_list:
        if comp_list[1] == new_record:
            error(8)
            return
        ID.append(int(line[0]))
#check first available ID
    for i in range(1, 1000):
        if i not in ID:
            NEW_ID = i
            break
    new_record.insert(0, str(NEW_ID))
    SQL_REQ("INSERT INTO dbo.Company_List(company_ID, company_name, designated, regular, trailer, truck, car, activity, insurance) VALUES (?,?,?,?,?,?,?,?,?)", new_record, "W")
    admin_company_cancel()
    admin_company_insert(Admin_Company_Scroll.frame)

def admin_company_commit_changes(entries):
    global Admin_Company_Entries
    modified_list = []
    for row in entries:
        modified_row = []
        for entry in row:
            value = entry.get().strip()
            modified_row.append(value)
        if modified_row[8] is not None and modified_row[8] != "":
            try:
                modified_row[8] = datetime.strptime(modified_row[8], "%Y-%m-%d").date()
            except:
                error(7)
                return
        if modified_row[7] == "1": modified_row[7] = True
        elif modified_row[7] == "0": modified_row[7] = False
        else:
            error(6)
            return
        for i in range(len(modified_row)):
            if 2<= i <=6:
                if modified_row[i] is not None and modified_row[i] != "":
                    try: int(modified_row[i])
                    except:
                        error(5)
                        return

        modified_list.append(modified_row)
    for widgets in Admin_Company_Scroll.frame.winfo_children(): widgets.destroy()
    # update SQL

    for line in modified_list:
        line = [value if value != "" else None for value in line]
        SQL_REQ("UPDATE dbo.Company_List SET company_name=?, designated=?, regular=?, trailer=?, truck=?, car=?, activity=?, insurance=? WHERE company_ID=?", (line[1], line[2], line[3], line[4], line[5], line[6], line[7], line[8] , line[0]), "W")


    Admin_Company_Entries = admin_company_insert(Admin_Company_Scroll.frame, modified_list)

def admin_company_add():
    Admin_Info_Frame.pack_forget()
    Admin_Company_Add_Button.pack_forget()
    Admin_Company_Cancel_Button.pack(side=tk.LEFT, fill=tk.Y, padx=2, pady=2)
    Admin_Company_Save_Button.config(command=admin_company_register, fg=conf["widget_sel_fg"])
    Admin_Company_Add_Entry_fr.pack(fill=tk.X, side=tk.LEFT, expand=1)


def admin_company_cancel():
    Admin_Company_Cancel_Button.pack_forget()
    Admin_Company_Add_Entry_fr.pack_forget()
    Admin_Company_Save_Button.config(command=lambda:admin_company_commit_changes(Admin_Company_Entries), fg=conf["submenu_fg"])
    Admin_Company_Add_Button.pack(side=tk.LEFT, fill=tk.Y, padx=2, pady=2)
    Admin_Info_Frame.pack(fill=tk.X, side=tk.LEFT, expand=1)


#######################################
# Accounts
#######################################
def adm_account_change(event, column_name, id, obj):
    if column_name == "activity":
        stat = int(obj.get())
        SQL_REQ(f"UPDATE dbo.authentication SET {column_name}=? WHERE ID=?", (str(stat), str(id)), "W")
    elif column_name == "password":
        pass_class = PasswordDatabase()
        new_pass = pass_class.hash_password(obj.get()).decode("utf-8")
        SQL_REQ(f"UPDATE dbo.authentication SET {column_name}=? WHERE ID=?", (new_pass, str(id)), "W")
        adm_account_display()
    else:
        SQL_REQ(f"UPDATE dbo.authentication SET {column_name}=? WHERE ID=?", (obj.get(), str(id)), "W")
        adm_account_display()
def adm_account_add_button(*args):
    Admin_Account_Add_Button.pack_forget()
    Admin_Account_Save_Button.pack(side=tk.LEFT, fill=tk.Y, padx=2, pady=2)
    Admin_Account_Cancel_Button.pack(side=tk.LEFT, fill=tk.Y, padx=2, pady=2)
    Admin_Account_Info_Frame.pack_forget()
    Admin_Account_Manual_Frame.pack(fill=tk.BOTH, side=tk.LEFT)

def adm_account_add_cancel(*args):
    global activity_check
    Admin_Account_Cancel_Button.pack_forget()
    Admin_Account_Save_Button.pack_forget()
    Admin_Account_Manual_Frame.pack_forget()
    Admin_Account_Add_Button.pack(side=tk.LEFT, fill=tk.Y, padx=2, pady=2)
    Admin_Account_Info_Frame.pack(fill=tk.BOTH, side=tk.LEFT)
    Admin_Account_Add_log_entry.delete(0, tk.END)
    Admin_Account_Add_pas_entry.delete(0, tk.END)
    Admin_Account_Add_name_entry.delete(0, tk.END)
    Admin_Account_Add_rights_entry.delete(0, tk.END)
    activity_check.set(0)

def adm_account_add(*args):
    login = Admin_Account_Add_log_entry.get().strip()
    pas = Admin_Account_Add_pas_entry.get().strip()
    name = Admin_Account_Add_name_entry.get().strip()
    rights = Admin_Account_Add_rights_entry.get().strip()
    act = activity_check.get()
    if login is None or pas is None or name is None or rights is None:
        error(5)
        return
    try: int(rights)
    except ValueError:
        error(5)
        return
    account_add = PasswordDatabase()
    check = SQL_REQ("SELECT login FROM dbo.authentication WHERE login=?", (login,), "S_one")
    if check is None:
        try:
            account_add.register(login, pas, name, rights, int(act))
        except:
            error(13)
            debuger("Error in >> eccount_add.register(login=\'"+login+", pas=\'"+pas+", name=\'"+rights+", act=\'"+str(act))
            return
    else:
        error(12)
        return
    adm_account_display()
    adm_account_add_cancel()
def adm_account_display():
    def adm_entry_creator(record, counter):
        adm_acc_log_entry = tk.Entry(Admin_Account_Table_Frame, bg=conf["widget_bg"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, highlightbackground=conf["widget_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.CENTER, width=20)
        adm_acc_log_entry.grid(row=counter, column=0, sticky=tk.NSEW, padx=(0, 1), pady=(1, 0))
        adm_acc_log_entry.insert(0, record["login"])
        adm_acc_log_entry.bind("<Return>", lambda event, column_name="login", id=record["ID"], obj=adm_acc_log_entry: adm_account_change(event, column_name, id, obj))
        adm_acc_pas_entry = tk.Entry(Admin_Account_Table_Frame, bg=conf["widget_bg"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, highlightbackground=conf["widget_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.CENTER, width=50)
        adm_acc_pas_entry.grid(row=counter, column=1, sticky=tk.NSEW, padx=(0, 1), pady=(1, 0))
        adm_acc_pas_entry.insert(0, record["password"])
        adm_acc_pas_entry.bind("<Return>", lambda event, column_name="password", id=record["ID"], obj=adm_acc_pas_entry: adm_account_change(event, column_name, id, obj))
        adm_acc_name_entry = tk.Entry(Admin_Account_Table_Frame, bg=conf["widget_bg"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, highlightbackground=conf["widget_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.CENTER, width=20)
        adm_acc_name_entry.grid(row=counter, column=2, sticky=tk.NSEW, padx=(0, 1), pady=(1, 0))
        adm_acc_name_entry.insert(0, record["full_name"])
        adm_acc_name_entry.bind("<Return>", lambda event, column_name="full_name", id=record["ID"], obj=adm_acc_name_entry: adm_account_change(event, column_name, id, obj))
        adm_acc_rights_entry = tk.Entry(Admin_Account_Table_Frame, bg=conf["widget_bg"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, highlightbackground=conf["widget_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.CENTER, width=20)
        adm_acc_rights_entry.grid(row=counter, column=3, sticky=tk.NSEW, padx=(0, 1), pady=(1, 0))
        adm_acc_rights_entry.insert(0, record["rights"])
        adm_acc_rights_entry.bind("<Return>", lambda event, column_name="rights", id=record["ID"], obj=adm_acc_rights_entry: adm_account_change(event, column_name, id, obj))
        activity = tk.BooleanVar()
        adm_acc_activity_checkbx = tk.Checkbutton(Admin_Account_Table_Frame, bg=conf["widget_bg"], onvalue=True, offvalue=False, variable=activity, justify=tk.CENTER, width=20, command=lambda: adm_account_change(None, "activity", record["ID"], activity))
        adm_acc_activity_checkbx.grid(row=counter, column=4, sticky=tk.NSEW, padx=(0, 1), pady=(1, 0))
        if record["activity"]: activity.set(1)
        else: activity.set(0)

    for widget in Admin_Account_Table_Frame.winfo_children(): widget.destroy()
    adm_acc_log_lb = tk.Label(Admin_Account_Table_Frame, text="Login", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=15)
    adm_acc_log_lb.grid(row=0, column=0, sticky=tk.EW, padx=(0, 1))
    adm_acc_pas_lb = tk.Label(Admin_Account_Table_Frame, text="Password", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=35)
    adm_acc_pas_lb.grid(row=0, column=1, sticky=tk.EW, padx=(0, 1))
    adm_acc_name_lb = tk.Label(Admin_Account_Table_Frame, text="Name", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=20)
    adm_acc_name_lb.grid(row=0, column=2, sticky=tk.EW, padx=(0, 1))
    adm_acc_rights_lb = tk.Label(Admin_Account_Table_Frame, text="Rights", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=10)
    adm_acc_rights_lb.grid(row=0, column=3, sticky=tk.EW, padx=(0, 1))
    adm_acc_activity_lb = tk.Label(Admin_Account_Table_Frame, text="Activity", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"], width=10)
    adm_acc_activity_lb.grid(row=0, column=4, sticky=tk.EW, padx=(0, 1))
    query = SQL_REQ("SELECT * FROM dbo.authentication", (), "S_all_D")
    acc_list = list()
    if query:
        column_name = [column[0] for column in query[1]]
        for record in query[0]:
            rec = {column: value for column, value in zip(column_name, record)}
            acc_list.append(rec)
    else: return
    counter = 1
    for record in acc_list:
        adm_entry_creator(record, counter)
        counter += 1

Admin_Account_Table_Frame = tk.Frame(Admin_Account_Frame, bg=conf["window_bg"], highlightthickness=0, width=600)
Admin_Account_Table_Frame.pack(side=tk.TOP)
adm_account_display()
#Frame for accounts list
Admin_Account_List_Frame = tk.Frame(Admin_Account_Frame, highlightthickness=0, relief=tk.RAISED, bg=conf["window_bg"])
Admin_Account_List_Frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1)
#Frame for Buttons and Entry for edit
Admin_Account_Buttons_Frame = tk.Frame(Admin_Account_Frame, highlightthickness=3, relief=tk.RAISED, bg=conf["submenu_bg"], highlightbackground=conf["submenu_sel_bg"], height=conf["admin_company_button_height"])
Admin_Account_Buttons_Frame.pack(fill=tk.X, side=tk.BOTTOM)



Admin_Account_Refresh_Button = tk.Button(Admin_Account_Buttons_Frame, text="REFRESH", command=adm_account_display, width=20, fg=conf["submenu_fg"])
Admin_Account_Refresh_Button.pack(side=tk.LEFT, fill=tk.Y,  padx=2, pady=2)
Admin_Account_Add_Button = tk.Button(Admin_Account_Buttons_Frame, text="ADD", command=adm_account_add_button, width=20, fg=conf["submenu_fg"])
Admin_Account_Add_Button.pack(side=tk.LEFT, fill=tk.Y,  padx=2, pady=2)
Admin_Account_Save_Button = tk.Button(Admin_Account_Buttons_Frame, text="SAVE", command=adm_account_add, width=20, fg=conf["submenu_fg"])
Admin_Account_Cancel_Button = tk.Button(Admin_Account_Buttons_Frame, text="Cancel", command=adm_account_add_cancel, width=20, fg=conf["submenu_fg"])
#frame and entries for new account
Admin_Account_Manual_Frame = tk.Frame(Admin_Account_Buttons_Frame, highlightthickness=0, bg=conf["submenu_bg"])
Admin_Account_Add_log_lb = tk.Label(Admin_Account_Manual_Frame, text="Login:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Account_Add_log_lb.pack(side=tk.LEFT)
Admin_Account_Add_log_entry = tk.Entry(Admin_Account_Manual_Frame, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=20)
Admin_Account_Add_log_entry .pack(side=tk.LEFT)
Admin_Account_Add_pas_lb = tk.Label(Admin_Account_Manual_Frame, text="Password:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Account_Add_pas_lb.pack(side=tk.LEFT)
Admin_Account_Add_pas_entry = tk.Entry(Admin_Account_Manual_Frame, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=20)
Admin_Account_Add_pas_entry .pack(side=tk.LEFT)
Admin_Account_Add_name_lb = tk.Label(Admin_Account_Manual_Frame, text="Name:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Account_Add_name_lb.pack(side=tk.LEFT)
Admin_Account_Add_name_entry = tk.Entry(Admin_Account_Manual_Frame, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=20)
Admin_Account_Add_name_entry .pack(side=tk.LEFT)
Admin_Account_Add_rights_lb = tk.Label(Admin_Account_Manual_Frame, text="Rights:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Account_Add_rights_lb.pack(side=tk.LEFT)
Admin_Account_Add_rights_entry = tk.Entry(Admin_Account_Manual_Frame, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=3)
Admin_Account_Add_rights_entry .pack(side=tk.LEFT)
Admin_Account_Add_act_lb = tk.Label(Admin_Account_Manual_Frame, text="Activity:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Account_Add_act_lb.pack(side=tk.LEFT)
activity_check = tk.BooleanVar()
Admin_Account_Add_act_checkbtn = tk.Checkbutton(Admin_Account_Manual_Frame, bg=conf["submenu_bg"], onvalue=True, offvalue=False, variable=activity_check, justify=tk.CENTER)
Admin_Account_Add_act_checkbtn.pack(side=tk.LEFT, padx=2)


Admin_Account_Info_Frame = tk.Frame(Admin_Account_Buttons_Frame, highlightthickness=0, bg=conf["submenu_bg"])
Admin_Account_Info_Frame.pack(fill=tk.Y, side=tk.LEFT)

#rights explanation
rights_info = "Rights: 1-All, 2-GN, 3-Parking+History"

Admin_Account_Info_lb = tk.Label(Admin_Account_Info_Frame, text=rights_info, bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"])
Admin_Account_Info_lb.pack(side=tk.LEFT)

#################################################################################
# Admin Vendors
#################################################################################
def Adm_Ven_Insert(masta):
    global Admin_Vendor_MaxID
    row_num=0
    def adm_ven_select(event, ven_name,id, obj):
        global Admin_Vendor_obj
        global Admin_Vendor_Var
        if Admin_Vendor_obj is not None: Admin_Vendor_obj.config(bg=conf["widget_bg"])
        Admin_Vendor_Var = [id, ven_name]
        Admin_Vendor_obj = obj
        Admin_Ven_Add_entry.delete(0, tk.END)
        Admin_Ven_Add_entry.insert(0, ven_name)

    def adm_ven_creator(id, ven, row_num):
        global Admin_Vendor_Var
        adm_ven_unit = tk.Label(masta.frame, text=ven, bg=conf["widget_bg"], highlightcolor=conf["entry_sel_frame"], highlightthickness=1, highlightbackground=conf["widget_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=14)
        adm_ven_unit.grid(row=row_num, column=0, sticky=tk.NSEW, padx=(0, 1), pady=(0, 1))
        adm_ven_unit.bind("<Button-1>", lambda event, ven_name=ven, id=id, obj=adm_ven_unit: adm_ven_select(event, ven_name, id, obj))
        adm_ven_unit.bind("<Enter>", lambda event: adm_ven_unit.config(bg=conf["widget_sel_bg"]))
        adm_ven_unit.bind("<Leave>", lambda event: adm_ven_unit.config(bg=conf["widget_bg"]) if (Admin_Vendor_Var is None or Admin_Vendor_Var[1] != ven) else None)
        print(ven)

    masta.delete()
    masta.refresh()
    query = SQL_REQ("SELECT * FROM dbo.Car_Vendors ORDER BY Vendor", (), "S_all")
    if query:
        ven_list = {k:v for k, v in query}
    else:
        error(15)
        return
    for row_num, (id, vendor) in enumerate(ven_list.items()): adm_ven_creator(id, vendor, row_num)
    Admin_Vendor_MaxID = max(ven_list.keys())
    masta.refresh()
    Admin_Ven_Add_entry.focus_set()

def adm_ven_add_button(*args):
    global Admin_Vendor_MaxID
    global Admin_Vendor_Var
    new_ven = Admin_Ven_Add_entry.get().strip()
    if new_ven == "": return
    else:
        if Admin_Vendor_Var is None:
            check = SQL_REQ("SELECT * FROM dbo.Car_vendors WHERE Vendor=?", (new_ven,), "S_one")
            if not check:
                new_rec = [Admin_Vendor_MaxID+1, new_ven]
                SQL_REQ("INSERT INTO dbo.Car_Vendors (ID, Vendor) VALUES (?,?)", new_rec, "W")
                adm_ven_refresh()
            else:
                error("Vendor's name already exists!")
                return
        else:
            SQL_REQ("UPDATE dbo.Car_Vendors SET Vendor=? WHERE ID=?", (new_ven, Admin_Vendor_Var[0]), "W")
            adm_ven_refresh()
def adm_ven_refresh():
    global Admin_Vendor_obj
    global Admin_Vendor_Var
    Admin_Ven_Add_entry.delete(0, tk.END)
    if Admin_Vendor_obj is not None:
        Admin_Vendor_obj.config(bg=conf["widget_bg"])
        Admin_Vendor_obj = None
        Admin_Vendor_Var = None
    Adm_Ven_Insert(Admin_Ven_List_Frame)
def adm_ven_delete_button():
    ven = Admin_Ven_Add_entry.get().strip()
    check = SQL_REQ("SELECT Vendor FROM dbo.Car_Vendors WHERE Vendor=?", (ven,), "S_one")
    if check:
        SQL_REQ("DELETE FROM dbo.Car_Vendors WHERE Vendor=?", (ven,), "W")
        adm_ven_refresh()
    else:
        error(11)
        adm_ven_refresh()

Admin_Ven_Frame = tk.Frame(adm_main_frame, bg=conf["window_bg"], highlightthickness=0)
Admin_Ven_Frame.pack_propagate(0)



Admin_Ven_Table_Frame = tk.Frame(Admin_Ven_Frame, bg=conf["window_bg"], highlightthickness=0, width=145)
Admin_Ven_Table_Frame.pack(side=tk.TOP, fill=tk.Y, expand=1)
Admin_Ven_Table_Frame.pack_propagate(0)


Adm_Ven_lb = tk.Label(Admin_Ven_Table_Frame, text="Vendor:", bg=conf["header_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["header_fg"])
Adm_Ven_lb.pack(side=tk.TOP, fill=tk.BOTH, pady=(0, 1))

#Scrollable Frame for vendor list
Admin_Ven_List_Frame = scroller(Admin_Ven_Table_Frame)
#Admin_Ven_List_Frame = tk.Frame(Admin_Ven_Frame, highlightthickness=0, relief=tk.RAISED, bg=conf["window_bg"])
Admin_Ven_List_Frame.pack(fill=tk.BOTH, side=tk.TOP, expand=1)

#Frame for Buttons and Entry for edit
Admin_Ven_Buttons_Frame = tk.Frame(Admin_Ven_Frame, highlightthickness=3, relief=tk.RAISED, bg=conf["submenu_bg"], highlightbackground=conf["submenu_sel_bg"], height=conf["admin_company_button_height"])
Admin_Ven_Buttons_Frame.pack(fill=tk.X, side=tk.BOTTOM)
Admin_Ven_Buttons_Frame.pack_propagate(0)

#Add button and entry For Vendor
Admin_Ven_Add_Button = tk.Button(Admin_Ven_Buttons_Frame, text="ADD", command=adm_ven_add_button, width=20, fg=conf["submenu_fg"])
Admin_Ven_Add_Button.pack(side=tk.LEFT, fill=tk.Y,  padx=2, pady=2)
Admin_Ven_DELETE_Button = tk.Button(Admin_Ven_Buttons_Frame, text="DELETE", command=adm_ven_delete_button, width=20, fg=conf["submenu_fg"])
Admin_Ven_DELETE_Button.pack(side=tk.LEFT, fill=tk.Y,  padx=2, pady=2)
Admin_Ven_ref_Button = tk.Button(Admin_Ven_Buttons_Frame, text="REFRESH", command=adm_ven_refresh, width=20, fg=conf["submenu_fg"])
Admin_Ven_ref_Button.pack(side=tk.LEFT, fill=tk.Y,  padx=2, pady=2)
Admin_Ven_entry_lb = tk.Label(Admin_Ven_Buttons_Frame, text="Vendor:", bg=conf["submenu_bg"], font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["submenu_fg"], width=10)
Admin_Ven_entry_lb.pack(side=tk.LEFT)
Admin_Ven_Add_entry = tk.Entry(Admin_Ven_Buttons_Frame, bg=conf["window_bg"], bd=0, font=(conf["entry_font"], conf["notebook_tab_size"]), fg=conf["widget_fg"], justify=tk.LEFT, width=30)
Admin_Ven_Add_entry.pack(side=tk.LEFT)
Admin_Ven_Add_entry.bind("<Return>", lambda event: adm_ven_add_button())





root.mainloop()
